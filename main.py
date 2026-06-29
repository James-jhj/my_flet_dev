import flet as ft
import flet.canvas as cv
import flet_audio as ftaudio
from flet_audio import AudioState
import asyncio
import threading
import time
import json
import os
import platform
import requests
import re
import mutagen
import html
import datetime
import math
import asyncio
import datetime as dt
import calendar
from datetime import datetime, timedelta
from pathlib import Path
from lunardate import LunarDate
from urllib.parse import quote
from typing import Optional
from zhdate import ZhDate
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import timezone, timedelta
from chinese_calendar import is_workday as cn_is_workday

import hashlib
import subprocess
import uuid
import sys

# ========== 2. 版本信息 ==========
APP_VERSION = "1.0.136"
APP_VERSION_CODE = 136
# =============================

# ========== 3. 设备绑定功能 ==========

def get_device_id():
    """获取设备唯一标识"""
    if platform.system() == "Windows":
        # Windows：获取硬盘序列号
        try:
            result = subprocess.run(['wmic', 'diskdrive', 'get', 'serialnumber'], 
                                    capture_output=True, text=True, timeout=5)
            lines = result.stdout.strip().split('\n')
            if len(lines) > 1 and lines[-1].strip():
                return hashlib.md5(lines[-1].strip().encode()).hexdigest()
        except:
            pass
        # 降级：使用计算机名 + 用户名
        return hashlib.md5(f"{os.getlogin()}_{platform.node()}".encode()).hexdigest()
    
    else:
        # Android：首次运行生成ID并存储，之后读取
        try:
            # 获取应用私有存储目录，这是 Android 上 App 的专属地盘
            app_data_dir = os.getenv("FLET_APP_STORAGE_DATA", "")
            if not app_data_dir:
                # 兼容本地运行
                app_data_dir = "."

            device_id_file = os.path.join(app_data_dir, "device_id.json")

            # 1. 尝试读取已保存的ID
            if os.path.exists(device_id_file):
                try:
                    with open(device_id_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        return data.get('device_id')
                except:
                    pass

            # 2. 首次运行，生成新ID并保存
            new_device_id = hashlib.md5(str(uuid.uuid4()).encode()).hexdigest()
            
            # 3. 保存这个ID到文件
            try:
                os.makedirs(app_data_dir, exist_ok=True)
                with open(device_id_file, 'w', encoding='utf-8') as f:
                    json.dump({'device_id': new_device_id}, f)
            except Exception as e:
                print(f"保存设备ID失败: {e}")

            return new_device_id
        except:
            pass


def get_auth_file_path():
    """获取授权文件路径（Android 兼容）"""
    app_data_dir = os.getenv("FLET_APP_STORAGE_DATA", "")
    if app_data_dir:
        os.makedirs(app_data_dir, exist_ok=True)
        return os.path.join(app_data_dir, "device_auth.json")
    return "device_auth.json"

def is_device_authorized():
    """检查设备是否已授权（首次运行自动授权）"""
    current_device_id = get_device_id()
    auth_file = get_auth_file_path()
    
    # 检查授权文件是否存在
    if os.path.exists(auth_file):
        try:
            with open(auth_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('device_id') == current_device_id
        except:
            pass
    
    # 首次运行，自动授权当前设备
    try:
        with open(auth_file, 'w', encoding='utf-8') as f:
            json.dump({'device_id': current_device_id}, f)
        return True
    except:
        return False

def show_unauthorized_page(page, device_id=None):
    """显示未授权页面"""
    if device_id is None:
        device_id = get_device_id()
    
    page.clean()
    page.title = "设备未授权"
    page.window_width = 400
    page.window_height = 400
    page.bgcolor = ft.Colors.WHITE
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER
    page.vertical_alignment = ft.MainAxisAlignment.CENTER
    
    async def copy_to_clipboard(page, text):
        """异步复制文本到剪贴板"""
        try:
            clipboard = ft.Clipboard()
            await clipboard.set(text)
            # 显示成功提示
            snack = ft.SnackBar(content=ft.Text("✅ 设备ID已复制，请发送给管理员"), duration=3000)
            page.overlay.append(snack)
            snack.open = True
            page.update()
            return True
        except Exception as e:
            print(f"复制失败: {e}")
            return False

    def copy_device_id(e):
        """复制设备ID"""
        # 创建异步任务
        asyncio.create_task(copy_to_clipboard(page, device_id))

    page.add(
        ft.Column([
            ft.Icon(ft.Icons.WARNING, size=80, color=ft.Colors.RED_700),
            ft.Text("设备未授权", size=24, weight=ft.FontWeight.BOLD, color=ft.Colors.RED_700),
            ft.Text("当前设备未获得使用授权", size=14, color=ft.Colors.GREY_600),
            ft.Container(height=10),
            ft.Text("请将以下设备ID发送给管理员:", size=12, color=ft.Colors.GREY_600),
            ft.Container(
                content=ft.Text(device_id, size=11, selectable=True),
                padding=8,
                bgcolor=ft.Colors.GREY_100,
                border_radius=5,
                width=320,
            ),
            ft.Button("复制设备ID", on_click=copy_device_id),
            ft.Container(height=10),
            ft.Text("管理员授权后即可使用", size=12, color=ft.Colors.GREY_500),
        ], spacing=15, horizontal_alignment=ft.CrossAxisAlignment.CENTER)
    )
    page.update()
# ==================   添加设备授权功能  ============================

# ========== 平台检测（放在这里） ==========
IS_WINDOWS = platform.system() == "Windows"

if IS_WINDOWS:
    ANDROID_NOTIFY_AVAILABLE = False
    print("✅ Windows 平台，使用假通知模块")
    
    class Notification:
        def __init__(self, title="", message="", notification_id=0, ongoing=False):
            self.title = title
            self.message = message
        
        def send(self):
            print(f"[通知] {self.title}: {self.message}")
            return False
        
        def cancel(self):
            pass
else:
    try:
        from android_notify import Notification
        ANDROID_NOTIFY_AVAILABLE = True
        print("✅ android_notify 导入成功")
    except ImportError as e:
        ANDROID_NOTIFY_AVAILABLE = False
        print(f"❌ android_notify 导入失败: {e}")
        
        class Notification:
            def __init__(self, title="", message="", notification_id=0, ongoing=False):
                self.title = title
                self.message = message
            
            def send(self):
                return False
            
            def cancel(self):
                pass

# 根据平台决定是否启用网易云模块
if not IS_WINDOWS:
    PYCNM_AVAILABLE = False
    PLAYWRIGHT_AVAILABLE = False
    print("Android平台，网易云音乐和Playwright模块已禁用")
else:
    # 尝试导入 playwright，失败时设置标志
    try:
        from playwright.sync_api import sync_playwright
        PLAYWRIGHT_AVAILABLE = True
    except ImportError:
        PLAYWRIGHT_AVAILABLE = False
        print("警告: playwright 模块不可用，音乐下载功能将使用降级方案")
    
    # 尝试导入 pyncm 和 plyer
    try:
        from plyer import notification
        from pyncm import apis
        from pyncm.apis.login import LoginViaAnonymousAccount
        PYCNM_AVAILABLE = True
        print("pyncm 模块可用")
    except ImportError:
        PYCNM_AVAILABLE = False
        print("警告: pyncm 模块不可用")


class Transaction:
    """记账记录"""
    def __init__(self, id: str, date: str, type: str, category: str, amount: float, note: str = ""):
        self.id = id
        self.date = date  # 格式: YYYY-MM-DD
        self.type = type  # "income" 或 "expense"
        self.category = category  # 分类
        self.amount = amount
        self.note = note
    
    def to_dict(self):
        return {
            "id": self.id,
            "date": self.date,
            "type": self.type,
            "category": self.category,
            "amount": self.amount,
            "note": self.note,
        }
    
    @classmethod
    def from_dict(cls, data):
        return cls(
            data["id"],
            data["date"],
            data["type"],
            data["category"],
            data["amount"],
            data.get("note", ""),
        )

class SearchableDropdown(ft.Column):
    """可搜索的下拉选择框（使用 Overlay 实现悬浮）"""
    def __init__(self, page, label, options, value=None, on_change=None, **kwargs):
        super().__init__(**kwargs)
        self._page = page
        self.options = options
        self.on_change_callback = on_change
        self._overlay_container = None
        self._is_open = False  # 添加状态标记
        
        # 文本输入框
        self.text_field = ft.TextField(
            label=label,
            value=value,
            height=45,
            expand=True,
            read_only=True,   # 添加只读属性
            on_click=self._on_text_click,  # 点击时阻止焦点
            on_blur=self.on_blur,  # 添加失去焦点事件
            #on_change=self.on_text_change,
            #on_focus=self.on_focus,
            suffix=ft.IconButton(ft.Icons.ARROW_DROP_DOWN, on_click=self.toggle_dropdown),
            **kwargs
        )
        
        from flet import Border, BorderSide
        border = Border(
            left=BorderSide(1, ft.Colors.GREY_300),
            top=BorderSide(1, ft.Colors.GREY_300),
            right=BorderSide(1, ft.Colors.GREY_300),
            bottom=BorderSide(1, ft.Colors.GREY_300),
        )
        
        # 下拉列表容器
        self.dropdown_container = ft.Container(
            content=ft.Column([], spacing=2, scroll=ft.ScrollMode.AUTO),
            #width=300,
            expand=True,
            height=50,
            bgcolor=ft.Colors.WHITE,
            border=border,
            border_radius=4,
            shadow=ft.BoxShadow(blur_radius=10, color=ft.Colors.BLACK12),
        )
        
        self.controls = [self.text_field]
    
    def _on_text_click(self, e):
        """点击文本框时，打开下拉框"""
        # 直接打开下拉框
        self.toggle_dropdown(e)
    
    def on_text_change(self, e):
        """文本变化时过滤选项"""
        search_text = self.text_field.value.lower()
        print(f"[搜索] 输入: {search_text}")
        print(f"[搜索] 所有选项: {self.options}")
        filtered = [opt for opt in self.options if search_text in opt.lower()]
        print(f"[搜索] 过滤结果: {filtered}")
        
        # ========== 更新内容，但不重新创建 Overlay ==========
        self.update_dropdown_content(filtered)
        
        # 如果下拉框已经打开，更新显示
        if self._is_open and self._overlay_container and self._overlay_container in self._page.overlay:
            self._page.update()
        # 如果下拉框未打开且有关键字，自动打开
        elif search_text and len(search_text) > 0:
            self.show_dropdown()
        
        if self.on_change_callback:
            value = self.text_field.value
            if value and value.strip():
                self.on_change_callback(value)
            else:
                self.on_change_callback(None)
    
    def on_focus(self, e):
        """获得焦点时显示下拉列表"""
        self.show_dropdown()

    def on_blur(self, e):
        pass
    
    def toggle_dropdown(self, e):
        """切换下拉列表显示"""
        if self._is_open and self._overlay_container and self._overlay_container in self._page.overlay:
            self.hide_dropdown()
        else:
            self.show_dropdown()
    
    def show_dropdown(self):
        """显示下拉列表（使用 Overlay 悬浮）"""
        self.update_dropdown_content(self.options)
        
        if self._is_open and self._overlay_container and self._overlay_container in self._page.overlay:
            return
        
        self._is_open = True
        
        # 使用 Column + Row 让下拉框出现在文本框下方
        self._overlay_container = ft.Container(
            content=ft.Column([
                ft.Container(expand=True, on_click=lambda e: self.hide_dropdown()),
                ft.Row([
                    ft.Container(width=10),
                    ft.Container(
                        content=self.dropdown_container,
                        expand=True,  # 宽度填满剩余空间
                    ),
                    ft.Container(width=10),
                ]),
                ft.Container(height=106, on_click=lambda e: self.hide_dropdown()),   # 如果太高就调小一点
            ]),
            expand=True,
            bgcolor=ft.Colors.TRANSPARENT,
        )
        self._page.overlay.append(self._overlay_container)
        self.dropdown_container.visible = True
        self._page.update()
    
    def hide_dropdown(self):
        """隐藏下拉列表"""
        self._is_open = False
        if self._overlay_container and self._overlay_container in self._page.overlay:
            self._page.overlay.remove(self._overlay_container)
            self._overlay_container = None
            self._page.update()
    
    def update_dropdown_content(self, options):
        """更新下拉列表内容"""
        self.dropdown_container.content.controls.clear()
        
        if not options:
            self.dropdown_container.height = 50
            return
        
        for i, opt in enumerate(options):
            btn = ft.Container(
                content=ft.Row([
                    ft.Text(opt, size=14, color=ft.Colors.BLACK),
                ], alignment=ft.MainAxisAlignment.START),
                on_click=lambda e, val=opt: self.select_option(val),
                ink=True,
                expand=True,
                height=40,
            )
            self.dropdown_container.content.controls.append(btn)
            
            if i < len(options) - 1:
                divider = ft.Divider(height=1, color=ft.Colors.GREY_200)
                self.dropdown_container.content.controls.append(divider)
        
        # ========== 高度 = 子项高度 * 子项个数 + 分割线高度 ==========
        total_items = len(options)
        item_height = 40
        divider_height = 1
        
        # 总高度 = 选项数 * 选项高度 + (选项数-1) * 分割线高度
        total_height = total_items * item_height + (total_items - 1) * divider_height
        
        # 加上上下内边距（如果有）
        total_height += 10

        print(f"[高度计算] 选项数: {total_items}, 高度: {total_height}")
        
        # 限制最大高度（防止超出屏幕）
        max_height = 150
        if total_height > max_height:
            total_height = max_height - 15  # 电脑月份刚刚好高度
        else:
            total_height -=20               # 电脑年份刚刚好高度
        
        # 确保最小高度
        min_height = 50
        if total_height < min_height:
            total_height = min_height
        
        self.dropdown_container.height = total_height
        print(f"[高度计算] 选项数: {total_items}, 高度: {total_height}")
    
    def select_option(self, value):
        self.text_field.value = value
        self.hide_dropdown()
        if self.on_change_callback:
            if value and value.strip():
                self.on_change_callback(value)
            else:
                self.on_change_callback(None)
        self._page.update()
    
    @property
    def value(self):
        return self.text_field.value
    
    @value.setter
    def value(self, val):
        self.text_field.value = val
        self.text_field.update()

class SearchableDropdownFl(ft.Column):
    """可搜索的下拉选择框（使用 Overlay 实现悬浮）"""
    def __init__(self, page, label, options, value=None, on_change=None, **kwargs):
        super().__init__(**kwargs)
        self._page = page
        self.options = options
        self._filtered_options = options  # 初始化
        self.on_change_callback = on_change
        self._overlay_container = None
        self._is_open = False
        self._bottom_offset = 120

        # ========== 手动记录焦点状态 ==========
        self._has_focus = False
        
        # 文本输入框
        self.text_field = ft.TextField(
            label=label,
            value=value,
            height=56,
            expand=True,
            on_change=self.on_text_change,
            on_focus=self._on_focus,      # 获得焦点时记录
            on_blur=self._on_blur,        # 失去焦点时记录
            on_click=self.toggle_dropdown,  # 添加这行
            suffix=ft.IconButton(ft.Icons.ARROW_DROP_DOWN, on_click=self.toggle_dropdown),
            **kwargs
        )
        
        from flet import Border, BorderSide
        border = Border(
            left=BorderSide(1, ft.Colors.GREY_300),
            top=BorderSide(1, ft.Colors.GREY_300),
            right=BorderSide(1, ft.Colors.GREY_300),
            bottom=BorderSide(1, ft.Colors.GREY_300),
        )
        
        # 下拉列表容器
        self.dropdown_container = ft.Container(
            content=ft.Column([], spacing=2, scroll=ft.ScrollMode.AUTO),
            #width=300,
            expand=True,
            height=50,
            bgcolor=ft.Colors.WHITE,
            border=border,
            border_radius=4,
            shadow=ft.BoxShadow(blur_radius=10, color=ft.Colors.BLACK12),
        )
        
        self.controls = [self.text_field]

    def _on_focus(self, e):
        """获得焦点时记录状态"""
        self._has_focus = True
        print(f"[焦点状态] 获得焦点: {self._has_focus}")
        # 调用原有的 on_focus 逻辑
        self.on_focus(e)
    
    def _on_blur(self, e):
        #失去焦点时记录状态
        self._has_focus = False
        print(f"[焦点状态] 失去焦点: {self._has_focus}")
    
    def on_focus(self, e):
        """获得焦点时，设置底部偏移为100（键盘弹出）"""

        # 获取当前下拉框高度
        dropdown_height = self.dropdown_container.height
        
        # 多个选项（高度>100，=135），底部偏移120
        self._bottom_offset = 120
        
        # 如果下拉框已打开，刷新显示
        if self._is_open and self._overlay_container and self._overlay_container in self._page.overlay:
            self._update_overlay_height()
        else:
            # 未打开则自动打开
            self.show_dropdown()
    
    def on_text_change(self, e):
        """文本变化时过滤选项"""
        search_text = self.text_field.value.lower()
        
        # ========== 如果文本框为空，重置并显示完整列表 ==========
        if not search_text or len(search_text) == 0:
            self._filtered_options = None
            if self._has_focus:
                self._bottom_offset = 120
            else:
                self._bottom_offset = 120
            
            # 更新显示完整列表
            self.update_dropdown_content(self.options)
            
            # ========== 强制重新显示下拉框 ==========
            # 先移除旧的 Overlay
            if self._overlay_container and self._overlay_container in self._page.overlay:
                self._page.overlay.remove(self._overlay_container)
                self._overlay_container = None
            
            # 重新打开下拉框
            self._is_open = False  # 重置状态，让 show_dropdown 重新创建
            self.show_dropdown()
            
            if self.on_change_callback:
                self.on_change_callback("")
            return
        
        # ========== 有搜索内容，进行过滤 ==========
        filtered = [opt for opt in self.options if search_text in opt.lower()]

        # 获取当前下拉框高度
        dropdown_height = self.dropdown_container.height

        # 更新下拉框内容
        self.update_dropdown_content(filtered)
        
        # 保存过滤结果
        self._filtered_options = filtered

        # ========== 根据筛选结果处理 ==========
        if len(filtered) == 0:
            # 没有匹配结果，隐藏下拉框
            self.hide_dropdown()
            if self.on_change_callback:
                value = self.text_field.value
                if value and value.strip():
                    self.on_change_callback(value)
                else:
                    self.on_change_callback(None)
            return
        
        # ========== 单个选项时，底部偏移205 ==========
        if len(filtered) == 1:
            self._bottom_offset = 205   # 单个选项
        else:
            self._bottom_offset = 120   # 多个选项
        
        # ========== 强制重新创建 Overlay ==========
        if self._is_open:
            if self._overlay_container and self._overlay_container in self._page.overlay:
                self._page.overlay.remove(self._overlay_container)
                self._overlay_container = None
            if len(filtered) > 0:
                self.show_dropdown()
            else:
                self.hide_dropdown()
        elif search_text and len(search_text) > 0 and len(filtered) > 0:
            self.show_dropdown()
        elif len(filtered) == 0:
            self.hide_dropdown()
        
        if self.on_change_callback:
            value = self.text_field.value
            if value and value.strip():
                self.on_change_callback(value)
            else:
                self.on_change_callback(None)
    
    def toggle_dropdown(self, e):
        """切换下拉列表显示（点击箭头时触发）"""

        # ========== 阻止事件冒泡，避免触发 on_blur ==========
        if hasattr(e, 'stop_propagation'):
            e.stop_propagation()

        # ========== 点击箭头时，让文本框获得焦点 ==========
        self.text_field.focus()
        self._has_focus = True

        print(f"点击下拉框图标按钮获取焦点: {self._has_focus}")

        # ========== 根据焦点状态决定偏移 ==========
        if self._has_focus:
            # 有焦点，键盘弹出
            dropdown_height = self.dropdown_container.height
            is_android = platform.system() == "Linux"
            if is_android:
                # 手机端 + 文本框获得焦点（键盘弹出）
                if dropdown_height == 50:
                    self._bottom_offset = 205 # 只有1个子项时的偏移量
                elif dropdown_height == 135:
                    self._bottom_offset = 120 # 有多个子项时的偏移量
            else:
                # 键盘未弹出或电脑端
                self._bottom_offset = 120     # 正常情况下的偏移量
        else:
            # 无焦点，使用默认偏移
            self._bottom_offset = 404     # 正常情况下的偏移量

        if self._is_open and self._overlay_container and self._overlay_container in self._page.overlay:
            self.hide_dropdown()
            # 隐藏后重新获取焦点，让键盘保持打开
            self.text_field.focus()
        else:
            self._filtered_options = None
            self.text_field.value = ""
            self.text_field.update()
            self.show_dropdown()
    
    def show_dropdown(self):
        """显示下拉列表（使用 Overlay 悬浮）"""
        print(f"[创建Overlay] 底部偏移最新: {self._bottom_offset}")
        
        # 先移除旧的 Overlay
        if self._overlay_container and self._overlay_container in self._page.overlay:
            self._page.overlay.remove(self._overlay_container)
            self._overlay_container = None
        
        # ========== 判断使用过滤结果还是完整列表 ==========
        if hasattr(self, '_filtered_options') and self._filtered_options:
            # 有过滤结果，使用过滤后的
            self.update_dropdown_content(self._filtered_options)
        else:
            # 没有过滤结果，使用完整列表
            self.update_dropdown_content(self.options)
        
        if self._is_open and self._overlay_container and self._overlay_container in self._page.overlay:
            return
        
        self._is_open = True

        # ========== 使用最新的 _bottom_offset ==========
        print(f"[创建Overlay] 底部偏移最新: {self._bottom_offset}")
        
        # 创建 Overlay 容器
        self._overlay_container = ft.Container(
            content=ft.Column([
                ft.Container(expand=True, on_click=lambda e: self.hide_dropdown()),
                ft.Row([
                    ft.Container(width=30),  # 左边距
                    ft.Container(
                        content=self.dropdown_container,
                        expand=True,  # 宽度填满剩余空间
                    ),
                    ft.Container(width=30),  # 右边距
                ]),
                ft.Container(height=self._bottom_offset, on_click=lambda e: self.hide_dropdown()),
            ]),
            expand=True,
            bgcolor=ft.Colors.TRANSPARENT,
        )
        
        # 添加到 Overlay
        self._page.overlay.append(self._overlay_container)
        self.dropdown_container.visible = True
        
        # 更新页面
        self._page.update()
    
    def _update_overlay_height(self):
        """更新 Overlay 中的底部高度"""
        if not self._overlay_container or self._overlay_container not in self._page.overlay:
            return
        
        # 重新创建 Overlay 容器
        self._overlay_container = ft.Container(
            content=ft.Column([
                ft.Container(expand=True, on_click=lambda e: self.hide_dropdown()),
                ft.Row([
                    ft.Container(expand=True),
                    self.dropdown_container,
                    ft.Container(expand=True),
                ]),
                ft.Container(height=self._bottom_offset, on_click=lambda e: self.hide_dropdown()),
            ]),
            expand=True,
            bgcolor=ft.Colors.TRANSPARENT,
        )
        self._page.overlay.append(self._overlay_container)
        self._page.update()
    
    def hide_dropdown(self):
        """隐藏下拉列表"""
        self._is_open = False
        if self._overlay_container and self._overlay_container in self._page.overlay:
            self._page.overlay.remove(self._overlay_container)
            self._overlay_container = None
            self._page.update()
    
    def update_dropdown_content(self, options):
        """更新下拉列表内容"""
        self.dropdown_container.content.controls.clear()
        
        if not options:
            self.dropdown_container.height = 50
            return
        
        for i, opt in enumerate(options):
            btn = ft.Container(
                content=ft.Row([
                    ft.Text(opt, size=14, color=ft.Colors.BLACK),
                ], alignment=ft.MainAxisAlignment.START),
                on_click=lambda e, val=opt: self.select_option(val),
                ink=True,
                expand=True,
                height=40,
            )
            self.dropdown_container.content.controls.append(btn)
            
            if i < len(options) - 1:
                divider = ft.Divider(height=1, color=ft.Colors.GREY_200)
                self.dropdown_container.content.controls.append(divider)
        
        # ========== 高度 = 子项高度 * 子项个数 + 分割线高度 ==========
        total_items = len(options)
        item_height = 40
        divider_height = 1
        
        # 总高度 = 选项数 * 选项高度 + (选项数-1) * 分割线高度
        total_height = total_items * item_height + (total_items - 1) * divider_height
        
        # 加上上下内边距（如果有）
        total_height += 10

        print(f"[高度计算] 选项数: {total_items}, 高度: {total_height}")
        
        # 限制最大高度（防止超出屏幕）
        max_height = 150
        if total_height > max_height:
            total_height = max_height - 15  # 电脑展开下拉框刚刚好高度
        
        # 确保最小高度
        min_height = 50
        if total_height < min_height:
            total_height = min_height
        
        self.dropdown_container.height = total_height
        print(f"[高度计算] 选项数: {total_items}, 高度: {total_height}")
    
    def select_option(self, value):
        self.text_field.value = value
        self.hide_dropdown()
        if self.on_change_callback:
            if value and value.strip():
                self.on_change_callback(value)
            else:
                self.on_change_callback(None)
        self._page.update()
    
    @property
    def value(self):
        return self.text_field.value
    
    @value.setter
    def value(self, val):
        self.text_field.value = val
        self.text_field.update()

class SearchableDropdownEvtTp(ft.Column):
    """可搜索的下拉选择框（使用 Overlay 实现悬浮，位置自动适配）"""
    def __init__(self, page, label, options, value=None, on_change=None, **kwargs):
        super().__init__(**kwargs)
        self._page = page
        self.options = options
        self.on_change_callback = on_change
        self._overlay_container = None
        
        # 文本输入框
        self.text_field = ft.TextField(
            label=label,
            value=value,
            height=56,
            expand=True,
            read_only=True,  # 添加只读属性
            on_click=self._on_text_click,  # 点击时阻止焦点
            on_blur=self.on_blur,  # 添加失去焦点事件
            #on_change=self.on_text_change,
            #on_focus=self.on_focus,
            suffix=ft.IconButton(ft.Icons.ARROW_DROP_DOWN, on_click=self.toggle_dropdown),
            **kwargs
        )
        
        from flet import Border, BorderSide
        border = Border(
            left=BorderSide(1, ft.Colors.GREY_300),
            top=BorderSide(1, ft.Colors.GREY_300),
            right=BorderSide(1, ft.Colors.GREY_300),
            bottom=BorderSide(1, ft.Colors.GREY_300),
        )
        
        # 下拉列表容器
        self.dropdown_container = ft.Container(
            content=ft.Column([], spacing=2, scroll=ft.ScrollMode.AUTO),
            #width=300,
            expand=True,
            height=50,
            bgcolor=ft.Colors.WHITE,
            border=border,
            border_radius=4,
            shadow=ft.BoxShadow(blur_radius=10, color=ft.Colors.BLACK12),
        )
        
        self.controls = [self.text_field]

    def _on_text_click(self, e):
        """点击文本框时，打开下拉框"""
        # 直接打开下拉框
        self.toggle_dropdown(e)

    def on_text_change(self, e):
        search_text = self.text_field.value.lower()
        filtered = [opt for opt in self.options if search_text in opt.lower()]
        self.update_dropdown_content(filtered)
        if self._overlay_container and self._overlay_container in self._page.overlay:
            self._page.update()
        
        if self.on_change_callback:
            value = self.text_field.value
            if value and value.strip():
                self.on_change_callback(value)
            else:
                self.on_change_callback(None)
    
    def on_focus(self, e):
        self.show_dropdown()

    def on_blur(self, e):
        pass
    
    def toggle_dropdown(self, e):
        if self._overlay_container and self._overlay_container in self._page.overlay:
            self.hide_dropdown()
        else:
            self.show_dropdown()
    
    def show_dropdown(self):
        """显示下拉列表（使用 Overlay 悬浮）"""
        self.update_dropdown_content(self.options)
        
        if self._overlay_container and self._overlay_container in self._page.overlay:
            return
        
        # ========== 使用 Column + Row 让下拉框出现在文本框下方 ==========
        # 使用弹性布局，让下拉框在文本框正下方
        self._overlay_container = ft.Container(
            content=ft.Column([
                # 上方空白（点击关闭）
                ft.Container(expand=True, on_click=lambda e: self.hide_dropdown()),
                # 下拉框（在 Row 中居中）
                ft.Row([
                    ft.Container(width=30),  # 左边距
                    ft.Container(
                        content=self.dropdown_container,
                        expand=True,  # 宽度填满剩余空间
                    ),
                    ft.Container(width=30),  # 右边距
                ]),
                # 下方空白
                ft.Container(height=341, on_click=lambda e: self.hide_dropdown()),
            ]),
            expand=True,
            bgcolor=ft.Colors.TRANSPARENT,
        )
        self._page.overlay.append(self._overlay_container)
        self.dropdown_container.visible = True
        self._page.update()
    
    def hide_dropdown(self):
        if self._overlay_container and self._overlay_container in self._page.overlay:
            self._page.overlay.remove(self._overlay_container)
            self._overlay_container = None
            self._page.update()
    
    def update_dropdown_content(self, options):
        """更新下拉列表内容"""
        self.dropdown_container.content.controls.clear()
        
        if not options:
            self.dropdown_container.height = 50
            return
        
        for i, opt in enumerate(options):
            btn = ft.Container(
                content=ft.Row([
                    ft.Text(opt, size=14, color=ft.Colors.BLACK),
                ], alignment=ft.MainAxisAlignment.START),
                on_click=lambda e, val=opt: self.select_option(val),
                ink=True,
                expand=True,
                height=40,
            )
            self.dropdown_container.content.controls.append(btn)
            
            if i < len(options) - 1:
                divider = ft.Divider(height=1, color=ft.Colors.GREY_200)
                self.dropdown_container.content.controls.append(divider)
        
        # ========== 高度 = 子项高度 * 子项个数 + 分割线高度 ==========
        total_items = len(options)
        item_height = 40
        divider_height = 1
        
        # 总高度 = 选项数 * 选项高度 + (选项数-1) * 分割线高度
        total_height = total_items * item_height + (total_items - 1) * divider_height
        
        # 加上上下内边距（如果有）
        total_height += 10

        print(f"[高度计算] 选项数: {total_items}, 高度: {total_height}")
        
        # 限制最大高度（防止超出屏幕）
        max_height = 300
        if total_height > max_height:
            total_height = max_height
        else:
            total_height +=14  # 电脑刚刚好高度，手机调高一点+2
        
        # 确保最小高度
        min_height = 50
        if total_height < min_height:
            total_height = min_height
        
        self.dropdown_container.height = total_height
        print(f"[高度计算] 选项数: {total_items}, 高度: {total_height}")
    
    def select_option(self, value):
        self.text_field.value = value
        self.hide_dropdown()
        if self.on_change_callback:
            if value and value.strip():
                self.on_change_callback(value)
            else:
                self.on_change_callback(None)
        self._page.update()
    
    @property
    def value(self):
        return self.text_field.value
    
    @value.setter
    def value(self, val):
        self.text_field.value = val
        self.text_field.update()

class SearchableDropdownEvtLf(ft.Column):
    """可搜索的下拉选择框（使用 Overlay 实现悬浮，位置自动适配）"""
    def __init__(self, page, label, options, value=None, on_change=None, **kwargs):
        super().__init__(**kwargs)
        self._page = page
        self.options = options
        self.on_change_callback = on_change
        self._overlay_container = None
        
        # 文本输入框
        self.text_field = ft.TextField(
            label=label,
            value=value,
            height=56,
            expand=True,
            read_only=True,  # 添加只读属性
            on_click=self._on_text_click,  # 点击时阻止焦点
            on_blur=self.on_blur,  # 添加失去焦点事件
            #on_change=self.on_text_change,
            #on_focus=self.on_focus,
            suffix=ft.IconButton(ft.Icons.ARROW_DROP_DOWN, on_click=self.toggle_dropdown),
            **kwargs
        )
        
        from flet import Border, BorderSide
        border = Border(
            left=BorderSide(1, ft.Colors.GREY_300),
            top=BorderSide(1, ft.Colors.GREY_300),
            right=BorderSide(1, ft.Colors.GREY_300),
            bottom=BorderSide(1, ft.Colors.GREY_300),
        )
        
        # 下拉列表容器
        self.dropdown_container = ft.Container(
            content=ft.Column([], spacing=2, scroll=ft.ScrollMode.AUTO),
            expand=True,
            #width=300,
            height=50,
            bgcolor=ft.Colors.WHITE,
            border=border,
            border_radius=4,
            shadow=ft.BoxShadow(blur_radius=10, color=ft.Colors.BLACK12),
        )
        
        self.controls = [self.text_field]

    def _on_text_click(self, e):
        """点击文本框时，打开下拉框"""
        # 直接打开下拉框
        self.toggle_dropdown(e)
    
    def on_text_change(self, e):
        search_text = self.text_field.value.lower()
        filtered = [opt for opt in self.options if search_text in opt.lower()]
        self.update_dropdown_content(filtered)
        if self._overlay_container and self._overlay_container in self._page.overlay:
            self._page.update()
        
        if self.on_change_callback:
            value = self.text_field.value
            if value and value.strip():
                self.on_change_callback(value)
            else:
                self.on_change_callback(None)
    
    def on_focus(self, e):
        self.show_dropdown()

    def on_blur(self, e):
        pass
    
    def toggle_dropdown(self, e):
        if self._overlay_container and self._overlay_container in self._page.overlay:
            self.hide_dropdown()
        else:
            self.show_dropdown()
    
    def show_dropdown(self):
        """显示下拉列表（使用 Overlay 悬浮）"""
        self.update_dropdown_content(self.options)
        
        if self._overlay_container and self._overlay_container in self._page.overlay:
            return
        
        # ========== 使用 Column + Row 让下拉框出现在文本框下方 ==========
        # 使用弹性布局，让下拉框在文本框正下方
        self._overlay_container = ft.Container(
            content=ft.Column([
                # 上方空白（点击关闭）
                ft.Container(expand=True, on_click=lambda e: self.hide_dropdown()),
                # 下拉框（在 Row 中居中）
                ft.Row([
                    ft.Container(width=30),  # 左边距
                    ft.Container(
                        content=self.dropdown_container,
                        expand=True,  # 宽度填满剩余空间
                    ),
                    ft.Container(width=30),  # 右边距
                ]),
                # 下方空白
                ft.Container(height=309, on_click=lambda e: self.hide_dropdown()),
            ]),
            expand=True,
            bgcolor=ft.Colors.TRANSPARENT,
        )
        self._page.overlay.append(self._overlay_container)
        self.dropdown_container.visible = True
        self._page.update()
    
    def hide_dropdown(self):
        if self._overlay_container and self._overlay_container in self._page.overlay:
            self._page.overlay.remove(self._overlay_container)
            self._overlay_container = None
            self._page.update()
    
    def update_dropdown_content(self, options):
        """更新下拉列表内容"""
        self.dropdown_container.content.controls.clear()
        
        if not options:
            return
        
        for i, opt in enumerate(options):
            # ========== 使用 Container 包裹，expand=True 让整行可点击 ==========
            btn = ft.Container(
                content=ft.Row([
                    ft.Text(opt, size=14, color=ft.Colors.BLACK),
                ], alignment=ft.MainAxisAlignment.START),
                #padding=(12, 8, 12, 8),
                on_click=lambda e, val=opt: self.select_option(val),
                ink=True,
                expand=True,  # 整行展开
                height=40,
            )
            self.dropdown_container.content.controls.append(btn)
            
            if i < len(options) - 1:
                divider = ft.Divider(height=1, color=ft.Colors.GREY_200)
                self.dropdown_container.content.controls.append(divider)
        
        # 计算高度
        import platform
        is_android = platform.system() == "Linux"
        item_height = 42 if is_android else 35
        total_items = len(options)
        content_height = total_items * item_height + (total_items - 1) * 1 + 20
        
        min_height = 80 if is_android else 80
        max_height = 320 if is_android else 300
        
        if content_height < min_height:
            self.dropdown_container.height = min_height
        elif content_height > max_height:
            self.dropdown_container.height = max_height
        else:
            self.dropdown_container.height = content_height - 17  # 历法高度调试
    
    def select_option(self, value):
        self.text_field.value = value
        self.hide_dropdown()
        if self.on_change_callback:
            if value and value.strip():
                self.on_change_callback(value)
            else:
                self.on_change_callback(None)
        self._page.update()
    
    @property
    def value(self):
        return self.text_field.value
    
    @value.setter
    def value(self, val):
        self.text_field.value = val
        self.text_field.update()

class SearchableDropdownEvtWeek(ft.Column):
    """可搜索的下拉选择框（使用 Overlay 实现悬浮，位置自动适配）"""
    def __init__(self, page, label, options, value=None, on_change=None, **kwargs):
        super().__init__(**kwargs)
        self._page = page
        self.options = options
        self.on_change_callback = on_change
        self._overlay_container = None
        
        # 文本输入框
        self.text_field = ft.TextField(
            label=label,
            value=value,
            height=56,
            expand=True,
            read_only=True,  # 添加只读属性
            on_click=self._on_text_click,  # 点击时阻止焦点
            on_blur=self.on_blur,  # 添加失去焦点事件
            #on_change=self.on_text_change,
            #on_focus=self.on_focus,
            suffix=ft.IconButton(ft.Icons.ARROW_DROP_DOWN, on_click=self.toggle_dropdown),
            **kwargs
        )
        
        from flet import Border, BorderSide
        border = Border(
            left=BorderSide(1, ft.Colors.GREY_300),
            top=BorderSide(1, ft.Colors.GREY_300),
            right=BorderSide(1, ft.Colors.GREY_300),
            bottom=BorderSide(1, ft.Colors.GREY_300),
        )
        
        # 下拉列表容器
        self.dropdown_container = ft.Container(
            content=ft.Column([], spacing=2, scroll=ft.ScrollMode.AUTO),
            #width=300,
            expand=True,
            height=50,
            bgcolor=ft.Colors.WHITE,
            border=border,
            border_radius=4,
            shadow=ft.BoxShadow(blur_radius=10, color=ft.Colors.BLACK12),
        )
        
        self.controls = [self.text_field]
    
    def _on_text_click(self, e):
        """点击文本框时，打开下拉框"""
        # 直接打开下拉框
        self.toggle_dropdown(e)

    def on_text_change(self, e):
        search_text = self.text_field.value.lower()
        filtered = [opt for opt in self.options if search_text in opt.lower()]
        self.update_dropdown_content(filtered)
        if self._overlay_container and self._overlay_container in self._page.overlay:
            self._page.update()
        
        if self.on_change_callback:
            value = self.text_field.value
            if value and value.strip():
                self.on_change_callback(value)
            else:
                self.on_change_callback(None)
    
    def on_focus(self, e):
        self.show_dropdown()

    def on_blur(self, e):
        pass
    
    def toggle_dropdown(self, e):
        if self._overlay_container and self._overlay_container in self._page.overlay:
            self.hide_dropdown()
        else:
            self.show_dropdown()
    
    def show_dropdown(self):
        """显示下拉列表（使用 Overlay 悬浮）"""
        self.update_dropdown_content(self.options)
        
        if self._overlay_container and self._overlay_container in self._page.overlay:
            return
        
        # ========== 使用 Column + Row 让下拉框出现在文本框下方 ==========
        # 使用弹性布局，让下拉框在文本框正下方
        self._overlay_container = ft.Container(
            content=ft.Column([
                # 上方空白（点击关闭）
                ft.Container(expand=True, on_click=lambda e: self.hide_dropdown()),
                # 下拉框（在 Row 中居中）
                ft.Row([
                    ft.Container(width=30),  # 左边距
                    ft.Container(
                        content=self.dropdown_container,
                        expand=True,  # 宽度填满剩余空间
                    ),
                    ft.Container(width=30),  # 右边距
                ]),
                # 下方空白
                ft.Container(height=137, on_click=lambda e: self.hide_dropdown()),
            ]),
            expand=True,
            bgcolor=ft.Colors.TRANSPARENT,
        )
        self._page.overlay.append(self._overlay_container)
        self.dropdown_container.visible = True
        self._page.update()
    
    def hide_dropdown(self):
        if self._overlay_container and self._overlay_container in self._page.overlay:
            self._page.overlay.remove(self._overlay_container)
            self._overlay_container = None
            self._page.update()
    
    def update_dropdown_content(self, options):
        """更新下拉列表内容"""
        self.dropdown_container.content.controls.clear()
        
        if not options:
            self.dropdown_container.height = 50
            return
        
        for i, opt in enumerate(options):
            btn = ft.Container(
                content=ft.Row([
                    ft.Text(opt, size=14, color=ft.Colors.BLACK),
                ], alignment=ft.MainAxisAlignment.START),
                on_click=lambda e, val=opt: self.select_option(val),
                ink=True,
                expand=True,
                height=40,
            )
            self.dropdown_container.content.controls.append(btn)
            
            if i < len(options) - 1:
                divider = ft.Divider(height=1, color=ft.Colors.GREY_200)
                self.dropdown_container.content.controls.append(divider)
        
        # ========== 高度 = 子项高度 * 子项个数 + 分割线高度 ==========
        total_items = len(options)
        item_height = 40
        divider_height = 1
        
        # 总高度 = 选项数 * 选项高度 + (选项数-1) * 分割线高度
        total_height = total_items * item_height + (total_items - 1) * divider_height
        
        # 加上上下内边距（如果有）
        total_height += 10

        print(f"[高度计算] 选项数: {total_items}, 高度: {total_height}")
        
        # 限制最大高度（防止超出屏幕）
        max_height = 300
        if total_height > max_height:
            total_height = max_height
        else:
            total_height +=20  # 电脑刚刚好高度
        
        # 确保最小高度
        min_height = 50
        if total_height < min_height:
            total_height = min_height
        
        self.dropdown_container.height = total_height
        print(f"[高度计算] 选项数: {total_items}, 高度: {total_height}")
    
    def select_option(self, value):
        self.text_field.value = value
        self.hide_dropdown()
        if self.on_change_callback:
            if value and value.strip():
                self.on_change_callback(value)
            else:
                self.on_change_callback(None)
        self._page.update()
    
    @property
    def value(self):
        return self.text_field.value
    
    @value.setter
    def value(self, val):
        self.text_field.value = val
        self.text_field.update()

class SmoothMarqueeText(ft.Stack):
    """平滑滚动字幕控件 - 修复文本重叠问题"""
    
    def __init__(
        self,
        text: str = "",
        width: int = 240, # 原始宽度是： 300，调小一点是为了适应手机屏幕，调小方向是对的。
        height: int = 60,
        speed: float = 0.8,
        fps: int = 60,
        gap: int = None,  # 改为 None，表示自动计算
        font_size: int = 14,
        font_weight: ft.FontWeight = ft.FontWeight.BOLD,
        color: str = ft.Colors.BLUE_700,
        bgcolor: str = ft.Colors.TRANSPARENT,
        direction: str = "rtl",
        auto_start: bool = False,
        show_message=None,
    ):
        super().__init__()
        self.width = width
        self.height = height
        self.speed = speed
        self.fps = fps
        self.gap = gap  # 如果为 None，则动态计算
        self.font_size = font_size
        self.font_weight = font_weight
        self.color = color
        self.bgcolor = bgcolor
        self.direction = direction
        self.auto_start = auto_start
        
        # 内部变量
        self._texts = []
        self._offset = 0
        self._is_playing = False
        self._task = None
        self._initialized = False
        self._last_update_time = 0
        self._current_text_width = 0  # 存储当前文本宽度
        self.show_message = show_message  # 保存回调函数

        #self._gap_warning_printed = False  # 初始化标志
        self._warning_printed = {'gt500': False, 'gt300': False, 'gt150': False, 'else': False}
        
        # 创建画布
        self.canvas = cv.Canvas(
            width=width,
            height=height,
        )
        
        # 创建容器
        self.container = ft.Container(
            content=self.canvas,
            width=width,
            height=height,
            bgcolor=bgcolor,
            clip_behavior=ft.ClipBehavior.HARD_EDGE,
        )
        
        self.controls = [self.container]
        
        if text:
            self._texts.append(text.strip())

    def get_text_width(self, text: str) -> float:
        """计算文本宽度（更精确的估算）"""
        width = 0
        for char in text:
            if '\u4e00' <= char <= '\u9fff':  # 中文字符
                width += self.font_size
            elif char.isupper():
                width += self.font_size * 0.75  # 大写字母
            elif char.islower():
                width += self.font_size * 0.55  # 小写字母
            elif char.isdigit():
                width += self.font_size * 0.65  # 数字
            elif char in '.-，。！？':
                width += self.font_size * 0.4   # 标点符号
            else:
                width += self.font_size * 0.5
        return width
    
    def _get_effective_gap(self):
        global _warning_printed
        """获取实际使用的间隙值 - 根据文本长度动态调整"""
        if self.gap is not None:
            return self.gap
        
        # 自动模式：间隙等于当前文本宽度
        if self._texts:
            total_width = sum(self.get_text_width(t) for t in self._texts)
            
            # 根据文本长度动态调整间隙
            if total_width > 500:  # 长文本（超过500像素）
                # 长文本：减去80像素，让文本稍微重叠，避免滚动间隙过大，目前长歌曲名就是走的这个，电脑(-80)刚刚好前面的歌曲名刚消失，后面的歌曲名称就出现了，手机待确定
                gap = total_width - 80
                #if not self._warning_printed['gt500'] and self.show_message:
                    #self.show_message(f"歌曲长度大于500测试：{total_width}")
                    #self._warning_printed['gt500'] = True
            elif total_width > 300:  # 中等文本（300-500像素）， 缩短前后2个歌曲名称中间的空格方法-手机调试扩大一倍，现在手机歌曲长度到这里了大于300
                # 中等文本：减去50像素
                gap = total_width - 50
                #if not self._warning_printed['gt300'] and self.show_message:
                    #self.show_message(f"歌曲长度大于300测试：{total_width}")
                    #self._warning_printed['gt300'] = True
            elif total_width > 150:  # 较短文本（150-300像素），目前短歌曲名就是走的这个，电脑(+45)刚刚好前面的歌曲名刚消失，后面的歌曲名称就出现了，手机待确定
                # 较短文本：间隙等于文本宽度
                gap = total_width - 5  # 手机现在设置 - 25 刚刚好，歌曲长度再长一点，就 - 20或- 10试试，慢慢微调，手机歌曲长度约285~300之间
                #if not self._warning_printed['gt150'] and self.show_message:
                    #self.show_message(f"歌曲长度大于150测试：{total_width}")
                    #self._warning_printed['gt150'] = True
            else:  # 很短文本（小于150像素）
                # 很短文本：间隙 = 文本宽度 + 30，让滚动更平滑
                gap = total_width - 30  # 手机现在设置 - 30 刚刚好，歌曲长度再长一点，就 - 20或- 10试试，慢慢微调
                #if not self._warning_printed['else'] and self.show_message:
                    #self.show_message(f"其他歌曲长度测试：{total_width}")
                    #self._warning_printed['else'] = True

            return max(10, gap)  # 确保间隙至少为10像素
            
        return 80  # 默认值
    
    def set_text(self, text: str, append: bool = False):
        """设置或添加文本"""
        if not append:
            self._texts.clear()
            self._offset = 0

        if text and text.strip():
            self._texts.append(text.strip())
            # 更新当前文本宽度
            self._current_text_width = sum(self.get_text_width(t) for t in self._texts)
        
        if self._initialized:
            if self.auto_start and not self._is_playing and self._texts:
                self.start()
            else:
                self._draw_frame()
    
    def clear_texts(self):
        """清除所有文本"""
        self._texts.clear()
        self._offset = 0
        if self._initialized:
            self._draw_frame()
    
    def _draw_frame(self):
        """绘制当前帧 - 修复空文本问题"""
        if not self._initialized:
            return
        
        # 如果没有文本，清空画布并返回
        if not self._texts:
            self.canvas.shapes.clear()
            self.canvas.update()
            return
        
        self.canvas.shapes.clear()
        
        # 计算所有文本的宽度
        text_data = []
        total_width = 0
        for text in self._texts:
            text_width = self.get_text_width(text)
            text_data.append((text, text_width))
            total_width += text_width
        
        # ========== 新增：如果是"🎵 未播放"或包含"已暂停"，居中显示，不滚动 ==========
        if len(self._texts) == 1:
            text = self._texts[0]
            # 判断是否是停止或暂停状态
            if text == "🎵 未播放" or text.startswith("已暂停"):
                # 居中显示
                total_width = self.get_text_width(text)
                x = (self.width - total_width) / 2
                y = (self.height - self.font_size) / 2
                self.canvas.shapes.append(
                    cv.Text(
                        x,
                        y + self.font_size,
                        text,
                        ft.TextStyle(
                            size=self.font_size,
                            weight=self.font_weight,
                            color=self.color,
                        ),
                    )
                )
                try:
                    self.canvas.update()
                except Exception as e:
                    print(f"更新画布失败: {e}")
                return
        
        # 获取间隙值（可能是动态计算的）
        gap = self._get_effective_gap()
        
        # 每个副本的总宽度 = 文本总宽度 + 间隙
        unit_width = total_width + gap
        
        if unit_width <= 0:
            return
        
        # 计算需要显示多少个文本才能填满屏幕 + 2个确保平滑
        num_copies = max(3, int(self.width / unit_width) + 3)
        
        # 构建循环文本列表
        all_texts = []
        for i in range(num_copies):
            for text, w in text_data:
                all_texts.append((text, w))
        
        # 计算绘制位置
        y = (self.height - self.font_size) / 2
        
        # 确保 offset 在有效范围内
        if unit_width > 0:
            self._offset = self._offset % unit_width
            if self._offset < 0:
                self._offset += unit_width
        
        if self.direction == "ltr":
            # LTR 模式：从左向右滚动
            start_x = -self._offset
            x = start_x
            for text, w in all_texts:
                if x + w > 0 and x < self.width:
                    self.canvas.shapes.append(
                        cv.Text(
                            x,
                            y + self.font_size,
                            text,
                            ft.TextStyle(
                                size=self.font_size,
                                weight=self.font_weight,
                                color=self.color,
                            ),
                        )
                    )
                x += w + gap
                
                if x > self.width + unit_width:
                    break
        else:  # RTL - 从右向左滚动
            start_x = self.width - self._offset
            x = start_x
            for text, w in all_texts:
                if x < self.width and x + w > 0:
                    self.canvas.shapes.append(
                        cv.Text(
                            x,
                            y + self.font_size,
                            text,
                            ft.TextStyle(
                                size=self.font_size,
                                weight=self.font_weight,
                                color=self.color,
                            ),
                        )
                    )
                x -= w + gap
                
                if x < -unit_width:
                    break
        
        try:
            self.canvas.update()
        except Exception as e:
            print(f"更新画布失败: {e}")
    
    async def _animation_loop(self):
        """动画循环 - 修复延迟问题"""
        if not self._initialized:
            return
        
        total_text_width = sum(self.get_text_width(t) for t in self._texts)
        gap = self._get_effective_gap()
        unit_width = total_text_width + gap
        
        if unit_width <= 0:
            return
        
        # 计算完成一个完整周期需要的时间（秒）
        # 速度单位：像素/秒
        speed_px_per_sec = self.speed * 60
        cycle_time = unit_width / speed_px_per_sec
        
        last_time = asyncio.get_event_loop().time()
        last_offset = self._offset
        
        while self._is_playing and self._initialized:
            current_time = asyncio.get_event_loop().time()
            delta_time = current_time - last_time
            last_time = current_time
            
            if delta_time > 0.05:
                delta_time = 0.05
            
            # 基于时间计算应该移动的距离
            distance = speed_px_per_sec * delta_time
            self._offset += distance
            
            # 关键：使用浮点数取模，保持连续性
            if unit_width > 0:
                self._offset = self._offset % unit_width
            
            # 检测是否完成了一个完整周期
            # 如果 offset 回绕了（从接近 unit_width 变成接近 0）
            if last_offset > unit_width * 0.8 and self._offset < unit_width * 0.2:
                #print(f"[滚动] 完成一个周期，准备无缝衔接")
                pass
                #if self.show_message:
                    #self.show_message(f"[滚动] 完成一个周期，准备无缝衔接")
            
            last_offset = self._offset
            
            self._draw_frame()
            await asyncio.sleep(1.0 / self.fps)
    
    def start(self):
        """开始滚动"""
        if not self._initialized:
            return
        
        if self._is_playing:
            self.stop()
        if not self._texts:
            return
        
        #self._offset = 0  # 重置偏移量
        self._is_playing = True
        self._task = asyncio.create_task(self._animation_loop())
    
    def stop(self):
        """停止滚动"""
        self._is_playing = False
        if self._task:
            self._task.cancel()
            self._task = None
    
    def update_text(self, text: str):
        """更新文本（替换）"""
        if len(self._texts) == 1 and self._texts[0] == text:
            return
        
        was_playing = self._is_playing
        self.stop()
        
        self._texts.clear()
        if text and text.strip():
            self._texts.append(text.strip())
        
        self._offset = 0
        
        if self._initialized:
            self._draw_frame()
        
        if was_playing and self._texts:
            self.start()
    
    def update_properties(self, color=None, speed=None):
        """更新属性"""
        if color:
            self.color = color
        if speed:
            self.speed = speed
        if self._initialized:
            self._draw_frame()
    
    def will_unmount(self):
        """控件销毁时停止动画"""
        self.stop()

class AnalogClock(ft.Container):
    def __init__(self, main_page, size=160):
        super().__init__()
        self.main_page = main_page
        self.size = size
        self.canvas = cv.Canvas(width=size, height=size)
        self.content = self.canvas
        self.width = size
        self.height = size
        self.bgcolor = ft.Colors.WHITE
        self.border_radius = 10
        
    def update_clock(self):
        import datetime as dt
        now = dt.datetime.now()
        #print(f"时钟更新: {now.strftime('%H:%M:%S')}")  # 调试用
        self.canvas.shapes.clear()
        
        radius = self.size // 2
        cx = radius
        cy = radius
        
        # 外圆
        self.canvas.shapes.append(
            cv.Circle(cx, cy, radius-2,
                     paint=ft.Paint(style=ft.PaintingStyle.STROKE, stroke_width=2))
        )
        
        # 12个数字标记
        for hour_num in range(1, 13):
            angle = math.radians(hour_num * 30 - 90)
            num_radius = radius - 20
            x = cx + num_radius * math.cos(angle)
            y = cy + num_radius * math.sin(angle)
            self.canvas.shapes.append(
                cv.Circle(x, y, 3, paint=ft.Paint(color=ft.Colors.BLUE_800))
            )
        
        # 60个刻度线
        for i in range(60):
            angle = math.radians(i * 6 - 90)
            if i % 5 == 0:
                start_x = cx + (radius-15) * math.cos(angle)
                start_y = cy + (radius-15) * math.sin(angle)
                end_x = cx + (radius-5) * math.cos(angle)
                end_y = cy + (radius-5) * math.sin(angle)
                self.canvas.shapes.append(
                    cv.Line(start_x, start_y, end_x, end_y,
                           paint=ft.Paint(stroke_width=2.5, color=ft.Colors.BLACK))
                )
            else:
                start_x = cx + (radius-10) * math.cos(angle)
                start_y = cy + (radius-10) * math.sin(angle)
                end_x = cx + (radius-5) * math.cos(angle)
                end_y = cy + (radius-5) * math.sin(angle)
                self.canvas.shapes.append(
                    cv.Line(start_x, start_y, end_x, end_y,
                           paint=ft.Paint(stroke_width=1, color=ft.Colors.GREY_500))
                )
        
        # 指针
        hour = now.hour % 12
        minute = now.minute
        second = now.second
        
        hour_angle = math.radians((hour + minute/60) * 30 - 90)
        minute_angle = math.radians(minute * 6 - 90)
        second_angle = math.radians(second * 6 - 90)
        
        # 时针
        hour_len = radius * 0.45
        hour_end_x = cx + hour_len * math.cos(hour_angle)
        hour_end_y = cy + hour_len * math.sin(hour_angle)
        self.canvas.shapes.append(
            cv.Line(cx, cy, hour_end_x, hour_end_y,
                   paint=ft.Paint(stroke_width=3.5, color=ft.Colors.BLACK))
        )
        
        # 分针
        minute_len = radius * 0.65
        minute_end_x = cx + minute_len * math.cos(minute_angle)
        minute_end_y = cy + minute_len * math.sin(minute_angle)
        self.canvas.shapes.append(
            cv.Line(cx, cy, minute_end_x, minute_end_y,
                   paint=ft.Paint(stroke_width=2.5, color=ft.Colors.BLUE_800))
        )
        
        # 秒针
        second_len = radius * 0.75
        second_end_x = cx + second_len * math.cos(second_angle)
        second_end_y = cy + second_len * math.sin(second_angle)
        self.canvas.shapes.append(
            cv.Line(cx, cy, second_end_x, second_end_y,
                   paint=ft.Paint(stroke_width=1.5, color=ft.Colors.RED))
        )
        
        # 中心点
        self.canvas.shapes.append(
            cv.Circle(cx, cy, 4, paint=ft.Paint(color=ft.Colors.RED_700))
        )
        
        # 关键：强制刷新 canvas 和页面
        self.canvas.update()
        if self.main_page:
            self.main_page.update()

        # 强制刷新整个页面
        if self.main_page:
            self.main_page.update()  # 调用两次确保刷新

class Event:
    def __init__(self, id: str, name: str, birth_date: str, calendar_type: str, event_type: str = "birthday", sound_file: str = "", repeat_type: str = "yearly", reminders: list = None, weekdays: list = None):  # 新增 reminders 参数
        self.id = id
        self.name = name
        self.birth_date = birth_date if birth_date else ""  # 允许空字符串
        self.calendar_type = calendar_type
        self.event_type = event_type        # "birthday" 或 "event" 或 "monthly" 或 "once"
        self.repeat_type = repeat_type      # "yearly" 或 "monthly" 或 "once"
        self.sound_file = sound_file
        self.reminded_this_year = False
        self.last_remind_year = 0
        self.last_remind_month = 0          # 用于每月提醒
        self.completed = False              # 标记一次性事件是否已完成
        self.reminders = reminders if reminders else []  # 提醒时间列表
        self.weekdays = weekdays if weekdays else []     # 每周提醒的星期几 (1-7)
        self.workday_only = False  # 新增：是否只在法定工作日提醒
    
    def to_dict(self):
        return {
            "id": self.id,
            "name": self.name,
            "birth_date": self.birth_date if self.birth_date else "",  # 允许空字符串
            "calendar_type": self.calendar_type,
            "event_type": self.event_type,
            "repeat_type": self.repeat_type,
            "sound_file": self.sound_file,
            "reminded_this_year": self.reminded_this_year,
            "last_remind_year": self.last_remind_year,
            "completed": getattr(self, 'completed', False) ,        # 一次性事件完成标记
            "reminders": getattr(self, 'reminders', []),            # 新增
            "workday_only": getattr(self, 'workday_only', False),
        }
    
    @classmethod
    def from_dict(cls, data):
        if "name" not in data:
            return None
        
        # 先处理 birth_date
        birth_date = data.get("birth_date", "")
        event_type = data.get("event_type", "birthday")
        
        # 如果是每天事件且 birth_date 为空或无效，设置为空字符串
        if event_type == "daily" and (not birth_date or birth_date == "01-01"):
            birth_date = ""
        
        event = cls(
            data["id"], 
            data["name"], 
            birth_date,  # 使用处理后的 birth_date
            data["calendar_type"],
            event_type,
            data.get("sound_file", ""),
            data.get("repeat_type", "yearly"),
            data.get("reminders", []),
        )
        event.reminded_this_year = data.get("reminded_this_year", False)
        event.last_remind_year = data.get("last_remind_year", 0)
        event.last_remind_month = data.get("last_remind_month", 0)
        event.completed = data.get("completed", False)
        event.workday_only = data.get("workday_only", False)
        return event
    
    def is_event_on_date(self, date):
        """判断事件是否在指定日期发生"""
        # 每天事件
        if self.repeat_type == "daily" or self.event_type == "daily":
            return True
        
        # 每周事件
        if self.repeat_type == "weekly" or self.event_type == "weekly":
            target_weekday = int(self.birth_date)  # 1-7
            return date.isoweekday() == target_weekday
        
        # 每月事件
        if self.repeat_type == "monthly" or self.event_type == "monthly":
            target_day = int(self.birth_date)
            return date.day == target_day
        
        # 一次性事件
        if self.repeat_type == "once":
            event_date = datetime.strptime(self.birth_date, "%Y-%m-%d").date()
            return event_date == date
        
        # 生日/纪念日（每年重复）
        # 阳历
        if self.calendar_type == "solar":
            parts = self.birth_date.split("-")
            return date.month == int(parts[1]) and date.day == int(parts[2])
        else:
            # 农历需要转换，暂时跳过或使用原逻辑
            try:
                lunar = LunarDate(date.year, int(self.birth_date.split("-")[1]), int(self.birth_date.split("-")[2]))
                solar = lunar.toSolarDate()
                return solar == date
            except:
                return False

    def get_next_date_info(self):
        """获取下一个发生日期的信息（通用）"""
        today = datetime.now().date()
        current_year = today.year
        current_month = today.month
        
        # 每天提醒
        if self.repeat_type == "daily":
            # 每天都是今天，不需要日期
            today = datetime.now().date()
            return (today.month, today.day, today.year, 0, 0)
        
        # 每周提醒
        if self.repeat_type == "weekly":
            # birth_date 格式为 "1" 表示周一
            target_weekday = int(self.birth_date)  # 1-7
            today_weekday = datetime.now().isoweekday()  # 1=周一, 7=周日
            
            
            if target_weekday == today_weekday:
                days_until = 0
                today = datetime.now().date()
                return (today.month, today.day, today.year, 0, days_until)
            elif target_weekday > today_weekday:
                days_until = target_weekday - today_weekday
            else:
                days_until = (7 - today_weekday) + target_weekday
            
            next_date = datetime.now().date() + timedelta(days=days_until)
        
            #print(f"[每周事件] {self.name}, 今天星期: {today_weekday}, 目标星期: {target_weekday}, 剩余天数: {days_until}")

            return (next_date.month, next_date.day, next_date.year, 0, days_until)

        # 一次性事件
        if self.repeat_type == "once":
            # birth_date 格式为 "YYYY-MM-DD"
            parts = self.birth_date.split("-")
            event_year = int(parts[0])
            event_month = int(parts[1])
            event_day = int(parts[2])
            
            event_date = datetime(event_year, event_month, event_day).date()
            
            if self.completed:
                # 已完成的事件，返回一个很大的天数，表示不再提醒
                return (event_month, event_day, event_year, event_year, 9999)  # 修复：第4个参数返回 event_year
            
            if event_date < today:
                # 已经过期，返回负数天数
                days_until = (event_date - today).days
                return (event_month, event_day, event_year, event_year, days_until)  # 修复：第4个参数返回 event_year
            else:
                days_until = (event_date - today).days
                return (event_month, event_day, event_year, event_year, days_until)  # 修复：第4个参数返回 event_year

        # 每月提醒
        if self.repeat_type == "monthly":
            # birth_date 格式为 "15" 表示每月15号
            month_day = int(self.birth_date)
            
            # 检查本月是否已经过了
            try:
                this_month_date = datetime(current_year, current_month, month_day).date()
            except ValueError:
                # 处理2月30日等无效日期
                if current_month == 2 and month_day > 28:
                    month_day = 28
                this_month_date = datetime(current_year, current_month, month_day).date()
            
            if this_month_date < today:
                # 下个月的同一天
                if current_month == 12:
                    next_month = 1
                    next_year = current_year + 1
                else:
                    next_month = current_month + 1
                    next_year = current_year
                try:
                    next_date = datetime(next_year, next_month, month_day).date()
                except ValueError:
                    # 处理下个月没有这一天的情况（如1月31日 -> 2月28日）
                    if next_month == 2 and month_day > 28:
                        month_day = 28
                    next_date = datetime(next_year, next_month, month_day).date()
                days_until = (next_date - today).days
                return (next_month, month_day, next_year, 0, days_until)
            else:
                days_until = (this_month_date - today).days
                return (current_month, month_day, current_year, 0, days_until)

        # 每年提醒（原有的逻辑）
        elif  self.calendar_type == "solar":
            # 阳历生日/事件
            parts = self.birth_date.split("-")
            birth_month = int(parts[1])
            birth_day = int(parts[2])
            birth_year = int(parts[0])
            
            try:
                this_year_date = datetime(current_year, birth_month, birth_day).date()
                
                if this_year_date < today:
                    next_year_date = datetime(current_year + 1, birth_month, birth_day).date()
                    days_until = (next_year_date - today).days
                    return (next_year_date.month, next_year_date.day, current_year + 1, birth_year, days_until)
                else:
                    days_until = (this_year_date - today).days
                    return (birth_month, birth_day, current_year, birth_year, days_until)
            except ValueError:
                return (1, 1, current_year, birth_year, 365)
        else:
            # 农历生日/事件
            parts = self.birth_date.split("-")
            lunar_year = int(parts[0])
            lunar_month = int(parts[1])
            lunar_day = int(parts[2])
            
            try:
                this_year_lunar = LunarDate(current_year, lunar_month, lunar_day)
                solar_date = this_year_lunar.toSolarDate()
                
                if solar_date < today:
                    next_year_lunar = LunarDate(current_year + 1, lunar_month, lunar_day)
                    next_solar = next_year_lunar.toSolarDate()
                    days_until = (next_solar - today).days
                    return (next_solar.month, next_solar.day, current_year + 1, lunar_year, days_until)
                else:
                    days_until = (solar_date - today).days
                    return (solar_date.month, solar_date.day, current_year, lunar_year, days_until)
            except Exception as e:
                print(f"农历转换错误: {e}")
                return (1, 1, current_year, lunar_year, 365)

class LyricsDownloader:
    def __init__(self, page=None, show_snack_bar=None):
        self.session = requests.Session()
        self.page = page  # 保存 page 引用
        self.show_snack_bar = show_snack_bar if show_snack_bar else lambda msg: print(f"[消息] {msg}")
        #self.show_snack_bar = show_snack_bar  # 保存提示函数
        self.user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Mozilla/5.0 (Linux; Android 13; SM-S911B) AppleWebKit/537.36',
        ]
        self.session.headers.update({
            'Referer': 'https://www.gequbao.com/',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9',
        })

    def get_random_ua(self):
        import random
        return random.choice(self.user_agents)

    def get_online_play_url(self, song_name, artist=""):
        """获取在线播放URL（不下载文件）"""
        try:
            # 尝试使用网易云音乐API
            from pyncm import apis
            from pyncm.apis.login import LoginViaAnonymousAccount
            
            LoginViaAnonymousAccount()
            
            keyword = f"{song_name} {artist}".strip() if artist else song_name
            result = apis.cloudsearch.GetSearchResult(keyword=keyword, stype=1, limit=3)
            
            if result.get('result', {}).get('songs'):
                song = result['result']['songs'][0]
                song_id = song['id']
                found_song_name = song['name']
                found_artist = song['ar'][0]['name']
                

                # 1. 【主要方案】尝试获取真实的CDN链接（这是最可靠的）
                audio_info = apis.track.GetTrackAudio(song_id)
                real_url = audio_info.get('data', [{}])[0].get('url')
                if real_url:
                    print(f"[网易云] 获取到真实CDN链接: {real_url[:100]}...")
                    return {
                        'url': real_url,
                        'name': found_song_name,
                        'artist': found_artist,
                        'id': song_id
                    }
                else:
                    # 使用网易云外链（稳定）
                    play_url = f"https://music.163.com/song/media/outer/url?id={song_id}.mp3"
                    return {
                        'url': play_url,
                        'name': found_song_name,
                        'artist': found_artist,
                        'id': song_id
                    }
            return None
        except Exception as e:
            print(f"获取在线播放URL失败: {e}")
            return None

    def get_mp3_url_simple(self, song_url):
        """Windows/Mac平台：如果是Android系统直接跳过，无法下载！"""
        if not PYCNM_AVAILABLE:
            print("python 模块不可用，跳过")
            return None
        
        """Android平台：简单方法，直接从HTML中提取MP3链接，失败时使用网易云音乐兜底"""
        mp3_url = None
        
        # 方法1：从歌曲宝HTML中提取MP3链接
        try:
            headers = {'User-Agent': self.get_random_ua()}
            response = self.session.get(song_url, headers=headers, timeout=15)
            response.encoding = 'utf-8'
            html_content = response.text
            
            # 查找MP3链接
            mp3_match = re.search(r'https?://[^\s"\']+\.mp3', html_content)
            if mp3_match:
                mp3_url = mp3_match.group(0)
                # 修复：先检查 mp3_url 是否为 None
                if mp3_url:
                    self._safe_show_message(f"✅ 从页面获取到MP3链接")
                    print(f"[简单方法] 从HTML提取到MP3链接: {mp3_url[:100]}...")
                    return mp3_url
            
            # 查找M4A链接
            m4a_match = re.search(r'https?://[^\s"\']+\.m4a', html_content)
            if m4a_match:
                mp3_url = m4a_match.group(0)
                if mp3_url:
                    self._safe_show_message(f"✅ 从页面获取到M4A链接")
                    print(f"[简单方法] 从HTML提取到M4A链接: {mp3_url[:100]}...")
                    return mp3_url
                    
        except Exception as e:
            print(f"[简单方法] 从歌曲宝提取链接失败: {e}")
            self._safe_show_message(f"⚠️ 从歌曲宝提取失败: {str(e)[:50]}")
        
        # 方法2：从歌曲宝页面提取歌曲名称，然后使用网易云音乐下载
        print("[简单方法] 歌曲宝链接提取失败，尝试使用网易云音乐兜底...")
        self._safe_show_message("🔄 尝试网易云音乐...")
        
        try:
            # 先从歌曲宝页面提取歌曲名称和歌手
            headers = {'User-Agent': self.get_random_ua()}
            response = self.session.get(song_url, headers=headers, timeout=15)
            response.encoding = 'utf-8'
            html_content = response.text
            
            # 提取歌曲名称和歌手
            song_name = None
            artist = None
            
            # 方法：从title标签提取
            title_match = re.search(r'<title>(.+?)</title>', html_content)
            if title_match:
                title = title_match.group(1)
                # 格式通常是 "歌曲名 - 歌手名 - 歌曲宝"
                if ' - ' in title:
                    parts = title.split(' - ')
                    if len(parts) >= 2:
                        song_name = parts[0].strip()
                        artist = parts[1].strip()
                        print(f"[简单方法] 从页面提取到: {song_name} - {artist}")
            
            if not song_name:
                print("[简单方法] 无法从页面提取歌曲信息")
                self._safe_show_message("❌ [简单方法] 无法从页面提取歌曲信息")
                return None
            
            # 使用网易云音乐搜索并下载
            print(f"[网易云兜底] 正在搜索: {song_name} - {artist}")
            
            # 尝试导入 pyncm
            try:
                from pyncm import apis
                from pyncm.apis.login import LoginViaAnonymousAccount
                
                # 匿名登录
                LoginViaAnonymousAccount()
                print("[网易云兜底] 匿名登录成功")
                
                # 搜索歌曲
                result = apis.cloudsearch.GetSearchResult(
                    keyword=f"{song_name} {artist}" if artist else song_name,
                    stype=1,
                    limit=3
                )
                
                if not result.get('result', {}).get('songs'):
                    print("[网易云兜底] 未找到相关歌曲")
                    self._safe_show_message("❌ 网易云未找到歌曲")
                    return None
                
                # 取第一首搜索结果
                song = result['result']['songs'][0]
                song_id = song['id']
                found_song_name = song['name']
                found_artist = song['ar'][0]['name']
                print(f"[网易云兜底] 找到歌曲: {found_song_name} - {found_artist} (ID: {song_id})")
                
                # 获取下载链接
                audio_info = apis.track.GetTrackAudio(song_id)
                
                if not audio_info.get('data') or not audio_info['data'][0].get('url'):
                    print("[网易云兜底] 无法获取下载链接，可能需VIP")
                    self._safe_show_message("❌ 网易云链接获取失败（可能需要VIP）")
                    return None
                
                mp3_url = audio_info['data'][0]['url']
                if mp3_url:
                    mp3_url = re.sub(r'\?.*$', '', mp3_url)
                    self._safe_show_message(f"✅ 网易云获取到链接")
                    print(f"[网易云兜底] 获取到MP3链接: {mp3_url[:100]}...")
                    return mp3_url
                
            except ImportError:
                print("[网易云兜底] pyncm 未安装")
                # self._safe_show_message("❌ 网易云模块未安装")
                return None
            except Exception as e:
                print(f"[网易云兜底] 出错: {e}")
                self._safe_show_message(f"❌ 网易云出错: {str(e)[:50]}")
                return None
                
        except Exception as e:
            print(f"[简单方法] 网易云兜底失败: {e}")
            self._safe_show_message(f"❌ 兜底失败: {str(e)[:50]}")
            return None

    def get_mp3_url_playwright(self, song_url):
        """Windows/Mac平台：使用playwright获取MP3链接"""
        if not PLAYWRIGHT_AVAILABLE:
            print("playwright 模块不可用，跳过")
            return None
    
        mp3_url = None
    
        try:
            from playwright.sync_api import sync_playwright
            
            with sync_playwright() as p:
                # 查找系统浏览器路径（优先Edge，其次Chrome）
                browser_paths = [
                    r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
                    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
                    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                ]
                
                browser_exe = None
                for path in browser_paths:
                    if os.path.exists(path):
                        browser_exe = path
                        break
                
                if browser_exe:
                    browser = p.chromium.launch(headless=True, executable_path=browser_exe)
                    print(f"✓ 使用系统浏览器: {browser_exe}")
                else:
                    browser = p.chromium.launch(headless=True)
                    print("⚠️ 使用内置 Chromium")
                
                page = browser.new_page()
                
                # 监听网络请求（捕获MP3请求）
                def handle_request(request):
                    nonlocal mp3_url
                    # 检查是否为MP3请求（特别是来自 kuwo.cn 的）
                    if '.mp3' in request.url and ('kuwo' in request.url.lower() or 'music' in request.url.lower()):
                        mp3_url = re.sub(r'\?.*$', '', request.url)
                        print(f"✅ 拦截到MP3请求: {mp3_url[:80]}...")
                
                # 监听响应（备用方案）
                def handle_response(response):
                    nonlocal mp3_url
                    if '.mp3' in response.url and ('kuwo' in response.url.lower() or 'music' in response.url.lower()):
                        mp3_url = re.sub(r'\?.*$', '', response.url)
                        print(f"✅ 从响应中捕获到MP3链接: {mp3_url[:80]}...")
                
                # 绑定事件
                page.on('request', handle_request)
                page.on('response', handle_response)
                
                # 访问歌曲页面
                print(f"正在访问: {song_url}")
                page.goto(song_url, wait_until="domcontentloaded", timeout=15000)
                
                # 查找并点击下载按钮（使用多种方式）
                try:
                    # 方式1：通过文本查找
                    download_btn = page.locator("text=下载歌曲").first
                    if download_btn and download_btn.is_visible():
                        download_btn.click()
                        print("已点击「下载歌曲」按钮")
                except:
                    try:
                        # 方式2：通过CSS选择器查找
                        download_btn = page.locator("a:has-text('下载歌曲')").first
                        if download_btn and download_btn.is_visible():
                            download_btn.click()
                            print("已通过CSS选择器点击下载按钮")
                    except:
                        print("未找到下载按钮，尝试其他方式...")
                        # 方式3：尝试点击播放按钮触发
                        try:
                            play_btn = page.locator("audio").first
                            if play_btn:
                                play_btn.click()
                                print("已点击播放按钮")
                        except:
                            pass
                
                # 等待MP3请求（最多等待30秒）
                print("等待MP3链接...")
                for _ in range(60):
                    if mp3_url:
                        break
                    page.wait_for_timeout(500)
                
                # 如果还没有找到，尝试刷新页面
                if not mp3_url:
                    print("未拦截到请求，尝试刷新页面...")
                    page.reload(wait_until="domcontentloaded")
                    page.wait_for_timeout(3000)
                    
                    # 再次尝试点击下载按钮
                    try:
                        page.locator("text=下载歌曲").first.click()
                        page.wait_for_timeout(3000)
                    except:
                        pass
                    
                    # 再次等待
                    for _ in range(60):
                        if mp3_url:
                            break
                        page.wait_for_timeout(500)
                
                browser.close()
                
        except Exception as e:
            print(f"浏览器获取失败: {e}")
            import traceback
            traceback.print_exc()
        
        return mp3_url

    def get_mp3_url_auto(self, song_url):
        """自动根据平台选择方法获取MP3链接"""
        print(f"[get_mp3_url_auto] 开始执行")
        print(f"[平台检测] 当前系统: {platform.system()}")
        print(f"[get_mp3_url_auto] song_url: {song_url}")
        
        if platform.system() != "Windows":
            print("[下载] 安卓平台：暂时不支持下载功能")
            self._safe_show_message("📱 Android版本暂不支持下载音乐，请手动选择音乐文件")
            return None
        else:
            print("[下载] 桌面平台：使用 Playwright 获取链接")
            mp3_url = self.get_mp3_url_playwright(song_url)
            if not mp3_url:
                print("[下载] Playwright 方法失败，降级到简单方法")
                mp3_url = self.get_mp3_url_simple(song_url)
            print(f"[下载] 方法返回: {mp3_url}")
            return mp3_url

    def search_and_get_lyrics(self, song_name, artist=""):
        """根据歌名和歌手搜索并获取歌词 - 针对歌曲宝优化"""
        try:
            # 构建搜索关键词 - 优先使用歌曲名，歌手名作为辅助
            if artist and artist != "未知歌手":
                keyword = f"{song_name} {artist}".strip()
            else:
                keyword = song_name.strip()
            
            print(f"[搜索歌词] 关键词: {keyword}")
            
            # URL编码关键词
            from urllib.parse import quote
            encoded_keyword = quote(keyword)
            search_url = f"https://www.gequbao.com/s/{encoded_keyword}"
            headers = {'User-Agent': self.get_random_ua()}
            
            # 1. 搜索歌曲，获取第一个结果的URL
            response = self.session.get(search_url, headers=headers, timeout=15)
            response.encoding = 'utf-8'
            if response.status_code != 200:
                print(f"搜索失败，状态码: {response.status_code}")
                if self.show_snack_bar:
                    self.show_snack_bar(f"搜索失败，状态码: {response.status_code}")
                return None

            # 提取歌曲详情页URL (使用非贪婪匹配)
            match = re.search(r'<a href="(/music/\d+)"', response.text)
            if not match:
                print("未在搜索结果中找到歌曲链接")
                return None
                
            song_url = "https://www.gequbao.com" + match.group(1)
            print(f"找到歌曲页面: {song_url}")

            # 2. 访问歌曲详情页，获取HTML
            response2 = self.session.get(song_url, headers=headers, timeout=15)
            response2.encoding = 'utf-8'
            html_content = response2.text

            # 3. 【核心】使用正则直接提取时间标签和对应的歌词
            # 匹配格式如 [00:00.0]此生不换 - 青鸟飞鱼
            lrc_pattern = re.compile(r'\[(\d{2}:\d{2}\.\d+)\]([^\n<]+)')
            matches = re.findall(lrc_pattern, html_content)

            if matches:
                # 将匹配到的内容组合成完整的LRC字符串
                lrc_lines = []
                for time_tag, text in matches:
                    # 清理歌词文本中的HTML实体（如 &quot; 等）
                    clean_text = html.unescape(text.strip())
                    # 清理 HTML 标签
                    clean_text = re.sub(r'<br\s*/?>', '', clean_text, flags=re.IGNORECASE)
                    clean_text = re.sub(r'<[^>]+>', '', clean_text)
                    if clean_text:  # 确保不是空行
                        lrc_lines.append(f"[{time_tag}]{clean_text}")
                
                if lrc_lines:
                    print(f"成功解析到 {len(lrc_lines)} 行歌词")
                    return '\n'.join(lrc_lines)
                else:
                    print("解析到的歌词为空")
            else:
                print("未在页面中找到时间标签格式的歌词")
            
            return None

        except Exception as e:
            print(f"获取歌词过程中出错: {e}")
            if self.show_snack_bar:
                self.show_snack_bar(f"获取歌词过程中出错: {e}")
            return None
    
    def _safe_show_message(self, message):
        """安全地显示消息（线程安全）"""
        print(f"[LyricsDownloader] {message}")
        if self.show_snack_bar and self.page:
            # 使用 threading.Timer 在主线程中延迟执行
            def show():
                self.show_snack_bar(message)
            threading.Timer(0.1, show).start()

    def download_lyrics_for_music(self, sound_file_path, song_name=None, artist=None):
        """为本地音乐文件下载歌词"""
        lrc_path = os.path.splitext(sound_file_path)[0] + ".lrc"
        
        # 如果歌词已存在，跳过
        if os.path.exists(lrc_path):
            #print(f"歌词已存在: {lrc_path}")
            #self._safe_show_message(f"⚠️ 歌词已存在: {os.path.basename(lrc_path)}")
            #self.show_snack_bar(f"⚠️ 歌词已存在: {lrc_path}")
            return True
        
        # 如果没有提供歌名，从文件名解析
        if not song_name:
            base_name = os.path.basename(sound_file_path)
            base_name = os.path.splitext(base_name)[0]
            if " - " in base_name:
                parts = base_name.split(" - ")
                if len(parts) >= 2:
                    # 尝试智能判断：通常文件名格式是 "歌曲名 - 歌手名"
                    # 假设第一部分是歌曲名，第二部分是歌手名
                    song_name = parts[0].strip()
                    artist = parts[1].strip()
                    print(f"[文件名解析] 歌曲: {song_name}, 歌手: {artist}")
                else:
                    song_name = base_name
                    artist = ""
            else:
                song_name = base_name
                artist = ""
        
        print(f"正在搜索歌词: {song_name} - {artist}")
        
        lyrics = self.search_and_get_lyrics(song_name, artist)
        if lyrics:
            try:
                # 保存前再次清理歌词中的 HTML 标签
                cleaned_lines = []
                for line in lyrics.split('\n'):
                    # 清理每一行的 HTML 标签
                    clean_line = re.sub(r'<br\s*/?>', '', line, flags=re.IGNORECASE)
                    clean_line = re.sub(r'<[^>]+>', '', clean_line)
                    clean_line = html.unescape(clean_line)
                    if clean_line.strip():
                        cleaned_lines.append(clean_line)
                
                cleaned_lyrics = '\n'.join(cleaned_lines)
                
                with open(lrc_path, 'w', encoding='utf-8') as f:
                    f.write(cleaned_lyrics)
                print(f"歌词已保存: {lrc_path}")
                if self.show_snack_bar:
                    self.show_snack_bar(f"歌词已保存: {os.path.basename(lrc_path)}")
                return True
            except Exception as e:
                print(f"保存歌词文件失败: {e}")
                if self.show_snack_bar:
                    self.show_snack_bar(f"保存歌词文件失败: {e}")
        else:
            print("未能从歌曲宝获取歌词")
            if self.show_snack_bar:
                self.show_snack_bar(f"未能从歌曲宝获取歌词")
        
        return False


def get_data_file_path(filename):
    app_data_dir = os.getenv("FLET_APP_STORAGE_DATA")
    if app_data_dir:
        os.makedirs(app_data_dir, exist_ok=True)
        return os.path.join(app_data_dir, filename)
    else:
        return filename
    
def main(page: ft.Page):

    """入口：检查设备授权"""
    
    # ========== 预设授权设备列表 ==========
    # 先运行一次程序，从控制台获取设备ID，然后填在这里
    ALLOWED_DEVICES = [
        "6472c4db5200105e8788ba00aee9fe84",  # 开发者的window ID
        "819374e1a2b43595a5da70474fcc7e4f",  # 开发者的手机 ID1
        "49d8a4e0bdcd63998fb17e6033727d64",  # 开发者的手机 ID2
        #"",  # 可以添加多个
    ]
    
    # 获取当前设备ID
    current_device_id = get_device_id()
    print(f"[设备授权] 当前设备ID: {current_device_id}")
    
    # 检查设备是否在授权列表中
    if current_device_id not in ALLOWED_DEVICES:
        # 未授权，显示未授权页面
        show_unauthorized_page(page, current_device_id)
        return
    
    # 设备已授权，继续执行原来的主程序
    print("[设备授权] ✅ 设备已授权，启动主程序")


    # 设备已授权，进入主程序逻辑
    def show_snack_bar_new(page, message, is_error=False):
        """显示 SnackBar（兼容不同 Flet 版本）"""
        try:
            if hasattr(page, 'show_snack_bar'):
                page.show_snack_bar(
                    ft.SnackBar(
                        content=ft.Text(message),
                        bgcolor=ft.Colors.RED_700 if is_error else ft.Colors.GREEN_700,
                    )
                )
            else:
                snack = ft.SnackBar(
                    content=ft.Text(message),
                    bgcolor=ft.Colors.RED_700 if is_error else ft.Colors.GREEN_700,
                    open=True,
                )
                page.overlay.append(snack)
                page.update()
                def close_snack():
                    time.sleep(3)
                    if snack in page.overlay:
                        page.overlay.remove(snack)
                        page.update()
                threading.Thread(target=close_snack, daemon=True).start()
        except Exception as e:
            print(f"显示 SnackBar 失败: {e}")
            
    # 在函数最开始声明所有需要使用的全局变量
    global current_audio, is_playing, current_music_file,current_playing_event_id,current_music_state,music_state_update_callback
    global lyrics_fullscreen_container, auto_scroll_task, current_position_sec,current_lyrics , events  # 添加 events
    global scroll_timer,scroll_position,scroll_text_length, original_music_text  # 添加 original_music_text
    global last_check_date,reminder_flags,music_title_container, main_content, marquee_text # 添加这两个变量
    global selected_date,three_days_events, date_text,current_view   # 添加 date_text
    global month_text, current_year, current_month, today_circle_button  # 添加 today_circle_button
    global music_control_container, playback_buttons, music_section_container  # 修改这里
    global sent_notifications,events_list,filter_date
    global transactions  # 添加这行
    global current_page, floating_add_button,show_scroll_top_btn  # 添加这行，用于记录当前页面
    global auto_fullscreen_lyrics,hide_progress_timer,current_selected_lunar,last_card_update_time  # 添加这行
    global SLIDER_WIDTH, progress_slider, progress_bubble, progress_bubble_container, slider_wrapper,card_duration_texts
    global sent_reminders,sent_music_notifications


    page.window_icon = "icon.png"
    page.title = "事件提醒助手"
    page.bgcolor = ft.Colors.WHITE
    page.window_width = 550
    page.window_height = 800
    page.window_resizable = True
    #page.scroll = ft.ScrollMode.AUTO
    page.theme_mode = ft.ThemeMode.LIGHT

    # ========== 添加这段代码来修复状态栏 ==========
    # 创建样式：透明背景 + 深色图标
    my_overlay_style = ft.SystemOverlayStyle(
        status_bar_color=ft.Colors.TRANSPARENT,  # 关键：设为透明
        status_bar_icon_brightness=ft.Brightness.DARK,
    )
    # 应用到主题
    page.theme = ft.Theme(system_overlay_style=my_overlay_style)
    page.dark_theme = ft.Theme(system_overlay_style=my_overlay_style)

    
    # 请求 Android 存储权限
    def request_permissions():
        if hasattr(page, 'request_permission'):
            try:
                page.request_permission("android.permission.READ_EXTERNAL_STORAGE")
                page.request_permission("android.permission.READ_MEDIA_AUDIO")
                # 添加这行：请求通知权限（Android 13+ 必需）
                page.request_permission("android.permission.POST_NOTIFICATIONS")
                print("已请求存储权限")
            except Exception as e:
                print(f"权限请求失败: {e}")
    
    page.on_ready = request_permissions

    # ========== 初始化通知功能 ==========
    # 初始化通知渠道
    if ANDROID_NOTIFY_AVAILABLE and platform.system() == "Linux":
        try:
            init_notify = Notification()
            init_notify.channel_name = "事件提醒助手"
            init_notify.channel_description = "事件提醒助手通知渠道"
            init_notify.importance = "low"
            print("[通知] ✅ 通知渠道已初始化")
        except Exception as e:
            print(f"[通知] 渠道初始化失败: {e}")

    sent_notifications = set()  # 记录已发送的通知，格式: "事件ID_提醒时间_日期"

    reminder_flags = {}  # 存储提醒标记

    card_duration_texts = {}  # {event_id: Text控件}

    three_days_events = []  # 存储3日内事件列表

    sent_reminders = set()  # 记录已发送的提醒

    sent_music_notifications = set()  # 记录已发送的音乐通知

    # ========== 在文件顶部添加全局变量 ==========
    current_selected_lunar = ""  # 存储当前选中的农历日期

    # 添加一个时间戳变量，控制刷新频率
    last_card_update_time = 0

    # ========== 隐藏进度文本的定时器 ==========
    hide_progress_timer = None

    auto_fullscreen_lyrics = False  # 记录是否需要在下次播放时自动打开全屏歌词

    # 初始化 filter_date
    filter_date = None

    # 是否显示回到顶部按钮
    show_scroll_top_btn = False

    # 初始化当前页面
    current_page = "main"  # "main" 或 "accounting"

    current_display_view = "main"  # main: 全部/今日事件, warning: 预警事件

    # 在函数外部定义全局变量
    selected_date = None  # 选中的日期，初始为None
    
    events = {}
    selected_event = None
    current_view = "daily"  # 可选值: "today", "three_days", "all", "daily", "weekly"
    current_date = datetime.now().date()
    dialog_container = None

    debug_mode = True  # 开启调试模式
    last_check_date = None  # 记录上次检查的日期
    
    # 记录程序启动时间
    start_time = datetime.now()
    
    # 音乐控制变量
    current_audio = None
    current_music_file = None
    is_playing = False
    is_playing_lock = threading.Lock()  # 添加锁
    saved_sound_file = None
    music_playing_lock = threading.Lock()  # 添加音乐播放锁

    # 音乐播放控制变量
    current_duration = 0
    current_position = 0
    current_lyrics = []

    # 在适当位置添加这些变量（在其他变量附近）
    lyrics_fullscreen_container = None  # 全屏歌词容器
    auto_scroll_task = None  # 自动滚动任务
    current_position_sec = 0  # 当前播放位置（秒）

    # 添加这个字典来存储每个事件的循环状态
    event_loop_states = {}  # {event_id: bool}

    # 添加新的状态管理
    current_playing_event_id = None  # 当前播放的事件ID
    current_music_state = "stopped"  # 播放状态: playing, paused, stopped
    music_state_update_callback = None  # 用于更新UI的回调函数

    scroll_timer = None
    scroll_position = 0
    scroll_text_length = 0
    original_music_text = ""

    # 启动时间显示
    #start_time_text = ft.Text(value=f"🚀 启动时间: {start_time.strftime('%H:%M:%S')}", size=12, color=ft.Colors.GREY_600)
    start_time_text = ft.Text(value=f"🚀 启动时间: {start_time.strftime('%Y年%m月%d日 %H:%M:%S')}", size=12, color=ft.Colors.GREY_600)
    run_time_text = ft.Text(value="⏱️ 运行时间: 00:00:00", size=12, color=ft.Colors.GREEN_600)  # 新增
    # 当前日期时间显示
    current_datetime_text = ft.Text(value="📅 当前时间：",size=12, color=ft.Colors.BLUE_700)

    # ========== 记账分类定义（放在这里） ==========

    # 初始化记账数据
    transactions = []

    # 收入分类（预设）
    INCOME_CATEGORIES = [
        "工资收入",
        "报销收入",
        "奖金收入",
        "兼职收入",
        "投资收入",
        "红包收入",
        "退款",
        "其他收入",
    ]
    
    # 支出分类（预设）
    EXPENSE_CATEGORIES = [
        "餐饮",
        "水电费",
        "电话费",
        "生活费",
        "房贷",
        "车贷",
        "网贷",
        "零食",
        "母婴用品",
        "购物",
        "娱乐",
        "交通",
        "医疗",
        "教育",
        "服饰",
        "美容",
        "宠物",
        "社交",
        "旅游",
        "其他支出",
    ]

    
    def debug_log(msg):
        """调试日志函数"""
        if debug_mode:
            print(f"[DEBUG {datetime.now().strftime('%H:%M:%S')}] {msg}")

    # 测试按钮的回调函数，仅做测试用途
    def test_notification(e):
        """测试通知功能"""
        show_bottom_message("正在测试通知...")
        result = show_notification(page, "🔔 测试通知", f"当前时间: {datetime.now().strftime('%H:%M:%S')}")
        if result:
            show_bottom_message("✅ 通知已发送")
        else:
            show_bottom_message("❌ 通知发送失败")
    
    def is_workday(date):
        """判断是否为工作日（使用chinese-days库）"""
        try:
            return cn_is_workday(date)
        except Exception as e:
            print(f"判断工作日失败: {e}")
            # 降级：简单判断周末
            return date.weekday() < 5

    # ========== 通知功能开始 ==========
    def show_notification(page, title: str, message: str, notification_id: int = None, ongoing: bool = False):
        """发送系统通知
        Args:
            page: Flet Page 对象
            title: 通知标题
            message: 通知内容
            notification_id: 通知ID（保留参数，暂未使用）
            ongoing: 是否持续通知（保留参数，暂未使用）
        """
        print(f"[通知] 发送: {title} - {message}")

        # ========== Windows 平台直接返回 ==========
        if IS_WINDOWS:
            print(f"[通知] Windows 平台，通知已跳过: {title}")
            return False
        
        if not ANDROID_NOTIFY_AVAILABLE:
            print("[通知] ❌ android_notify 不可用")
            return False

        try:
            n = Notification(title=title, message=message)
            n.send()
            print("[通知] ✅ 发送成功")
            return True
        except Exception as e:
            print(f"[通知] ❌ 发送失败: {e}")
            return False


    def cancel_notification(notification_id: int):
        """取消通知（使用 android-notify）"""
        #print(f"[通知] 尝试取消通知 ID: {notification_id}")

        # ========== Windows 平台静默返回 ==========
        if IS_WINDOWS:
            return
        
        if platform.system() != "Linux":
            print("[通知] 非 Android 平台，跳过取消")
            return
        
        if not ANDROID_NOTIFY_AVAILABLE:
            print("[通知] android_notify 不可用，跳过取消")
            return
        
        try:
            # android-notify 的取消通知方法
            # 方法1：创建一个相同 ID 的通知并取消
            n = Notification(notification_id=notification_id)
            n.cancel()
            print(f"[通知] ✅ 已取消通知 ID: {notification_id}")
        except AttributeError:
            try:
                # 方法2：使用 NotificationManager 直接取消
                from android import activity
                from android.app import NotificationManager
                
                notification_manager = activity.getSystemService("notification")
                notification_manager.cancel(notification_id)
                print(f"[通知] ✅ 已取消通知 ID: {notification_id} (方法2)")
            except Exception as e2:
                print(f"[通知] ❌ 取消通知失败: {e2}")
        except Exception as e:
            print(f"[通知] ❌ 取消通知失败: {e}")


    # 音乐播放通知ID（plyer 不需要）
    MUSIC_NOTIFICATION_ID = 8888
    EVENT_NOTIFICATION_ID = 9999
    BACKGROUND_NOTIFICATION_ID = 7777


    def update_music_notification(song_name: str, is_playing: bool = True):
        """更新音乐播放通知（带去重）"""
        # ========== Windows 平台直接返回 ==========
        if IS_WINDOWS:
            return
    
        if not is_playing:
            return
        
        # ========== 去重检查 ==========
        notification_key = f"{song_name}_{datetime.now().strftime('%Y%m%d%H%M')}"
        if notification_key in sent_music_notifications:
            return  # 已发送过，跳过
        
        sent_music_notifications.add(notification_key)
        
        status = "▶️ 播放中" if is_playing else "⏸️ 已暂停"
        show_notification( page,"🎵 事件提醒助手", f"{status}: {song_name}",notification_id=MUSIC_NOTIFICATION_ID,)


    def show_event_notification(event_name: str, event_type: str, days_left: int = 0):
        """显示事件提醒通知"""
        # ========== Windows 平台直接返回 ==========
        if IS_WINDOWS:
            return
    
        if days_left == 0:
            title = "🎉 今日事件提醒"
            message = f"{event_name} 就在今天！"
        elif days_left == 1:
            title = "⏰ 事件提醒"
            message = f"{event_name} 明天就到啦！"
        else:
            title = "⏰ 事件提醒"
            message = f"{event_name} 还有 {days_left} 天"
        
        show_notification(page, title, message, notification_id=EVENT_NOTIFICATION_ID)


    def show_background_notification():
        """显示后台运行通知（持久）"""
        # ========== Windows 平台跳过 ==========
        if IS_WINDOWS:
            print("[通知] Windows 平台，后台通知已跳过")
            return
    
        show_notification(page,"🔔 事件提醒助手", "应用正在后台运行，监控您的提醒事件\n点击打开应用",notification_id=BACKGROUND_NOTIFICATION_ID,ongoing=True)
    # ========== 通知功能结束 ==========

    def load_events():
        try:
            json_path = get_data_file_path("events.json")
            if os.path.exists(json_path):
                with open(json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    for event_data in data:
                        try:
                            event = Event.from_dict(event_data)
                            if event:
                                events[event.id] = event
                        except:
                            continue
                print(f"加载 {len(events)} 个事件")
                
                # ========== 方案四：启动时重置跨年提醒标记 ==========
                current_year = datetime.now().year
                modified = False
                
                #print(f"[启动重置] 当前年份: {current_year}")
                
                for event_id, event in events.items():
                    #print(f"[启动重置] 检查事件: {event.name}, last_remind_year={event.last_remind_year}")
                    
                    # 如果 last_remind_year 小于当前年份，说明是去年的标记，需要重置
                    if event.last_remind_year > 0 and event.last_remind_year < current_year:
                        #print(f"[启动重置] ✓ 重置事件 {event.name} 的提醒状态 (从 {event.last_remind_year} 到 0)")
                        event.last_remind_year = 0
                        event.reminded_this_year = False
                        modified = True
                    elif event.last_remind_year == current_year:
                        #print(f"[启动重置] 事件 {event.name} 今年已提醒过，保持状态")
                        pass
                    else:
                        #print(f"[启动重置] 事件 {event.name} 状态正常")
                        pass
                
                if modified:
                    save_events()
                    #print(f"[启动重置] 已完成跨年提醒标记重置")
                else:
                    #print(f"[启动重置] 无需重置")
                    pass
                    
        except Exception as e:
            print(f"加载失败: {e}")
    
    def save_events(trigger_check=False):
        """保存事件到文件（安全版本）
        Args:
            trigger_check: 是否触发生日检查（编辑/新增事件时设为True）
        """
        try:
            json_path = get_data_file_path("events.json")
            
            # 如果原文件存在，先备份
            if os.path.exists(json_path):
                backup_path = json_path + ".bak"
                try:
                    import shutil
                    shutil.copy2(json_path, backup_path)
                    print(f"已备份到: {backup_path}")
                except:
                    pass
            
            # 写入新文件
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump([e.to_dict() for e in events.values()], f, ensure_ascii=False, indent=2)
            
            # 验证保存成功
            if os.path.exists(json_path) and os.path.getsize(json_path) > 0:
                print(f"已保存 {len(events)} 个事件到 {json_path}")
                # 不要在这里调用 check_events()，避免递归
            else:
                print("⚠️ 保存的文件可能为空")
            
            # ========== 根据当前视图刷新对应的视图 ==========
            #refresh_current_view_by_state()
                    
        except Exception as e:
            print(f"保存失败: {e}")
            show_snack_bar(f"保存失败: {str(e)}")
            # 如果保存失败，尝试恢复备份
            backup_path = json_path + ".bak"
            if os.path.exists(backup_path):
                try:
                    import shutil
                    shutil.copy2(backup_path, json_path)
                    show_snack_bar("已从备份恢复")
                except:
                    pass
    
    

    def delete_event(event_id):
        """删除事件（优化版）"""
        if event_id not in events:
            show_bottom_message("未找到该事件")
            return
        
        event = events[event_id]
        name = event.name
        
        dialog_container = None
        
        def close_dialog():
            nonlocal dialog_container
            if dialog_container and dialog_container in page.overlay:
                page.overlay.remove(dialog_container)
                dialog_container = None
                page.update()
        
        def confirm_delete(e):
            close_dialog()
            try:
                del events[event_id]
                save_events()
                #refresh_events_list()

                # ========== 更新 three_days_events ==========
                three_days_events.clear()
                for evt in events.values():
                    if evt.event_type == "daily" or evt.event_type == "weekly":
                        continue
                    month, day, year, base_year, days_until = evt.get_next_date_info()
                    if evt.repeat_type == "once" and (evt.completed or days_until < 0):
                        continue
                    if 0 < days_until <= 3:
                        three_days_events.append((evt, days_until))

                # 删除成功后，更新顶部日期文本显示
                update_three_days_events()
                today = datetime.now().date()
                update_date_text_with_events(today, three_days_events)

                # ========== 根据当前视图刷新对应的视图 ==========
                refresh_current_view_by_state()

                # ========== 导入后重新检查视图 ==========
                determine_startup_view()

                show_bottom_message(f"已删除「{name}」")
            except Exception as ex:
                show_bottom_message(f"删除失败: {str(ex)}")
            page.update()
        
        def cancel_delete(e):
            close_dialog()
            show_bottom_message(f"已取消删除「{name}」")
            page.update()
        
        # 对话框内容 WARNING_AMBER_ROUNDED WARNING CANCEL
        dialog_content = ft.Container(
            content=ft.Column([
                # 顶部图标（带背景圆）
                ft.Container(
                    content=ft.Icon(ft.Icons.WARNING_AMBER_ROUNDED, size=55, color=ft.Colors.RED_700),
                    padding=10,
                    bgcolor=ft.Colors.RED_50,
                    border_radius=50,
                ),
                ft.Text("确认删除", size=22, weight=ft.FontWeight.BOLD, color=ft.Colors.RED_700, text_align=ft.TextAlign.CENTER),
                ft.Divider(height=1, color=ft.Colors.GREY_300),
                ft.Text(f"确定要删除「{name}」吗？", size=14, color=ft.Colors.GREY_700, text_align=ft.TextAlign.CENTER),
                ft.Text("此操作不可撤销！", size=12, color=ft.Colors.RED_500, text_align=ft.TextAlign.CENTER),
                ft.Divider(height=1, color=ft.Colors.GREY_300),
                # 按钮区域 - 简化版
                ft.Row([
                    ft.ElevatedButton(
                        "取消", 
                        on_click=cancel_delete, 
                        expand=True,
                        style=ft.ButtonStyle(bgcolor=ft.Colors.GREY_100, color=ft.Colors.GREY_700),
                    ),
                    ft.ElevatedButton(
                        "确认删除", 
                        on_click=confirm_delete, 
                        expand=True,
                        style=ft.ButtonStyle(bgcolor=ft.Colors.RED_700, color=ft.Colors.WHITE),
                    ),
                ], spacing=12, alignment=ft.MainAxisAlignment.CENTER),
            ], spacing=15, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            width=320,
            padding=20,
            bgcolor=ft.Colors.WHITE,
            border_radius=16,
        )
        
        dialog_container = ft.Container(
            content=ft.Column([
                ft.Container(expand=True),  # 上方弹性空间
                ft.Row([
                    ft.Container(expand=True),  # 左侧弹性空间
                    dialog_content,
                    ft.Container(expand=True),  # 右侧弹性空间
                ]),
                ft.Container(expand=True),  # 下方弹性空间
            ]),
            expand=True,
            bgcolor=ft.Colors.BLACK26,
            on_click=close_dialog,
        )
        
        page.overlay.append(dialog_container)
        page.update()

    def format_time(seconds):
        """格式化时间显示 mm:ss"""
        minutes = int(seconds // 60)
        secs = int(seconds % 60)
        return f"{minutes}:{secs:02d}"
    

    # ========== 添加歌词相关函数 ==========
    def parse_lyrics_to_lines(file_path, offset=-0.5):
        """解析LRC歌词文件为带时间戳的行列表
        offset: 时间偏移量（秒），负数表示提前显示，正数表示延后显示
            例如 -0.5 表示歌词提前0.5秒显示
        """
        lyrics_lines = []
        try:
            lrc_path = os.path.splitext(file_path)[0] + ".lrc"
            #print(f"[解析歌词] 尝试读取: {lrc_path}")
            if os.path.exists(lrc_path):
                with open(lrc_path, "r", encoding="utf-8") as f:
                    content = f.read()
                    for line in content.split('\n'):
                        line = line.strip()
                        if not line:
                            continue
                        match = re.match(r'\[(\d{2}):(\d{2}(?:\.\d+)?)\](.*)', line)
                        if match:
                            minutes = int(match.group(1))
                            seconds = float(match.group(2))
                            time_sec = minutes * 60 + seconds
                            # 应用偏移量（负数 = 提前显示）
                            adjusted_time = max(0, time_sec + offset)
                            text = match.group(3).strip()
                            # 清理 HTML 标签（如 <br>、<br/>、<br /> 等）
                            text = re.sub(r'<br\s*/?>', '', text, flags=re.IGNORECASE)
                            # 清理其他可能的 HTML 标签
                            text = re.sub(r'<[^>]+>', '', text)
                            # 去除多余的空格
                            text = re.sub(r'\s+', ' ', text).strip()
                            # 解码 HTML 实体
                            text = html.unescape(text)
                            if text:
                                lyrics_lines.append((adjusted_time, text))
                lyrics_lines.sort(key=lambda x: x[0])
                #print(f"成功加载 {len(lyrics_lines)} 行歌词（偏移 {offset} 秒）")
            else:
                print(f"[解析歌词] 歌词文件不存在: {lrc_path}")
                pass
        except Exception as e:
            print(f"加载歌词文件失败: {e}")
        return lyrics_lines
    
    def update_lyrics_display(position_sec, lyrics_list, lyrics_widgets, is_fullscreen=False):
        """根据播放位置更新歌词显示 - 显示两行，当前行高亮"""
        line1_text, line2_text = lyrics_widgets
        
        # 修改：当音乐停止或未播放时
        if current_music_state == "stopped":
            line1_text.value = "🎵 未播放"
            line1_text.color = ft.Colors.GREY_600
            line2_text.value = ""
            line1_text.update()
            line2_text.update()
            return
        
        # 修改：当没有歌词数据时（音乐正在播放但没有歌词）
        if not lyrics_list or len(lyrics_list) == 0:
            line1_text.value = "📝 本地无歌词或未在线搜索到歌词"
            line1_text.color = ft.Colors.GREY_600
            line1_text.weight = ft.FontWeight.NORMAL
            line1_text.size = 16
            line2_text.value = "💡 提示：可以手动添加 .lrc 歌词文件到音乐同目录"
            line2_text.color = ft.Colors.GREY_500
            line2_text.weight = ft.FontWeight.NORMAL
            line2_text.size = 14
            line1_text.update()
            line2_text.update()
            return
        
        # 找到当前播放的歌词行索引
        current_index = -1
        for i, (time_sec, text) in enumerate(lyrics_list):
            if position_sec >= time_sec:
                current_index = i
            else:
                break
        
        if current_index >= 0:
            current_text = lyrics_list[current_index][1]
            
            if current_index + 1 < len(lyrics_list):
                next_text = lyrics_list[current_index + 1][1]
                # 第一行：当前歌词（高亮）
                line1_text.value = f"{current_text}"
                line1_text.color = ft.Colors.BLUE_700
                line1_text.weight = ft.FontWeight.BOLD
                line1_text.size = 16
                # 第二行：下一句歌词（普通）
                line2_text.value = next_text
                line2_text.color = ft.Colors.GREY_600
                line2_text.weight = ft.FontWeight.NORMAL
                line2_text.size = 14
            else:
                # 没有下一句，只显示当前歌词
                line1_text.value = f"{current_text}"
                line1_text.color = ft.Colors.BLUE_700
                line1_text.weight = ft.FontWeight.BOLD
                line1_text.size = 16
                line2_text.value = ""
            
            line1_text.update()
            line2_text.update()
    
    def show_fullscreen_lyrics():
        """显示全屏歌词（当前歌词永远居中 - 带闪烁标题）"""
        global lyrics_fullscreen_container, auto_scroll_task, current_lyrics, current_position_sec, current_playing_event_id, events
        
        # 获取当前播放的歌曲名称
        song_title = "歌词"
        if current_playing_event_id and current_playing_event_id in events:
            event = events[current_playing_event_id]
            if event.sound_file and os.path.exists(event.sound_file):
                base_name = os.path.basename(event.sound_file)
                song_title = os.path.splitext(base_name)[0]
        else:
            if current_music_file and os.path.exists(current_music_file):
                base_name = os.path.basename(current_music_file)
                song_title = os.path.splitext(base_name)[0]
        
        # 创建播放/暂停按钮
        play_button = ft.IconButton(
            icon=ft.Icons.PAUSE if current_music_state == "playing" else ft.Icons.PLAY_ARROW,
            icon_size=30,
        )
        
        def on_play_button_click(e):
            global current_audio, current_music_state
            if not current_audio:
                show_snack_bar("没有正在播放的音乐")
                return
            
            if current_music_state == "playing":
                asyncio.create_task(current_audio.pause())
                play_button.icon = ft.Icons.PLAY_ARROW
            elif current_music_state == "paused":
                asyncio.create_task(current_audio.resume())
                play_button.icon = ft.Icons.PAUSE
            play_button.update()
            page.update()
        
        play_button.on_click = on_play_button_click
        
        # ========== 创建闪烁标题文本 ==========
        title_text = ft.Text(
            f"{song_title}",
            size=18,
            weight=ft.FontWeight.BOLD,
            color=ft.Colors.BLUE_700,
            overflow=ft.TextOverflow.ELLIPSIS,
        )
        
        # ========== 闪烁动画任务 ==========
        flash_task = None
        flash_colors = [
            ft.Colors.BLUE_700,
            ft.Colors.RED_700,
            ft.Colors.GREEN_700,
            ft.Colors.PURPLE_700,
            ft.Colors.ORANGE_700,
            ft.Colors.PINK_700,
            ft.Colors.TEAL_700,
            ft.Colors.INDIGO_700,
        ]
        color_index = 0
        
        async def flash_title():
            """标题闪烁动画 - 只变化颜色"""
            nonlocal color_index
            while True:
                if not lyrics_fullscreen_container or lyrics_fullscreen_container not in page.overlay:
                    break
                try:
                    # 切换到下一个颜色
                    color_index = (color_index + 1) % len(flash_colors)
                    title_text.color = flash_colors[color_index]
                    title_text.update()
                    page.update()
                    await asyncio.sleep(0.5)  # 每0.5秒变化一次
                except Exception as e:
                    print(f"[闪烁] 错误: {e}")
                    break
        
        # 创建 ListView
        lyrics_list_view = ft.ListView(
            spacing=10,
            padding=20,
            auto_scroll=False,
        )
        
        # ========== 计算可视区域高度 ==========
        try:
            if hasattr(page, 'window_height') and page.window_height:
                #import platform
                is_android = platform.system() == "Linux"
                
                if is_android:
                    # 手机：屏幕高度 - 状态栏(约35px) - 顶部栏(约80px) + 微调
                    # 手机状态栏占用了屏幕高度，所以需要少减一些
                    list_view_height = page.window_height - 240  # 从180改为240
                else:
                    # Windows：窗口高度 - 顶部栏
                    list_view_height = page.window_height - 180
            else:
                list_view_height = 500
        except:
            list_view_height = 500

        print(f"[全屏歌词] 平台: {platform.system()}")
        print(f"[全屏歌词] 窗口高度: {page.window_height if hasattr(page, 'window_height') else '未知'}")
        print(f"[全屏歌词] ListView 高度: {list_view_height}")
        
        item_height = 50
        padding_count = int((list_view_height / 2) / item_height) + 2
        
        print(f"[歌词填充] 可视高度: {list_view_height}")
        
        # ========== 存储所有歌词项 ==========
        lyric_items = []
        
        # 顶部空白填充
        for idx in range(padding_count):
            empty_item = ft.Container(
                content=ft.Text("", size=16),
                height=item_height,
            )
            lyrics_list_view.controls.append(empty_item)
        
        # 实际歌词
        for i, (time_sec, text) in enumerate(current_lyrics):
            lyric_item = ft.Container(
                content=ft.Text(
                    text,
                    size=16,
                    color=ft.Colors.GREY_700,
                    text_align=ft.TextAlign.CENTER,
                ),
                padding=8,
                height=item_height,
            )
            lyric_items.append(lyric_item)
            lyrics_list_view.controls.append(lyric_item)
        
        # 底部空白填充
        for idx in range(padding_count):
            empty_item = ft.Container(
                content=ft.Text("", size=16),
                height=item_height,
            )
            lyrics_list_view.controls.append(empty_item)
        
        # 创建全屏容器
        lyrics_fullscreen_container = ft.Container(
            content=ft.Column([
                ft.Container(height=10),
                ft.Container(
                    content=ft.Row([
                        ft.IconButton(
                            icon=ft.Icons.ARROW_BACK,
                            icon_size=30,
                            on_click=lambda e: close_fullscreen_lyrics(user_closed=True)  # 标记为用户关闭
                        ),
                        title_text,  # 使用闪烁标题
                        play_button,
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                    padding=10,
                ),
                ft.Divider(),
                ft.Container(
                    content=lyrics_list_view,
                    expand=True,
                ),
            ]),
            bgcolor=ft.Colors.WHITE,
            left=0,
            top=0,
            right=0,
            bottom=0,
            on_click=lambda e: close_fullscreen_lyrics(user_closed=True),  # 标记为用户关闭
        )
        
        lyrics_fullscreen_container.data = {
            'lyric_items': lyric_items,
            'list_view': lyrics_list_view,
            'padding_count': padding_count,
            'item_height': item_height,
            'list_view_height': list_view_height,
            'is_scrolling': False,
            'actual_item_height': None,
            'measured': False,
            'title_text': title_text,  # 保存标题引用
            'flash_task': None,
        }
        
        # ========== 测量实际歌词项高度 ==========
        async def measure_item_height():
            """测量第一个歌词项的实际高度"""
            try:
                await asyncio.sleep(0.3)
                if lyric_items and len(lyric_items) > 0:
                    first_item = lyric_items[0]
                    if hasattr(first_item, 'height') and first_item.height:
                        actual_height = first_item.height + 10
                        if actual_height > 0:
                            lyrics_fullscreen_container.data['actual_item_height'] = actual_height
                            lyrics_fullscreen_container.data['measured'] = True
                            print(f"[测量] 实际歌词项高度: {actual_height}")
                            return True
            except Exception as e:
                print(f"[测量] 测量失败: {e}")
            return False
        
        # ========== 滚动到指定歌词并居中 ==========
        async def scroll_to_lyric_index(real_index, duration=300):
            if not current_lyrics or real_index < 0 or real_index >= len(current_lyrics):
                return
            
            if lyrics_fullscreen_container.data.get('is_scrolling', False):
                return
            
            padding_count = lyrics_fullscreen_container.data['padding_count']
            list_view_height = lyrics_fullscreen_container.data['list_view_height']
            
            actual_item_height = lyrics_fullscreen_container.data.get('actual_item_height')
            if actual_item_height and actual_item_height > 0:
                item_height = actual_item_height
            else:
                item_height = 55
            
            actual_index = real_index + padding_count
            total_items = len(current_lyrics) + padding_count * 2
            
            target_offset = actual_index * item_height - (list_view_height / 2) + (item_height / 2)
            
            if target_offset < 0:
                target_offset = 0
            max_offset = total_items * item_height - list_view_height
            if target_offset > max_offset:
                target_offset = max_offset
            
            #print(f"[歌词滚动] real_index={real_index}, item_height={item_height:.1f}")
            #print(f"[歌词滚动] target_offset={target_offset:.1f}, max_offset={max_offset:.1f}")
            
            lyrics_fullscreen_container.data['is_scrolling'] = True
            
            try:
                list_view = lyrics_fullscreen_container.data['list_view']
                await list_view.scroll_to(
                    offset=target_offset,
                    duration=duration,
                    curve=ft.AnimationCurve.EASE_IN_OUT,
                )
                await asyncio.sleep(duration / 1000 + 0.05)
            finally:
                lyrics_fullscreen_container.data['is_scrolling'] = False
            
            page.update()
        
        # ========== 更新歌词高亮 ==========
        def update_lyric_highlight(real_index):
            if not lyrics_fullscreen_container or not lyrics_fullscreen_container.data:
                return
            
            lyric_items = lyrics_fullscreen_container.data['lyric_items']
            for i, item in enumerate(lyric_items):
                if i == real_index:
                    if hasattr(item.content, 'color'):
                        item.content.color = ft.Colors.BLUE_700
                        item.content.weight = ft.FontWeight.BOLD
                        item.content.size = 18
                    item.bgcolor = ft.Colors.BLUE_50
                else:
                    if hasattr(item.content, 'color'):
                        item.content.color = ft.Colors.GREY_700
                        item.content.weight = ft.FontWeight.NORMAL
                        item.content.size = 16
                    item.bgcolor = None
            page.update()
        
        # ========== 立即滚动到当前歌词位置 ==========
        async def scroll_to_current_immediately():
            if not current_lyrics:
                return
            
            await measure_item_height()
            
            if current_position_sec <= 0:
                current_index = 0
            else:
                current_index = -1
                for i, (time_sec, text) in enumerate(current_lyrics):
                    if current_position_sec >= time_sec:
                        current_index = i
                    else:
                        break
                if current_index < 0:
                    current_index = 0
            
            update_lyric_highlight(current_index)
            await scroll_to_lyric_index(current_index, duration=0)
            
            # ========== 启动闪烁动画 ==========
            flash_task = asyncio.create_task(flash_title())
            lyrics_fullscreen_container.data['flash_task'] = flash_task
        
        # ========== 自动滚动 ==========
        async def auto_scroll_to_current():
            last_index = -1
            await asyncio.sleep(0.5)
            
            while lyrics_fullscreen_container and lyrics_fullscreen_container in page.overlay:
                await asyncio.sleep(0.2)
                
                if current_position_sec >= 0 and current_lyrics:
                    current_index = -1
                    for i, (time_sec, text) in enumerate(current_lyrics):
                        if current_position_sec >= time_sec:
                            current_index = i
                        else:
                            break
                    
                    if current_index < 0:
                        current_index = 0
                    
                    if current_index >= 0 and current_index != last_index:
                        last_index = current_index
                        update_lyric_highlight(current_index)
                        await scroll_to_lyric_index(current_index, duration=300)
        
        if auto_scroll_task:
            auto_scroll_task.cancel()
        auto_scroll_task = asyncio.create_task(auto_scroll_to_current())
        
        page.overlay.append(lyrics_fullscreen_container)
        page.update()
        
        asyncio.create_task(scroll_to_current_immediately())


    def toggle_music_from_fullscreen():
        """从全屏歌词页面控制音乐暂停/继续"""
        global current_audio, current_music_state, lyrics_fullscreen_container
        
        if not current_audio:
            show_snack_bar("没有正在播放的音乐")
            return
        
        if current_music_state == "playing":
            # 正在播放 -> 暂停
            asyncio.create_task(current_audio.pause())
            # 更新按钮图标（通过重新创建全屏容器或直接修改）
            if lyrics_fullscreen_container and lyrics_fullscreen_container in page.overlay:
                try:
                    # 方法：找到顶部栏的播放按钮并更新
                    top_bar = lyrics_fullscreen_container.content.controls[1]  # 索引1是顶部栏容器
                    if top_bar and hasattr(top_bar, 'content') and top_bar.content:
                        row = top_bar.content
                        if len(row.controls) > 2:
                            play_button = row.controls[2]
                            play_button.icon = ft.Icons.PLAY_ARROW
                            page.update()
                except Exception as e:
                    print(f"更新按钮图标失败: {e}")
        elif current_music_state == "paused":
            # 已暂停 -> 继续播放
            asyncio.create_task(current_audio.resume())
            # 更新按钮图标
            if lyrics_fullscreen_container and lyrics_fullscreen_container in page.overlay:
                try:
                    top_bar = lyrics_fullscreen_container.content.controls[1]
                    if top_bar and hasattr(top_bar, 'content') and top_bar.content:
                        row = top_bar.content
                        if len(row.controls) > 2:
                            play_button = row.controls[2]
                            play_button.icon = ft.Icons.PAUSE
                            page.update()
                except Exception as e:
                    print(f"更新按钮图标失败: {e}")
    
    def close_fullscreen_lyrics(user_closed=False):
        """关闭全屏歌词
        Args:
            user_closed: 是否由用户手动关闭（True表示用户点击关闭，False表示程序自动关闭）
        """
        global lyrics_fullscreen_container, auto_scroll_task, auto_fullscreen_lyrics
        
        if lyrics_fullscreen_container and lyrics_fullscreen_container in page.overlay:
            # 停止闪烁动画
            if lyrics_fullscreen_container.data and lyrics_fullscreen_container.data.get('flash_task'):
                try:
                    lyrics_fullscreen_container.data['flash_task'].cancel()
                except:
                    pass
            
            # 如果是由用户手动关闭，重置自动打开标志
            if user_closed:
                auto_fullscreen_lyrics = False
                print("[自动全屏] 用户手动关闭全屏歌词，取消自动打开")
            
            if hasattr(lyrics_fullscreen_container, 'data'):
                lyrics_fullscreen_container.data = None
            page.overlay.remove(lyrics_fullscreen_container)
            lyrics_fullscreen_container = None
            if auto_scroll_task:
                auto_scroll_task.cancel()
                auto_scroll_task = None
            page.update()
    
    def create_lyrics_display():
        """创建可点击的歌词显示控件"""
        # 第一行歌词（高亮）- 修改初始值
        line1_text = ft.Text(
            value="🎵 未播放",
            size=14,
            weight=ft.FontWeight.BOLD,
            color=ft.Colors.GREY_600,
            text_align=ft.TextAlign.CENTER,
        )
        
        # 第二行歌词（普通）
        line2_text = ft.Text(
            value="",
            size=12,
            color=ft.Colors.GREY_500,
            text_align=ft.TextAlign.CENTER,
        )
        
        # 添加点击事件
        def on_lyrics_click(e):
            global current_lyrics, current_position_sec, current_music_state
            if current_music_state != "stopped" and current_lyrics and len(current_lyrics) > 0:
                # 用户手动打开，重置自动标志
                global auto_fullscreen_lyrics
                auto_fullscreen_lyrics = False
                show_fullscreen_lyrics()
            else:
                show_snack_bar("没有歌词数据或音乐未播放")
        
        # 创建一个带墨水效果的可点击容器
        lyrics_text_container = ft.Container(
            content=ft.Column([
                line1_text,
                line2_text,
            ], spacing=3, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            on_click=on_lyrics_click,
            padding=10,
            ink=True,
            border_radius=8,
            bgcolor=ft.Colors.TRANSPARENT,
            width=float("inf"),
            height=78,  # 添加固定高度
        )
        
        return lyrics_text_container, (line1_text, line2_text)

    def on_music_state_change(event_id, state):
        """当音乐状态改变时，刷新UI"""
        global current_playing_event_id, current_music_state  # 添加这行
        
        print(f"[on_music_state_change] 收到回调 - event_id: {event_id}, state: {state}")
        
        # 更新全局状态
        current_playing_event_id = event_id
        current_music_state = state
        
        # 更新顶部当前播放信息
        try:
            update_current_playing_info()
        except Exception as e:
            print(f"更新播放信息失败: {e}")
        # 刷新事件列表，更新卡片中的状态
        print(f"[on_music_state_change] 准备刷新事件列表")
        refresh_events_list()
        # 强制更新页面
        page.update()
        print(f"[on_music_state_change] 刷新完成")
    
    # 在音乐控制区域添加一个按钮来显示/打开歌词文件
    def show_lyrics_file_location(e):
        """显示歌词文件位置（支持 Windows 和 Android）"""
        if current_music_file and os.path.exists(current_music_file):
            lrc_path = os.path.splitext(current_music_file)[0] + ".lrc"
            abs_lrc_path = os.path.abspath(lrc_path)
            
            if os.path.exists(abs_lrc_path):
                print(f"[歌词文件] 完整路径: {abs_lrc_path}")
                
                if platform.system() == "Windows":
                    # Windows 可以打开文件夹
                    try:
                        open_file_location(current_music_file)
                        show_snack_bar(f"已打开文件夹，歌词文件: {abs_lrc_path}")
                    except Exception as ex:
                        show_snack_bar(f"无法打开文件夹: {str(ex)}")
                        # 降级：只显示路径
                        show_snack_bar(f"歌词文件路径: {abs_lrc_path}")
                elif platform.system() == "Linux":
                    # Android 平台：复制路径到剪贴板
                    try:
                        page.set_clipboard(abs_lrc_path)
                        show_snack_bar(f"📱 歌词路径已复制到剪贴板")
                        print(f"[Android] 歌词文件路径: {abs_lrc_path}")
                    except:
                        show_snack_bar(f"📱 歌词文件路径: {abs_lrc_path}")
                else:
                    show_snack_bar(f"歌词文件路径: {abs_lrc_path}")
            else:
                show_snack_bar(f"歌词文件不存在: {abs_lrc_path}")
        else:
            show_snack_bar("没有正在播放的音乐")

    # 添加一个辅助函数来打开文件所在文件夹
    def open_file_location(file_path):
        """打开文件所在的文件夹（支持多平台）"""
        import subprocess
        import platform
        
        try:
            abs_path = os.path.abspath(file_path)
            folder_path = os.path.dirname(abs_path)
            
            print(f"[打开文件夹] 尝试打开: {folder_path}")
            
            system = platform.system()
            
            if system == "Windows":
                os.startfile(folder_path)
                print(f"[打开文件夹] Windows 打开成功")
            elif system == "Darwin":  # macOS
                subprocess.Popen(['open', folder_path])
                print(f"[打开文件夹] macOS 打开成功")
            elif system == "Linux":
                # Android 不支持直接打开文件管理器，只打印
                print(f"[打开文件夹] Android 不支持自动打开，路径: {folder_path}")
                # 不抛出异常，只打印
            else:
                print(f"[打开文件夹] 未知平台: {system}")
        except Exception as e:
            print(f"[打开文件夹] 打开失败: {e}")
            # 不抛出异常，避免影响主流程

    # 添加一个带锁的播放函数
    def play_music_with_lock(sound_file, loop=False, event_name=None, event_id=None):
        """带线程锁的播放函数，防止多个播放请求同时执行"""
        global current_audio, is_playing, current_music_file, current_duration, current_lyrics
        print(f"[play_music_with_lock] 接收到参数 - event_name: {event_name}, event_id: {event_id}, loop: {loop}")
        with is_playing_lock:
            # 明确使用关键字参数传递
            play_music(sound_file=sound_file, loop=loop, event_name=event_name, event_id=event_id)

    def play_music(sound_file, loop=False, event_name=None, event_id=None):
        global current_audio, is_playing, current_music_file, current_duration, current_lyrics
        global current_playing_event_id, current_music_state, music_state_update_callback
        global current_lyrics  # 添加 current_lyrics
        global music_section_container, playback_buttons
        
        print(f"[play_music] 接收到参数 - event_name: {event_name}, event_id: {event_id},sound_file: {sound_file}")

        # ========== 关键修复：保存原始参数供循环使用 ==========
        original_event_name = event_name
        original_event_id = event_id
        original_sound_file = sound_file
        original_loop = loop

        # 如果回调为 None，尝试重新设置
        if music_state_update_callback is None:
            print("[play_music] 回调为 None，尝试重新设置")
            set_music_state_update_callback()
        
        if not sound_file or not os.path.exists(sound_file):
            show_snack_bar("音乐文件不存在")
            return
        
        # 转换为绝对路径
        abs_sound_file = os.path.abspath(sound_file)
        abs_lrc_path = os.path.splitext(abs_sound_file)[0] + ".lrc"
        
        # ========== 添加歌词路径打印（绝对路径） ==========
        #print(f"[歌词路径] ========================================")
        #print(f"[歌词路径] 音乐文件绝对路径: {abs_sound_file}")
        #print(f"[歌词路径] 歌词文件绝对路径: {abs_lrc_path}")
        #print(f"[歌词路径] 歌词文件是否存在: {os.path.exists(abs_lrc_path)}")
        
        # 打印所在目录
        music_dir = os.path.dirname(abs_sound_file)
        #print(f"[歌词路径] 音乐文件所在目录: {music_dir}")
        
        # 列出目录下的所有歌词文件
        if os.path.exists(music_dir):
            lrc_files = [f for f in os.listdir(music_dir) if f.endswith('.lrc')]
            if lrc_files:
                #print(f"[歌词路径] 目录下找到的歌词文件: {lrc_files}")
                pass
            else:
                #print(f"[歌词路径] 目录下没有找到任何 .lrc 文件")
                pass
        
        #print(f"[歌词路径] ========================================")
        
        # 也可以显示在界面上（可选）
        #show_snack_bar(f"歌词路径: {abs_lrc_path}")

        # ========== 注释掉自动打开文件夹的代码（这会导致 Android 崩溃） ==========
        # open_file_location(sound_file)  # 不要自动打开，让用户手动点击按钮打开
        
        # 创建实例时传入 page 和 show_snack_bar
        lyrics_downloader = LyricsDownloader(page=page, show_snack_bar=show_snack_bar)
        # 从文件名提取歌曲名和歌手名用于搜索歌词
        base_name = os.path.basename(sound_file)
        name_without_ext = os.path.splitext(base_name)[0]
        song_name_for_search = None
        artist_for_search = None
        
        if " - " in name_without_ext:
            parts = name_without_ext.split(" - ")
            if len(parts) >= 2:
                # 假设第一部分是歌曲名，第二部分是歌手名
                song_name_for_search = parts[0].strip()
                artist_for_search = parts[1].strip()
                print(f"[播放] 从文件名解析: 歌曲={song_name_for_search}, 歌手={artist_for_search}")
        
        # 使用解析出的歌曲信息搜索歌词
        if song_name_for_search:
            lyrics_downloader.download_lyrics_for_music(sound_file, song_name_for_search, artist_for_search)
        else:
            lyrics_downloader.download_lyrics_for_music(sound_file)
        
        # 完全清理旧的音频控件
        if current_audio:
            try:
                # 先暂停
                async def cleanup():
                    try:
                        await current_audio.pause()
                    except:
                        pass
                asyncio.create_task(cleanup())
                
                # 移除控件
                if current_audio in page.services:
                    page.services.remove(current_audio)
                if current_audio in page.overlay:
                    page.overlay.remove(current_audio)
                page.update()
            except Exception as e:
                print(f"清理旧控件出错: {e}")
            finally:
                current_audio = None
                is_playing = False
        
        # 等待一下确保清理完成
        time.sleep(0.1)
        
        # 获取歌词
        lyrics_downloader = LyricsDownloader()
        lyrics_downloader.download_lyrics_for_music(sound_file)
        #current_lyrics = parse_lyrics_to_lines(sound_file)  # 使用新的解析函数
        # 获取歌词，-0.3表示提前0.3秒显示
        current_lyrics = parse_lyrics_to_lines(sound_file, offset=-0.3)
        #print(f"[play_music] 解析后 current_lyrics 长度: {len(current_lyrics)}")  # 添加这行
        #print(f"[play_music] current_lyrics 内存地址: {id(current_lyrics)}")  # 添加这行
        
        # 获取时长
        try:
            from mutagen.mp3 import MP3
            current_duration = MP3(sound_file).info.length
            #print(f"[播放] 音乐时长: {current_duration} 秒")
        except:
            # 尝试其他格式
            try:
                from mutagen.wave import WAVE
                current_duration = WAVE(sound_file).info.length
                #print(f"[播放] 音乐时长(WAVE): {current_duration} 秒")
            except:
                current_duration = 180  # 默认3分钟
                print(f"[播放] 无法获取时长，使用默认值: {current_duration} 秒")

        # ========== 获取音乐显示名称 ==========
        music_display_name = os.path.splitext(os.path.basename(sound_file))[0]

        #show_snack_bar(f"播放音乐: {sound_file}")
        current_music_file = sound_file

        # 记录当前播放的事件ID（可能为None）
        current_playing_event_id = event_id
        current_music_state = "playing"
        
        progress_slider.value = 0
        progress_text.value = f"0:00 / {format_time(current_duration)}"

        # 重置歌词显示
        line1_text, line2_text = lyrics_display_widgets
        if current_lyrics and len(current_lyrics) > 0:
            # 有歌词，显示第一句
            line1_text.value = f"🎵 {current_lyrics[0][1]}"
            line1_text.color = ft.Colors.BLUE_700
            line1_text.weight = ft.FontWeight.BOLD
            line1_text.size = 16
            if len(current_lyrics) > 1:
                line2_text.value = current_lyrics[1][1]
                line2_text.color = ft.Colors.GREY_600
                line2_text.weight = ft.FontWeight.NORMAL
                line2_text.size = 14
            else:
                line2_text.value = ""
        else:
            # 没有歌词，显示友好提示
            line1_text.value = "📝 本地无歌词或未在线搜索到歌词"
            line1_text.color = ft.Colors.GREY_600
            line1_text.weight = ft.FontWeight.NORMAL
            line1_text.size = 16
            line2_text.value = "💡 提示：可以手动添加 .lrc 歌词文件到音乐同目录"
            line2_text.color = ft.Colors.GREY_500
            line2_text.weight = ft.FontWeight.NORMAL
            line2_text.size = 14
        line1_text.update()
        line2_text.update()

        progress_slider.update()
        progress_text.update()
        
        # 添加一个标志防止重复播放
        is_playing_new = False

        # ========== 添加监控任务变量 ==========
        monitor_task = None
        # 添加一个变量来存储当前位置（局部变量）
        local_position_sec = 0  # 改为局部变量
        
        # 记录当前播放的事件ID
        current_playing_event_id = event_id
        current_music_state = "playing"

        # 强制显示音乐区域（试听模式也要显示）
        if music_section_container:
            music_section_container.visible = True
            music_section_container.update()
            #print("[play_music] 已显示音乐区域")
        
        if playback_buttons:
            playback_buttons.visible = True
            playback_buttons.update()
            #print("[play_music] 已显示播放按钮")
        
        # 设置状态（即使是试听，也要设置状态）
        current_playing_event_id = event_id
        current_music_state = "playing"
        current_music_file = sound_file

        # 立即更新UI显示
        try:
            update_current_playing_info()
        except Exception as e:
            print(f"更新播放信息失败: {e}")

        # 通知UI更新
        print(f"[play_music] music_state_update_callback is {music_state_update_callback}")
        if music_state_update_callback:
            music_state_update_callback(event_id, "playing")
        else:
            print("[play_music] 警告: music_state_update_callback 为 None!")

        # ========== 修改 on_state_change，使用闭包变量 ==========
        def on_state_change(e):
            global current_audio, is_playing, current_playing_event_id, current_music_state
            global auto_fullscreen_lyrics  # 添加这行
            nonlocal is_playing_new, monitor_task, local_position_sec  # 使用 local_position_sec
            
            print(f"[播放状态] 状态改变: {e.state}")
            
            if e.state == AudioState.PLAYING:
                print("[播放状态] ✓ 音乐开始播放")
                is_playing = True
                is_playing_new = True
                current_music_state = "playing"
                if music_state_update_callback and current_playing_event_id:
                    music_state_update_callback(current_playing_event_id, "playing")

                # ========== 启动监控任务（播放开始时） ==========
                if monitor_task:
                    monitor_task.cancel()
                
                # ========== 修正后的监控代码 ==========
                async def check_end():
                    while True:
                        await asyncio.sleep(0.5)
                        try:
                            # 使用 local_position_sec 更新的位置值
                            if local_position_sec >= current_duration - 0.3:
                                # 播放完成，重置进度条
                                progress_slider.value = 0
                                progress_text.value = f"0:00 / {format_time(current_duration)}"
                                progress_slider.update()
                                progress_text.update()
                                print("监控检测到播放结束，进度条归零")
                                break
                        except Exception as ex:
                            print(f"监控错误: {ex}")
                            break
                
                monitor_task = asyncio.create_task(check_end())
                # ========== 监控代码结束 ==========

                if event_name:
                    music_name = get_music_name_from_file(sound_file) or os.path.basename(sound_file)
                    # ========== 先检查是否已发送 ==========
                    notification_key = f"{event_name}_{music_display_name}_{datetime.now().strftime('%Y%m%d%H%M')}"
                    if notification_key not in sent_music_notifications:
                        update_music_notification(f"{event_name} - {music_display_name}", is_playing=True)

                # ========== 新增：如果 auto_fullscreen_lyrics 为 True，自动打开全屏歌词 ==========
                if auto_fullscreen_lyrics:
                    print("[自动全屏] 检测到需要自动打开全屏歌词")
                    # 延迟一下，确保音乐已经开始播放
                    asyncio.create_task(auto_open_fullscreen_lyrics())
            
            elif e.state == AudioState.COMPLETED:
                print("[播放状态] 音乐播放完成")
                is_playing = False
                current_audio = None

                # ========== 1. 先从 card_duration_texts 中删除该卡片的引用 ==========
                if original_event_id and original_event_id in card_duration_texts:
                    try:
                        # 直接删除引用，让卡片无法被更新
                        del card_duration_texts[original_event_id]
                        print(f"[播放完成] 已删除卡片引用: {original_event_id}")
                    except Exception as e:
                        print(f"[播放完成] 删除引用失败: {e}")
                
                # ========== 2. 再设置状态 ==========
                current_music_state = "stopped"
                
                # ========== 3. 然后刷新整个列表，让卡片显示总时长 ==========
                # 延迟一点点刷新，确保状态已经更新
                async def refresh_after_complete():
                    await asyncio.sleep(0.1)
                    # 刷新当前视图，重新创建卡片
                    refresh_current_view_by_state()
                
                asyncio.create_task(refresh_after_complete())

                # 如果是试听模式，清除事件ID
                if current_playing_event_id is None or current_playing_event_id not in events:
                    current_playing_event_id = None

                # 重置进度条
                progress_slider.value = 0
                #progress_text.value = f"0:00 / {format_time(current_duration)}"
                progress_text.value = "0:00 / 0:00"

                # 重置歌词
                line1_text, line2_text = lyrics_display_widgets
                line1_text.value = "🎵 未播放"
                line1_text.color = ft.Colors.GREY_600
                line2_text.value = ""
                line1_text.update()
                line2_text.update()

                progress_slider.update()
                progress_text.update()

                # ========== 关键：检测全屏歌词是否打开，如果打开则记录状态 ==========
                if lyrics_fullscreen_container and lyrics_fullscreen_container in page.overlay:
                    print("[自动全屏] 检测到全屏歌词正在显示，将在循环播放时自动重新打开")
                    auto_fullscreen_lyrics = True
                    # 关闭全屏歌词
                    close_fullscreen_lyrics()
                else:
                    auto_fullscreen_lyrics = False

                # ========== 关键：调用更新UI函数，隐藏音乐区域 ==========
                update_current_playing_info()
                
                # 取消通知
                cancel_notification(MUSIC_NOTIFICATION_ID)
                
                # 刷新页面
                page.update()

                # ========== 动态检查循环状态 ==========
                should_loop = event_loop_states.get(original_event_id, False)
                print(f"[播放状态] 检查循环状态: original_loop={original_loop}, current_loop={should_loop}")

                if should_loop and is_playing_new:
                    print(f"[播放状态] 循环播放模式，准备重新播放 (event_id={original_event_id})")
                    play_music(
                        sound_file=original_sound_file, 
                        loop=should_loop,
                        event_name=original_event_name, 
                        event_id=original_event_id
                    )
                else:
                    # 非循环模式：清除循环状态
                    if original_event_id:
                        event_loop_states[original_event_id] = False
                        print(f"[播放完成] 清除事件 {original_event_id} 的循环状态")
                        
                    # 非循环模式才重置状态
                    current_music_state = "stopped"
                    if music_state_update_callback and current_playing_event_id:
                        music_state_update_callback(current_playing_event_id, "stopped")
                    current_playing_event_id = None
                    auto_fullscreen_lyrics = False  # 重置标志

                    # 额外重置一次，确保歌词显示正确
                    # 重置歌词
                    line1_text, line2_text = lyrics_display_widgets
                    line1_text.value = "🎵 未播放"
                    line1_text.color = ft.Colors.GREY_600
                    line2_text.value = ""
                    line1_text.update()
                    line2_text.update()

                cancel_notification(MUSIC_NOTIFICATION_ID)
            elif e.state == AudioState.PAUSED:
                print("[播放状态] 音乐已暂停")
                current_music_state = "paused"
                if music_state_update_callback and current_playing_event_id:
                    music_state_update_callback(current_playing_event_id, "paused")

                if event_name:
                    update_music_notification(f"{event_name} - {music_display_name}", is_playing=False)

            
            elif e.state == AudioState.STOPPED:
                print("[播放状态] 音乐已停止")
                current_music_state = "stopped"
                if music_state_update_callback and current_playing_event_id:
                    music_state_update_callback(current_playing_event_id, "stopped")
                current_playing_event_id = None
                auto_fullscreen_lyrics = False  # 重置标志

                cancel_notification(MUSIC_NOTIFICATION_ID)

            else:
                print(f"[播放状态] 其他状态: {e.state}")

        # ========== 新增：自动打开全屏歌词的函数 ==========
        async def auto_open_fullscreen_lyrics():
            """自动打开全屏歌词（延迟执行，确保音乐已开始播放）"""
            await asyncio.sleep(0.3)  # 等待音乐开始播放
            if current_music_state == "playing" and current_lyrics and len(current_lyrics) > 0:
                print("[自动全屏] 执行自动打开全屏歌词")
                show_fullscreen_lyrics()
            else:
                print(f"[自动全屏] 条件不满足: state={current_music_state}, lyrics={len(current_lyrics) if current_lyrics else 0}")
        
        def on_position_change(e):
            nonlocal local_position_sec
            global last_card_update_time, current_position_sec,card_duration_texts
            
            if e.position is not None:
                local_position_sec = e.position / 1000

                # ========== 如果音乐已停止，不更新卡片时长 ==========
                if current_music_state == "stopped":
                    # 更新全局位置（但不更新卡片）
                    current_position_sec = local_position_sec
                    return

                # ========== 获取总时长（使用 current_duration 数字） ==========
                total_duration_sec = current_duration  # 这是数字（秒）
                
                if current_duration > 0:
                    progress = (local_position_sec / current_duration) * 100
                    progress = max(0, min(100, progress))
                    progress_slider.value = progress
                    progress_slider.update()
                    
                    # 如果进度文本可见，更新时间
                    if progress_text.visible:
                        progress_text.value = f"{format_time(local_position_sec)} / {format_time(current_duration)}"
                        progress_text.update()
                
                # 更新全局位置
                current_position_sec = local_position_sec
                
                # 更新歌词显示
                if current_lyrics:
                    update_lyrics_display(local_position_sec, current_lyrics, lyrics_display_widgets, is_fullscreen=False)

                # ========== 只有当音乐正在播放或暂停时，才更新卡片时长 ==========
                # ========== 更新卡片时长 ==========
                if current_music_state in ["playing", "paused"]:
                    if current_playing_event_id and current_playing_event_id in card_duration_texts:
                        if total_duration_sec and total_duration_sec > 0 and local_position_sec > 0:
                            current_pos_str = format_time(local_position_sec)
                            total_duration_str = format_time(total_duration_sec)
                            duration_text = card_duration_texts[current_playing_event_id]
                            
                            try:
                                if hasattr(duration_text, 'page') and duration_text.page is not None:
                                    # ========== 如果当前位置接近总时长（差0.5秒内），显示总时长 ==========
                                    if local_position_sec >= total_duration_sec - 0.5:
                                        duration_text.value = f"⏱️ {total_duration_str}"
                                        duration_text.color = ft.Colors.GREY_500
                                        duration_text.weight = ft.FontWeight.NORMAL
                                    else:
                                        duration_text.value = f"⏱️ {current_pos_str} / {total_duration_str}"
                                        duration_text.color = ft.Colors.BLUE_700
                                        duration_text.weight = ft.FontWeight.BOLD
                                    duration_text.update()
                                else:
                                    if current_playing_event_id in card_duration_texts:
                                        del card_duration_texts[current_playing_event_id]
                            except Exception as ex:
                                print(f"[时长更新] 更新失败: {ex}")
                                if current_playing_event_id in card_duration_texts:
                                    del card_duration_texts[current_playing_event_id]
                # 更新全局位置
                current_position_sec = local_position_sec

                # 更新歌词显示
                if current_lyrics:
                    update_lyrics_display(local_position_sec, current_lyrics, lyrics_display_widgets, is_fullscreen=False)
                
                # 如果全屏歌词打开，也更新
                if lyrics_fullscreen_container and lyrics_fullscreen_container in page.overlay:
                    pass
        
        audio = ftaudio.Audio(
            src=sound_file,
            autoplay=True,
            volume=1,
            balance=0,
            on_loaded=lambda _: print("音乐加载完成"),
            on_state_change=on_state_change,
            on_position_change=on_position_change,
        )
        
        page.services.append(audio)
        current_audio = audio
        is_playing = True
        #show_snack_bar(f"正在播放: {os.path.basename(sound_file)}")

    def stop_music():
        global current_audio, is_playing, current_music_file, current_lyrics
        global current_playing_event_id, current_music_state, music_section_container, playback_buttons,card_duration_texts
        global sent_music_notifications
        
        print("停止音乐")
        
        # 添加去重标志
        if hasattr(stop_music, '_is_stopping') and stop_music._is_stopping:
            print("[停止音乐] 已经在停止中，跳过")
            return
        
        stop_music._is_stopping = True

        cancel_notification(MUSIC_NOTIFICATION_ID)
        
        try:
            # 只有当前有音乐播放时才执行停止逻辑
            if current_audio is None and current_music_file is None:
                print("[停止音乐] 没有正在播放的音乐，跳过")
                return
            
            # 保存要清除的事件ID
            clearing_event_id = current_playing_event_id
            
            # 立即清除状态，防止后续回调
            # 清空音乐文件路径（关键）
            current_music_file = None
            current_playing_event_id = None
            current_music_state = "stopped"
            is_playing = False
            current_lyrics = []

            # ========== 关键：调用更新函数来刷新UI ==========
            update_current_playing_info()  # 添加这行

            # 清除循环状态
            if clearing_event_id:
                event_loop_states[clearing_event_id] = False
                print(f"[停止音乐] 清除事件 {clearing_event_id} 的循环状态")
            
            # 异步停止音频
            async def stop_async():
                global current_audio
                try:
                    if current_audio:
                        try:
                            await current_audio.pause()
                        except:
                            pass
                        
                        try:
                            if current_audio in page.services:
                                page.services.remove(current_audio)
                        except:
                            pass
                        
                        try:
                            if current_audio in page.overlay:
                                page.overlay.remove(current_audio)
                        except:
                            pass
                        
                        current_audio = None
                        page.update()
                        
                except Exception as e:
                    print(f"停止音乐出错: {e}")
            
            asyncio.create_task(stop_async())

            # 隐藏音乐区域
            if music_section_container:
                music_section_container.visible = False
                music_section_container.update()
            if playback_buttons:
                playback_buttons.visible = False
                playback_buttons.update()
            
            # 重置状态
            current_music_file = None
            is_playing = False
            current_lyrics = []
            
            # 关闭全屏歌词
            if lyrics_fullscreen_container and lyrics_fullscreen_container in page.overlay:
                close_fullscreen_lyrics()
            
            # 重置UI显示
            try:
                progress_slider.value = 0
                progress_text.value = "0:00 / 0:00"
                
                line1_text, line2_text = lyrics_display_widgets
                line1_text.value = "🎵 未播放"
                line1_text.color = ft.Colors.GREY_600
                line1_text.size = 16
                line2_text.value = ""
                line1_text.update()
                line2_text.update()
                
                progress_slider.update()
                progress_text.update()
                page.update()
                
            except Exception as e:
                print(f"重置UI出错: {e}")
            
            show_snack_bar("音乐已停止")
            
            # ========== 刷新事件列表，但不要再触发 UI 更新 ==========
            # ========== 根据当前视图刷新对应的视图 ==========
            refresh_current_view_by_state()
            #refresh_events_list()

            # ========== 清除卡片时长显示 ==========
            if clearing_event_id and clearing_event_id in card_duration_texts:
                try:
                    duration_text = card_duration_texts[clearing_event_id]
                    if hasattr(duration_text, 'page') and duration_text.page is not None:
                        # 只显示总时长
                        total_duration = get_music_duration_display(current_music_file) if current_music_file else ""
                        duration_text.value = f"⏱️ {total_duration}" if total_duration else ""
                        duration_text.color = ft.Colors.GREY_500
                        duration_text.weight = ft.FontWeight.NORMAL
                        duration_text.update()
                    else:
                        if clearing_event_id in card_duration_texts:
                            del card_duration_texts[clearing_event_id]
                except Exception as e:
                    print(f"[停止音乐] 清除时长显示失败: {e}")
                    if clearing_event_id in card_duration_texts:
                        del card_duration_texts[clearing_event_id]
            
            # 注意：不要再调用 music_state_update_callback，因为我们已经手动设置了UI
            # 如果需要通知其他组件，可以考虑，但会导致重复更新

            sent_music_notifications.clear()
            
        finally:
            stop_music._is_stopping = False

    def refresh_current_view_by_state():
        """根据当前视图刷新对应的视图"""
        global current_view
        
        #print(f"[刷新视图] 当前视图: {current_view}")
        
        if current_view == "all":
            display_all_events()
        elif current_view == "today":
            show_today_events()
        elif current_view == "three_days":
            show_three_days_events()
        elif current_view == "daily":
            show_daily_events()
        elif current_view == "weekly":
            show_weekly_events()
        elif current_view == "birthday":
            show_birthday_events()
        elif current_view == "event":
            show_event_events()
        elif current_view == "once":
            show_once_events()
        elif current_view == "monthly":
            show_monthly_events()
        else:
            refresh_events_list()
            
    def pause_music(e):
        global current_audio, is_playing, current_music_state
        
        if not current_audio:
            show_snack_bar("没有正在播放的音乐")
            return
        
        try:
            if current_music_state == "playing":
                # 正在播放 -> 暂停
                print(f"[暂停/继续] 暂停音乐")
                asyncio.create_task(current_audio.pause())
                # show_snack_bar("⏸️ 音乐已暂停")
            elif current_music_state == "paused":
                # 已暂停 -> 继续播放
                print(f"[暂停/继续] 继续播放音乐")
                asyncio.create_task(current_audio.resume())
                # show_snack_bar("▶️ 音乐继续播放")
            else:
                print(f"[暂停/继续] 当前状态为 {current_music_state}，无法暂停/继续")
                # show_snack_bar(f"当前音乐已停止，请重新播放")

            # 延迟一下刷新当前视图（更新播放状态显示）
            threading.Timer(0.1, refresh_current_view_by_state).start()

        except Exception as ex:
            print(f"暂停/继续失败: {ex}")
            show_snack_bar(f"操作失败: {str(ex)}")

    def get_music_name_from_file(file_path):
        """从文件路径提取音乐名称（返回歌曲名，不是歌手名）"""
        if not file_path or not os.path.exists(file_path):
            return None
        base_name = os.path.basename(file_path)
        name_without_ext = os.path.splitext(base_name)[0]
        
        #print(f"[解析文件名] 原始: {name_without_ext}")
        
        # 如果有" - "分隔符，格式是 "歌曲名 - 歌手名"
        if " - " in name_without_ext:
            parts = name_without_ext.split(" - ")
            if len(parts) >= 2:
                # 返回第一部分作为歌曲名（这才是歌曲名称）
                song_name = parts[0].strip()
                print(f"[解析文件名] 歌曲名: {song_name}")
                return song_name
            return name_without_ext
        return name_without_ext
    
    def get_full_music_name(file_path):
        """获取完整的音乐名称（歌曲名 - 歌手名）"""
        if not file_path or not os.path.exists(file_path):
            return None
        base_name = os.path.basename(file_path)
        return os.path.splitext(base_name)[0]

    def update_event_count():
        count_text.value = f"📊 事件总数: {len(events)}"
        count_text.update()

    def show_events_by_type(event_type):
        """根据类型显示事件"""
        global current_view

        # 安全关闭菜单
        if hasattr(on_date_text_click, 'menu_container'):
            try:
                if on_date_text_click.menu_container in page.overlay:
                    page.overlay.remove(on_date_text_click.menu_container)
            except:
                pass
            on_date_text_click.menu_container = None

        # 更新当前视图
        current_view = event_type
        
        # ========== 更新自定义下拉框的显示 ==========
        update_view_dropdown_display(event_type)
        
        if event_type == "today":
            # 显示今日事件
            current_view = "today"
            #refresh_events_list.view_dropdown.value = "today"
            show_today_events()
            show_bottom_message("📌 已切换到今日事件视图")
        elif event_type == "three_days":
            # 显示3日内事件
            current_view = "three_days"
            #refresh_events_list.view_dropdown.value = "three_days"
            show_three_days_events()
            show_bottom_message("🔔 已切换到预警事件视图")
        elif event_type == "all":
            # 显示全部事件
            current_view = "all"
            #refresh_events_list.view_dropdown.value = "all"
            display_all_events()
            show_bottom_message("📋 已切换到全部事件视图")
        elif event_type == "daily":
            current_view = "daily"
            #refresh_events_list.view_dropdown.value = "daily"
            show_daily_events()
            show_bottom_message("⏰  已切换到每日事件视图")
        elif event_type == "weekly":
            current_view = "weekly"
            #refresh_events_list.view_dropdown.value = "weekly"
            show_weekly_events()
            show_bottom_message("🔁 已切换到每周事件视图")

        page.update()
    
    def show_daily_events():
        """显示每日事件列表"""
        global current_view, events_list, card_duration_texts
        current_view = "daily"

        # ========== 清空旧的卡片引用 ==========
        card_duration_texts.clear()

        events_list.controls.clear()
        
        #print(f"[DEBUG] show_daily_events 被调用, current_view={current_view}")
        #print(f"[show_daily_events] 当前事件总数: {len(events)}")
        daily_events = []
    
        for event in events.values():
            if event.event_type == "daily":
                # 获取最早的提醒时间用于排序
                earliest_time = "23:59"  # 默认最大值
                if event.reminders:
                    times = [r.get("time", "23:59") for r in event.reminders if r.get("enabled")]
                    if times:
                        earliest_time = min(times)  # 取最早的时间
                daily_events.append({
                    "event": event,
                    "sort_time": earliest_time
                })
        
        #print(f"[show_daily_events] 每日事件数量: {len(daily_events)}")

        # 按提醒时间排序
        daily_events.sort(key=lambda x: x["sort_time"])
        
        # ========== 提取事件对象并调用置顶排序 ==========
        event_list = [item["event"] for item in daily_events]
        event_list = get_sorted_events_for_display(event_list)
        
        # ========== 始终显示标题行和下拉框 ==========
        if hasattr(refresh_events_list, 'view_dropdown'):
            title_text = f"⏰ 每日事件 {len(daily_events)} 个" if daily_events else "⏰ 每日事件 0 个"
            events_list.controls.append(ft.Row([
                ft.Text(title_text, size=14, weight=ft.FontWeight.BOLD, expand=True),
                refresh_events_list.view_dropdown,
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN))
            events_list.controls.append(ft.Divider(height=10))
        
        # 然后显示事件内容
        if not daily_events:
            events_list.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Text("✨ 暂无每日事件", size=14, color=ft.Colors.GREEN_700),
                    ], spacing=8, ),
                    padding=20,
                )
            )
        else:
            # ========== 调试：打印排序后的事件列表 ==========
            #print(f"[显示] 排序后事件列表:")
            #for idx, event in enumerate(event_list):
                #is_playing = (event.id == current_playing_event_id and current_music_state in ["playing", "paused"])
                #print(f"  {idx}: {event.name}, is_playing: {is_playing}")
            
            for event in event_list:
                display_event_card(event, is_filter_mode=True)
            
            if events_list.controls and isinstance(events_list.controls[-1], ft.Divider):
                events_list.controls.pop()
        
        page.update()
        #print(f"[show_daily_events] 刷新完成")
    
    def show_weekly_events():
        """显示每周事件列表"""
        global current_view, events_list, card_duration_texts
        current_view = "weekly"

        # ========== 清空旧的卡片引用 ==========
        card_duration_texts.clear()

        events_list.controls.clear()

        weekly_events = []
        today = datetime.now().date()
        
        for event in events.values():
            if event.event_type == "weekly":
                month, day, year, base_year, days_until = event.get_next_date_info()
                weekly_events.append({
                    "event": event,
                    "days_until": days_until
                })
        
        # 按剩余天数排序（由近到远）
        weekly_events.sort(key=lambda x: x["days_until"])

         # ========== 提取事件对象列表并调用排序函数 ==========
        event_list = [item["event"] for item in weekly_events]
        event_list = get_sorted_events_for_display(event_list)

        # ========== 始终显示标题行和下拉框 ==========
        if hasattr(refresh_events_list, 'view_dropdown'):
            title_text = f"🔁 每周事件 {len(weekly_events)} 个" if weekly_events else "🔁 每周事件 0 个"
            events_list.controls.append(ft.Row([
                ft.Text(title_text, size=14, weight=ft.FontWeight.BOLD, expand=True),
                refresh_events_list.view_dropdown,
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN))
            events_list.controls.append(ft.Divider(height=10))
        
        if not weekly_events:
            events_list.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Text("✨ 暂无每周事件", size=14, color=ft.Colors.GREEN_700),
                    ], spacing=8, ),
                    padding=20,
                )
            )
        else:
            for event in event_list:
                display_event_card(event, is_filter_mode=True)

            # 移除最后一个多余的分隔符
            if events_list.controls and isinstance(events_list.controls[-1], ft.Divider):
                events_list.controls.pop()
        
        page.update()

    def show_three_days_events():
        """显示3日内事件列表（预警事件）"""
        global current_view, events_list, card_duration_texts
        current_view = "three_days"

        # ========== 清空旧的卡片引用 ==========
        card_duration_texts.clear()

        events_list.controls.clear()
        
        today = datetime.now().date()
        three_days_events = []
        
        #print(f"[预警事件] 开始收集，当前日期: {today}")
        
        for event in events.values():
            # 跳过每天事件和每周事件
            if event.event_type == "daily" or event.event_type == "weekly":
                continue
            
            month, day, year, base_year, days_until = event.get_next_date_info()
            #print(f"[预警事件] 检查: {event.name}, 类型: {event.event_type}, 剩余天数: {days_until}")
            
            # 一次性事件特殊处理
            if event.repeat_type == "once":
                if event.completed or days_until < 0:
                    #print(f"[预警事件]   - 跳过（已完成或已过期）")
                    continue
            
            # 每月事件：检查剩余天数
            if event.event_type == "monthly":
                if 0 < days_until <= 3:
                    three_days_events.append((event, days_until))
                    #print(f"[预警事件]   - 添加每月事件到预警列表")
            
            # 生日/纪念日：检查剩余天数
            elif event.event_type in ["birthday", "event"]:
                if 0 < days_until <= 3:
                    three_days_events.append((event, days_until))
                    #print(f"[预警事件]   - 添加生日/纪念日到预警列表")
            
            # 一次性事件：检查剩余天数
            elif event.repeat_type == "once":
                if 0 < days_until <= 3:
                    three_days_events.append((event, days_until))
                    #print(f"[预警事件]   - 添加一次性事件到预警列表")
        
        #print(f"[预警事件] 共找到 {len(three_days_events)} 个预警事件")
        
        # 按剩余天数排序
        three_days_events.sort(key=lambda x: x[1])

        # ========== 提取事件对象列表并调用排序函数 ==========
        event_list = [item[0] for item in three_days_events]
        event_list = get_sorted_events_for_display(event_list)
        
        # 添加标题行
        if hasattr(refresh_events_list, 'view_dropdown'):
            title_text = f"🔔 预警事件 {len(three_days_events)} 个" if three_days_events else "🔔 预警事件 0 个"
            events_list.controls.append(ft.Row([
                ft.Text(title_text, size=14, weight=ft.FontWeight.BOLD, expand=True),
                refresh_events_list.view_dropdown,
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN))
            events_list.controls.append(ft.Divider(height=10))
        
        # 显示事件内容
        if not three_days_events:
            events_list.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Text("✨ 最近3天内没有事件", size=14, color=ft.Colors.GREEN_700),
                    ], spacing=8),
                    padding=20,
                )
            )
        else:
            # ========== 直接使用排序后的 event_list ==========
            for event in event_list:
                display_event_card(event, is_filter_mode=True)

            # 移除最后一个多余的分隔符
            if events_list.controls and isinstance(events_list.controls[-1], ft.Divider):
                events_list.controls.pop()
        
        page.update()
    
    # ===========================  记账功能添加 ===================================
    # 加载记账数据
    def load_accounting_data():
        global transactions
        try:
            json_path = get_data_file_path("accounting.json")
            if os.path.exists(json_path):
                with open(json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    transactions = [Transaction.from_dict(t) for t in data.get("transactions", [])]
            else:
                # 首次使用，创建空记录
                transactions = []
                save_accounting_data()
        except Exception as e:
            print(f"加载记账数据失败: {e}")
            transactions = []
    
    # 保存记账数据
    def save_accounting_data():
        global transactions
        try:
            json_path = get_data_file_path("accounting.json")
            data = {
                "transactions": [t.to_dict() for t in transactions],
            }
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"记账数据已保存，共 {len(transactions)} 条记录")
        except Exception as e:
            print(f"保存记账数据失败: {e}")

    def show_accounting_page(page: ft.Page):
        """显示记账页面（升级版：支持按月查询、编辑、删除）"""
        global transactions
        global current_page, floating_add_button, original_floating_add_click

        # 切换到记账页面
        current_page = "accounting"

        # ========== 筛选相关变量 ==========
        filter_income_categories = []  # 选中的收入分类
        filter_expense_categories = []  # 选中的支出分类
        is_filter_active = False  # 是否启用筛选

        # 保存原来的点击事件，并替换为记账页面的添加菜单
        original_floating_add_click = floating_add_button.on_click
        floating_add_button.on_click = lambda e: show_accounting_add_menu()

        # 隐藏主界面的返回今日按钮（如果需要）
        #today_circle_button.visible = False

        # 当前选中的年月
        current_year = datetime.now().year
        current_month = datetime.now().month
        selected_date = datetime.now()

        # ========== 新增：区间查询变量 ==========
        query_mode = "month"  # "month" 或 "range"
        start_date = datetime.now().replace(day=1).date()
        end_date = datetime.now().date()

        # 记录列表容器（用于滚动）
        records_list = ft.Column(spacing=5, expand=True)  # 移除 scroll，由外层控制

        # 加载数据
        load_accounting_data()

        # ========== 在函数顶部定义滚动状态变量 ==========
        show_scroll_top_btn = False  # 定义在函数顶部，所有内部函数都可以访问

        # ========== 先定义滚动事件处理函数 ==========
        def on_scroll_changed(e):
            """滚动事件回调"""
            nonlocal show_scroll_top_btn
            
            # 获取滚动位置
            scroll_offset = e.pixels if hasattr(e, 'pixels') else 0
            
            # 调试打印
            #print(f"[滚动事件] offset: {scroll_offset}")
            
            # 只要滚动超过20像素就显示回到顶部按钮
            if scroll_offset > 20 and not show_scroll_top_btn:
                show_scroll_top_btn = True
                accounting_scroll_top_button.visible = True   # 使用局部变量
                page.update()
            elif scroll_offset <= 20 and show_scroll_top_btn:
                show_scroll_top_btn = False
                accounting_scroll_top_button.visible = False  # 使用局部变量
                page.update()

        # ========== 创建记账界面独立的回到顶部按钮（局部变量） ==========
        accounting_scroll_top_button = ft.Container(
            content=ft.Icon(ft.Icons.ARROW_UPWARD, size=28, color=ft.Colors.BLUE_700),
            width=50,  # 与 today_circle_button 一致
            height=50,  # 与 today_circle_button 一致
            bgcolor=ft.Colors.WHITE,
            border_radius=25,  # 50/2 = 25
            ink=True,
            on_click=lambda e: asyncio.create_task(accounting_scroll_to_top(e)),
            tooltip="回到顶部",
            shadow=ft.BoxShadow(
                spread_radius=1,
                blur_radius=8,
                color=ft.Colors.BLACK12,
                offset=ft.Offset(0, 2),
            ),
            visible=False,
        )

        # ========== 滚动到顶部函数 ==========
        async def accounting_scroll_to_top(e):
            """滚动到顶部"""
            nonlocal show_scroll_top_btn
            
            # 关键：使用 scroll_container 而不是 records_list
            if hasattr(scroll_container, 'scroll_to'):
                await scroll_container.scroll_to(offset=0, duration=500, curve=ft.AnimationCurve.EASE_IN_OUT)
                # 滚动到顶部后隐藏按钮
                show_scroll_top_btn = False
                scroll_top_button.visible = False
                page.update()
            else:
                # 备用方案：尝试使用 page.scroll_to
                try:
                    page.scroll_to(offset=0, duration=500)
                    show_scroll_top_btn = False
                    scroll_top_button.visible = False
                    page.update()
                except:
                    print("滚动失败")

        # ========== 创建可滚动的容器（在 on_scroll_changed 定义之后） ==========
        scroll_container = ft.Column(
            [records_list],
            scroll=ft.ScrollMode.AUTO,
            expand=True,
            on_scroll=on_scroll_changed,
        )

        def delete_transaction(transaction_id, transaction_name):
            """删除记录（带确认对话框）"""
            
            # 找到要删除的记录
            transaction_to_delete = None
            for t in transactions:
                if t.id == transaction_id:
                    transaction_to_delete = t
                    break
            
            if not transaction_to_delete:
                show_bottom_message("未找到该记录")
                return
            
            dialog_container = None
            
            def close_dialog():
                nonlocal dialog_container
                if dialog_container and dialog_container in page.overlay:
                    page.overlay.remove(dialog_container)
                    dialog_container = None
                    page.update()
            
            def confirm_delete(e):
                close_dialog()
                global transactions
                transactions = [t for t in transactions if t.id != transaction_id]
                save_accounting_data()
                refresh_records_list()
                refresh_summary()
                show_bottom_message(f"已删除{transaction_to_delete.category}记录")
            
            def cancel_delete(e):
                close_dialog()
                show_bottom_message(f"已取消删除")
                page.update()
            
            # 确定显示内容
            is_income = transaction_to_delete.type == "income"
            type_text = "收入" if is_income else "支出"
            amount_text = f"{transaction_to_delete.category} - ¥{abs(transaction_to_delete.amount):,.2f}"
            
            # 对话框内容
            dialog_content = ft.Container(
                content=ft.Column([
                    # 顶部图标（带背景圆）
                    ft.Container(
                        content=ft.Icon(ft.Icons.WARNING_AMBER_ROUNDED, size=55, color=ft.Colors.RED_700),
                        padding=10,
                        bgcolor=ft.Colors.RED_50,
                        border_radius=50,
                    ),
                    ft.Text("确认删除", size=22, weight=ft.FontWeight.BOLD, color=ft.Colors.RED_700, text_align=ft.TextAlign.CENTER),
                    ft.Divider(height=1, color=ft.Colors.GREY_300),
                    ft.Text(f"确定要删除这条{type_text}记录吗？", size=14, color=ft.Colors.GREY_700, text_align=ft.TextAlign.CENTER),
                    ft.Text(amount_text, size=13, color=ft.Colors.BLUE_700, text_align=ft.TextAlign.CENTER),
                    ft.Text(transaction_to_delete.date, size=12, color=ft.Colors.GREY_500, text_align=ft.TextAlign.CENTER),
                    ft.Text("此操作不可撤销！", size=12, color=ft.Colors.RED_500, text_align=ft.TextAlign.CENTER),
                    ft.Divider(height=1, color=ft.Colors.GREY_300),
                    # 按钮区域
                    ft.Row([
                        ft.ElevatedButton(
                            "取消", 
                            on_click=cancel_delete, 
                            expand=True,
                            style=ft.ButtonStyle(bgcolor=ft.Colors.GREY_100, color=ft.Colors.GREY_700),
                        ),
                        ft.ElevatedButton(
                            "确认删除", 
                            on_click=confirm_delete, 
                            expand=True,
                            style=ft.ButtonStyle(bgcolor=ft.Colors.RED_700, color=ft.Colors.WHITE),
                        ),
                    ], spacing=12, alignment=ft.MainAxisAlignment.CENTER),
                ], spacing=15, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                width=320,
                padding=20,
                bgcolor=ft.Colors.WHITE,
                border_radius=16,
            )
            
            dialog_container = ft.Container(
                content=ft.Column([
                    ft.Container(expand=True),  # 上方弹性空间
                    ft.Row([
                        ft.Container(expand=True),  # 左侧弹性空间
                        dialog_content,
                        ft.Container(expand=True),  # 右侧弹性空间
                    ]),
                    ft.Container(expand=True),  # 下方弹性空间
                ]),
                expand=True,
                bgcolor=ft.Colors.BLACK26,
                on_click=close_dialog,
            )
            
            page.overlay.append(dialog_container)
            page.update()
        
        def edit_transaction(transaction):
            """编辑记录（与添加事件界面风格一致）"""
            edit_dialog_container = None
            
            def close_edit_dialog():
                nonlocal edit_dialog_container
                if edit_dialog_container and edit_dialog_container in page.overlay:
                    page.overlay.remove(edit_dialog_container)
                    edit_dialog_container = None
                    page.update()
            
            categories = INCOME_CATEGORIES if transaction.type == "income" else EXPENSE_CATEGORIES

            # ========== 确定标题 ==========
            if transaction.type == "income":
                title_text = "编辑收入"
                title_icon = "💰"
            else:
                title_text = "编辑支出"
                title_icon = "💸"

            def on_date_field_blur(e):
                # 名称输入框失去焦点时的操作
                pass

            date_field = ft.TextField(
                label="日期",
                value=transaction.date,
                read_only=True,
                expand=True,
                on_blur=on_date_field_blur,  # 添加失去焦点事件
            )

            category_field = SearchableDropdownFl(
                page=page,  # 传入 page
                label="分类",
                options=categories,
                value=transaction.category if transaction else None,
                on_change=lambda e: print(f"选择: {e}"),
            )
            
            def on_amount_blur(e):
                # 名称输入框失去焦点时的操作
                pass

            amount_field = ft.TextField(
                label="金额",
                value=str(abs(transaction.amount)),
                keyboard_type=ft.KeyboardType.NUMBER,
                expand=True,
                on_blur=on_amount_blur,  # 添加失去焦点事件
            )

            def on_note_blur(e):
                # 名称输入框失去焦点时的操作
                pass
            
            note_field = ft.TextField(
                label="备注",
                value=transaction.note,
                expand=True,
                multiline=True,
                max_lines=3,
                on_blur=on_note_blur,  # 添加失去焦点事件
            )
            
            # ========== 修复日期选择器 ==========
            # 解析当前日期，用于初始化日期选择器
            current_date_value = None
            try:
                if transaction.date:
                    current_date_value = datetime.strptime(transaction.date, "%Y-%m-%d")
            except:
                pass
            
            date_picker = ft.DatePicker(
                first_date=datetime(2020, 1, 1),
                last_date=datetime(2030, 12, 31),
                value=current_date_value,  # 设置初始值为当前记录的日期
                on_change=lambda e: on_date_selected(e),
            )

            def on_date_selected(e):
                if date_picker.value:
                    # 添加8小时时区转换，解决手机端日期少一天的问题
                    #local_date = date_picker.value + timedelta(hours=8)
                    local_date = date_picker.value + timedelta(days=1)
                    date_field.value = local_date.strftime("%Y-%m-%d")
                    date_field.update()
                    page.update()
                    
            date_field.on_click = lambda e: page.show_dialog(date_picker)
            
            def save_edit(e):
                try:
                    amount = float(amount_field.value)
                    if amount <= 0:
                        show_bottom_message("金额必须大于0", is_error=True)
                        return
                    transaction.date = date_field.value
                    transaction.category = category_field.value
                    transaction.amount = amount
                    transaction.note = note_field.value
                    save_accounting_data()
                    refresh_records_list()
                    refresh_summary()
                    show_bottom_message("已更新记录")
                    close_edit_dialog()
                except ValueError:
                    show_bottom_message("请输入有效的金额", is_error=True)
            
            # 顶部按钮栏（与添加事件一致）
            top_bar = ft.Row([
                ft.IconButton(
                    icon=ft.Icons.CLOSE,
                    icon_size=24,
                    icon_color=ft.Colors.RED_700,
                    tooltip="取消",
                    on_click=lambda e: close_edit_dialog(),
                ),
                ft.Text(f"{title_icon} {title_text}", size=18, weight=ft.FontWeight.BOLD, expand=True, text_align=ft.TextAlign.CENTER),
                ft.IconButton(
                    icon=ft.Icons.CHECK,
                    icon_size=24,
                    icon_color=ft.Colors.GREEN_700,
                    tooltip="保存",
                    on_click=save_edit,
                ),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.CENTER)
            
            # 可滚动的内容区域
            scrollable_content = ft.Column([
                ft.Container(height=1),
                date_field,
                category_field,
                amount_field,
                note_field,
            ], spacing=15, scroll=ft.ScrollMode.AUTO)
            
            # 整体布局（带边框和阴影）
            dialog_content = ft.Column([
                top_bar,
                ft.Divider(height=5),
                ft.Container(
                    content=scrollable_content,
                    expand=True,
                ),
            ], spacing=10, height=420)
            
            edit_dialog_container = ft.Container(
                content=ft.Container(
                    content=dialog_content,
                    bgcolor=ft.Colors.WHITE,
                    padding=20,
                    border_radius=12,
                    #border=ft.border.all(1, ft.Colors.BLUE_200),
                    shadow=ft.BoxShadow(
                        spread_radius=1,
                        blur_radius=15,
                        color=ft.Colors.BLACK12,
                    ),
                    expand=True,
                ),
                left=20,
                top=50,
                right=20,
                bottom=50,
            )

            page.overlay.append(edit_dialog_container)
            page.update()

        async def export_filtered_accounting(e):
            """导出当前筛选后的记账数据到Excel"""
            global transactions
            
            # ========== 获取当前列表的数据（与 refresh_records_list 相同的筛选逻辑） ==========
            # 根据查询模式筛选记录
            if query_mode == "month":
                month_str = f"{current_year}-{current_month:02d}"
                base_records = [t for t in transactions if t.date.startswith(month_str)]
            else:
                start_str = start_date.strftime("%Y-%m-%d")
                end_str = end_date.strftime("%Y-%m-%d")
                base_records = [t for t in transactions if start_str <= t.date <= end_str]
            
            # 应用分类筛选
            if is_filter_active:
                filtered_records = []
                for t in base_records:
                    if t.type == "income":
                        if filter_income_categories and t.category in filter_income_categories:
                            filtered_records.append(t)
                    else:
                        if filter_expense_categories and t.category in filter_expense_categories:
                            filtered_records.append(t)
                base_records = filtered_records
            
            if not base_records:
                show_bottom_message("当前没有可导出的记录", is_error=True)
                return
            
            # ========== 创建Excel文件 ==========
            try:
                temp_dir = get_data_file_path("")
                temp_file = os.path.join(temp_dir, f"accounting_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                
                wb = Workbook()
                ws = wb.active
                ws.title = "记账明细"
                
                # 写入表头
                headers = ["日期", "类型", "分类", "金额", "备注"]
                ws.append(headers)
                
                # 设置表头样式
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color="CCE6FF", end_color="CCE6FF", fill_type="solid")
                
                # 写入数据
                for t in base_records:
                    type_str = "收入" if t.type == "income" else "支出"
                    ws.append([
                        t.date,
                        type_str,
                        t.category,
                        t.amount,
                        t.note,
                    ])
                
                # 调整列宽
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 8
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 30
                
                # 保存临时文件
                wb.save(temp_file)
                
                # 读取文件内容
                with open(temp_file, 'rb') as f:
                    file_bytes = f.read()
                
                # 创建 FilePicker
                file_picker = ft.FilePicker()
                page.services.append(file_picker)
                page.update()
                
                # 选择保存位置
                result = await file_picker.save_file(
                    file_name=f"记账明细_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    src_bytes=file_bytes,
                    dialog_title="保存记账明细"
                )
                
                # 移除 FilePicker
                page.services.remove(file_picker)
                page.update()
                
                # 删除临时文件
                os.remove(temp_file)
                
                if result:
                    # 显示导出的记录数量
                    total_income = sum(t.amount for t in base_records if t.type == "income")
                    total_expense = sum(t.amount for t in base_records if t.type == "expense")
                    show_bottom_message(f"✅ 成功导出 {len(base_records)} 条记录 (收入: ¥{total_income:,.2f}, 支出: ¥{total_expense:,.2f})")
                else:
                    show_bottom_message("已取消导出")
                
                page.update()
                
            except Exception as ex:
                show_bottom_message(f"导出失败: {str(ex)}", is_error=True)
                print(f"导出错误: {ex}")
                import traceback
                traceback.print_exc()

        def show_date_picker_bottom_sheet():
            """显示底部日期选择器"""
            
            # 当前选中的年份和月份
            selected_year = current_year
            selected_month = current_month
            
            # 自定义区间日期
            custom_start = start_date.strftime("%Y-%m-%d")
            custom_end = end_date.strftime("%Y-%m-%d")
            
            # 当前选中的页签：0=按月查询，1=自定义区间
            current_tab = 0


            # 年份列表（往前5年，往后1年）
            year_options = [str(y) for y in range(current_year - 5, current_year + 2)]

            def on_year_change(value):
                """年份选择变化"""
                nonlocal selected_year
                if not value or not value.strip():
                    return
                selected_year = int(value)
                # 更新月份下拉框的值
                month_dropdown.value = str(selected_month)
                page.update()

            # 年份下拉框（使用 SearchableDropdown）
            year_dropdown = SearchableDropdown(
                page=page,
                label="年份",
                options=year_options,
                value=str(selected_year),
                on_change=on_year_change,
            )


            # ========== 月份选择 ==========
            month_options = [f"{i}" for i in range(1, 13)]

            def on_month_change(value):
                """月份选择变化"""
                nonlocal selected_month
                if not value or not value.strip():
                    return
                # 从 "1月" 中提取数字
                selected_month = int(value.replace("月", ""))
                # 切换到按月查询
                query_mode = "month"
                current_year = selected_year
                current_month = selected_month
                month_text.value = f"{current_year}年{current_month}月"
                refresh_summary()
                refresh_records_list()
                update_query_mode_display()
                page.update()

            month_dropdown = SearchableDropdown(
                page=page,
                label="月份",
                options=month_options,
                value=str(selected_month),
                on_change=on_month_change,
            )


            # ========== 按月查询内容 ==========
            monthly_content = ft.Container(
                content=ft.Stack([
                    # 顶部：快捷按钮行
                    ft.Container(
                        content=ft.Row([
                            ft.TextButton(
                                "上月",
                                on_click=lambda e: select_month(
                                    current_year if current_month > 1 else current_year - 1,
                                    current_month - 1 if current_month > 1 else 12
                                ),
                                style=ft.ButtonStyle(
                                    color=ft.Colors.BLUE_700,
                                    #bgcolor=ft.Colors.BLUE_50,
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                ),
                            ),
                            ft.TextButton(
                                "本月",
                                on_click=lambda e: [select_month(current_year, datetime.now().month)],
                                style=ft.ButtonStyle(
                                    color=ft.Colors.BLUE_700,
                                    #bgcolor=ft.Colors.BLUE_700,
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                ),
                            ),
                            ft.TextButton(
                                "近三月",
                                on_click=lambda e: select_near_months(3),
                                style=ft.ButtonStyle(
                                    color=ft.Colors.BLUE_700,
                                    #bgcolor=ft.Colors.BLUE_50,
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                ),
                            ),
                            ft.TextButton(
                                "近一年",
                                on_click=lambda e: select_near_months(12),
                                style=ft.ButtonStyle(
                                    color=ft.Colors.BLUE_700,
                                    #bgcolor=ft.Colors.BLUE_50,
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                ),
                            ),
                        ], spacing=8, alignment=ft.MainAxisAlignment.CENTER),
                        padding=20,
                        top=0,
                        left=0,
                        right=0,
                    ),
                    # 第二行：年份和月份下拉框
                    ft.Container(
                        content=ft.Row([
                            ft.Container(year_dropdown, expand=True),  # 平分空间
                            ft.Container(width=10),  # 间距
                            ft.Container(month_dropdown, expand=True),  # 平分空间
                        ], spacing=5, alignment=ft.MainAxisAlignment.CENTER),
                        padding=10,
                        top=55,
                        left=0,
                        right=0,
                    ),
                    # 确定按钮（在底部）
                    ft.Container(
                        content=ft.ElevatedButton(
                            "确定",
                            on_click=lambda e: select_month(selected_year, selected_month),
                            expand=True,
                            style=ft.ButtonStyle(
                                bgcolor=ft.Colors.BLUE_700,
                                color=ft.Colors.WHITE,
                                shape=ft.RoundedRectangleBorder(radius=8),
                            ),
                        ),
                        padding=10,
                        bottom=0,
                        left=0,
                        right=0,
                    ),
                ]),
                expand=True,
            )

            def select_near_months(n):
                """查询最近 n 个月的收支明细（区间查询）"""
                nonlocal start_date, end_date, query_mode
                
                now = datetime.now()
                start = now - timedelta(days=30 * n)
                end = now
                
                start_date = start.date()
                end_date = end.date()
                query_mode = "range"
                
                # 根据 n 值显示不同的标签
                if n == 3:
                    label = "近三月"
                elif n == 12:
                    label = "近一年"
                else:
                    label = f"近{n}个月"
                
                refresh_summary()
                refresh_records_list()
                update_query_mode_display()
                close_bottom_sheet()
                show_bottom_message(f"已查询 {label} ({start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')})")
            
            def apply_custom_range(e):
                """应用自定义区间"""
                try:
                    start = datetime.strptime(start_date_field_bottom.value, "%Y-%m-%d").date()
                    end = datetime.strptime(end_date_field_bottom.value, "%Y-%m-%d").date()
                    if start > end:
                        show_bottom_message("开始日期不能大于结束日期", is_error=True)
                        return
                    nonlocal start_date, end_date, query_mode
                    start_date = start
                    end_date = end
                    query_mode = "range"
                    refresh_summary()
                    refresh_records_list()
                    update_query_mode_display()
                    close_bottom_sheet()
                    show_bottom_message(f"已查询 {start.strftime('%Y-%m-%d')} ~ {end.strftime('%Y-%m-%d')}")
                except ValueError:
                    show_bottom_message("日期格式错误", is_error=True)
            
            # ========== 选择月份函数 ==========
            def select_month(year, month):
                """选择月份"""
                nonlocal selected_year, selected_month
                selected_year = year
                selected_month = month
                
                # 判断是上月、本月还是其他月份
                now = datetime.now()
                if year == now.year and month == now.month:
                    month_label = "本月"
                elif year == now.year and month == now.month - 1:
                    month_label = "上月"
                elif year == now.year and month == now.month + 1:
                    month_label = "下月"
                else:
                    month_label = f"{year}年{month}月"
                
                # 切换到按月查询
                switch_to_month_mode(None)
                # 更新日期
                nonlocal current_year, current_month
                current_year = year
                current_month = month
                month_text.value = f"{year}年{month}月"
                refresh_summary()
                refresh_records_list()
                update_query_mode_display()
                close_bottom_sheet()
                show_bottom_message(f"已切换到 {month_label}")
            
            def on_start_date_field_bottom_blur(e):
                # 名称输入框失去焦点时的操作
                pass

            # ========== 自定义区间 ==========
            start_date_field_bottom = ft.TextField(
                label="开始日期",
                value=custom_start,
                read_only=True,
                expand=True,
                height=45,
                on_click=lambda e: show_date_picker("start"),
                on_blur=on_start_date_field_bottom_blur,  # 添加失去焦点事件
            )
            
            def on_end_date_field_bottom_blur(e):
                # 名称输入框失去焦点时的操作
                pass

            end_date_field_bottom = ft.TextField(
                label="结束日期",
                value=custom_end,
                read_only=True,
                expand=True,
                height=45,
                on_click=lambda e: show_date_picker("end"),
                on_blur=on_end_date_field_bottom_blur,  # 添加失去焦点事件
            )

             # ========== 自定义区间内容 ==========
            range_content = ft.Container(
                content=ft.Stack([
                    # 日期选择行（在顶部）
                    ft.Container(
                        content=ft.Row([
                            start_date_field_bottom,
                            ft.Text("~", size=14, color=ft.Colors.GREY_700),
                            end_date_field_bottom,
                        ], spacing=5),
                        padding=10,
                        top=0,
                        left=0,
                        right=0,
                    ),
                    # 确定按钮（在底部）
                    ft.Container(
                        content=ft.ElevatedButton(
                            "确定",
                            on_click=apply_custom_range,
                            expand=True,
                            style=ft.ButtonStyle(
                                bgcolor=ft.Colors.BLUE_700,
                                color=ft.Colors.WHITE,
                                shape=ft.RoundedRectangleBorder(radius=8),
                                #padding=(0, 12, 0, 12),
                            ),
                        ),
                        padding=10,
                        bottom=0,
                        left=0,
                        right=0,
                    ),
                ]),
                expand=True,
                visible=False,
            )
            
            # 日期选择器（用于区间选择）
            temp_date_picker = ft.DatePicker(
                first_date=datetime(2020, 1, 1),
                last_date=datetime(2030, 12, 31),
                on_change=lambda e: on_temp_date_selected(e),
            )
            page.overlay.append(temp_date_picker)
            
            picker_target = {"type": "start"}
            
            def show_date_picker(target_type):
                picker_target["type"] = target_type
                page.show_dialog(temp_date_picker)
            
            def on_temp_date_selected(e):
                if temp_date_picker.value:
                    local_date = temp_date_picker.value + timedelta(days=1)
                    date_str = local_date.strftime("%Y-%m-%d")
                    if picker_target["type"] == "start":
                        start_date_field_bottom.value = date_str
                    else:
                        end_date_field_bottom.value = date_str
                    start_date_field_bottom.update()
                    end_date_field_bottom.update()
                    page.update()
            
            def close_bottom_sheet():
                if overlay_container in page.overlay:
                    page.overlay.remove(overlay_container)
                    page.update()
            
            # ========== 切换页签 ==========
            def switch_tab(index):
                nonlocal current_tab
                current_tab = index
                
                # ========== 更新页签样式 ==========
                if index == 0:
                    # 按月查询选中
                    tab_month.content = ft.Text("按月查询", size=14, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700)
                    tab_month.bgcolor = ft.Colors.TRANSPARENT
                    # 加下划线：在下方添加一个容器
                    tab_month.border = ft.border.Border(
                        bottom=ft.border.BorderSide(2, ft.Colors.BLUE_700)
                    )
                    tab_month.border_radius = 0
                    
                    tab_range.content = ft.Text("自定义区间", size=14, weight=ft.FontWeight.NORMAL, color=ft.Colors.GREY_600)
                    tab_range.bgcolor = ft.Colors.TRANSPARENT
                    tab_range.border = None
                    tab_range.border_radius = 0
                    
                    monthly_content.visible = True
                    range_content.visible = False
                    
                else:
                    # 自定义区间选中
                    tab_range.content = ft.Text("自定义区间", size=14, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700)
                    tab_range.bgcolor = ft.Colors.TRANSPARENT
                    tab_range.border = ft.border.Border(
                        bottom=ft.border.BorderSide(2, ft.Colors.BLUE_700)
                    )
                    tab_range.border_radius = 0
                    
                    tab_month.content = ft.Text("按月查询", size=14, weight=ft.FontWeight.NORMAL, color=ft.Colors.GREY_600)
                    tab_month.bgcolor = ft.Colors.TRANSPARENT
                    tab_month.border = None
                    tab_month.border_radius = 0
                    
                    monthly_content.visible = False
                    range_content.visible = True
                
                page.update()
            
            # ========== 页签 ==========
            tab_month = ft.Container(
                content=ft.Text("按月查询", size=14, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700),
                bgcolor=ft.Colors.TRANSPARENT,
                border=ft.border.Border(
                    bottom=ft.border.BorderSide(2, ft.Colors.BLUE_700)
                ),
                border_radius=0,
                ink=True,
                on_click=lambda e: switch_tab(0),
            )
            
            tab_range = ft.Container(
                content=ft.Text("自定义区间", size=14, weight=ft.FontWeight.NORMAL, color=ft.Colors.GREY_600),
                bgcolor=ft.Colors.TRANSPARENT,
                border=None,
                border_radius=0,
                ink=True,
                on_click=lambda e: switch_tab(1),
            )
            
            # ========== 创建底部弹出内容 ==========
            bottom_content = ft.Container(
                content=ft.Column([
                    # 顶部手柄
                    ft.Container(
                        content=ft.Container(
                            width=40,
                            height=4,
                            bgcolor=ft.Colors.GREY_300,
                            border_radius=2,
                        ),
                        alignment=ft.Alignment(0, 0),
                        padding=10,
                    ),
                    # 页签
                    ft.Row([
                        tab_month,
                        ft.Container(width=80),
                        tab_range,
                    ], spacing=0, alignment=ft.MainAxisAlignment.START),
                    ft.Divider(height=1),
                    # 内容区域（expand=True 让内容占满剩余空间）
                    ft.Container(
                        content=ft.Column([
                            monthly_content,
                            range_content,
                        ], spacing=0, expand=True),
                        expand=True,
                    ),
                ], spacing=10, expand=True),  # 整个 Column expand
                padding=10,
                bgcolor=ft.Colors.WHITE,
            )
            
            sheet_container = ft.Container(
                content=bottom_content,
                left=0,
                right=0,
                bottom=0,
                height=450,
                bgcolor=ft.Colors.WHITE,
                shadow=ft.BoxShadow(
                    spread_radius=1,
                    blur_radius=20,
                    color=ft.Colors.BLACK26,
                    offset=ft.Offset(0, -4),
                ),
            )
            
            # 背景遮罩和底部弹出（使用 Stack）
            overlay_container = ft.Container(
                content=ft.Stack([
                    ft.Container(
                        expand=True,
                        bgcolor=ft.Colors.BLACK26,
                        on_click=lambda e: close_bottom_sheet(),
                    ),
                    sheet_container,
                ]),
                expand=True,
            )
            
            page.overlay.append(overlay_container)
            page.update()

        def refresh_summary():
            """刷新统计卡片"""
            summary_container.controls.clear()

             # 根据查询模式计算收支
            base_records = []  # 初始化为空列表
            if query_mode == "month":
                month_str = f"{current_year}-{current_month:02d}"
                month_income = sum(t.amount for t in transactions if t.type == "income" and t.date.startswith(month_str))
                month_expense = sum(t.amount for t in transactions if t.type == "expense" and t.date.startswith(month_str))
                month_balance = month_income - month_expense
                date_label = f"📅 {current_year}年{current_month}月"
            else:
                # 区间查询
                start_str = start_date.strftime("%Y-%m-%d")
                end_str = end_date.strftime("%Y-%m-%d")
                month_income = sum(t.amount for t in transactions if t.type == "income" and start_str <= t.date <= end_str)
                month_expense = sum(t.amount for t in transactions if t.type == "expense" and start_str <= t.date <= end_str)
                month_balance = month_income - month_expense
                date_label = f"📅 {start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"

            # ========== 应用分类筛选 ==========
            if is_filter_active:
                filtered_records = []
                for t in base_records:
                    if t.type == "income":
                        if filter_income_categories and t.category in filter_income_categories:
                            filtered_records.append(t)
                    else:
                        if filter_expense_categories and t.category in filter_expense_categories:
                            filtered_records.append(t)
                base_records = filtered_records
            
            # 计算总收支
            total_income = sum(t.amount for t in transactions if t.type == "income")
            total_expense = sum(t.amount for t in transactions if t.type == "expense")
            total_balance = total_income - total_expense
            
            # ========== 日期标题（可点击，弹出日期选择器） ==========
            if query_mode == "month":
                # 判断是否是上月、本月等
                now = datetime.now()
                if current_year == now.year and current_month == now.month:
                    date_label_text = "本月"
                elif current_year == now.year and current_month == now.month - 1:
                    date_label_text = "上月"
                elif current_year == now.year and current_month == now.month + 1:
                    date_label_text = "下月"
                else:
                    date_label_text = f"{current_year}年{current_month}月"
            else:
                # 区间查询
                if (end_date - start_date).days <= 35:  # 约1个月
                    date_label_text = f"近一月"
                elif (end_date - start_date).days <= 95:  # 约3个月
                    date_label_text = f"近三月"
                elif (end_date - start_date).days <= 370:  # 约1年
                    date_label_text = f"近一年"
                else:
                    date_label_text = f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"

            # ========== 可点击的日期标题 ==========
            date_label = ft.Row([
                ft.Text(date_label_text, size=14, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700),
                ft.Icon(ft.Icons.ARROW_DROP_DOWN, size=20, color=ft.Colors.BLUE_700),  # 使用图标
            ], spacing=2, alignment=ft.MainAxisAlignment.START)

            # 让月份文本可点击
            date_label_container = ft.Container(
                content=date_label,
                on_click=lambda e: show_date_picker_bottom_sheet(),
                ink=True,
                padding=5,
                border_radius=4,
            )

            summary_container.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Divider(height=1),

                        date_label_container,
                        ft.Row([
                            ft.Column([
                                ft.Text("收入", size=12, color=ft.Colors.GREY_600),
                                ft.Text(f"¥ {month_income:,.2f}", size=18, weight=ft.FontWeight.BOLD, color=ft.Colors.GREEN_700),
                            ], expand=True, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                            ft.Column([
                                ft.Text("支出", size=12, color=ft.Colors.GREY_600),
                                ft.Text(f"¥ {month_expense:,.2f}", size=18, weight=ft.FontWeight.BOLD, color=ft.Colors.RED_700),
                            ], expand=True, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                            ft.Column([
                                ft.Text("结余", size=12, color=ft.Colors.GREY_600),
                                ft.Text(f"¥ {month_balance:,.2f}", size=18, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700),
                            ], expand=True, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                        ], spacing=5),
                        ft.Divider(height=1),
                        ft.Row([
                            ft.Text(f"累计结余: ¥ {total_balance:,.2f}", size=12, color=ft.Colors.GREY_600),
                        ], alignment=ft.MainAxisAlignment.END),
                    ], spacing=8),
                    padding=12,
                    bgcolor=ft.Colors.TRANSPARENT,
                    border_radius=10,
                )
            )
            page.update()

        def update_filter_button():
            """更新筛选按钮显示"""
            if is_filter_active:
                total = len(filter_income_categories) + len(filter_expense_categories)
                filter_btn.text = f"筛选 ▾ ({total})"
                filter_btn.style = ft.ButtonStyle(color=ft.Colors.BLUE_700, text_style=ft.TextStyle(size=11, weight=ft.FontWeight.BOLD))
            else:
                filter_btn.text = "筛选 ▾"
                filter_btn.style = ft.ButtonStyle(color=ft.Colors.GREY_600, text_style=ft.TextStyle(size=11))
            filter_btn.update()
        
        def show_filter_dialog(e):
            """显示分类筛选对话框（支持多选）"""
            
            # 当前选中的分类（使用集合方便操作）
            selected_income = set(filter_income_categories)
            selected_expense = set(filter_expense_categories)
            
            # ========== 创建收入分类复选框列表（带分割线） ==========
            income_checkboxes = []
            for i, cat in enumerate(INCOME_CATEGORIES):
                cb = ft.Checkbox(
                    label=cat,
                    value=cat in selected_income,
                    active_color=ft.Colors.GREEN_700,
                )
                income_checkboxes.append(cb)

            # 全选/取消全选功能
            def select_all_income(e):
                for cb in income_checkboxes:
                    cb.value = True
                dialog_container.update()
                page.update()
            
            def deselect_all_income(e):
                for cb in income_checkboxes:
                    cb.value = False
                dialog_container.update()
                page.update()
            
            def select_all_expense(e):
                for cb in expense_checkboxes:
                    cb.value = True
                dialog_container.update()
                page.update()
            
            def deselect_all_expense(e):
                for cb in expense_checkboxes:
                    cb.value = False
                dialog_container.update()
                page.update()
            
            # 创建收入分类容器（带分割线）
            income_container = ft.Column([
                ft.Row([
                    ft.Text("💰 收入分类", size=16, weight=ft.FontWeight.BOLD, color=ft.Colors.GREEN_700),
                    ft.Container(expand=True),
                    ft.TextButton("全选", on_click=select_all_income, style=ft.ButtonStyle(text_style=ft.TextStyle(size=11))),
                    ft.TextButton("取消", on_click=deselect_all_income, style=ft.ButtonStyle(text_style=ft.TextStyle(size=11))),
                ], spacing=5),
            ], spacing=2)
            
            # 添加收入分类选项（带分割线）
            for i, cb in enumerate(income_checkboxes):
                income_container.controls.append(cb)
                # 在选项之间添加分割线（最后一个不加）
                if i < len(income_checkboxes) - 1:
                    income_container.controls.append(ft.Divider(height=1, color=ft.Colors.GREY_200))
            
            # ========== 创建支出分类复选框列表（带分割线） ==========
            expense_checkboxes = []
            for i, cat in enumerate(EXPENSE_CATEGORIES):
                cb = ft.Checkbox(
                    label=cat,
                    value=cat in selected_expense,
                    active_color=ft.Colors.RED_700,
                )
                expense_checkboxes.append(cb)
            
            # 创建支出分类容器（带分割线）
            expense_container = ft.Column([
                ft.Row([
                    ft.Text("💸 支出分类", size=16, weight=ft.FontWeight.BOLD, color=ft.Colors.RED_700),
                    ft.Container(expand=True),
                    ft.TextButton("全选", on_click=select_all_expense, style=ft.ButtonStyle(text_style=ft.TextStyle(size=11))),
                    ft.TextButton("取消", on_click=deselect_all_expense, style=ft.ButtonStyle(text_style=ft.TextStyle(size=11))),
                ], spacing=5),
            ], spacing=2)
            
            # 添加支出分类选项（带分割线）
            for i, cb in enumerate(expense_checkboxes):
                expense_container.controls.append(cb)
                # 在选项之间添加分割线（最后一个不加）
                if i < len(expense_checkboxes) - 1:
                    expense_container.controls.append(ft.Divider(height=1, color=ft.Colors.GREY_200))
            
            
            
            # 确认筛选
            def apply_filter(e):
                nonlocal filter_income_categories, filter_expense_categories, is_filter_active
                
                # 收集选中的分类
                filter_income_categories = [cb.label for cb in income_checkboxes if cb.value]
                filter_expense_categories = [cb.label for cb in expense_checkboxes if cb.value]
                
                is_filter_active = True
                
                # 关闭对话框
                close_filter_dialog()
                
                # 刷新列表
                refresh_records_list()
                refresh_summary()
                
                # 更新筛选按钮文字
                update_filter_button()
                
                total_selected = len(filter_income_categories) + len(filter_expense_categories)
                show_bottom_message(f"已筛选 {total_selected} 个分类")
            
            # 清除筛选
            def clear_filter(e):
                nonlocal filter_income_categories, filter_expense_categories, is_filter_active
                filter_income_categories = []
                filter_expense_categories = []
                is_filter_active = False
                close_filter_dialog()
                refresh_records_list()
                refresh_summary()
                update_filter_button()
                show_bottom_message("已清除筛选")
            
            # 关闭对话框
            def close_filter_dialog():
                if dialog_container in page.overlay:
                    page.overlay.remove(dialog_container)
                    page.update()
            
            # 创建对话框内容
            filter_content = ft.Column([
                ft.Container(
                    content=ft.Row([
                        ft.IconButton(
                            ft.Icons.CLOSE,
                            icon_size=24,
                            icon_color=ft.Colors.RED_700,
                            on_click=lambda e: close_filter_dialog(),
                        ),
                        ft.Text("选择筛选条件", size=20, weight=ft.FontWeight.BOLD, expand=True, text_align=ft.TextAlign.CENTER),
                        ft.IconButton(
                            ft.Icons.CHECK,
                            icon_size=24,
                            icon_color=ft.Colors.GREEN_700,
                            on_click=apply_filter,
                            tooltip="应用筛选",
                        ),
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ),
                ft.Divider(),
                ft.Container(
                    content=ft.Column([
                        income_container,
                        ft.Divider(height=10),
                        expense_container,
                    ], spacing=8, scroll=ft.ScrollMode.AUTO),
                    height=400,
                    padding=5,
                ),
                ft.Divider(),
                ft.Row([
                    ft.TextButton("清除筛选", on_click=clear_filter, style=ft.ButtonStyle(color=ft.Colors.RED_700)),
                    ft.Container(expand=True),
                ], spacing=10),
            ], spacing=10)
            
            dialog_container = ft.Container(
                content=ft.Container(
                    content=filter_content,
                    bgcolor=ft.Colors.WHITE,
                    padding=20,
                    border_radius=16,
                    shadow=ft.BoxShadow(
                        spread_radius=1,
                        blur_radius=20,
                        color=ft.Colors.BLACK12,
                        offset=ft.Offset(0, 4),
                    ),
                    expand=True,
                ),
                left=20,
                top=30,
                right=20,
                bottom=30,
            )
            
            page.overlay.append(dialog_container)
            page.update()

        def refresh_records_list():
            """刷新记录列表"""
            records_list.controls.clear()
            
            # ========== 根据查询模式筛选记录 ==========
            if query_mode == "month":
                month_str = f"{current_year}-{current_month:02d}"
                month_records = [t for t in transactions if t.date.startswith(month_str)]
            else:
                # 区间查询
                start_str = start_date.strftime("%Y-%m-%d")
                end_str = end_date.strftime("%Y-%m-%d")
                month_records = [t for t in transactions if start_str <= t.date <= end_str]
            

            # ========== 应用分类筛选 ==========
            filtered_records = month_records  # 默认使用全部
            if is_filter_active:
                filtered_records = []
                for t in month_records:
                    if t.type == "income":
                        # 收入：检查是否在选中的收入分类中
                        if filter_income_categories and t.category in filter_income_categories:
                            filtered_records.append(t)
                    else:
                        # 支出：检查是否在选中的支出分类中
                        if filter_expense_categories and t.category in filter_expense_categories:
                            filtered_records.append(t)

            filtered_records.sort(key=lambda x: x.date, reverse=True)

            # ========== 计算统计信息 ==========
            total_income = sum(t.amount for t in filtered_records if t.type == "income")
            total_expense = sum(t.amount for t in filtered_records if t.type == "expense")
            total_balance = total_income - total_expense
            
            if not filtered_records:
                records_list.controls.append(
                    ft.Container(
                        content=ft.Text("暂无记录，点击 + 添加", size=14, color=ft.Colors.GREY_500),
                        padding=20,
                    )
                )
                # 即使没有记录，也显示统计信息
                records_list.controls.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Divider(height=1),
                            ft.Row([
                                ft.Text("📊 统计汇总", size=14, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700),
                                ft.Container(expand=True),
                                ft.Text(f"共 {len(filtered_records)} 笔交易", size=11, color=ft.Colors.GREY_500),
                            ]),
                            ft.Row([
                                ft.Column([
                                    ft.Text("收入", size=12, color=ft.Colors.GREY_600),
                                    ft.Text(f"¥ {total_income:,.2f}", size=16, weight=ft.FontWeight.BOLD, color=ft.Colors.GREEN_700),
                                ], expand=True, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                                ft.Column([
                                    ft.Text("支出", size=12, color=ft.Colors.GREY_600),
                                    ft.Text(f"¥ {total_expense:,.2f}", size=16, weight=ft.FontWeight.BOLD, color=ft.Colors.RED_700),
                                ], expand=True, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                                ft.Column([
                                    ft.Text("结余", size=12, color=ft.Colors.GREY_600),
                                    ft.Text(f"¥ {total_balance:,.2f}", size=16, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700),
                                ], expand=True, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                            ], spacing=5),
                        ], spacing=8),
                        padding=15,
                        bgcolor=ft.Colors.TRANSPARENT,
                        border_radius=10,
                    )
                )
                page.update()
                return
            
            # 显示记录卡片
            for index, t in enumerate(filtered_records):
                is_income = t.type == "income"
                amount_color = ft.Colors.GREEN_700 if is_income else ft.Colors.RED_700
                amount_prefix = "+" if is_income else "-"
                
                record_card = ft.Container(
                    content=ft.Row([
                        ft.Column([
                            ft.Row([
                                ft.Icon(ft.Icons.ARROW_UPWARD if is_income else ft.Icons.ARROW_DOWNWARD, 
                                    size=16, color=amount_color),
                                ft.Text(t.category, size=14, weight=ft.FontWeight.BOLD),
                            ], spacing=5),
                            ft.Text(t.date, size=11, color=ft.Colors.GREY_500),
                            ft.Text(t.note, size=11, color=ft.Colors.GREY_500) if t.note else ft.Container(),
                        ], expand=True),
                        ft.Row([
                            ft.Text(f"{amount_prefix}¥ {abs(t.amount):,.2f}", size=14, weight=ft.FontWeight.BOLD, color=amount_color),
                            ft.IconButton(ft.Icons.EDIT, icon_size=18, icon_color=ft.Colors.BLUE_400, 
                                        on_click=lambda e, tr=t: edit_transaction(tr)),
                            ft.IconButton(ft.Icons.DELETE, icon_size=18, icon_color=ft.Colors.RED_400,
                                        on_click=lambda e, tr=t: delete_transaction(tr.id, tr.category)),
                        ], spacing=0),
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                    padding=10,
                    border=ft.border.Border(bottom=ft.border.BorderSide(1, ft.Colors.GREY_200)) if index < len(filtered_records) - 1 else None,
                    ink=True,
                    on_click=lambda e, tr=t: edit_transaction(tr),
                )
                records_list.controls.append(record_card)
            
            # ========== 添加底部统计汇总 ==========
            records_list.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Divider(height=1),
                        ft.Row([
                            ft.Text("📊 统计汇总", size=14, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700),
                            ft.Container(expand=True),
                            ft.Text(f"共 {len(filtered_records)} 笔交易", size=11, color=ft.Colors.GREY_500),
                        ]),
                        ft.Row([
                            ft.Column([
                                ft.Text("收入", size=12, color=ft.Colors.GREY_600),
                                ft.Text(f"¥ {total_income:,.2f}", size=16, weight=ft.FontWeight.BOLD, color=ft.Colors.GREEN_700),
                            ], expand=True, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                            ft.Column([
                                ft.Text("支出", size=12, color=ft.Colors.GREY_600),
                                ft.Text(f"¥ {total_expense:,.2f}", size=16, weight=ft.FontWeight.BOLD, color=ft.Colors.RED_700),
                            ], expand=True, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                            ft.Column([
                                ft.Text("结余", size=12, color=ft.Colors.GREY_600),
                                ft.Text(f"¥ {total_balance:,.2f}", size=16, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700),
                            ], expand=True, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                        ], spacing=5),
                    ], spacing=8),
                    padding=15,
                    bgcolor=ft.Colors.TRANSPARENT,
                    border_radius=10,
                )
            )

            # ========== 关键：在底部添加内边距，防止被悬浮按钮遮挡 ==========
            # 计算悬浮按钮的高度（50px按钮 + 12px间距 + 50px添加按钮 = 112px + 额外安全边距）
            records_list.controls.append(ft.Container(height=130))

            # ========== 安全地重置滚动位置 ==========
            async def reset_scroll():
                try:
                    # 检查 scroll_container 是否还存在且有效
                    if scroll_container and hasattr(scroll_container, 'page') and scroll_container.page is not None:
                        if hasattr(scroll_container, 'scroll_to'):
                            await scroll_container.scroll_to(offset=0, duration=0)
                except Exception as e:
                    # 忽略错误，控件可能已被销毁
                    print(f"滚动重置失败: {e}")
                    pass
            
            # 使用 asyncio.create_task 并忽略错误
            try:
                asyncio.create_task(reset_scroll())
            except:
                pass
            
            page.update()
                
        def change_month_acct(delta):
            """切换月份"""
            nonlocal current_year, current_month, selected_date, query_mode
            query_mode = "month"
            current_month += delta
            if current_month > 12:
                current_month = 1
                current_year += 1
            elif current_month < 1:
                current_month = 12
                current_year -= 1
            selected_date = datetime(current_year, current_month, 1)
            
            # 检查当前月份是否是本月
            now = datetime.now()
            is_current_month = (current_year == now.year and current_month == now.month)
            
            # 控制回到本月按钮的显示
            back_to_today_btn.visible = not is_current_month

            # 更新查询模式显示
            update_query_mode_display()
            
            refresh_summary()
            refresh_records_list()
            month_text.value = f"{current_year}年{current_month}月"
            page.update()

        # ========== 年份切换函数 ==========
        def change_year_acct(delta):
            """切换年份（记账本）"""
            nonlocal current_year, current_month, selected_date, query_mode
            query_mode = "month"
            current_year += delta
            # 限制年份范围
            if current_year < 2000:
                current_year = 2000
            elif current_year > 2100:
                current_year = 2100
            
            selected_date = datetime(current_year, current_month, 1)
            
            now = datetime.now()
            is_current_month = (current_year == now.year and current_month == now.month)
            back_to_today_btn.visible = not is_current_month
            
            update_query_mode_display()
            refresh_summary()
            refresh_records_list()
            month_text.value = f"{current_year}年{current_month}月"
            page.update()
        
        def go_to_current_month(e):
            """回到当前月份"""
            nonlocal current_year, current_month, selected_date, query_mode
            query_mode = "month"
            now = datetime.now()
            current_year = now.year
            current_month = now.month
            selected_date = now
            
            # 隐藏回到本月按钮
            back_to_today_btn.visible = False

            update_query_mode_display()
            
            refresh_summary()
            refresh_records_list()
            month_text.value = f"{current_year}年{current_month}月"
            page.update()
            show_bottom_message("已回到本月")

        # ========== 新增：区间查询相关函数 ==========
        def update_query_mode_display():
            """更新查询模式显示"""
            if query_mode == "month":
                range_row.visible = False
                month_row.visible = True
            else:
                range_row.visible = False
                month_row.visible = True
            # 更新所有控件
            try:
                range_row.update()
                month_row.update()
                page.update()
            except:
                pass

        def switch_to_month_mode(e):
            """切换到按月查询"""
            nonlocal query_mode
            query_mode = "month"
            update_query_mode_display()
            refresh_summary()
            refresh_records_list()
            show_bottom_message("已切换到按月查询")

        def switch_to_range_mode(e):
            """切换到区间查询"""
            nonlocal query_mode
            query_mode = "range"
            update_query_mode_display()
            refresh_summary()
            refresh_records_list()

            # ========== 自动触发区间查询（使用 asyncio） ==========
            async def auto_query():
                await asyncio.sleep(0.1)  # 等待界面更新
                try:
                    start = datetime.strptime(start_date_field.value, "%Y-%m-%d").date()
                    end = datetime.strptime(end_date_field.value, "%Y-%m-%d").date()
                    if start <= end:
                        # 直接调用查询逻辑，不通过 apply_range_query（避免模式切换问题）
                        nonlocal start_date, end_date, query_mode
                        start_date = start
                        end_date = end
                        query_mode = "range"
                        refresh_summary()
                        refresh_records_list()
                        update_query_mode_display()
                        show_bottom_message(f"已查询 {start.strftime('%Y-%m-%d')} ~ {end.strftime('%Y-%m-%d')}")
                    else:
                        show_bottom_message("开始日期不能大于结束日期", is_error=True)
                except Exception as ex:
                    print(f"自动查询失败: {ex}")
            
            asyncio.create_task(auto_query())

            show_bottom_message("已切换到区间查询")

        def apply_range_query(e):
            """应用区间查询"""
            nonlocal start_date, end_date, query_mode
            query_mode = "range"
            
            # 从日期字段获取值
            try:
                start_date = datetime.strptime(start_date_field.value, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date_field.value, "%Y-%m-%d").date()
                
                if start_date > end_date:
                    show_bottom_message("开始日期不能大于结束日期", is_error=True)
                    return
                
                refresh_summary()
                refresh_records_list()
                update_query_mode_display()
                show_bottom_message(f"已查询 {start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}")
            except ValueError:
                show_bottom_message("日期格式错误，请选择有效日期", is_error=True)
        

        def on_start_date_field_blur(e):
            # 名称输入框失去焦点时的操作
            pass

        # ========== 区间查询控件 ==========
        start_date_field = ft.TextField(
            label="开始日期",
            value=datetime.now().replace(day=1).strftime("%Y-%m-%d"),
            read_only=True,
            expand=True,
            on_click=lambda e: page.show_dialog(start_date_picker),
            on_blur=on_start_date_field_blur,  # 添加失去焦点事件
        )

        def on_end_date_field_blur(e):
            # 名称输入框失去焦点时的操作
            pass

        end_date_field = ft.TextField(
            label="结束日期",
            value=datetime.now().strftime("%Y-%m-%d"),
            read_only=True,
            expand=True,
            on_click=lambda e: page.show_dialog(end_date_picker),
            on_blur=on_end_date_field_blur,  # 添加失去焦点事件
        )

        range_row = ft.Row([
            start_date_field,
            ft.Text("~", size=16, color=ft.Colors.GREY_700),
            end_date_field,
        ], spacing=5, visible=False)

        # ========== 构建界面 ==========
        month_text = ft.Text(f"{current_year}年{current_month}月", size=18, weight=ft.FontWeight.BOLD)
        month_row = ft.Row(
            [
                # 年份减按钮
                ft.IconButton(
                    icon=ft.Icons.KEYBOARD_DOUBLE_ARROW_LEFT,
                    icon_size=20,
                    icon_color=ft.Colors.GREY_700,
                    on_click=lambda e: change_year_acct(-1),
                    tooltip="上一年",
                ),
                # 月份减按钮
                ft.IconButton(
                    icon=ft.Icons.KEYBOARD_ARROW_LEFT,
                    icon_size=24,
                    icon_color=ft.Colors.GREY_700,
                    on_click=lambda e: change_month_acct(-1),
                    tooltip="上个月",
                ),
                # 月份文本
                ft.Container(
                    content=month_text,
                    padding=10,  # 简单的整数值
                ),
                # 月份加按钮
                ft.IconButton(
                    icon=ft.Icons.KEYBOARD_ARROW_RIGHT,
                    icon_size=24,
                    icon_color=ft.Colors.GREY_700,
                    on_click=lambda e: change_month_acct(1),
                    tooltip="下个月",
                ),
                # 年份加按钮
                ft.IconButton(
                    icon=ft.Icons.KEYBOARD_DOUBLE_ARROW_RIGHT,
                    icon_size=20,
                    icon_color=ft.Colors.GREY_700,
                    on_click=lambda e: change_year_acct(1),
                    tooltip="下一年",
                ),
            ],
            alignment=ft.MainAxisAlignment.CENTER,
            spacing=8,
        )
        
        # ========== 区间查询控件 ==========
        start_date_picker = ft.DatePicker(
            first_date=datetime(2020, 1, 1),
            last_date=datetime(2030, 12, 31),
            value=datetime.now(),
            on_change=lambda e: on_start_date_selected(e),
        )

        end_date_picker = ft.DatePicker(
            first_date=datetime(2020, 1, 1),
            last_date=datetime(2030, 12, 31),
            value=datetime.now(),
            on_change=lambda e: on_end_date_selected(e),
        )

        def on_start_date_selected(e):
            if start_date_picker.value:
                local_date = start_date_picker.value + timedelta(days=1)
                start_date_field.value = local_date.strftime("%Y-%m-%d")
                start_date_field.update()
                page.update()

        def on_end_date_selected(e):
            if end_date_picker.value:
                local_date = end_date_picker.value + timedelta(days=1)
                end_date_field.value = local_date.strftime("%Y-%m-%d")
                end_date_field.update()
                page.update()

                # ========== 自动触发区间查询（使用 asyncio） ==========
                async def auto_query():
                    await asyncio.sleep(0.1)  # 等待界面更新
                    try:
                        start = datetime.strptime(start_date_field.value, "%Y-%m-%d").date()
                        end = datetime.strptime(end_date_field.value, "%Y-%m-%d").date()
                        if start <= end:
                            # 直接调用查询逻辑，不通过 apply_range_query（避免模式切换问题）
                            nonlocal start_date, end_date, query_mode
                            start_date = start
                            end_date = end
                            query_mode = "range"
                            refresh_summary()
                            refresh_records_list()
                            update_query_mode_display()
                            show_bottom_message(f"已查询 {start.strftime('%Y-%m-%d')} ~ {end.strftime('%Y-%m-%d')}")
                        else:
                            show_bottom_message("开始日期不能大于结束日期", is_error=True)
                    except Exception as ex:
                        print(f"自动查询失败: {ex}")
                
                asyncio.create_task(auto_query())

        # ========== 添加收支记录对话框 ==========
        def show_add_transaction_dialog(transaction_type="expense"):
            """添加收支记录对话框（与添加事件界面风格一致）"""
            dialog_container = None
    
            def close_dialog():
                nonlocal dialog_container
                if dialog_container and dialog_container in page.overlay:
                    page.overlay.remove(dialog_container)
                    dialog_container = None
                    page.update()

             # ========== 确定标题和图标 ==========
            if transaction_type == "income":
                title_text = "添加收入"
                title_icon = "💰"
            else:
                title_text = "添加支出"
                title_icon = "💸"
            
            def on_date_field_blur(e):
                # 名称输入框失去焦点时的操作
                pass

            # 日期字段
            date_field = ft.TextField(
                label="日期",
                value=datetime.now().strftime("%Y-%m-%d"),
                read_only=True,
                expand=True,
                on_blur=on_date_field_blur,  # 添加失去焦点事件
            )
            
            # 根据收支类型显示不同的分类列表
            categories = INCOME_CATEGORIES if transaction_type == "income" else EXPENSE_CATEGORIES

            category_field = SearchableDropdownFl(
                page=page,  # 传入 page
                label="分类",
                options=categories,
                value=categories[0] if categories else None,
                on_change=lambda e: print(f"选择: {e}"),
            )
            

            def on_amount_field_blur(e):
                # 名称输入框失去焦点时的操作
                pass

            amount_field = ft.TextField(
                label="金额",
                hint_text="请输入金额",
                keyboard_type=ft.KeyboardType.NUMBER,
                expand=True,
                on_blur=on_amount_field_blur,  # 添加失去焦点事件
            )

            def on_note_field_blur(e):
                # 名称输入框失去焦点时的操作
                pass
            
            note_field = ft.TextField(
                label="备注",
                hint_text="可选",
                expand=True,
                multiline=True,
                max_lines=3,
                on_blur=on_note_field_blur,  # 添加失去焦点事件
            )
            
            # 日期选择器
            date_picker = ft.DatePicker(
                first_date=datetime(2020, 1, 1),
                last_date=datetime(2030, 12, 31),
                on_change=lambda e: on_date_selected(e),
            )

            def on_date_selected(e):
                if date_picker.value:
                    # 添加8小时时区转换，解决手机端日期少一天的问题
                    #local_date = date_picker.value + timedelta(hours=8)
                    local_date = date_picker.value + timedelta(days=1)
                    date_field.value = local_date.strftime("%Y-%m-%d")
                    date_field.update()
                    page.update()
                    
            date_field.on_click = lambda e: page.show_dialog(date_picker)
            
            def save_transaction(e):
                try:
                    amount = float(amount_field.value)
                    if amount <= 0:
                        show_bottom_message("金额必须大于0", is_error=True)
                        return
                    
                    transaction_id = str(int(datetime.now().timestamp() * 1000))
                    new_transaction = Transaction(
                        id=transaction_id,
                        date=date_field.value,
                        type=transaction_type,
                        category=category_field.value,
                        amount=amount,
                        note=note_field.value,
                    )
                    transactions.append(new_transaction)
                    save_accounting_data()
                    show_bottom_message(f"已添加{'收入' if transaction_type == 'income' else '支出'}: ¥{amount:,.2f}")
                    close_dialog()
                    refresh_records_list()
                    refresh_summary()
                except ValueError:
                    show_bottom_message("请输入有效的金额", is_error=True)
            
            def cancel_click(e):
                close_dialog()
                show_bottom_message("已取消")
            
            # 顶部按钮栏（与添加事件一致）
            top_bar = ft.Row([
                ft.IconButton(
                    icon=ft.Icons.CLOSE,
                    icon_size=24,
                    icon_color=ft.Colors.RED_700,
                    tooltip="取消",
                    on_click=cancel_click,
                ),
                ft.Text(f"{title_icon} {title_text}", size=18, weight=ft.FontWeight.BOLD, expand=True, text_align=ft.TextAlign.CENTER),
                ft.IconButton(
                    icon=ft.Icons.CHECK,
                    icon_size=24,
                    icon_color=ft.Colors.GREEN_700,
                    tooltip="保存",
                    on_click=save_transaction,
                ),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.CENTER)
            
            # 可滚动的内容区域
            scrollable_content = ft.Column([
                ft.Container(height=1),
                date_field,
                category_field,
                amount_field,
                note_field,
            ], spacing=15, scroll=ft.ScrollMode.AUTO)
            
            # 整体布局（带边框和阴影，与添加事件一致）
            dialog_content = ft.Column([
                top_bar,
                ft.Divider(height=5),
                ft.Container(
                    content=scrollable_content,
                    expand=True,
                ),
            ], spacing=10, height=420)
            
            dialog_container = ft.Container(
                content=ft.Container(
                    content=dialog_content,
                    bgcolor=ft.Colors.WHITE,
                    padding=20,
                    border_radius=12,
                    #border=ft.border.all(1, ft.Colors.BLUE_200),  # 添加蓝色边框
                    shadow=ft.BoxShadow(
                        spread_radius=1,
                        blur_radius=15,
                        color=ft.Colors.BLACK12,
                    ),
                    expand=True,
                ),
                left=20,
                top=50,
                right=20,
                bottom=50,
            )
            
            page.overlay.append(dialog_container)
            page.update()

        # ========== 添加菜单 ==========
        def show_accounting_add_menu():
            """显示记账添加菜单"""
            menu_container = None

            def close_menu():
                nonlocal menu_container
                if menu_container and menu_container in page.overlay:
                    page.overlay.remove(menu_container)
                    menu_container = None
                    page.update()
            
            menu_content = ft.Container(
                content=ft.Column([
                    # 顶部图标
                    ft.Container(
                        content=ft.Icon(ft.Icons.ADD_CIRCLE, size=48, color=ft.Colors.BLUE_700),
                        padding=10,
                        bgcolor=ft.Colors.BLUE_50,
                        border_radius=50,
                    ),
                    ft.Text("添加记录", size=20, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_800),
                    ft.Text("请选择记录类型", size=12, color=ft.Colors.GREY_500),
                    ft.Divider(height=1, color=ft.Colors.GREY_200),
                    ft.ElevatedButton(
                        "💰 收入", 
                        on_click=lambda e: [close_menu(), show_add_transaction_dialog("income")], 
                        icon=ft.Icons.ARROW_UPWARD,
                        style=ft.ButtonStyle(
                            bgcolor=ft.Colors.GREEN_700,
                            color=ft.Colors.WHITE,
                            shape=ft.RoundedRectangleBorder(radius=8),
                        ),
                        expand=True,
                    ),
                    ft.ElevatedButton(
                        "💸 支出", 
                        on_click=lambda e: [close_menu(), show_add_transaction_dialog("expense")], 
                        icon=ft.Icons.ARROW_DOWNWARD,
                        style=ft.ButtonStyle(
                            bgcolor=ft.Colors.RED_700,
                            color=ft.Colors.WHITE,
                            shape=ft.RoundedRectangleBorder(radius=8),
                        ),
                        expand=True,
                    ),
                    ft.Divider(height=1, color=ft.Colors.GREY_200),
                    ft.TextButton(
                        "取消", 
                        on_click=lambda e: close_menu(),
                        style=ft.ButtonStyle(
                            color=ft.Colors.GREY_600,
                        ),
                        expand=True,
                    ),
                ], spacing=12, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                width=300,
                padding=20,
                bgcolor=ft.Colors.WHITE,
                border_radius=20,
                shadow=ft.BoxShadow(
                    spread_radius=1,
                    blur_radius=15,
                    color=ft.Colors.BLACK12,
                    offset=ft.Offset(0, 4),
                ),
            )
            menu_container = ft.Container(
                content=ft.Column([ft.Container(expand=True), ft.Row([ft.Container(expand=True), menu_content, ft.Container(expand=True)]), ft.Container(expand=True)]),
                expand=True, bgcolor=ft.Colors.BLACK26, on_click=lambda e: close_menu(),
            )
            page.overlay.append(menu_container)
            page.update()
    
        def back_to_main():
            global current_page
            current_page = "main"

            # 恢复原来的点击事件
            floating_add_button.on_click = original_floating_add_click

            # ========== 先重置所有滚动状态 ==========
            nonlocal show_scroll_top_btn
            show_scroll_top_btn = False
            scroll_top_button.visible = False  # 全局的回到顶部按钮
            
            # 清除页面
            page.clean()
            
            # 重新添加主界面
            page.add(main_stack)

            # 强制滚动到顶部
            async def reset_main_scroll():
                try:
                    # 尝试重置主界面的滚动
                    if hasattr(scrollable_content, 'scroll_to'):
                        await scrollable_content.scroll_to(offset=0, duration=0)
                except:
                    pass
                
                # 再次确保按钮隐藏
                scroll_top_button.visible = False
                page.update()
            
            asyncio.create_task(reset_main_scroll())

            # 只有当音乐正在播放或暂停时才刷新播放信息
            if current_music_state in ["playing", "paused"] and current_music_file:
                update_current_playing_info()
            else:
                # 确保音乐区域隐藏
                if music_section_container:
                    music_section_container.visible = False
                    music_section_container.update()
                if playback_buttons:
                    playback_buttons.visible = False
                    playback_buttons.update()
                marquee_text.update_text("🎵 未播放")
                marquee_text.color = ft.Colors.GREY_600

            page.update()

        # ========== 初始化界面 ==========
        load_accounting_data()
        
        # 统计卡片容器
        summary_container = ft.Column(spacing=10)
        
        back_btn = ft.IconButton(ft.Icons.ARROW_BACK, on_click=lambda e: back_to_main())
        
        refresh_summary()
        refresh_records_list()
        
        # ========== 创建筛选按钮 ==========
        filter_btn = ft.TextButton(
            content=ft.Row([
                ft.Text("筛选", size=14, color=ft.Colors.BLUE_700),
                ft.Icon(ft.Icons.ARROW_DROP_DOWN, size=20, color=ft.Colors.BLUE_700),
            ], spacing=2, alignment=ft.MainAxisAlignment.CENTER),
            on_click=show_filter_dialog,
            style=ft.ButtonStyle(
                color=ft.Colors.BLUE_700,
                #padding=(8, 4, 8, 4),  # (左, 上, 右, 下)
            ),
        )
        
        # ========== 创建导出按钮 ==========
        export_btn = ft.TextButton(
            "导出 📤",
            on_click=lambda e: asyncio.create_task(export_filtered_accounting(e)),
            style=ft.ButtonStyle(color=ft.Colors.GREEN_700, text_style=ft.TextStyle(14)),
        )

        # ========== 关键修改：使用 Stack + Column 固定标题，内容滚动 ==========
        # 固定标题区域
        fixed_header = ft.Container(
            content=ft.Column([
                ft.Container(height=16),
                ft.Row([
                    ft.Container(
                        content=back_btn,
                        width=40,
                    ),
                    ft.Text("💰 账单", size=20, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700, expand=True, text_align=ft.TextAlign.CENTER),
                    ft.Container(width=40),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ft.Divider(),
                month_row,
                #range_row,
                summary_container,
                ft.Divider(),
                ft.Row([
                    ft.Icon(ft.Icons.LIST, size=18, color=ft.Colors.BLUE_700),
                    ft.Text("记录列表", size=16, weight=ft.FontWeight.BOLD),
                    ft.Container(expand=True),
                    filter_btn,  # 添加筛选按钮
                    export_btn,  # 添加导出按钮
                ], spacing=5),
                ft.Divider(),
            ], spacing=8),
            bgcolor=ft.Colors.WHITE,
        )

        # ========== 可滚动的内容区域（添加滚动监听） ==========
        # 使用 ListView 或 Column 并包装滚动事件
        scrollable_records = ft.Container(
            content=scroll_container,  # 使用 scroll_container
            expand=True,
            #padding=ft.padding.only(left=5, right=5),
        )

        # 创建回到本月按钮（与回到今天按钮风格一致）
        back_to_today_btn = ft.Container(
            content=ft.Icon(ft.Icons.TODAY, size=24, color=ft.Colors.BLUE_700),
            width=50,
            height=50,
            bgcolor=ft.Colors.WHITE,
            border_radius=25,
            ink=True,
            on_click=go_to_current_month,
            tooltip="回到本月",
            #border=ft.border.all(1, ft.Colors.BLUE_200),
            shadow=ft.BoxShadow(
                spread_radius=1,
                blur_radius=8,
                color=ft.Colors.BLACK12,
                offset=ft.Offset(0, 2),
            ),
            visible=False,  # 初始隐藏
        )
        
        # 悬浮按钮组（垂直排列）
        floating_buttons = ft.Column(
            [
                accounting_scroll_top_button,  # 初始隐藏
                back_to_today_btn,             # 初始隐藏
                floating_add_button,
            ],
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            spacing=12,
        )
        
        # 使用 Column 布局：固定头部 + 可滚动内容
        accounting_page = ft.Column(
            [
                fixed_header,           # 固定标题
                scrollable_records,     # 可滚动的内容
            ],
            expand=True,
            spacing=0,
        )
        
        # 使用 Stack 布局，将悬浮按钮放在右下角
        accounting_stack = ft.Stack(
            [
                accounting_page,  # 原有的记账页面内容
                ft.Container(
                    content=floating_buttons,
                    right=20,
                    bottom=20,
                ),
            ],
            expand=True,
        )
        
        page.clean()
        page.add(accounting_stack)

        # 初始化查询模式显示
        update_query_mode_display()

        page.update()
    
    def update_view_dropdown_display(view_type):
        """更新视图下拉框的显示文本"""
        # 视图选项映射
        view_display_map = {
            "today": "📌 今日事件",
            "three_days": "🔔 预警事件",
            "daily": "⏰ 每日事件",
            "weekly": "🔁 每周事件",
            "monthly": "🔄 每月事件",
            "birthday": "🎉 生日",
            "event": "💝 纪念日",
            "once": "🎯 一次性事件",
            "all": "📋 全部事件",
        }
        
        # 更新 PopupMenuButton 的显示文本
        if hasattr(refresh_events_list, 'view_dropdown'):
            view_popup = refresh_events_list.view_dropdown
            if hasattr(view_popup, 'content') and view_popup.content:
                display_text = view_display_map.get(view_type, "📋 全部事件")
                view_popup.content.controls[0].value = display_text
                view_popup.update()

    def on_date_text_click(e):
        """点击日期文本时显示事件选择菜单"""
        print(f"[点击事件] 日期文本被点击！")
        
        # 获取数据
        data = None
        if hasattr(e.control, 'data') and e.control.data:
            data = e.control.data
        elif hasattr(date_text, 'data') and date_text.data:
            data = date_text.data
        
        if not data:
            print(f"[点击事件] 无法获取事件数据")
            show_bottom_message("当前没有事件")
            return
        
        today_count = data.get('today_count', 0)
        three_days_count = data.get('three_days_count', 0)
        daily_count = data.get('daily_count', 0)
        weekly_count = data.get('weekly_count', 0)
        
        print(f"[点击事件] 今日:{today_count}, 预警:{three_days_count}, 每日:{daily_count}, 每周:{weekly_count}")
        
        # 关闭菜单的函数
        def close_menu():
            if hasattr(on_date_text_click, 'menu_container'):
                if on_date_text_click.menu_container:
                    try:
                        if on_date_text_click.menu_container in page.overlay:
                            page.overlay.remove(on_date_text_click.menu_container)
                    except Exception as ex:
                        print(f"关闭菜单出错: {ex}")
                    on_date_text_click.menu_container = None
                    page.update()
        
        # 创建安全的回调函数
        def create_callback(event_type):
            def callback(e):
                close_menu()
                show_events_by_type(event_type)
                # ========== 新增：更新自定义下拉框的显示 ==========
                update_view_dropdown_display(event_type)
            return callback
        
        # 创建菜单内容
        menu_items_content = []
        
        # 今日事件按钮
        if today_count > 0:
            menu_items_content.append(
                ft.ElevatedButton(
                    f"📌 今日事件 ({today_count})",
                    on_click=create_callback("today"),
                    expand=True,
                    style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_50, color=ft.Colors.BLUE_700),
                )
            )
        
        # 预警事件按钮
        if three_days_count > 0:
            menu_items_content.append(
                ft.ElevatedButton(
                    f"🔔 预警事件 ({three_days_count})",
                    on_click=create_callback("three_days"),
                    expand=True,
                    style=ft.ButtonStyle(bgcolor=ft.Colors.ORANGE_50, color=ft.Colors.ORANGE_700),
                )
            )
        
        # 每日事件按钮
        if daily_count > 0:
            menu_items_content.append(
                ft.ElevatedButton(
                    f"⏰ 每日事件 ({daily_count})",
                    on_click=create_callback("daily"),
                    expand=True,
                    style=ft.ButtonStyle(bgcolor=ft.Colors.PURPLE_50, color=ft.Colors.PURPLE_700),
                )
            )
        
        # 每周事件按钮
        if weekly_count > 0:
            menu_items_content.append(
                ft.ElevatedButton(
                    f"🔁 每周事件 ({weekly_count})",
                    on_click=create_callback("weekly"),
                    expand=True,
                    style=ft.ButtonStyle(bgcolor=ft.Colors.TEAL_50, color=ft.Colors.TEAL_700),
                )
            )
        
        # 创建菜单容器
        menu_content = ft.Container(
            content=ft.Column([
                # 顶部装饰条
                ft.Container(
                    height=4,
                    width=60,
                    bgcolor=ft.Colors.BLUE_700,
                    border_radius=2,
                ),
                ft.Container(height=10),
                # 图标
                ft.Icon(ft.Icons.EVENT_NOTE, size=48, color=ft.Colors.BLUE_700),
                ft.Text("事件选择", size=18, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_800),
                ft.Text("请选择要查看的事件类型", size=12, color=ft.Colors.GREY_500),
                ft.Divider(height=1, color=ft.Colors.GREY_200),
                ft.Column(menu_items_content, spacing=10),
                ft.Divider(height=1, color=ft.Colors.GREY_200),
                ft.Row([
                    ft.FilledButton(
                        "全部事件",
                        on_click=create_callback("all"),
                        expand=True,
                        icon=ft.Icons.VIEW_LIST,
                    ),
                    ft.OutlinedButton(
                        "取消",
                        on_click=lambda e: close_menu(),
                        expand=True,
                        icon=ft.Icons.CLOSE,
                    ),
                ], spacing=12),
            ], spacing=12, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            width=340,
            padding=20,
            bgcolor=ft.Colors.WHITE,
            border_radius=24,
            shadow=ft.BoxShadow(
                spread_radius=1,
                blur_radius=20,
                color=ft.Colors.BLACK12,
                offset=ft.Offset(0, 4),
            ),
        )
        
        menu_container = ft.Container(
            content=ft.Column([
                ft.Container(expand=True),
                ft.Row([
                    ft.Container(expand=True),
                    menu_content,
                    ft.Container(expand=True),
                ]),
                ft.Container(expand=True),
            ]),
            expand=True,
            bgcolor=ft.Colors.BLACK26,
            on_click=lambda e: close_menu(),
        )
        
        on_date_text_click.menu_container = menu_container
        page.overlay.append(menu_container)
        page.update()

    def update_date_text_with_events(today_date, three_days_events_list):
        """更新日期显示，包含事件数量"""
        global current_date, events
        
        # ========== 统计各类事件数量 ==========
        # 今日事件（生日、纪念日、一次性事件）
        today_events_count = 0
        # 预警事件（未来3天的生日、纪念日、一次性事件）
        three_days_count = 0
        # 每日事件数量
        daily_events_count = 0
        # 每周事件数量
        weekly_events_count = 0
        
        for event in events.values():
            # 统计每日事件
            if event.event_type == "daily":
                daily_events_count += 1
                continue
            
            # 统计每周事件
            if event.event_type == "weekly":
                weekly_events_count += 1
                continue
            
            # 统计今日事件（生日、纪念日、一次性事件）
            month, day, year, base_year, days_until = event.get_next_date_info()
            if month == today_date.month and day == today_date.day:
                if event.repeat_type == "once":
                    if not event.completed and days_until >= 0:
                        today_events_count += 1
                else:
                    today_events_count += 1
        
        # 统计预警事件
        for event, days_until in three_days_events_list:
            if event.event_type != "daily" and event.event_type != "weekly":
                three_days_count += 1
        
        # 构建显示文本
        text_parts = []
        if today_events_count > 0:
            text_parts.append(f"今日 {today_events_count} 个")
        if three_days_count > 0:
            text_parts.append(f"预警 {three_days_count} 个")
        if daily_events_count > 0:
            text_parts.append(f"每日 {daily_events_count} 个")
        if weekly_events_count > 0:
            text_parts.append(f"每周 {weekly_events_count} 个")
        
        # 获取农历和星期信息
        try:
            now = datetime.now()
            lunar = LunarDate.fromSolarDate(now.year, now.month, now.day)
            lunar_month_str = number_to_chinese_month(lunar.month)
            lunar_day_str = number_to_chinese_day(lunar.day)
            lunar_str = f"农历{lunar_month_str}{lunar_day_str}"
        except:
            lunar_str = "农历计算失败"
        
        weekdays = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
        weekday_str = weekdays[today_date.weekday()]
        
        # 基础日期信息
        #base_text = f"{today_date.year}年{today_date.month:02d}月{today_date.day:02d}日 {weekday_str} {lunar_str}"
        base_text = ""
        
        # 更新文本内容
        if text_parts:
            display_text = f"📌 {' ,'.join(text_parts)}"
            date_text.content.value = display_text
            date_text.content.color = ft.Colors.BLUE_700
            date_text.content.weight = ft.FontWeight.BOLD
            date_text.tooltip = "点击查看事件分类"
            date_text.on_click = on_date_text_click
        else:
            display_text = "近期暂无事件发生"
            date_text.content.value = display_text
            date_text.content.color = ft.Colors.GREY_600
            date_text.content.weight = ft.FontWeight.NORMAL
            date_text.tooltip = "暂无事件"
            date_text.on_click = on_date_text_click
        
        # 存储所有事件数量供点击使用
        date_text.data = {
            'today_count': today_events_count,
            'three_days_count': three_days_count,
            'daily_count': daily_events_count,
            'weekly_count': weekly_events_count,
            'three_days_events': three_days_events_list
        }
        
        #print(f"[日期显示] 今日:{today_events_count}, 预警:{three_days_count}, 每日:{daily_events_count}, 每周:{weekly_events_count}")
        date_text.update()
    
    # ========== 在刷新事件列表的函数中，添加排序逻辑 ==========


    # ========== 通用置顶排序函数 ==========
    def get_sorted_events_for_display(events_list):
        """
        对事件列表进行排序，播放中的排在第一位
        events_list: 事件对象列表
        返回排序后的事件对象列表
        """
        if not events_list:
            return events_list
        
        playing_events = []
        other_events = []
        
        # ========== 调试信息 ==========
        #print(f"[置顶调试] current_playing_event_id: {current_playing_event_id}")
        #print(f"[置顶调试] current_music_state: {current_music_state}")
        
        for event in events_list:
            is_playing = (event.id == current_playing_event_id and current_music_state in ["playing", "paused"])
            #print(f"[置顶调试] 事件: {event.name}, id: {event.id}, is_playing: {is_playing}")
            
            if is_playing:
                playing_events.append(event)
            else:
                other_events.append(event)
        
        result = playing_events + other_events
        #print(f"[置顶调试] 排序后第一个事件: {result[0].name if result else '无'}")
        
        return result

    def display_all_events():
        """显示全部事件"""
        global current_view, current_playing_event_id, current_music_state, card_duration_texts
        current_view = "all"

        # ========== 清空旧的卡片引用 ==========
        card_duration_texts.clear()
        
        events_list.controls.clear()
        
        # ========== 添加标题行 ==========
        title_text = f"📋 全部事件 {len(events)} 个" if events else "📋 全部事件 0 个"
        events_list.controls.append(ft.Row([
            ft.Text(title_text, size=14, weight=ft.FontWeight.BOLD, expand=True),
            refresh_events_list.view_dropdown,
        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN))
        events_list.controls.append(ft.Divider(height=10))
        
        if not events:
            events_list.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Text("✨ 暂无事件", size=14, color=ft.Colors.GREEN_700),
                    ], spacing=8),
                    padding=20,
                )
            )
            page.update()
            return
        
        #events_list.controls.append(ft.Text(f"✨ 全部事件有 {len(events)} 个", size=14, color=ft.Colors.GREEN_700))
        #events_list.controls.append(ft.Divider(height=5))
        
        # ========== 分离播放中的事件和其他事件 ==========
        playing_event_info = None
        other_events_info = []
        
        today = datetime.now().date()
        
        for event in events.values():
            month, day, year, base_year, days_until = event.get_next_date_info()
            
            # 计算年龄/年份显示
            if event.event_type == "birthday":
                if base_year > 0 and base_year <= today.year:
                    age_text = f"🎉 {today.year - base_year}岁"
                else:
                    age_text = "🎉 生日"
            elif event.event_type == "monthly":
                age_text = "🔄 每月提醒"
            elif event.event_type == "daily":
                age_text = "⏰ 每天提醒"
                # 每日事件：获取最早的提醒时间用于排序
                earliest_time = "23:59"
                if event.reminders:
                    times = [r.get("time", "23:59") for r in event.reminders if r.get("enabled")]
                    if times:
                        earliest_time = min(times)
                days_until = earliest_time  # 字符串
            elif event.event_type == "weekly":
                age_text = "🔁 每周提醒"
            elif event.repeat_type == "once":
                age_text = ""
            else:
                if base_year > 0 and base_year <= today.year:
                    years_passed = today.year - base_year + 1
                    age_text = f"💝 第{years_passed}年"
                else:
                    age_text = "💝 纪念日"
            
            event_info = {
                "event": event,
                "month": month,
                "day": day,
                "age_text": age_text,
                "days_until": days_until,
                "base_year": base_year,
                "event_type": event.event_type
            }
            
            # 判断是否是播放中的事件
            if event.id == current_playing_event_id and current_music_state in ["playing", "paused"]:
                playing_event_info = event_info
            else:
                other_events_info.append(event_info)
        
        # ========== 先显示播放中的事件 ==========
        if playing_event_info:
            print(f"[置顶] 播放中事件: {playing_event_info['event'].name}")
            display_event_card(playing_event_info["event"], is_filter_mode=True)
        
        # ========== 对其他事件排序（按事件类型分组排序） ==========
        # 分离每日事件和其他事件
        daily_events = []
        normal_events = []
        
        for info in other_events_info:
            if info["event_type"] == "daily":
                daily_events.append(info)
            else:
                normal_events.append(info)
        
        # 每日事件按提醒时间排序
        daily_events.sort(key=lambda x: x["days_until"])  # days_until 是字符串 "HH:MM"
        
        # 其他事件按剩余天数排序
        normal_events.sort(key=lambda x: x["days_until"])  # days_until 是整数
        
        # 合并：每日事件在前，其他事件在后
        other_events_info = daily_events + normal_events
        
        for info in other_events_info:
            display_event_card(info["event"], is_filter_mode=True)

        # 移除最后一个多余的分隔符
        if events_list.controls and isinstance(events_list.controls[-1], ft.Divider):
            events_list.controls.pop()
        
        page.update()

    def show_today_events():
        """显示今日事件列表"""
        global current_view, events_list, card_duration_texts
        current_view = "today"

        # ========== 清空旧的卡片引用 ==========
        card_duration_texts.clear()

        events_list.controls.clear()
        
        today = datetime.now().date()
        today_events = []
        
        for event in events.values():
            if event.event_type == "daily" or event.event_type == "weekly":
                continue
            month, day, year, base_year, days_until = event.get_next_date_info()
            if month == today.month and day == today.day:
                if event.repeat_type == "once":
                    if not event.completed and days_until >= 0:
                        today_events.append(event)
                else:
                    today_events.append(event)

        # ========== 调用置顶排序 ==========
        today_events = get_sorted_events_for_display(today_events)
        
        # 先添加标题行（包含下拉框），始终显示
        if hasattr(refresh_events_list, 'view_dropdown'):
            events_list.controls.append(ft.Row([
                ft.Text(f"📌 今日事件 {len(today_events)} 个", size=14, weight=ft.FontWeight.BOLD, expand=True),
                refresh_events_list.view_dropdown,
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN))
            events_list.controls.append(ft.Divider(height=10))
        
        # 然后显示事件内容
        if not today_events:
            events_list.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Text("🎉 今日没有事件", size=14, color=ft.Colors.GREEN_700),
                    ], spacing=8, horizontal_alignment=ft.CrossAxisAlignment.START),
                    padding=20,
                )
            )
        else:
            # ========== 直接使用排序后的 today_events ==========
            for event in today_events:
                display_event_card(event, is_filter_mode=True)

            # 移除最后一个多余的分隔符
            if events_list.controls and isinstance(events_list.controls[-1], ft.Divider):
                events_list.controls.pop()
        
        page.update()
    
    def get_music_duration_display(file_path):
        """获取音乐文件的时长，返回格式化的字符串"""
        if not file_path or not os.path.exists(file_path):
            return ""
        
        try:
            from mutagen.mp3 import MP3
            duration = MP3(file_path).info.length
            return format_time(duration)
        except:
            try:
                from mutagen.wave import WAVE
                duration = WAVE(file_path).info.length
                return format_time(duration)
            except:
                try:
                    from mutagen.flac import FLAC
                    duration = FLAC(file_path).info.length
                    return format_time(duration)
                except:
                    return ""

    def display_event_card(event, is_filter_mode=False, custom_days_until=None):
        """显示单个事件卡片"""
        global current_playing_event_id, current_music_state, current_position_sec, events_list,card_duration_texts  # 添加 events_list
        
        today = datetime.now().date()
        now = datetime.now()
        base_year = 0
        month, day = 1, 1

        # 在 display_event_card 中，判断是否是播放中的事件
        is_playing_event = (event.id == current_playing_event_id and current_music_state in ["playing", "paused"])

        # 如果是播放中的事件，使用特殊背景色
        if is_playing_event:
            bg_color = ft.Colors.BLUE_50  # 浅蓝色背景
            #border = ft.border.all(1, ft.Colors.BLUE_300)  # 蓝色边框
        else:
            bg_color = ft.Colors.WHITE
            border = None

        #bg_color = ft.Colors.WHITE

        # 优先使用自定义天数
        if custom_days_until is not None:
            days_until = custom_days_until
        else:
            month, day, year, base_year, days_until = event.get_next_date_info()

        # ========== 确定状态文本和颜色 ==========
        if custom_days_until is not None:
            # 筛选模式
            if days_until == -1:
                status_text = "已过期"
                status_color = ft.Colors.GREY_500
            elif days_until == 0:
                status_text = "今天"
                status_color = ft.Colors.RED_700
            else:
                status_text = f"还剩 {days_until} 天"
                status_color = ft.Colors.BLUE_700

        else:

        # ========== 统一背景色和状态文字颜色 ==========
        # 所有卡片使用统一的白色背景，状态文字使用灰色

            status_color = ft.Colors.GREY_600
            status_text = ""
            
            # ========== 每天事件特殊处理（放在最前面） ==========
            if event.event_type == "daily":
                is_workday_only = getattr(event, 'workday_only', False)
                
                if is_workday_only:
                    # 工作日提醒：计算下一个工作日的提醒时间
                    now = datetime.now()
                    now_time = now.strftime("%H:%M")
                    is_today_workday = is_workday(now)
                    
                    # 获取第一个提醒时间
                    reminder_time = None
                    if event.reminders:
                        for reminder in event.reminders:
                            if reminder.get("enabled"):
                                reminder_time = reminder.get("time", "")
                                break
                    
                    if reminder_time:
                        reminder_hour, reminder_minute = map(int, reminder_time.split(":"))
                        
                        # 计算目标提醒的日期时间
                        target_datetime = None
                        
                        if is_today_workday and reminder_time > now_time:
                            # 今天是工作日且提醒时间还没到，使用今天
                            target_datetime = datetime(now.year, now.month, now.day, reminder_hour, reminder_minute)
                        else:
                            # 今天不是工作日或提醒时间已过，找下一个工作日
                            days_offset = 1
                            next_date = now + timedelta(days=days_offset)
                            while not is_workday(next_date):
                                days_offset += 1
                                next_date = now + timedelta(days=days_offset)
                            target_datetime = datetime(next_date.year, next_date.month, next_date.day, reminder_hour, reminder_minute)
                        
                        # 计算时间差
                        time_diff = target_datetime - now
                        
                        if time_diff.total_seconds() > 0:
                            total_seconds = int(time_diff.total_seconds())
                            days = total_seconds // 86400
                            hours = (total_seconds % 86400) // 3600
                            minutes = (total_seconds % 3600) // 60
                            
                            if days > 0:
                                if hours > 0 and minutes > 0:
                                    status_text = f"{days} 天 {hours} 小时 {minutes} 分钟后"
                                elif hours > 0:
                                    status_text = f"{days} 天 {hours} 小时 {minutes} 分钟后"
                                else:
                                    status_text = f"{days} 天 {hours} 小时 {minutes} 分钟后"
                            elif hours > 0:
                                if minutes > 0:
                                    status_text = f"{hours} 小时 {minutes} 分钟后"
                                else:
                                    status_text = f"{hours} 小时后"
                            else:
                                if minutes > 0:
                                    status_text = f"{minutes} 分钟后"
                                else:
                                    status_text = f"即将"
                            status_color = ft.Colors.BLUE_700
                        else:
                            status_text = f"已过"
                            status_color = ft.Colors.GREY_500
                    else:
                        status_text = "工作日"
                        status_color = ft.Colors.BLUE_700
                else:
                    # 普通每天提醒：原来的逻辑
                    if event.reminders:
                        now_time = now.strftime("%H:%M")
                        next_reminder_time = None
                        is_today_reminder = False
                        
                        for reminder in event.reminders:
                            if reminder.get("enabled"):
                                reminder_time = reminder.get("time", "")
                                if reminder_time:
                                    if reminder_time > now_time:
                                        next_reminder_time = reminder_time
                                        is_today_reminder = True
                                        break
                                    elif not next_reminder_time:
                                        next_reminder_time = reminder_time
                        
                        if next_reminder_time:
                            if is_today_reminder:
                                reminder_hour, reminder_minute = map(int, next_reminder_time.split(":"))
                                reminder_datetime = datetime(now.year, now.month, now.day, reminder_hour, reminder_minute)
                                time_diff = reminder_datetime - now
                                
                                if time_diff.total_seconds() > 0:
                                    hours = int(time_diff.total_seconds() // 3600)
                                    minutes = int((time_diff.total_seconds() % 3600) // 60)
                                    if hours > 0:
                                        if minutes > 0:
                                            status_text = f"{hours} 小时 {minutes} 分钟后"
                                        else:
                                            status_text = f"{hours} 小时后"
                                    else:
                                        status_text = f"{minutes} 分钟后"
                                    status_color = ft.Colors.BLUE_700
                                else:
                                    status_text = f"已过"
                                    status_color = ft.Colors.GREY_500
                            else:
                                status_text = f"明天"
                                status_color = ft.Colors.ORANGE_700
                        else:
                            status_text = "每天"
                            status_color = ft.Colors.PURPLE_700
                    else:
                        status_text = "每天"
                        status_color = ft.Colors.PURPLE_700
            
            # ========== 其他事件类型 ==========
            elif event.event_type == "weekly":
                if days_until == 0:
                    status_text = "今天"
                    status_color = ft.Colors.RED_700
                elif days_until == 1:
                    status_text = "明天"
                    status_color = ft.Colors.ORANGE_700
                else:
                    status_text = f"{days_until}天后"
                    status_color = ft.Colors.BLUE_700
            
            elif event.repeat_type == "once":
                if event.completed:
                    status_text = "已完成"
                    status_color = ft.Colors.GREY_500
                elif days_until < 0:
                    status_text = "已过期"
                    status_color = ft.Colors.GREY_500
                elif days_until == 0:
                    status_text = "今天"
                    status_color = ft.Colors.RED_700
                else:
                    status_text = f"{days_until}天后"
                    status_color = ft.Colors.ORANGE_700
            
            elif event.event_type == "monthly":
                if days_until == 0:
                    status_text = "今天"
                    status_color = ft.Colors.RED_700
                elif days_until == 1:
                    status_text = "明天"
                    status_color = ft.Colors.ORANGE_700
                else:
                    status_text = f"{days_until}天后"
                    status_color = ft.Colors.BLUE_700
            
            elif event.event_type == "birthday":
                if days_until == 0:
                    status_text = "今天"
                    status_color = ft.Colors.RED_700
                elif days_until <= 7:
                    status_text = f"{days_until}天后"
                    status_color = ft.Colors.ORANGE_700
                else:
                    status_text = f"{days_until}天后"
                    status_color = ft.Colors.BLUE_700
            
            elif event.event_type == "event":
                if days_until == 0:
                    status_text = "今天"
                    status_color = ft.Colors.RED_700
                elif days_until <= 7:
                    status_text = f"{days_until}天后"
                    status_color = ft.Colors.ORANGE_700
                else:
                    status_text = f"{days_until}天后"
                    status_color = ft.Colors.BLUE_700
            
            else:
                # 筛选模式或其他
                if days_until == 0:
                    status_text = "今天"
                    status_color = ft.Colors.RED_700
                else:
                    status_text = f"{days_until}天后"
                    status_color = ft.Colors.BLUE_700
        
        # 创建状态容器
        if status_text:
            status_container = ft.Container(
                content=ft.Text(status_text, size=12, weight=ft.FontWeight.BOLD, color=status_color),
                padding=5,
                bgcolor=ft.Colors.WHITE,
                border_radius=5,
            )
        else:
            status_container = ft.Container()
        
        # ========== 获取事件图标和显示日期 ==========
        calendar_icon = get_event_icon(event)
        display_date = get_display_date(event)
        type_name = get_event_type_name(event)
        age_text = get_age_text(event, today, base_year)

        # ========== 获取提醒时间 ==========
        reminder_time_str = ""
        if event.reminders:
            time_list = [r.get("time", "") for r in event.reminders if r.get("enabled")]
            if time_list:
                reminder_time_str = " ⏰ " + " ".join(time_list)

        # ========== 根据事件类型选择日期图标 ==========
        date_icon_map = {
            "birthday": "☀️" if event.calendar_type == "solar" else "🌙",  # 阳历生日/农历生日
            "event": "☀️" if event.calendar_type == "solar" else "🌙",     # 阳历纪念日/农历纪念日
            "daily": "⏰",    # 每日事件
            "weekly": "🔁",   # 每周事件（星期）
            "monthly": "🔄",  # 每月事件（日）
            "once": "☀️" if event.calendar_type == "solar" else "🌙",     # 一次性事件
        }
        
        date_icon = date_icon_map.get(event.event_type, "📅")
        
        # ========== 组合日期显示 ==========
        date_display_text = f"{date_icon} {display_date}{reminder_time_str}"
        
        # ========== 获取音乐名称和状态 ==========
        music_name = None
        music_duration_str = ""  # 显示当前播放位置/总时长
        music_status_icon = "🔇"
        music_status_text = "❌ 无音乐"
        music_status_color = ft.Colors.GREY_400
        
        if event.sound_file and os.path.exists(event.sound_file):
            music_name = get_full_music_name(event.sound_file)

            # ========== 获取歌曲总时长 ==========
            total_duration = get_music_duration_display(event.sound_file)
            
            # ========== 如果是当前播放的事件，显示实时位置 ==========
            if current_playing_event_id == event.id:
                # 直接获取当前的播放位置
                pos_sec = current_position_sec
                if pos_sec > 0 and total_duration:
                    music_duration_str = f"{format_time(pos_sec)} / {total_duration}"
                else:
                    music_duration_str = total_duration if total_duration else ""
                
                if current_music_state == "playing":
                    music_status_icon = "▶️"
                    music_status_text = "播放中"
                    music_status_color = ft.Colors.GREEN_700
                elif current_music_state == "paused":
                    music_status_icon = "⏸️"
                    music_status_text = "已暂停"
                    music_status_color = ft.Colors.ORANGE_700
                else:
                    music_status_icon = "🎵"
                    music_status_text = "未播放"
                    music_status_color = ft.Colors.GREY_500
            else:
                music_status_icon = "🎵"
                music_status_text = "未播放"
                music_status_color = ft.Colors.GREY_500
                music_duration_str = total_duration if total_duration else ""
        
        # ========== 根据事件类型选择标签图标 ==========
        type_icon_map = {
            "birthday": "🎉",
            "event": "💝",
            "daily": "⏰",
            "weekly": "🔁",
            "monthly": "🔄",
            "once": "🎯",
        }

        type_icon = type_icon_map.get(event.event_type, "🏷️")

        # ========== 创建动态音乐显示Row ==========
        # ========== 创建时长文本控件（保存引用） ==========
        duration_display = f"⏱️ {music_duration_str}" if music_duration_str else ""
    
        duration_text = ft.Text(
            duration_display, 
            size=9, 
            color=ft.Colors.BLUE_700 if current_playing_event_id == event.id else ft.Colors.GREY_500,
            weight=ft.FontWeight.BOLD if current_playing_event_id == event.id else ft.FontWeight.NORMAL,
        )
        
        # ========== 保存到全局字典中 ==========
        card_duration_texts[event.id] = duration_text
        
        # 🏷️
        music_info_row = ft.Row([
            ft.Text(f"{type_icon} {type_name}", size=10, color=ft.Colors.BLUE_400),
            ft.Container(width=8),
            ft.Text(music_status_icon, size=10),
            ft.Text(music_name if music_name else "无音乐", size=10, color=ft.Colors.GREY_600,
                    weight=ft.FontWeight.NORMAL if music_status_icon in ["🔇", "🎵"] else ft.FontWeight.BOLD),
            duration_text,  # 使用保存引用的控件
            ft.Text(music_status_text, size=9, color=music_status_color,
                    weight=ft.FontWeight.BOLD if music_status_icon in ["▶️", "⏸️"] else ft.FontWeight.NORMAL),
        ], spacing=3, vertical_alignment=ft.CrossAxisAlignment.CENTER)
        
        # ========== 获取循环状态 ==========
        loop_state = event_loop_states.get(event.id, False)
        loop_checkbox = ft.Checkbox(label="循环", value=loop_state, tooltip="勾选后循环播放")
        
        def on_loop_change(e, eid=event.id, cb=loop_checkbox):
            event_loop_states[eid] = cb.value
        loop_checkbox.on_change = on_loop_change
        
        # ========== 创建播放按钮 ==========
        def create_play_handler(e):
            if event.sound_file and os.path.exists(event.sound_file):
                should_loop = loop_checkbox.value
                event_loop_states[event.id] = should_loop
                
                if current_playing_event_id and current_playing_event_id != event.id:
                    if current_playing_event_id in event_loop_states:
                        event_loop_states[current_playing_event_id] = False
                
                if current_playing_event_id == event.id:
                    if current_music_state == "playing":
                        async def pause_music_handler():
                            if current_audio:
                                await current_audio.pause()
                        asyncio.create_task(pause_music_handler())
                        return
                    elif current_music_state == "paused":
                        async def resume_music_handler():
                            if current_audio:
                                await current_audio.resume()
                        asyncio.create_task(resume_music_handler())
                        return
                
                play_music_with_lock(event.sound_file, loop=should_loop, event_name=event.name, event_id=event.id)
            else:
                show_snack_bar("未设置音乐文件")
        
        play_button = ft.TextButton("🔊 播放", on_click=create_play_handler)
        
        # ========== 创建事件卡片 ==========
        event_card = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Column([
                        ft.Text(f"{calendar_icon} {event.name}", size=16, weight=ft.FontWeight.BOLD),
                        ft.Text(f"{date_display_text}", size=12, color=ft.Colors.GREY_600),
                        ft.Text(age_text, size=11, color=ft.Colors.ORANGE_700) if age_text else ft.Container(),
                        music_info_row,
                    ], expand=True),
                    status_container,
                ]),
                ft.Row([
                    ft.Row([loop_checkbox, play_button], spacing=5),
                    ft.Row([
                        ft.TextButton("✏️ 编辑", on_click=lambda e, eid=event.id: edit_event_dialog(eid)),
                        ft.TextButton("🗑️ 删除", on_click=lambda e, eid=event.id: delete_event(eid)),
                    ], spacing=10),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            ], spacing=5),
            padding=10, 
            bgcolor=bg_color, 
            border_radius=10,
            # ========== 新增：点击卡片进入编辑模式 ==========
            on_click=lambda e, eid=event.id: edit_event_dialog(eid),
            ink=True,  # 添加墨水效果，点击时有反馈
        )

        # 添加卡片和分隔符
        events_list.controls.append(event_card)
        events_list.controls.append(ft.Divider(height=1, color=ft.Colors.GREY_200))

    def get_event_icon(event):
        """获取事件图标"""
        if event.event_type == "daily":
            return "⏰"
        elif event.event_type == "weekly":
            return "🔁"
        elif event.event_type == "birthday":
            return "🎉" if event.calendar_type == "solar" else "🎋"
        elif event.event_type == "monthly":
            return "🔄"
        elif event.repeat_type == "once":
            return "🎯"
        else:
            return "💝" if event.calendar_type == "solar" else "💝"
        
    def get_event_type_name(event):
        """获取事件类型名称"""
        if event.event_type == "daily":
            if hasattr(event, 'workday_only') and event.workday_only:
                return "工作日"
            else:
                return "每天"
        elif event.event_type == "weekly":
            return "每周"
        elif event.event_type == "birthday":
            return "生日"
        elif event.event_type == "monthly":
            return "每月"
        elif event.repeat_type == "once":
            return "一次性"
        else:
            return "纪念日"
        
    def get_display_date(event):
        """获取事件显示日期"""
        month, day, year, base_year, _ = event.get_next_date_info()
        
        if event.event_type == "daily":
            # 检查是否开启了法定工作日提醒
            if hasattr(event, 'workday_only') and event.workday_only:
                # 显示工作日提醒
                if event.reminders:
                    time_list = [r.get("time", "") for r in event.reminders if r.get("enabled")]
                    time_str = " ".join(time_list)
                    return f"{time_str}"
                else:
                    return "工作日提醒"
            else:
                # 普通每天提醒
                if event.reminders:
                    time_list = [r.get("time", "") for r in event.reminders if r.get("enabled")]
                    time_str = " ".join(time_list)
                    return f"每天 {time_str}"
                else:
                    return "每天"
        elif event.event_type == "weekly":
            weekday_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]
            weekday_num = int(event.birth_date) if event.birth_date else 1
            if event.reminders:
                time_list = [r.get("time", "") for r in event.reminders if r.get("enabled")]
                return f"{weekday_names[weekday_num]} {' '.join(time_list)}"
            return f"{weekday_names[weekday_num]}"
        elif event.event_type == "birthday":
            if event.calendar_type == "solar":
                lunar_parts = event.birth_date.split("-")
                return f"阳历 {int(lunar_parts[0])}年{int(lunar_parts[1])}月{int(lunar_parts[2])}日"
            else:
                lunar_parts = event.birth_date.split("-")
                return f"农历 {int(lunar_parts[0])}年{int(lunar_parts[1])}月{int(lunar_parts[2])}日"
        elif event.event_type == "monthly":
            day_num = int(event.birth_date)
            if event.reminders:
                time_list = [r.get("time", "") for r in event.reminders if r.get("enabled")]
                return f"{day_num}日 {' '.join(time_list)}"
            return f"{day_num}日"
        elif event.repeat_type == "once":
            if event.calendar_type == "solar":
                lunar_parts = event.birth_date.split("-")
                return f"阳历 {int(lunar_parts[0])}年{int(lunar_parts[1])}月{int(lunar_parts[2])}日"
            else:
                lunar_parts = event.birth_date.split("-")
                return f"农历 {int(lunar_parts[0])}年{int(lunar_parts[1])}月{int(lunar_parts[2])}日"
        else:
            if event.calendar_type == "solar":
                lunar_parts = event.birth_date.split("-")
                return f"阳历 {int(lunar_parts[0])}年{int(lunar_parts[1])}月{int(lunar_parts[2])}日"
            else:
                lunar_parts = event.birth_date.split("-")
                return f"农历 {int(lunar_parts[0])}年{int(lunar_parts[1])}月{int(lunar_parts[2])}日"
            
    def get_age_text(event, today, base_year):
        """获取年龄或年份显示文本"""
        if event.event_type == "birthday":
            if base_year > 0 and base_year <= today.year:
                return f"🎉 {today.year - base_year}岁"
            else:
                return "🎉 生日"
        elif event.event_type == "monthly":
            return "🔄 每月提醒"
        elif event.event_type == "daily":
            # 检查是否开启了法定工作日提醒
            if getattr(event, 'workday_only', False):
                return "⏰ 工作日提醒"
            else:
                return "⏰ 每天提醒"
        elif event.event_type == "weekly":
            return "🔁 每周提醒"
        elif event.repeat_type == "once":
            if event.completed:
                date_parts = event.birth_date.split("-")
                return f"✅ 已完成于 {date_parts[0]}年{date_parts[1]}月{date_parts[2]}日"
            else:
                month, day, year, base_year, days_until = event.get_next_date_info()
                if days_until < 0:
                    date_parts = event.birth_date.split("-")
                    return f"⏰ 已过期 ({date_parts[0]}年{date_parts[1]}月{date_parts[2]}日)"
                elif days_until == 0:
                    return "🎯 今天执行"
                else:
                    date_parts = event.birth_date.split("-")
                    return f"⏰ {date_parts[0]}年{date_parts[1]}月{date_parts[2]}日"
        else:
            if base_year > 0 and base_year <= today.year:
                years_passed = today.year - base_year + 1
                if years_passed < 1:
                    years_passed = 1
                return f"💝 第{years_passed}年"
            else:
                return "💝 纪念日"
    
    def show_monthly_events():
        """显示每月事件列表"""
        global current_view, current_playing_event_id, current_music_state, card_duration_texts
        current_view = "monthly"
        
        # ========== 清空旧的卡片引用 ==========
        card_duration_texts.clear()

        events_list.controls.clear()
        
        monthly_events = []
        today = datetime.now().date()
        
        for event in events.values():
            if event.event_type == "monthly":
                month, day, year, base_year, days_until = event.get_next_date_info()
                monthly_events.append({
                    "event": event,
                    "days_until": days_until
                })
        
        # 按剩余天数排序（每月事件的剩余天数是指距离下一个提醒日的天数）
        monthly_events.sort(key=lambda x: x["days_until"])

        # ========== 调用排序函数 ==========
        event_list = [item["event"] for item in monthly_events]
        event_list = get_sorted_events_for_display(event_list)
        
        # ========== 始终显示标题行和下拉框 ==========
        if hasattr(refresh_events_list, 'view_dropdown'):
            title_text = f"🔄 每月事件 {len(monthly_events)} 个" if monthly_events else "🔄 每月事件 0 个"
            events_list.controls.append(ft.Row([
                ft.Text(title_text, size=14, weight=ft.FontWeight.BOLD, expand=True),
                refresh_events_list.view_dropdown,
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN))
            events_list.controls.append(ft.Divider(height=10))

        if not monthly_events:
            events_list.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Text("✨ 暂无每月事件", size=14, color=ft.Colors.GREEN_700),
                    ], spacing=8, ),
                    padding=20,
                )
            )
        else:
            for event in event_list:
                display_event_card(event, is_filter_mode=True)

            # 移除最后一个多余的分隔符
            if events_list.controls and isinstance(events_list.controls[-1], ft.Divider):
                events_list.controls.pop()

        page.update()

    def show_birthday_events():
        """显示生日事件列表"""
        global current_view, current_playing_event_id, current_music_state, card_duration_texts
        current_view = "birthday"

        # ========== 清空旧的卡片引用 ==========
        card_duration_texts.clear()

        events_list.controls.clear()
        
        birthday_events = []
        today = datetime.now().date()
        
        for event in events.values():
            if event.event_type == "birthday":
                month, day, year, base_year, days_until = event.get_next_date_info()
                birthday_events.append({
                    "event": event,
                    "days_until": days_until,
                    "base_year": base_year
                })
        
        # 按剩余天数排序
        birthday_events.sort(key=lambda x: x["days_until"])

        # ========== 调用排序函数 ==========
        event_list = [item["event"] for item in birthday_events]
        event_list = get_sorted_events_for_display(event_list)

        # ========== 调试：打印排序后的列表 ==========
        #print(f"[show_birthday_events] 排序后:")
        #for idx, e in enumerate(event_list):
            #is_playing = (e.id == current_playing_event_id and current_music_state in ["playing", "paused"])
            #print(f"  {idx}: {e.name}, is_playing: {is_playing}")

        # ========== 始终显示标题行和下拉框 ==========
        if hasattr(refresh_events_list, 'view_dropdown'):
            title_text = f"🎉 生日事件 {len(birthday_events)} 个" if birthday_events else "🎉 生日事件 0 个"
            events_list.controls.append(ft.Row([
                ft.Text(title_text, size=14, weight=ft.FontWeight.BOLD, expand=True),
                refresh_events_list.view_dropdown,
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN))
            events_list.controls.append(ft.Divider(height=10))
        
        if not birthday_events:
            events_list.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Text("✨ 暂无生日事件", size=14, color=ft.Colors.GREEN_700),
                    ], spacing=8, ),
                    padding=20,
                )
            )
        else:
            # ========== 直接使用排序后的 event_list ==========
            for event in event_list:
                display_event_card(event, is_filter_mode=True)

            # 移除最后一个多余的分隔符
            if events_list.controls and isinstance(events_list.controls[-1], ft.Divider):
                events_list.controls.pop()
        
        page.update()
        

    def show_event_events():
        """显示纪念日事件列表"""
        global current_view, current_playing_event_id, current_music_state, card_duration_texts
        current_view = "event"

        # ========== 清空旧的卡片引用 ==========
        card_duration_texts.clear()

        events_list.controls.clear()
        
        event_events_list = []
        today = datetime.now().date()
        
        for event in events.values():
            if event.event_type == "event":
                month, day, year, base_year, days_until = event.get_next_date_info()
                event_events_list.append({
                    "event": event,
                    "days_until": days_until,
                    "base_year": base_year
                })
        
        # 按剩余天数排序
        event_events_list.sort(key=lambda x: x["days_until"])

        # ========== 调用排序函数 ==========
        event_list = [item["event"] for item in event_events_list]
        event_list = get_sorted_events_for_display(event_list)

        # ========== 始终显示标题行和下拉框 ==========
        if hasattr(refresh_events_list, 'view_dropdown'):
            title_text = f"💝 纪念日事件 {len(event_events_list)} 个" if event_events_list else "💝 纪念日事件 0 个"
            events_list.controls.append(ft.Row([
                ft.Text(title_text, size=14, weight=ft.FontWeight.BOLD, expand=True),
                refresh_events_list.view_dropdown,
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN))
            events_list.controls.append(ft.Divider(height=10))
        
        if not event_events_list:
            events_list.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Text("✨ 暂无纪念日事件", size=14, color=ft.Colors.GREEN_700),
                    ], spacing=8, ),
                    padding=20,
                )
            )
        else:
            for event in event_list:
                display_event_card(event, is_filter_mode=True)

            # 移除最后一个多余的分隔符
            if events_list.controls and isinstance(events_list.controls[-1], ft.Divider):
                events_list.controls.pop()

        page.update()

    def show_once_events():
        """显示一次性事件列表"""
        global current_view, current_playing_event_id, current_music_state, card_duration_texts
        current_view = "once"

        # ========== 清空旧的卡片引用 ==========
        card_duration_texts.clear()

        events_list.controls.clear()
        
        once_events_list = []
        today = datetime.now().date()
        
        for event in events.values():
            if event.event_type == "once":
                month, day, year, base_year, days_until = event.get_next_date_info()
                once_events_list.append({
                    "event": event,
                    "days_until": days_until,
                    "base_year": base_year
                })
        
        # 按剩余天数排序
        once_events_list.sort(key=lambda x: x["days_until"])

        # ========== 调用排序函数 ==========
        event_list = [item["event"] for item in once_events_list]
        event_list = get_sorted_events_for_display(event_list)

        # ========== 始终显示标题行和下拉框 ==========
        if hasattr(refresh_events_list, 'view_dropdown'):
            title_text = f"🎯 一次性事件 {len(once_events_list)} 个" if once_events_list else "🎯 一次性事件 0 个"
            events_list.controls.append(ft.Row([
                ft.Text(title_text, size=14, weight=ft.FontWeight.BOLD, expand=True),
                refresh_events_list.view_dropdown,
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN))
            events_list.controls.append(ft.Divider(height=10))
        
        if not once_events_list:
            events_list.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Text("✨ 暂无一次性事件", size=14, color=ft.Colors.GREEN_700),
                    ], spacing=8, ),
                    padding=20,
                )
            )
        else:
            for event in event_list:
                display_event_card(event, is_filter_mode=True)

            # 移除最后一个多余的分隔符
            if events_list.controls and isinstance(events_list.controls[-1], ft.Divider):
                events_list.controls.pop()
        
        page.update()

    def on_view_change(e):
        """下拉框选择改变时的回调"""
        global current_view, previous_view

        # ========== 支持两种调用方式 ==========
        if hasattr(e, 'data'):
            # 从事件对象获取（Dropdown 调用）
            selected = e.data
        else:
            # 直接传入字符串（PopupMenuButton 调用）
            selected = e

        # 保存之前的视图
        previous_view = current_view
        current_view = selected

        print(f"[视图切换] 从 {previous_view} 切换到 {selected}")
    
        # 清空事件列表
        events_list.controls.clear()
        
        # 根据选择显示不同的事件列表
        if selected == "all":
            display_all_events()
            show_bottom_message("📋 已切换到全部事件视图")
        elif selected == "today":
            show_today_events()
            show_bottom_message("📌 已切换到今日事件视图")
        elif selected == "three_days":
            show_three_days_events()
            show_bottom_message("🔔 已切换到预警事件视图")
        elif selected == "daily":
            show_daily_events()
            show_bottom_message("⏰ 已切换到每日事件视图")
        elif selected == "weekly":
            show_weekly_events()
            show_bottom_message("🔁 已切换到每周事件视图")
        elif selected == "monthly":
            show_monthly_events()
            show_bottom_message("🔄 已切换到每月事件视图")
        elif selected == "birthday":
            show_birthday_events()
            show_bottom_message("🎉 已切换到生日事件视图")
        elif selected == "event":
            show_event_events()
            show_bottom_message("💝 已切换到纪念日事件视图")
        elif selected == "once":
            show_once_events()
            show_bottom_message("🎯 已切换到一次性事件视图")

        # 更新下拉框的显示值
        if hasattr(refresh_events_list, 'view_dropdown'):
            refresh_events_list.view_dropdown.value = selected
        
        page.update()

    def reset_to_all_events():
        """重置到全部事件视图"""
        global current_view
        current_view = "all"
        refresh_events_list()
        show_bottom_message("📋 已切换到全部事件视图")

    def get_view_title():
        """获取当前视图的标题"""
        global current_view
        
        if current_view == "all":
            return "全部事件视图"
        elif current_view == "today":
            return "今日事件视图"
        elif current_view == "three_days":
            return "预警事件视图"
        elif current_view == "daily":
            return "每日事件视图"
        elif current_view == "weekly":
            return "每周事件视图"
        elif current_view == "monthly":
            return "每月事件视图"
        elif current_view == "birthday":
            return "生日事件视图"
        elif current_view == "event":
            return "纪念日事件视图"
        elif current_view == "once":
            return "一次性事件视图"
        else:
            return "事件视图"

    def restore_previous_view():
        """恢复到之前的视图"""
        global current_view, previous_view
        
        print(f"[恢复视图] previous_view: {previous_view}, current_view: {current_view}")
        
        # 如果 previous_view 存在，恢复到该视图
        if previous_view:
            current_view = previous_view
        else:
            # 如果 previous_view 不存在，默认返回到全部事件
            current_view = "all"
        
        # 更新下拉框的值
        if hasattr(refresh_events_list, 'view_dropdown'):
            refresh_events_list.view_dropdown.value = current_view
        
        # 根据恢复的视图刷新事件列表
        if current_view == "all":
            display_all_events()
        elif current_view == "today":
            show_today_events()
        elif current_view == "three_days":
            show_three_days_events()
        elif current_view == "daily":
            show_daily_events()
        elif current_view == "weekly":
            show_weekly_events()
        elif current_view == "monthly":
            show_monthly_events()
        elif current_view == "birthday":
            show_birthday_events()
        elif current_view == "event":
            show_event_events()
        elif current_view == "once":
            show_once_events()
        else:
            display_all_events()
        
        show_bottom_message("已返回")
    
    # 重新计算 three_days_events
    def update_three_days_events():
        """更新3日内事件列表"""
        global three_days_events
        three_days_events = []
        for evt in events.values():
            if evt.event_type == "daily" or evt.event_type == "weekly":
                continue
            month, day, year, base_year, days_until = evt.get_next_date_info()
            if evt.repeat_type == "once" and (evt.completed or days_until < 0):
                continue
            if 0 < days_until <= 3:
                three_days_events.append((evt, days_until))
    
    def sort_events_by_playing(events_list):
        """将事件列表排序，播放中的排在第一位
        支持传入事件对象列表或字典列表（包含 'event' 键）
        """
        if not events_list:
            return events_list
        
        playing_events = []
        other_events = []
        
        for item in events_list:
            # ========== 判断是事件对象还是字典 ==========
            if hasattr(item, 'id'):  # 是事件对象
                event = item
            elif isinstance(item, dict) and 'event' in item:  # 是字典，包含 'event' 键
                event = item['event']
            else:
                other_events.append(item)
                continue
            
            # 判断是否是播放中的事件
            if event.id == current_playing_event_id and current_music_state in ["playing", "paused"]:
                playing_events.append(item)
            else:
                other_events.append(item)
        
        return playing_events + other_events

    def refresh_events_list(filter_date=None):
        global current_playing_event_id, current_music_state, three_days_events, current_view, current_selected_lunar, card_duration_texts
        
        # ========== 清空旧的卡片引用 ==========
        card_duration_texts.clear()
        
        # 清空事件列表控件
        events_list.controls.clear()
        today = datetime.now().date()
        
        # ========== 收集3日内事件（只统计生日、纪念日、一次性事件） ==========
        three_days_events = []
        for event in events.values():
            # 排除每天事件和每周事件
            if event.event_type == "daily" or event.event_type == "weekly":
                continue

            month, day, year, base_year, days_until = event.get_next_date_info()
            # 一次性事件特殊处理
            if event.repeat_type == "once" and (event.completed or days_until < 0):
                continue
            # 3日内（不包括今天）
            if 0 < days_until <= 3:
                three_days_events.append((event, days_until))
        
        # 更新 date_text 显示
        update_date_text_with_events(today, three_days_events)

        # ========== 筛选模式 ==========
        if filter_date is not None:
            filtered_events = []
            for event in events.values():
                # 排除每日事件
                if event.event_type == "daily" or event.repeat_type == "daily":
                    continue
                
                # 使用 is_event_on_date 方法判断事件是否在指定日期发生
                if event.is_event_on_date(filter_date):
                    days_until = (filter_date - today).days
                    if days_until < 0:
                        days_until = -1
                    filtered_events.append((event, days_until))
            
            # 调用排序函数
            event_list = [item[0] for item in filtered_events]
            event_list = sort_events_by_playing(event_list)
            
            # 重新构建 filtered_events
            filtered_events = []
            for event in event_list:
                days_until = (filter_date - today).days
                if days_until < 0:
                    days_until = -1
                filtered_events.append((event, days_until))
                
            # 显示筛选结果
            events_list.controls.clear()

            # 显示日期标题（阳历 + 农历）
            lunar_str = get_lunar_date_str(filter_date.year, filter_date.month, filter_date.day)
            date_title = f"🌙 农历 {lunar_str}"
            
            # 始终显示返回按钮/下拉框
            if hasattr(refresh_events_list, 'view_dropdown'):
                events_list.controls.append(ft.Row([
                    ft.Text(date_title, size=16, weight=ft.FontWeight.BOLD, expand=True),
                    refresh_events_list.view_dropdown,
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN))
                events_list.controls.append(ft.Divider(height=10))
            
            if not filtered_events:
                # 当天没有事件，显示提示和返回按钮
                events_list.controls.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Text("✨ 当天没有事件", size=14, color=ft.Colors.GREEN_700),
                            ft.Container(height=10),
                            ft.Button(
                                "📋 返回之前事件", 
                                on_click=lambda e: restore_previous_view(),
                                style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_700, color=ft.Colors.WHITE),
                            ),
                        ], spacing=8, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                        padding=20,
                    )
                )
            else:
                # 有事件，显示事件列表
                events_list.controls.append(ft.Text(f"✨ 当天有 {len(filtered_events)} 个事件", 
                                                    size=14, color=ft.Colors.GREEN_700))
                events_list.controls.append(ft.Divider(height=5))
                for event, days_until in filtered_events:
                    display_event_card(event, is_filter_mode=True, custom_days_until=days_until)
            
            update_event_count()
            page.update()
            return
        
        # ========== 非筛选模式 ==========
        if not events:
            events_list.controls.append(ft.Text("✨ 暂无事件，点击「+」添加", color=ft.Colors.GREY_500, size=14))
            page.update()
            return
        
        # ========== 确保下拉框存在 ==========
        if not hasattr(refresh_events_list, 'view_dropdown'):
            
            # 视图选项（使用统一定义）
            view_options = [
                ("all", "📋 全部事件"),
                ("today", "📌 今日事件"),
                ("three_days", "🔔 预警事件"),
                ("daily", "⏰ 每日事件"),
                ("weekly", "🔁 每周事件"),
                ("monthly", "🔄 每月事件"),
                ("birthday", "🎉 生日"),
                ("event", "💝 纪念日"),
                ("once", "🎯 一次性事件"),
            ]
            
            # 获取当前选中视图的显示文本
            def get_view_display_text(value):
                for v, text in view_options:
                    if v == value:
                        return text
                return "📋 全部事件"

            # 重新构建 items，在选项之间添加分割线
            popup_items = []
            for i, (value, text) in enumerate(view_options):
                popup_items.append(
                    ft.PopupMenuItem(
                        content=ft.Container(
                            content=ft.Text(text, size=14),
                            width=150,
                        ),
                        on_click=lambda e, val=value: select_view_popup(val),
                        height=40,
                    )
                )
                if i < len(view_options) - 1:
                    popup_items.append(
                        ft.PopupMenuItem(
                            content=ft.Divider(height=1, color=ft.Colors.GREY_300),
                            disabled=True,
                            height=2,
                        )
                    )

            view_popup = ft.PopupMenuButton(
                content=ft.Row([
                    ft.Text(get_view_display_text(current_view), size=14, weight=ft.FontWeight.BOLD),
                    ft.Icon(ft.Icons.ARROW_DROP_DOWN, size=18),
                ], spacing=5, alignment=ft.MainAxisAlignment.CENTER),
                items=popup_items,
                bgcolor=ft.Colors.WHITE,
            )

            def select_view_popup(value):
                view_popup.content.controls[0].value = get_view_display_text(value)
                on_view_change(value)
                page.update()

            refresh_events_list.view_dropdown = view_popup
            refresh_events_list.view_dropdown.value = current_view

        # ========== 获取视图标题 ==========
        def get_view_title():
            global current_view
            
            if current_view == "all":
                return f"📋 全部事件 ({len(events)}个)"
            elif current_view == "today":
                today = datetime.now().date()
                count = 0
                for event in events.values():
                    if event.event_type == "daily" or event.event_type == "weekly":
                        continue
                    month, day, year, base_year, days_until = event.get_next_date_info()
                    if month == today.month and day == today.day:
                        if event.repeat_type == "once":
                            if not event.completed and days_until >= 0:
                                count += 1
                        else:
                            count += 1
                return f"📌 今日事件 ({count}个)" if count > 0 else "📌 今日事件"
            elif current_view == "three_days":
                return "🔔 预警事件"
            elif current_view == "daily":
                daily_count = len([e for e in events.values() if e.event_type == "daily"])
                return f"⏰ 每日事件 ({daily_count}个)" if daily_count > 0 else "⏰ 每日事件"
            elif current_view == "weekly":
                weekly_count = len([e for e in events.values() if e.event_type == "weekly"])
                return f"🔁 每周事件 ({weekly_count}个)" if weekly_count > 0 else "🔁 每周事件"
            elif current_view == "monthly":
                monthly_count = len([e for e in events.values() if e.event_type == "monthly"])
                return f"🔄 每月事件 ({monthly_count}个)" if monthly_count > 0 else "🔄 每月事件"
            elif current_view == "birthday":
                birthday_count = len([e for e in events.values() if e.event_type == "birthday"])
                return f"🎉 生日 ({birthday_count}个)" if birthday_count > 0 else "🎉 生日"
            elif current_view == "event":
                event_count = len([e for e in events.values() if e.event_type == "event"])
                return f"💝 纪念日 ({event_count}个)" if event_count > 0 else "💝 纪念日"
            elif current_view == "once":
                once_count = len([e for e in events.values() if e.repeat_type == "once"])
                return f"🎯 一次性事件 ({once_count}个)" if once_count > 0 else "🎯 一次性事件"
            return "事件列表"

        # ========== 添加标题行 ==========
        events_list.controls.append(ft.Row([
            ft.Text(get_view_title(), size=14, weight=ft.FontWeight.BOLD, expand=True),
            refresh_events_list.view_dropdown,
        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN))
        events_list.controls.append(ft.Divider(height=10))
        
        update_event_count()
        page.update()

        # ========== 收集事件数据 ==========
        today_events = []
        all_events = []
        
        for event in events.values():
            month, day, year, base_year, days_until = event.get_next_date_info()
            
            # 根据事件类型计算年龄/年份显示
            if event.event_type == "birthday":
                if base_year > 0 and base_year <= today.year:
                    age = today.year - base_year
                    age_text = f"🎉 {age}岁"
                else:
                    age_text = "🎉 生日"
            elif event.event_type == "monthly":
                age_text = "🔄 每月提醒"
            elif event.event_type == "daily":
                age_text = "⏰ 每天提醒"
            elif event.event_type == "weekly":
                age_text = "🔁 每周提醒"
            elif event.repeat_type == "once":
                age_text = ""
            else:  # event
                if base_year > 0 and base_year <= today.year:
                    years_passed = today.year - base_year + 1
                    if years_passed < 1:
                        years_passed = 1
                    age_text = f"💝 第{years_passed}年"
                else:
                    age_text = "💝 纪念日"
            
            # 判断是否是今日事件
            is_today = False
            if event.event_type != "daily" and event.event_type != "weekly":
                is_today = (month == today.month and day == today.day)
            
            event_info = {
                "event": event, 
                "month": month, 
                "day": day, 
                "age_text": age_text, 
                "days_until": days_until, 
                "is_today": is_today,
                "base_year": base_year
            }
            all_events.append(event_info)
            if is_today:
                today_events.append(event_info)
        
        # 根据当前视图选择标题和显示内容
        if current_view == "today":
            display_events = today_events
            if not display_events:
                events_list.controls.append(ft.Text("🎉 今日没有事件", size=14, color=ft.Colors.GREEN_700))
                update_event_count()
                page.update()
                return
        elif current_view == "all":
            display_events = sorted(all_events, key=lambda x: x["days_until"])
        elif current_view == "daily":
            show_daily_events()
            return
        elif current_view == "weekly":
            show_weekly_events()
            return
        elif current_view == "monthly":
            show_monthly_events()
            return
        elif current_view == "three_days":
            show_three_days_events()
            return
        elif current_view == "birthday":
            show_birthday_events()
            return
        elif current_view == "event":
            show_event_events()
            return
        elif current_view == "once":
            show_once_events()
            return
        else:
            display_events = sorted(all_events, key=lambda x: x["days_until"])
        
        # ========== 显示事件卡片 ==========
        for info in display_events:
            display_event_card(info["event"], is_filter_mode=True)
        
        # 移除最后一个多余的分隔符
        if events_list.controls and isinstance(events_list.controls[-1], ft.Divider):
            events_list.controls.pop()
        
        update_event_count()
        page.update()
    
    def show_bottom_message(message, is_error=False):
        """显示底部消息（使用 SnackBar）"""
        print(f"[底部消息] {message}")
        
        # 根据消息类型设置颜色和图标
        if "✅" in message or "成功" in message or "完成" in message:
            color = ft.Colors.GREEN_700
            icon = "✅ "
        elif "❌" in message or "失败" in message or "错误" in message:
            color = ft.Colors.RED_700
            icon = "❌ "
        elif "⚠️" in message or "警告" in message:
            color = ft.Colors.ORANGE_700
            icon = "⚠️ "
        else:
            color = ft.Colors.BLUE_700
            icon = "ℹ️ "
        
        # 使用 SnackBar 显示消息
        try:
            if hasattr(page, 'show_snack_bar'):
                page.show_snack_bar(
                    ft.SnackBar(
                        content=ft.Text(f"{icon}{message}"),
                        bgcolor=color,
                        duration=3000,
                        behavior=ft.SnackBarBehavior.FLOATING,
                    )
                )
            else:
                snack = ft.SnackBar(
                    content=ft.Text(f"{icon}{message}"),
                    bgcolor=color,
                    open=True,
                    duration=3000,
                )
                page.overlay.append(snack)
                snack.open = True
                page.update()
                def close_snack():
                    time.sleep(3)
                    if snack in page.overlay:
                        page.overlay.remove(snack)
                        page.update()
                threading.Thread(target=close_snack, daemon=True).start()
        except Exception as e:
            print(f"显示 SnackBar 失败: {e}")
    
    # 保留原有的 show_snack_bar 作为兼容
    def show_snack_bar(message):
        """显示底部提示（兼容旧代码）"""
        show_bottom_message(message)
        # 整个函数已禁用，改用 show_bottom_message
        """
        print(f"[show_snack_bar] 调用显示: {message}")
        
        def close_sheet(e):
            sheet.open = False
            page.update()
        
        sheet = ft.BottomSheet(
            ft.Container(
                content=ft.Text(message, size=16),
                padding=20,
            ),
            open=True,
        )
        page.overlay.append(sheet)
        page.update()
        
        # 2秒后自动关闭
        def auto_close():
            time.sleep(2)
            sheet.open = False
            page.update()
        
        threading.Thread(target=auto_close, daemon=True).start()
        """

    def show_snack_bar2(message):
        """显示底部提示（兼容旧代码）"""
        print(f"[show_snack_bar] 调用显示: {message}")
        
        def close_sheet(e):
            sheet.open = False
            page.update()
        
        sheet = ft.BottomSheet(
            ft.Container(
                content=ft.Text(message, size=16),
                padding=20,
            ),
            open=True,
        )
        page.overlay.append(sheet)
        page.update()
        
        # 2秒后自动关闭
        def auto_close():
            time.sleep(2)
            sheet.open = False
            page.update()
        
        threading.Thread(target=auto_close, daemon=True).start()
    
    def close_dialog():
        nonlocal dialog_container
        if dialog_container and dialog_container in page.overlay:
            page.overlay.remove(dialog_container)
            dialog_container = None
            page.update()
    
    def edit_event_dialog(event_id):
        nonlocal selected_event
        selected_event = events.get(event_id)
        if selected_event:
            open_add_dialog(is_edit=True)
    
    # 在 open_add_dialog 函数开始处添加权限检查
    def open_add_dialog(is_edit=False):
        nonlocal dialog_container, selected_event
        close_dialog()

        # 在函数开头定义变量存储当前选中的事件类型
        
        event_type_selected = selected_event.event_type if selected_event else "birthday"
        calendar_selected = selected_event.calendar_type if selected_event else "solar"

        # Android 平台检查存储权限
        if platform.system() == "Linux":
            def check_storage_permission():
                if hasattr(page, 'can_access_storage'):
                    # 检查是否有访问权限
                    pass
            
            # 尝试请求权限
            if hasattr(page, 'request_permission'):
                try:
                    page.request_permission("android.permission.READ_EXTERNAL_STORAGE")
                    page.request_permission("android.permission.WRITE_EXTERNAL_STORAGE")
                    page.request_permission("android.permission.READ_MEDIA_AUDIO")
                    print("[Android] 已请求存储权限")
                except Exception as e:
                    print(f"[Android] 权限请求失败: {e}")
        
        # 创建 FilePicker 并添加到页面服务
        file_picker = ft.FilePicker()
        page.services.append(file_picker)
        
        # 显示选中的文件名
        selected_file_display = ft.Text(value="", size=12, color=ft.Colors.GREEN_700)

        # ========== 在这里添加新控件 ==========
    
        # 日期选择器
        # 先计算初始日期（如果是编辑模式）
        #if is_edit and selected_event and selected_event.repeat_type == "once":
        initial_date = None
        if is_edit and selected_event:
            if selected_event.event_type == "monthly":
                # 每月事件：使用当前月份 + 事件保存的日
                day_num = int(selected_event.birth_date) if selected_event.birth_date else 1
                now = datetime.now()
                try:
                    # 构造当前年月 + 保存的日
                    initial_date = datetime(now.year, now.month, day_num)
                except ValueError:
                    # 处理无效日期（如2月30日）
                    # 使用该月的最后一天
                    if day_num > 28:
                        # 计算该月的最后一天
                        if now.month == 2:
                            # 2月：判断闰年
                            import calendar
                            last_day = 29 if calendar.isleap(now.year) else 28
                            day_num = min(day_num, last_day)
                        elif now.month in [4, 6, 9, 11]:
                            day_num = min(day_num, 30)
                        initial_date = datetime(now.year, now.month, day_num)
            elif selected_event.repeat_type == "once" or selected_event.event_type in ["birthday", "event"]:
                # 一次性事件和生日/纪念日：使用事件保存的完整日期
                try:
                    date_parts = selected_event.birth_date.split("-")
                    if len(date_parts) == 3:
                        initial_date = datetime(int(date_parts[0]), int(date_parts[1]), int(date_parts[2]))
                except:
                    pass

        date_picker = ft.DatePicker(
            first_date=datetime(1900, 1, 1),
            last_date=datetime(2100, 12, 31),
            value=initial_date,  # 设置初始值
            on_change=lambda e: on_date_selected(e),
        )
        
        def on_date_display_field_blur(e):
            # 名称输入框失去焦点时的操作
            pass

        # 日期显示字段
        date_display_field = ft.TextField(
            label="日期",
            hint_text="点击选择日期",
            read_only=True,
            expand=True,
            on_click=lambda e: page.show_dialog(date_picker),  # 使用 show_dialog
            visible=True,  # 默认可见，但会根据事件类型动态调整
            on_blur=on_date_display_field_blur,  # 添加失去焦点事件
        )
        
        # 添加多个时间提醒的按钮和列表
        reminders_list = ft.Column(spacing=5)
        
        # 保存到函数属性中，供 save_click 使用
        open_add_dialog.reminders_list = reminders_list
        
        # 定义添加提醒时间的函数
        def add_reminder_time(time_str=None):
            """添加提醒时间"""
            print(f"[添加提醒时间] time_str={time_str}")  # 调试输出

            # 解析初始时间（如果有）
            initial_hour = 9
            initial_minute = 0
            if time_str:
                try:
                    parts = time_str.split(":")
                    if len(parts) == 2:
                        initial_hour = int(parts[0])
                        initial_minute = int(parts[1])
                except:
                    pass

            # 创建时间选择器
            time_picker = ft.TimePicker()
            
            def on_time_display_field_blur(e):
                # 名称输入框失去焦点时的操作
                pass

            # 时间显示字段
            time_display_field = ft.TextField(
                label="提醒时间",
                hint_text="点击选择时间（可选）",
                read_only=True,
                #width=120,
                expand=True,
                value=time_str if time_str else "",  # 直接设置显示值
                on_blur=on_time_display_field_blur,  # 添加失去焦点事件
            )

            # 如果传入了时间参数，设置显示值
            if time_str:
                time_display_field.value = time_str
                print(f"[添加提醒时间] 设置初始时间: {time_str}")

            def open_time_picker(e):
                """打开时间选择器，并用当前显示的时间初始化"""
                try:
                    # 从显示字段读取当前时间
                    current_time = time_display_field.value
                    if current_time:
                        # 如果已有时间，用该时间初始化选择器
                        h, m = map(int, current_time.split(":"))
                        from datetime import time
                        time_picker.value = time(h, m)
                    else:
                        # 如果没有时间，使用默认时间
                        from datetime import time
                        time_picker.value = time(initial_hour, initial_minute)
                    
                    # 显示时间选择器
                    page.show_dialog(time_picker)
                except (ValueError, TypeError, AttributeError):
                    # 如果解析失败，使用默认时间
                    from datetime import time
                    time_picker.value = time(initial_hour, initial_minute)
                    page.show_dialog(time_picker)

            def on_time_selected(e):
                """时间选择后的回调"""
                if time_picker.value:
                    time_str = time_picker.value.strftime("%H:%M")
                    time_display_field.value = time_str
                    time_display_field.update()
                    page.update()

            # 绑定事件
            time_picker.on_change = on_time_selected

            # 设置点击字段时打开选择器
            time_display_field.on_click = open_time_picker

            checkbox = ft.Checkbox(value=True, label="启用")

            delete_button = ft.IconButton(
                ft.Icons.DELETE_OUTLINE,
                icon_size=20,
                icon_color=ft.Colors.RED_400,
            )

            # 创建行
            row = ft.Row([
                time_display_field,
                checkbox,
                delete_button,
            ], spacing=5, vertical_alignment=ft.CrossAxisAlignment.CENTER)
            
            # 设置删除按钮的点击事件（此时 row 已经定义）
            delete_button.on_click = lambda e, r=row: remove_reminder_row(r)
            
            reminders_list.controls.append(row)
            page.update()
        
        def remove_reminder_row(row):
            reminders_list.controls.remove(row)
            page.update()
        
        # 如果是编辑模式，加载已有的提醒时间
        if is_edit and selected_event and selected_event.reminders:
            for reminder in selected_event.reminders:
                add_reminder_time(reminder.get("time", "09:00"))

        # 定义回调函数
        def on_date_selected(e):
            if date_picker.value:
                # 时区转换
                local_date = date_picker.value + timedelta(days=1)
                
                year = local_date.year
                month = local_date.month
                day = local_date.day
                
                # 增加判断，如果是每月事件，只需要显示一个日
                print(f'打印事件类型测试：{event_type_selected}')
                if event_type_selected == "monthly":
                    date_display_field.value = f"{day:02d}"
                else:
                    date_display_field.value = f"{year:04d}-{month:02d}-{day:02d}"

                # 直接更新整个页面
                page.update()


        # ========== 1. 先定义 update_date_visibility 函数 ==========
        def update_date_visibility(e=None):
            import traceback
            """根据事件类型切换显示不同的日期输入控件"""
            #print(f"[调试] ========== update_date_visibility 被调用 ==========")
            #print(f"[调试] event_type_selected = {event_type_selected}")
            #print(f"[调试] 调用栈: {traceback.extract_stack()[-2].name}")
            
            if event_type_selected == "daily":
                # 每天提醒：隐藏所有日期控件，显示工作日选项
                weekday_row.visible = False
                calendar_dropdown.visible = False
                repeat_type.visible = False
                date_display_field.visible = False   # 隐藏日期选择器显示字段
                #workday_only_checkbox.visible = True # 显示工作日选项
                hint_text.value = "💡 提示: 每天提醒，可设置具体时间。开启「仅在法定工作日提醒」后，只在工作日触发提醒"
                
            elif event_type_selected == "weekly":
                # 每周提醒：隐藏日期选择器，隐藏工作日选项
                weekday_row.visible = True      # 显示星期选择
                calendar_dropdown.visible = False   # 隐藏历法选择
                repeat_type.visible = False
                date_display_field.visible = False     # 隐藏日期选择器显示字段
                #workday_only_checkbox.visible = False  # 隐藏工作日选项
                hint_text.value = "💡 提示: 每周提醒，选择日期后每周同一天提醒"
                
            elif event_type_selected == "monthly":
                # 每月提醒：只显示日，隐藏工作日选项
                weekday_row.visible = False
                calendar_dropdown.visible = False
                repeat_type.visible = False
                date_display_field.visible = True      # 显示日期选择器
                #workday_only_checkbox.visible = False  # 隐藏工作日选项
                hint_text.value = "💡 提示: 每月固定日期提醒，只需选择每月几号"
                
            elif event_type_selected == "once":
                # 一次性事件：显示完整日期和日期选择器，隐藏工作日选项
                weekday_row.visible = False
                calendar_dropdown.visible = True
                repeat_type.visible = False
                date_display_field.visible = True      # 显示日期选择器
                #workday_only_checkbox.visible = False  # 隐藏工作日选项
                hint_text.value = "💡 提示: 一次性事件只在指定日期提醒一次"
                
            else:
                # 生日/纪念日：显示完整日期和日期选择器，隐藏工作日选项
                weekday_row.visible = False
                calendar_dropdown.visible = True
                repeat_type.visible = True
                date_display_field.visible = True      # 显示日期选择器
                #workday_only_checkbox.visible = False  # 隐藏工作日选项
                if event_type_selected == "birthday":
                    hint_text.value = "💡 提示: 农历生日会自动计算每年对应的阳历日期"
                else:
                    hint_text.value = "💡 提示: 纪念日每年重复提醒，可设置农历或阳历"
            
            #print(f"[调试] date_row.visible = {date_row.visible}")
            #print(f"[调试] monthly_day_row.visible = {monthly_day_row.visible}")
            page.update()

        
        # 事件类型选项
        event_type_options = ["🎉 生日", "💝 纪念日/事件", "🔄 每月提醒", "🎯 一次性事件", "⏰  每天提醒", "🔁 每周提醒"]
        event_type_keys = ["birthday", "event", "monthly", "once", "daily", "weekly"]

        def get_event_type_key(text):
            for i, t in enumerate(event_type_options):
                if t == text:
                    return event_type_keys[i]
            return "birthday"

        def get_event_type_text(key):
            for i, k in enumerate(event_type_keys):
                if k == key:
                    return event_type_options[i]
            return event_type_options[0]

        # 获取初始值
        initial_event_type = selected_event.event_type if selected_event else "birthday"
        initial_event_type_text = get_event_type_text(initial_event_type)
        
        event_type_selected = initial_event_type

        def on_event_type_change(e):
            """事件类型改变回调"""
            nonlocal event_type_selected  # 重要：使用 nonlocal 修改外部变量
            value = e.control.value if hasattr(e, 'control') else e
            event_type_selected = get_event_type_key(value)
            
            print(f"[事件类型变化] 选中文本: {value}, key: {event_type_selected}")
            print(f"[调试] 选择的事件类型是: {event_type_selected}")  # 添加调试
            
            # 更新名称字段标签
            if event_type_selected == "birthday":
                name_field.label = "姓名"
                calendar_dropdown.visible = True
                weekday_dropdown.visible = False
                repeat_type.visible = True
                repeat_type.value = "yearly"
                date_display_field.visible = True
                hint_text.value = "💡 提示: 农历生日会自动计算每年对应的阳历日期"
            elif event_type_selected == "event":
                name_field.label = "事件名称"
                calendar_dropdown.visible = True
                weekday_dropdown.visible = False
                repeat_type.visible = True
                date_display_field.visible = True
                hint_text.value = "💡 提示: 纪念日每年重复提醒，可设置农历或阳历"
            elif event_type_selected == "monthly":
                name_field.label = "事件名称"
                calendar_dropdown.visible = False
                weekday_dropdown.visible = False
                repeat_type.visible = False
                repeat_type.value = "monthly"
                date_display_field.visible = True
                hint_text.value = "💡 提示: 每月固定日期提醒，只需选择每月几号（如：15号）"
            elif event_type_selected == "once":
                name_field.label = "事件名称"
                calendar_dropdown.visible = True
                weekday_dropdown.visible = False
                repeat_type.visible = False
                repeat_type.value = "once"
                date_display_field.visible = True
                hint_text.value = "💡 提示: 一次性事件只在指定日期提醒一次，提醒后会自动标记为已完成"
            elif event_type_selected == "daily":
                name_field.label = "事件名称"
                calendar_dropdown.visible = False
                weekday_dropdown.visible = False
                repeat_type.visible = False
                repeat_type.value = "daily"
                date_display_field.visible = False
                hint_text.value = "💡 提示: 每天提醒，可设置具体时间（如：08:30、18:30）"
            elif event_type_selected == "weekly":
                name_field.label = "事件名称"
                calendar_dropdown.visible = False
                weekday_dropdown.visible = True
                repeat_type.visible = False
                repeat_type.value = "weekly"
                date_display_field.visible = False
                hint_text.value = "💡 提示: 每周固定日期提醒，选择星期几"
            
            update_date_visibility()
            page.update()

        # 创建事件类型下拉框
        event_type_dropdown = SearchableDropdownEvtTp(
            page=page,  # 传入 page
            label="事件类型",
            options=event_type_options,
            value=initial_event_type_text,
            on_change=on_event_type_change,
        )

        # ========== 历法下拉框 ==========
        calendar_options = ["☀️ 阳历", "🌙 农历"]
        calendar_keys = ["solar", "lunar"]

        def get_calendar_key(text):
            for i, t in enumerate(calendar_options):
                if t == text:
                    return calendar_keys[i]
            return "solar"

        def get_calendar_text(key):
            for i, k in enumerate(calendar_keys):
                if k == key:
                    return calendar_options[i]
            return calendar_options[0]

        initial_calendar = selected_event.calendar_type if selected_event else "solar"
        initial_calendar_text = get_calendar_text(initial_calendar)

        def on_calendar_change(value):
            """历法选择变化"""
            nonlocal calendar_selected
            # ========== 修复：使用 get_calendar_key 函数 ==========
            calendar_selected = get_calendar_key(value)
            print(f"[历法变化] 选中: {value} -> key: {calendar_selected}")

        calendar_dropdown = SearchableDropdownEvtLf(
            page=page,  # 传入 page
            label="历法",
            options=calendar_options,
            value=get_calendar_text(calendar_selected),
            on_change=on_calendar_change,
        )
        
        print(f"[调试] 初始事件类型: {initial_event_type}")  # 添加调试

        # ========== 星期下拉框 ==========
        weekday_options = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        weekday_keys = ["1", "2", "3", "4", "5", "6", "7"]

        def get_weekday_key(text):
            for i, t in enumerate(weekday_options):
                if t == text:
                    return weekday_keys[i]
            return "1"

        def get_weekday_text(key):
            for i, k in enumerate(weekday_keys):
                if k == key:
                    return weekday_options[i]
            return weekday_options[0]
        
        # 获取当前星期
        current_weekday = datetime.now().isoweekday()  # 返回 1-7，1=周一，7=周日

        # 获取初始星期值
        if selected_event and selected_event.event_type == "weekly" and selected_event.birth_date:
            initial_weekday = selected_event.birth_date
        else:
            initial_weekday = str(current_weekday)  # 确保是字符串
        initial_weekday_text = get_weekday_text(initial_weekday)

        # ========== 根据初始事件类型设置可见性 ==========
        weekday_row_visible = (initial_event_type == "weekly")

        print(f"[调试] weekday_row_visible: {weekday_row_visible}")  # 添加调试

        weekday_dropdown = SearchableDropdownEvtWeek(
            page=page,  # 传入 page
            label="星期",
            options=weekday_options,
            value=initial_weekday_text,
            on_change=lambda e: None,
        )

        # 星期行
        weekday_row = ft.Row(
            [weekday_dropdown],
            alignment=ft.MainAxisAlignment.CENTER,
            visible=weekday_row_visible,  # 根据初始事件类型设置
        )

        # 如果 weekday_row 隐藏，也隐藏内部的 dropdown
        if not weekday_row_visible:
            weekday_dropdown.visible = False

        # 在事件类型选择之后添加重复类型选择
        repeat_type = ft.Dropdown(
            label="重复类型",
            options=[
                ft.dropdown.Option("yearly", "📅 每年重复"),
                ft.dropdown.Option("monthly", "📆 每月重复")
            ],
            value=selected_event.repeat_type if selected_event and hasattr(selected_event, 'repeat_type') else "yearly",
            expand=True,
        )
        
        # 添加确认对话框函数
        def show_lyrics_confirm_dialog(music_file_path, original_dir, original_basename, target_path):
            """显示是否选择歌词文件的确认对话框"""
            
            def close_dialog():
                if hasattr(show_lyrics_confirm_dialog, 'dialog') and show_lyrics_confirm_dialog.dialog in page.overlay:
                    page.overlay.remove(show_lyrics_confirm_dialog.dialog)
                    page.update()
            
            def on_yes(e):
                close_dialog()
                # 用户选择是，打开文件选择器选择歌词文件
                asyncio.create_task(pick_lyrics_file_after_music(music_file_path, original_dir, original_basename, target_path))
            
            def on_no(e):
                close_dialog()
                selected_file_display.value = f"已保存音乐: {os.path.basename(target_path)}"
                page.update()
                show_snack_bar(f"音乐已保存，未添加歌词")
            
            def on_cancel(e):
                close_dialog()
            
            # 创建对话框容器
            dialog_content = ft.Container(
                content=ft.Column([
                    ft.Text("添加歌词", size=18, weight=ft.FontWeight.BOLD),
                    ft.Divider(height=5),
                    ft.Text("是否要添加歌词文件？", size=14),
                    ft.Text("如果选择「是」，请手动选择 .lrc 歌词文件", size=12, color=ft.Colors.GREY_600),
                    ft.Text("如果选择「否」，后续可在线搜索歌词", size=12, color=ft.Colors.GREY_600),
                    ft.Divider(height=10),
                    ft.Row([
                        ft.ElevatedButton("是", on_click=on_yes, expand=True),
                        #ft.Button("是", on_click=on_yes, expand=True, style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_700, color=ft.Colors.WHITE)),
                        ft.ElevatedButton("否", on_click=on_no, expand=True),
                        ft.TextButton("取消", on_click=on_cancel),
                    ], alignment=ft.MainAxisAlignment.CENTER, spacing=10),
                ], spacing=10, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                width=300,
                padding=20,
                bgcolor=ft.Colors.WHITE,
                border_radius=10,
            )
            
            dialog = ft.Container(
                content=ft.Column([
                    ft.Container(expand=True),
                    ft.Row([
                        ft.Container(expand=True),
                        dialog_content,
                        ft.Container(expand=True),
                    ]),
                    ft.Container(expand=True),
                ]),
                expand=True,
                bgcolor=ft.Colors.BLACK26,
                on_click=on_cancel,
            )
            
            show_lyrics_confirm_dialog.dialog = dialog
            page.overlay.append(dialog)
            page.update()

        async def pick_lyrics_file_after_music(music_file_path, original_dir, original_basename, target_path):
            """选择音乐后手动选择歌词文件（Android 兼容版 - 直接读取文件内容）"""
            
            print(f"[调试] pick_lyrics_file_after_music 被调用")
            print(f"[调试] 平台: {platform.system()}")
            
            # 创建 FilePicker
            if not hasattr(page, 'lyrics_file_picker'):
                page.lyrics_file_picker = ft.FilePicker()
                page.services.append(page.lyrics_file_picker)
                page.update()
            
            try:
                result = await page.lyrics_file_picker.pick_files(
                    allow_multiple=False,
                    allowed_extensions=["lrc"],
                    dialog_title="选择歌词文件 (.lrc)"
                )
                
                print(f"[调试] 选择结果类型: {type(result)}")
                print(f"[调试] 选择结果: {result}")
                
                if result and len(result) > 0:
                    lrc_file = result[0]
                    lrc_path = lrc_file.path
                    lrc_name = lrc_file.name
                    target_lrc_path = os.path.splitext(target_path)[0] + ".lrc"
                    
                    print(f"[调试] 选择的歌词文件: {lrc_path}")
                    print(f"[调试] 目标歌词路径: {target_lrc_path}")
                    
                    # Android 兼容方案：直接读取文件内容
                    try:
                        if platform.system() == "Windows":
                            # Windows 直接复制
                            import shutil
                            shutil.copy2(lrc_path, target_lrc_path)
                        else:
                            # Android：尝试多种方法复制
                            lrc_content = None
                            
                            # 方法1：直接读取路径
                            try:
                                with open(lrc_path, 'r', encoding='utf-8') as f:
                                    lrc_content = f.read()
                                print(f"[调试] 方法1成功: 直接读取路径")
                            except Exception as e1:
                                print(f"[调试] 方法1失败: {e1}")
                                
                                # 方法2：尝试使用 lrc_file 对象的其他属性
                                try:
                                    # 某些版本的 Flet 可能提供 bytes 属性
                                    if hasattr(lrc_file, 'bytes') and lrc_file.bytes:
                                        lrc_content = lrc_file.bytes.decode('utf-8')
                                        print(f"[调试] 方法2成功: 使用 bytes 属性")
                                except Exception as e2:
                                    print(f"[调试] 方法2失败: {e2}")
                            
                            # 如果成功读取到内容，写入目标文件
                            if lrc_content:
                                with open(target_lrc_path, 'w', encoding='utf-8') as f:
                                    f.write(lrc_content)
                                print(f"[调试] 歌词内容已写入: {len(lrc_content)} 字符")
                            else:
                                # 方法3：尝试二进制读写
                                try:
                                    with open(lrc_path, 'rb') as src:
                                        with open(target_lrc_path, 'wb') as dst:
                                            dst.write(src.read())
                                    print(f"[调试] 方法3成功: 二进制读写")
                                except Exception as e3:
                                    print(f"[调试] 方法3失败: {e3}")
                                    raise Exception("无法读取歌词文件内容")
                        
                        selected_file_display.value = f"已保存音乐和歌词: {os.path.basename(target_path)}"
                        show_snack_bar(f"✅ 歌词已添加: {lrc_name}")
                        
                        # 显示歌词预览
                        if os.path.exists(target_lrc_path):
                            show_lyrics_preview(target_lrc_path)
                            
                    except PermissionError as pe:
                        print(f"[调试] 权限错误: {pe}")
                        show_snack_bar(f"⚠️ 权限不足，请检查应用权限")
                    except Exception as copy_err:
                        print(f"[调试] 复制失败: {copy_err}")
                        show_snack_bar(f"⚠️ 复制歌词失败: {str(copy_err)}")
                else:
                    selected_file_display.value = f"已保存音乐: {os.path.basename(target_path)}"
                    show_snack_bar(f"音乐已保存，未添加歌词")
                    
            except Exception as ex:
                print(f"[调试] 选择歌词出错: {ex}")
                selected_file_display.value = f"已保存音乐: {os.path.basename(target_path)}"
                show_snack_bar(f"选择歌词失败: {str(ex)}")
            
            page.update()

        async def safe_read_file(file_path):
            """安全地读取文件内容（Android 兼容）"""
            try:
                # 尝试直接读取
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read()
            except:
                try:
                    # 尝试二进制读取然后解码
                    with open(file_path, 'rb') as f:
                        return f.read().decode('utf-8', errors='ignore')
                except:
                    return None
                
        # 处理文件选择
        # 修改 handle_pick_files 函数，增加 Android 平台的兼容性
        async def handle_pick_files(e):
            files = await file_picker.pick_files(allow_multiple=False, allowed_extensions=["mp3", "wav", "flac", "m4a"])
            if files:
                file = files[0]
                original_path = file.path
                
                # 创建持久音乐目录
                music_dir = os.path.join(os.path.dirname(get_data_file_path("")), "music")
                os.makedirs(music_dir, exist_ok=True)
                
                # 完全保持原始文件名
                original_filename = os.path.basename(original_path)
                target_path = os.path.join(music_dir, original_filename)
                
                # 复制音乐文件
                music_copied = False
                if os.path.exists(target_path):
                    try:
                        if os.path.getsize(original_path) == os.path.getsize(target_path):
                            music_field.value = target_path
                            music_copied = True
                            show_snack_bar(f"音乐文件已存在: {original_filename}")
                        else:
                            show_snack_bar(f"文件已存在，请手动重命名或选择其他文件")
                            return
                    except:
                        music_field.value = target_path
                        music_copied = True
                else:
                    # 复制音乐文件
                    try:
                        import shutil
                        if platform.system() == "Linux":  # Android
                            with open(original_path, 'rb') as src:
                                with open(target_path, 'wb') as dst:
                                    dst.write(src.read())
                        else:
                            shutil.copy2(original_path, target_path)
                        
                        music_field.value = target_path
                        music_copied = True
                        show_snack_bar(f"音乐已保存: {original_filename}")
                    except Exception as e:
                        show_snack_bar(f"复制文件失败: {str(e)}")
                        return
                
                # ========== 所有平台统一弹出对话框询问是否添加歌词 ==========
                if music_copied:
                    original_dir = os.path.dirname(original_path)
                    original_basename = os.path.splitext(original_filename)[0]
                    
                    # 显示确认对话框（所有平台都使用）
                    show_lyrics_confirm_dialog(music_field.value, original_dir, original_basename, target_path)
                
                page.update()

        # 在音乐按钮行添加一个"选择歌词"按钮
        async def pick_lyrics_file(e):
            """手动选择歌词文件"""
            files = await file_picker.pick_files(allow_multiple=False, allowed_extensions=["lrc"])
            if files:
                file = files[0]
                lrc_path = file.path
                
                # 获取当前音乐文件路径
                current_music = music_field.value.strip()
                if not current_music:
                    show_snack_bar("请先选择音乐文件")
                    return
                
                # 目标歌词路径
                target_lrc_path = os.path.splitext(current_music)[0] + ".lrc"
                
                try:
                    if platform.system() == "Linux":  # Android
                        with open(lrc_path, 'rb') as src:
                            with open(target_lrc_path, 'wb') as dst:
                                dst.write(src.read())
                    else:
                        import shutil
                        shutil.copy2(lrc_path, target_lrc_path)
                    
                    show_snack_bar(f"歌词已复制: {os.path.basename(target_lrc_path)}")
                    selected_file_display.value = f"音乐: {os.path.basename(current_music)}, 歌词已添加"
                    
                    # 显示歌词预览
                    if os.path.exists(target_lrc_path):
                        show_lyrics_preview(target_lrc_path)
                except Exception as e:
                    show_snack_bar(f"复制歌词失败: {str(e)}")

        # 添加歌词预览函数（放在 handle_pick_files 函数后面）
        def show_lyrics_preview(lrc_path):
            """显示歌词预览（Android 兼容）"""
            try:
                if not os.path.exists(lrc_path):
                    print(f"[歌词预览] 文件不存在: {lrc_path}")
                    return
                
                # 直接读取文件（已经通过前面的方法复制到应用目录了）
                with open(lrc_path, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                    
                preview_lines = []
                for line in lines:
                    line = line.strip()
                    if line and '[' in line and ']' in line:
                        # 提取歌词文本
                        match = re.search(r'\]\s*(.+)$', line)
                        if match:
                            text = match.group(1).strip()
                            # 过滤掉元数据行
                            if text and not text.startswith('[') and text not in ['', '作词', '作曲', '编曲', '制作']:
                                preview_lines.append(text)
                                if len(preview_lines) >= 5:
                                    break
                
                if preview_lines:
                    preview_text = "🎤 歌词预览:\n" + "\n".join(preview_lines[:3])
                    if len(preview_lines) > 3:
                        preview_text += f"\n... 共 {len(preview_lines)} 行"
                    #show_snack_bar2(preview_text)
                    print(preview_text)
                else:
                    # 如果提取不到歌词，显示文件基本信息
                    file_size = os.path.getsize(lrc_path)
                    show_snack_bar2(f"📝 歌词文件已添加 ({file_size} 字节)")
                    
            except Exception as e:
                print(f"[歌词预览] 显示失败: {e}")
        
        # 选择音乐文件的函数
        #def pick_music_file(e):
            #asyncio.create_task(handle_pick_files(e))
        
        def pick_music_file(e):
            """选择音乐文件 - 带确认对话框（保留复制到程序目录和歌词选择功能）"""
            
            # 存储对话框容器的引用
            menu_container = None
            
            def close_menu():
                nonlocal menu_container
                if menu_container and menu_container in page.overlay:
                    page.overlay.remove(menu_container)
                    menu_container = None
                    page.update()
            
            async def select_music_file():
                file_picker = None
                try:
                    # 创建 FilePicker
                    file_picker = ft.FilePicker()
                    page.services.append(file_picker)
                    page.update()
                    
                    result = await file_picker.pick_files(
                        allow_multiple=False,
                        allowed_extensions=["mp3", "wav", "flac", "m4a"],
                        dialog_title="选择音乐文件"
                    )
                    
                    # 移除 FilePicker
                    if file_picker and file_picker in page.overlay:
                        page.services.remove(file_picker)
                    page.update()
                    
                    if not result or len(result) == 0:
                        show_bottom_message("未选择音乐文件")
                        return
                    
                    # 获取原始文件信息
                    original_file = result[0]
                    
                    # 创建音乐目录（应用私有目录）
                    music_dir = os.path.join(os.path.dirname(get_data_file_path("")), "music")
                    os.makedirs(music_dir, exist_ok=True)
                    
                    # 获取原始文件名
                    if hasattr(original_file, 'name'):
                        original_filename = original_file.name
                    else:
                        original_filename = os.path.basename(original_file.path) if hasattr(original_file, 'path') else "music.mp3"
                    
                    # 目标路径
                    target_path = os.path.join(music_dir, original_filename)
                    
                    # 复制音乐文件到程序目录
                    if hasattr(original_file, 'path'):
                        # Windows：直接复制文件
                        import shutil
                        shutil.copy2(original_file.path, target_path)
                    elif hasattr(original_file, 'bytes'):
                        # 移动端：写入字节内容
                        with open(target_path, 'wb') as f:
                            f.write(original_file.bytes)
                    else:
                        show_bottom_message("无法读取文件")
                        return
                    
                    # 更新音乐文件路径
                    music_field.value = target_path
                    selected_file_display.value = f"已选择: {original_filename}"
                    show_bottom_message(f"音乐已保存: {original_filename}")
                    page.update()
                    
                    # ========== 询问是否添加歌词 ==========
                    # 获取原始文件所在目录（用于搜索同名的歌词文件）
                    original_dir = None
                    original_basename = None
                    
                    if hasattr(original_file, 'path'):
                        original_dir = os.path.dirname(original_file.path)
                        original_basename = os.path.splitext(os.path.basename(original_file.path))[0]
                    elif hasattr(original_file, 'name'):
                        # 移动端：没有原始路径，只能从文件名获取
                        original_basename = os.path.splitext(original_filename)[0]
                    
                    # 显示歌词确认对话框
                    show_lyrics_confirm_dialog(target_path, original_dir, original_basename, target_path)
                    
                except Exception as ex:
                    show_bottom_message(f"选择音乐失败: {str(ex)}")
                    print(f"选择音乐错误: {ex}")
                    import traceback
                    traceback.print_exc()
                finally:
                    if file_picker and file_picker in page.overlay:
                        page.overlay.remove(file_picker)
                    page.update()
            
            # 包装函数：先关闭菜单，再选择文件
            def on_select_music():
                close_menu()
                asyncio.create_task(select_music_file())
            
            def on_cancel():
                close_menu()
                show_bottom_message("已取消选择")
            
            # 创建菜单内容
            menu_content = ft.Container(
                content=ft.Column([
                    ft.Container(
                        content=ft.Icon(ft.Icons.MUSIC_NOTE, size=55, color=ft.Colors.BLUE_700),
                        padding=10,
                        bgcolor=ft.Colors.BLUE_50,
                        border_radius=50,
                    ),
                    ft.Text("选择音乐文件", size=22, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700, text_align=ft.TextAlign.CENTER),
                    ft.Divider(height=1, color=ft.Colors.GREY_300),
                    ft.Text("请选择音乐文件", size=14, color=ft.Colors.GREY_700, text_align=ft.TextAlign.CENTER),
                    ft.Text("支持格式: MP3, WAV, FLAC, M4A", size=12, color=ft.Colors.GREY_500, text_align=ft.TextAlign.CENTER),
                    ft.Divider(height=1, color=ft.Colors.GREY_300),
                    ft.Row([
                        ft.ElevatedButton(
                            "选择音乐", 
                            on_click=lambda e: on_select_music(), 
                            expand=True,
                            style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_700, color=ft.Colors.WHITE),
                        ),
                    ], alignment=ft.MainAxisAlignment.CENTER),
                    ft.Row([
                        ft.ElevatedButton(
                            "取消", 
                            on_click=lambda e: on_cancel(), 
                            expand=True,
                            style=ft.ButtonStyle(bgcolor=ft.Colors.GREY_100, color=ft.Colors.GREY_700),
                        ),
                    ], alignment=ft.MainAxisAlignment.CENTER),
                ], spacing=15, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                width=320,
                padding=20,
                bgcolor=ft.Colors.WHITE,
                border_radius=16,
            )
            
            menu_container = ft.Container(
                content=ft.Column([
                    ft.Container(expand=True),
                    ft.Row([
                        ft.Container(expand=True),
                        menu_content,
                        ft.Container(expand=True),
                    ]),
                    ft.Container(expand=True),
                ]),
                expand=True,
                bgcolor=ft.Colors.BLACK26,
                on_click=lambda e: close_menu(),
            )
            
            page.overlay.append(menu_container)
            page.update()

        
        # 清除音乐文件
        def clear_music(e):
            music_field.value = ""
            selected_file_display.value = ""
            page.update()
            show_snack_bar("已清除音乐文件路径")
        
        # 试听
        def test_play(e):
            global music_section_container,playback_buttons, current_music_state, current_playing_event_id, current_music_file
            file_path = music_field.value.strip()

            if not file_path:
                show_snack_bar("请输入音乐文件路径")
                return
            
            if not os.path.exists(file_path):
                show_snack_bar("音乐文件不存在，请选择有效的文件")
                return
            
            # 获取当前编辑的事件名称和ID（如果是编辑模式）
            test_event_name = None
            test_event_id = None

            if is_edit and selected_event:
                # 编辑模式：使用已有事件的名称和ID
                test_event_name = selected_event.name
                test_event_id  = selected_event.id
            elif name_field.value.strip():
                # 新增模式：使用输入的名称，ID 暂时为 None
                test_event_name = name_field.value.strip()
                test_event_id = None # 新增事件还没有 ID
            
            # 播放音乐，如果是新增模式则不传递 event_id
            if test_event_name:
                if test_event_id:
                    # 编辑模式：传递 event_id
                    play_music(file_path, loop=False, event_name=test_event_name, event_id=test_event_id)
                else:
                    # 新增模式：只传递 event_name
                    play_music(file_path, loop=False, event_name=test_event_name,event_id=None)
            else:
                play_music(file_path, loop=False, event_name="试听音乐", event_id=None)

            # 强制显示音乐区域
            if music_section_container:
                music_section_container.visible = True
                music_section_container.update()
            if playback_buttons:
                playback_buttons.visible = True
                playback_buttons.update()

            # 设置状态
            current_music_state = "playing"
            current_music_file = file_path
            if test_event_id:
                current_playing_event_id = test_event_id
            else:
                current_playing_event_id = None

            # 更新界面
            update_current_playing_info()
            page.update()

        def on_name_blur(e):
            # 名称输入框失去焦点时的操作
            pass

        # 定义所有控件
        # 名称输入框
        name_field = ft.TextField(
            label="姓名" if (selected_event and selected_event.event_type == "birthday") else "事件名称",
            value=selected_event.name if selected_event else "", 
            expand=True,
            on_blur=on_name_blur,  # 添加失去焦点事件
        )
        
        def on_music_blur(e):
            # 名称输入框失去焦点时的操作
            pass

        music_field = ft.TextField(
            label="音乐文件路径", 
            value=selected_event.sound_file if selected_event else "", 
            hint_text="可直接输入路径，或点击按钮选择",
            expand=True,
            on_blur=on_music_blur,  # 添加失去焦点事件
        )
        
        # 提示文本
        hint_text = ft.Text(
            "💡 提示: 农历生日会自动计算每年对应的阳历日期", 
            size=11, 
            color=ft.Colors.GREY_500
        )

        # 按钮行 - 换行显示（使用 Column 或 Wrap）
        music_buttons = ft.Row(
            controls=[
                ft.Button("📁 选择", on_click=pick_music_file, expand=True, style=ft.ButtonStyle(text_style=ft.TextStyle(size=12),)),
                #ft.Button("📝 歌词", on_click=lambda e: asyncio.create_task(pick_lyrics_file(e)), expand=True, style=ft.ButtonStyle(text_style=ft.TextStyle(size=12),)),
                ft.Button("🗑️ 清除", on_click=clear_music, expand=True, style=ft.ButtonStyle(text_style=ft.TextStyle(size=12),)),
                ft.Button("▶️ 试听", on_click=test_play, expand=True, style=ft.ButtonStyle(text_style=ft.TextStyle(size=12),)),
            ],
            spacing=5,
        )

        def on_search_keyword_blur(e):
            # 名称输入框失去焦点时的操作
            pass

        # ========== 音乐搜索相关控件 ==========
        search_keyword_field = ft.TextField(
            label="搜索歌曲", 
            hint_text="输入歌曲名或歌手名",
            expand=True,
            on_blur=on_search_keyword_blur,  # 添加失去焦点事件
        )

        search_btn = ft.Button("🔍 搜索", expand=True)
        search_results_dropdown = ft.Dropdown(
            label="搜索结果",
            hint_text="点击搜索后选择歌曲",
            expand=True,
            options=[],
        )
        download_btn = ft.Button("📥 下载并应用", expand=True)
        search_status = ft.Text("", size=11, color=ft.Colors.GREY_500)
        
        search_results = []
        
        # ========== 非 Android 平台才定义函数和绑定事件 ==========
        #print(f"测试打印平台： {platform.system()}")
        #print(f"测试打印平台： {IS_WINDOWS}")
        if IS_WINDOWS:
            # 定义搜索函数
            def do_search(e):
                keyword = search_keyword_field.value.strip()
                print(f"[搜索] 按钮被点击！关键词: '{keyword}'")
                
                if not keyword:
                    print("[搜索] 关键词为空，显示提示")
                    show_snack_bar("请输入歌曲名称")  # 直接调用，不在线程中
                    return
                
                search_btn.disabled = True
                search_btn.text = "搜索中..."
                search_status.value = "正在搜索..."
                search_status.color = ft.Colors.BLUE_700
                page.update()
                
                def search_thread():
                    nonlocal search_results
                    print(f"[搜索线程] 开始执行，关键词: {keyword}")
                    try:
                        downloader = LyricsDownloader()
                        search_url = f"https://www.gequbao.com/s/{keyword}"
                        print(f"[搜索线程] 请求URL: {search_url}")
                        
                        headers = {'User-Agent': downloader.get_random_ua()}
                        response = downloader.session.get(search_url, headers=headers, timeout=15)
                        response.encoding = 'utf-8'
                        print(f"[搜索线程] 响应状态码: {response.status_code}")
                        
                        if response.status_code == 200:
                            pattern = r'<a href="/music/(\d+)"[^>]*>.*?<span class="text-primary[^"]*"[^>]*>(.*?)</span>.*?<small class="text-jade[^"]*"[^>]*>(.*?)</small>'
                            matches = re.findall(pattern, response.text, re.DOTALL)
                            print(f"[搜索线程] 找到 {len(matches)} 个匹配项")
                            
                            search_results = []
                            options = []
                            for music_id, song_name, artist in matches[:10]:
                                song_name = re.sub(r'<[^>]+>', '', song_name).strip()
                                artist = re.sub(r'<[^>]+>', '', artist).strip()
                                if song_name:
                                    search_results.append({
                                        'id': music_id,
                                        'name': song_name,
                                        'artist': artist if artist else "未知歌手",
                                        'url': f"https://www.gequbao.com/music/{music_id}"
                                    })
                                    display_text = f"{song_name} - {artist}" if artist else song_name
                                    options.append(ft.dropdown.Option(music_id, display_text))
                                    print(f"[搜索线程] 歌曲: {display_text}")
                            
                            # 使用 threading.Timer 在主线程中更新UI
                            threading.Timer(0.1, lambda: update_search_results(options)).start()
                        else:
                            threading.Timer(0.1, lambda: show_snack_bar(f"搜索失败，状态码: {response.status_code}")).start()
                    except requests.exceptions.ConnectionError as e:
                        print(f"网络连接失败: {e}")
                        threading.Timer(0.1, lambda: show_snack_bar("网络连接失败，请检查网络")).start()
                    except requests.exceptions.Timeout as e:
                        print(f"请求超时: {e}")
                        threading.Timer(0.1, lambda: show_snack_bar("请求超时，请稍后重试")).start()
                    except Exception as e:
                        print(f"搜索出错: {e}")
                        threading.Timer(0.1, lambda: show_snack_bar(f"搜索出错: {str(e)}")).start()
                    finally:
                        threading.Timer(0.1, reset_search_btn).start()
                
                def update_search_results(options):
                    print(f"[UI更新] 更新搜索结果，共 {len(options)} 条")
                    search_results_dropdown.options = options
                    if options:
                        search_results_dropdown.disabled = False
                        search_status.value = f"找到 {len(options)} 首歌曲，请选择"
                        search_status.color = ft.Colors.GREEN_700
                    else:
                        search_results_dropdown.disabled = True
                        download_btn.disabled = True
                        search_status.value = "未找到相关歌曲"
                        search_status.color = ft.Colors.RED_700
                    # 使用同步 update 方法
                    search_results_dropdown.update()
                    search_status.update()
                    download_btn.update()
                    page.update()
                
                def reset_search_btn():
                    search_btn.disabled = False
                    search_btn.text = "🔍 搜索"
                    search_btn.update()
                    page.update()
                
                threading.Thread(target=search_thread, daemon=True).start()
            
            def on_result_select(e):
                print(f"[选择] 选中歌曲ID: {search_results_dropdown.value}")
                print(f"[选择] search_results 内容: {search_results}")
                
                if search_results_dropdown.value:
                    for song in search_results:
                        print(f"[选择] 比较: song['id']={song['id']} ({type(song['id'])}), 选中值={search_results_dropdown.value} ({type(search_results_dropdown.value)})")
                        if str(song['id']) == str(search_results_dropdown.value):  # 确保类型一致
                            download_btn.disabled = False
                            search_status.value = f"已选择: {song['name']} - {song['artist']}"
                            search_status.color = ft.Colors.BLUE_700
                            print(f"[选择] 找到匹配: {song['name']}")
                            break
                    else:
                        download_btn.disabled = True
                        search_status.value = "请重新搜索选择"
                        search_status.color = ft.Colors.RED_700
                        print(f"[选择] 未找到匹配的歌曲")
                else:
                    download_btn.disabled = True
                    search_status.value = ""
                
                download_btn.update()
                search_status.update()
                page.update()
            
            def do_download(e):
                selected_id = search_results_dropdown.value
                print(f"[下载] 开始下载，选中ID: {selected_id}")
                if not selected_id:
                    return
                
                selected_song = None
                for song in search_results:
                    if song['id'] == selected_id:
                        selected_song = song
                        break
                
                if not selected_song:
                    show_snack_bar("未找到选中的歌曲")
                    return
                
                download_btn.disabled = True
                download_btn.text = "下载中..."
                page.update()
                
                def download_thread():
                    try:
                        print("[下载线程] 开始执行")
                        downloader = LyricsDownloader(page=page, show_snack_bar=show_snack_bar)
                        song_url = selected_song['url']
                        print(f"[下载线程] 歌曲URL: {song_url}")
                        #mp3_url = downloader.get_mp3_url_simple(song_url)
                        # 测试下载歌曲宝的音乐
                        mp3_url = downloader.get_mp3_url_auto(song_url)
                        print(f"[下载线程] 获取到MP3链接: {mp3_url}")
                        
                        if not mp3_url:
                            threading.Timer(0.1, lambda: show_snack_bar("❌ 未能获取到MP3链接")).start()
                            threading.Timer(0.1, reset_download_button).start()
                            return
                        
                        # ========== 根据平台选择保存路径 ==========
                        if platform.system() == "Linux":
                            # 华为手机等Android设备 - 使用公共音乐目录
                            # 获取外部存储路径（通常是 /storage/emulated/0）
                            external_storage = os.environ.get("EXTERNAL_STORAGE", "/storage/emulated/0")
                            download_dir = Path(external_storage) / "Music" / "BirthdayReminder"
                            print(f"[下载线程] Android平台，保存到: {download_dir}")
                        else:
                            # Windows 电脑 - 使用用户音乐目录
                            download_dir = Path.home() / "Music" / "BirthdayReminder"
                            print(f"[下载线程] Windows平台，保存到: {download_dir}")
                        
                        # 创建目录
                        download_dir.mkdir(parents=True, exist_ok=True)
                        
                        # 清理文件名中的非法字符
                        filename = f"{selected_song['name']}-{selected_song['artist']}.mp3"
                        filename = re.sub(r'[\\/*?:"<>|]', '', filename)
                        filepath = download_dir / filename
                        print(f"[下载线程] 保存路径: {filepath}")
                        
                        threading.Timer(0.1, lambda: show_snack_bar(f"正在下载: {selected_song['name']}...")).start()

                        # 使用你提供的方法下载MP3文件
                        success = download_mp3_file_with_headers(mp3_url, filepath, downloader)
                        
                        if success:
                            # 更新音乐文件路径
                            threading.Timer(0.1, lambda: setattr(music_field, 'value', str(filepath))).start()
                            threading.Timer(0.1, lambda: setattr(selected_file_display, 'value', f"已选择: {filename}")).start()
                            
                            # 尝试下载歌词
                            lyrics = downloader.search_and_get_lyrics(selected_song['name'], selected_song['artist'])
                            if lyrics:
                                lrc_path = filepath.with_suffix('.lrc')
                                with open(lrc_path, 'w', encoding='utf-8') as f:
                                    f.write(lyrics)
                                print(f"[下载] 歌词已保存: {lrc_path}")
                            
                            threading.Timer(0.1, lambda: show_snack_bar(f"下载完成: {filename}")).start()
                        else:
                            threading.Timer(0.1, lambda: show_snack_bar("下载失败")).start()
                        
                        threading.Timer(0.1, reset_download_button).start()
                        
                    except Exception as e:
                        print(f"下载出错: {e}")
                        threading.Timer(0.1, lambda: show_snack_bar(f"下载失败: {str(e)}")).start()
                        threading.Timer(0.1, reset_download_button).start()
                
                def reset_download_button():
                    download_btn.disabled = False
                    download_btn.text = "📥 下载并应用"
                    download_btn.update()
                    page.update()
                
                threading.Thread(target=download_thread, daemon=True).start()

            # 绑定事件
            search_btn.on_click = do_search
            search_results_dropdown.on_change = on_result_select
            download_btn.on_click = do_download
        else:
            # Android 平台：禁用所有搜索相关控件
            search_keyword_field.disabled = True
            search_btn.disabled = True
            search_results_dropdown.disabled = True
            download_btn.disabled = True
            search_status.value = "📱 Android版本暂不支持在线下载，请手动选择音乐文件"
            search_status.color = ft.Colors.ORANGE_700
        
        def download_mp3_file_with_headers(mp3_url, filepath, downloader):
            """使用正确的请求头下载MP3文件"""
            try:
                # 使用动态UA
                headers = {
                    'User-Agent': downloader.get_random_ua(),
                    'Referer': 'https://www.gequbao.com/',
                    'Accept': '*/*',
                    'Accept-Encoding': 'gzip, deflate, br',
                    'Connection': 'keep-alive',
                }
                
                # 根据域名设置合适的Referer和Origin
                if 'kuwo.cn' in mp3_url:
                    headers['Referer'] = 'https://www.kuwo.cn/'
                    headers['Origin'] = 'https://www.kuwo.cn'
                    print("[下载] 检测到酷我音乐链接，使用专用headers")
                elif '163.com' in mp3_url or '126.net' in mp3_url:
                    headers['Referer'] = 'https://music.163.com/'
                    headers['Origin'] = 'https://music.163.com'
                    print("[下载] 检测到网易云音乐链接，使用专用headers")
                
                # 开始下载
                response = downloader.session.get(mp3_url, headers=headers, stream=True, timeout=60)
                
                # 检查状态码
                if response.status_code != 200:
                    print(f"[下载错误] HTTP状态码: {response.status_code}")
                    return False
                
                # 获取文件大小
                total_size = int(response.headers.get('content-length', 0))
                if total_size == 0:
                    print("[下载错误] 文件大小为0，链接可能无效")
                    return False
                
                print(f"[下载] 文件大小: {total_size / 1024 / 1024:.2f} MB")
                
                # 下载文件
                downloaded_size = 0
                with open(filepath, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                            downloaded_size += len(chunk)
                            # 每10MB打印一次进度
                            if downloaded_size % (10 * 1024 * 1024) < 8192:
                                progress = (downloaded_size / total_size) * 100
                                print(f"[下载进度] {progress:.1f}%")
                
                # 验证下载的文件大小
                file_size = filepath.stat().st_size
                if file_size == 0:
                    print("[下载错误] 下载的文件大小为0")
                    return False
                
                print(f"[下载] 下载完成: {filepath.name} ({file_size / 1024 / 1024:.2f} MB)")
                return True
                
            except Exception as e:
                print(f"[下载错误] {e}")
                return False

        # ========== 如果是编辑模式，覆盖默认值 ==========
        if is_edit and selected_event:
            # 设置事件类型（这会触发 on_event_type_select）
            event_type_selected = selected_event.event_type
            
            # 设置名称
            name_field.value = selected_event.name
            name_field.label = "姓名" if selected_event.event_type == "birthday" else "事件名称"
            
            # 设置历法
            calendar_selected = selected_event.calendar_type
            
            # 设置重复类型
            if hasattr(selected_event, 'repeat_type'):
                repeat_type.value = selected_event.repeat_type
            
            # 设置音乐文件
            music_field.value = selected_event.sound_file if selected_event.sound_file else ""
            
            # 根据事件类型设置日期
            # ========== 根据事件类型设置日期 ==========
            # 每日事件
            if selected_event.event_type == "daily":
                date_display_field.visible = False
                date_display_field.value = ""
            
            # 每周事件
            elif selected_event.event_type == "weekly":
                date_display_field.visible = False
                date_display_field.value = ""
                # 设置星期 添加安全检查，放在后面更新，因为这时候控件还没定义
                #if selected_event.birth_date:
                    #weekday_text = get_weekday_text(selected_event.birth_date)
                    #weekday_dropdown.value = weekday_text
            
            # 每月事件
            elif selected_event.event_type == "monthly":
                date_display_field.visible = True
                day_num = int(selected_event.birth_date) if selected_event.birth_date else 1
                date_display_field.value = f"{day_num:02d}"
                # 设置日期选择器的初始值
                try:
                    now = datetime.now()
                    initial_date = datetime(now.year, now.month, day_num)
                    date_picker.value = initial_date
                except:
                    pass
            
            # 一次性事件
            elif selected_event.repeat_type == "once":
                date_display_field.visible = True
                date_parts = selected_event.birth_date.split("-")
                if len(date_parts) == 3:
                    year = int(date_parts[0])
                    month = int(date_parts[1])
                    day = int(date_parts[2])
                    date_display_field.value = f"{year:04d}-{month:02d}-{day:02d}"
                    # 设置日期选择器的初始值
                    try:
                        initial_date = datetime(year, month, day)
                        date_picker.value = initial_date
                    except:
                        pass
            
            # 生日或纪念日
            else:
                date_display_field.visible = True
                date_parts = selected_event.birth_date.split("-")
                if len(date_parts) == 3:
                    year = int(date_parts[0])
                    month = int(date_parts[1])
                    day = int(date_parts[2])
                    date_display_field.value = f"{year:04d}-{month:02d}-{day:02d}"
                    # 设置日期选择器的初始值
                    try:
                        initial_date = datetime(year, month, day)
                        date_picker.value = initial_date
                    except:
                        pass

        # 定义取消函数（放在这里，在使用之前）
        def cancel_click(e):
            close_dialog()
        
        # 在保存时使用 event_type
        def save_click(e):
            nonlocal event_type_selected, calendar_selected, dialog_container
            name = name_field.value.strip()
            if not name:
                show_snack_bar("请输入名称")
                #show_snack_bar_new(page, "⚠️ 请输入名称", is_error=True)
                return
            
            print(f"[保存] 事件类型: {event_type_selected}")

            # 获取工作日选项的值（使用专门保存的 Switch 变量）
            workday_only = False
            if hasattr(open_add_dialog, 'workday_only_switch'):
                workday_only = open_add_dialog.workday_only_switch.value
                print(f"[保存] 工作日选项: {workday_only}")
            
            # ========== 收集提醒时间（关键修复） ==========
            reminders = []
            # 尝试从 open_add_dialog.reminders_list 获取
            if hasattr(open_add_dialog, 'reminders_list') and open_add_dialog.reminders_list:
                for row in open_add_dialog.reminders_list.controls:
                    if len(row.controls) >= 2:
                        time_display_field = row.controls[0]
                        checkbox = row.controls[1]
                        if checkbox.value and time_display_field.value:
                            reminders.append({"time": time_display_field.value, "enabled": True})
                            print(f"[保存] 添加提醒时间: {time_display_field.value}")
            
            print(f"[保存] 总共收集到 {len(reminders)} 个提醒时间")

            repeat = repeat_type.value if event_type_selected != "monthly" else "monthly"

            # ========== 从 date_display_field 获取日期 ==========
            year = 1990
            month = 1
            day = 1
            
            if date_display_field.value and date_display_field.value != "点击选择日期":
                try:
                    if event_type_selected == "monthly":
                        day = date_display_field.value
                    else:
                        date_parts = date_display_field.value.split("-")
                        year = int(date_parts[0])
                        month = int(date_parts[1])
                        day = int(date_parts[2])
                        print(f"[保存] 从日期选择器获取: {year}-{month}-{day}")
                except:
                    print(f"[保存] 解析日期失败，使用默认值")
            
            # ========== 根据事件类型处理日期 ==========
            if event_type_selected == "daily":
                # 每天提醒：不需要日期，设置为空字符串或默认值
                birth_date = ""  # 设置为空，表示不需要日期
                calendar_type_value = "solar"
                repeat_type_value = "daily"
                
            elif event_type_selected == "weekly":
                weekday_text = weekday_dropdown.value  # 获取显示文本
                # 每周提醒：保存星期几
                weekday = get_weekday_key(weekday_text)
                if not weekday:
                    show_snack_bar("请选择星期几")
                    return
                birth_date = weekday  # 保存 "1" 表示周一
                calendar_type_value = "solar"
                repeat_type_value = "weekly"
                
            elif event_type_selected == "monthly":
                # 每月提醒：使用 day
                birth_date = day
                calendar_type_value = "solar"
                repeat_type_value = "monthly"

            elif event_type_selected == "once":
                # 一次性事件：使用完整的年月日
                event_date = datetime(year, month, day).date()
                today = datetime.now().date()
                
                if event_date < today:
                    show_snack_bar("一次性事件的日期不能早于今天")
                    return
                birth_date = f"{year:04d}-{month:02d}-{day:02d}"
                calendar_type_value = calendar_selected
                repeat_type_value = "once"
                
            elif event_type_selected == "birthday":
                birth_date = f"{year}-{month:02d}-{day:02d}"
                calendar_type_value = calendar_selected
                repeat_type_value = "yearly"
                
            else:  # event
                birth_date = f"{year}-{month:02d}-{day:02d}"
                calendar_type_value = calendar_selected
                repeat_type_value = "yearly"

            # 收集提醒时间
            reminders = []
            if hasattr(open_add_dialog, 'reminders_list') and open_add_dialog.reminders_list:
                for row in open_add_dialog.reminders_list.controls:
                    time_display_field = row.controls[0]
                    checkbox = row.controls[1]
                    if checkbox.value and time_display_field.value:
                        reminders.append({"time": time_display_field.value, "enabled": True})
            
            # 保存事件
            if is_edit and selected_event:
                try:
                    reset_all_reminders()
                    selected_event.workday_only = workday_only
                    selected_event.last_remind_year = 0
                    selected_event.reminded_this_year = False
                    selected_event.name = name
                    selected_event.birth_date = birth_date
                    selected_event.calendar_type = calendar_type_value
                    selected_event.event_type = event_type_selected
                    selected_event.repeat_type = repeat_type_value
                    selected_event.sound_file = music_field.value.strip()
                    selected_event.reminders = reminders
                    if repeat_type_value == "once":
                        selected_event.completed = False
                    save_events(trigger_check=False)

                    # ========== 强制触发检查 ==========
                    async def do_check():
                        await asyncio.sleep(0.3)
                        print(f"[强制检查] 开始检查，事件数: {len(events)}")
                        
                        # 重新加载事件数据
                        load_events()
                        print(f"[强制检查] 重新加载后事件数: {len(events)}")
                        
                        # 调用 check_events
                        check_time_reminders()
                        check_events()
                        page.update()
                    
                    asyncio.create_task(do_check())
                    
                    # ========== 重新收集3日内事件 ==========
                    three_days_events = []
                    for evt in events.values():
                        if evt.event_type == "daily" or evt.event_type == "weekly":
                            continue
                        month, day, year, base_year, days_until = evt.get_next_date_info()
                        if evt.repeat_type == "once" and (evt.completed or days_until < 0):
                            continue
                        if 0 < days_until <= 3:
                            three_days_events.append((evt, days_until))
                    
                    # 更新顶部日期文本显示
                    update_three_days_events()
                    today = datetime.now().date()
                    update_date_text_with_events(today, three_days_events)

                    # 根据当前视图刷新对应的视图
                    refresh_current_view_by_state()

                    # ========== 保存后重新检查视图 ==========
                    determine_startup_view()

                    close_dialog()
                    show_snack_bar(f"已更新「{name}」")
                except Exception as e:
                    print(f"更新失败: {e}")
                    show_snack_bar(f"更新失败: {str(e)}")
            else:
                try:
                    event_id = str(int(datetime.now().timestamp()))
                    new_event = Event(
                        event_id, name, birth_date, calendar_type_value, 
                        event_type_selected, music_field.value.strip(), repeat_type_value,
                        reminders=reminders
                    )
                    new_event.workday_only = workday_only
                    if repeat_type_value == "once":
                        new_event.completed = False
                    events[event_id] = new_event
                    save_events(trigger_check=False)

                    # ========== 强制触发检查 ==========
                    async def do_check():
                        await asyncio.sleep(0.3)
                        print(f"[强制检查] 开始检查，事件数: {len(events)}")
                        
                        # 重新加载事件数据
                        load_events()
                        print(f"[强制检查] 重新加载后事件数: {len(events)}")
                        
                        # 调用 check_events
                        check_time_reminders()
                        check_events()
                        page.update()
                    
                    asyncio.create_task(do_check())
                    
                    # ========== 重新收集3日内事件 ==========
                    three_days_events = []
                    for evt in events.values():
                        if evt.event_type == "daily" or evt.event_type == "weekly":
                            continue
                        month, day, year, base_year, days_until = evt.get_next_date_info()
                        if evt.repeat_type == "once" and (evt.completed or days_until < 0):
                            continue
                        if 0 < days_until <= 3:
                            three_days_events.append((evt, days_until))
                    
                    # 更新顶部日期文本显示
                    update_three_days_events()
                    today = datetime.now().date()
                    update_date_text_with_events(today, three_days_events)
                    
                    # 根据当前视图刷新对应的视图
                    refresh_current_view_by_state()

                    # ========== 保存后重新检查视图 ==========
                    determine_startup_view()

                    close_dialog()
                    show_snack_bar(f"已添加「{name}」")
                except Exception as e:
                    print(f"添加失败: {e}")
                    show_snack_bar(f"添加失败: {str(e)}")
            
            async def delayed_check():
                await asyncio.sleep(0.5)
                check_events()
            
            asyncio.create_task(delayed_check())


        # 先确定初始值(法定工作日)
        initial_workday_only = False
        if is_edit and selected_event:
            initial_workday_only = getattr(selected_event, 'workday_only', False)

        workday_only_switch = ft.Switch(
            value=initial_workday_only,
            active_color=ft.Colors.BLUE_600,
            inactive_thumb_color=ft.Colors.GREY_500,
            inactive_track_color=ft.Colors.GREY_300,
            active_track_color=ft.Colors.BLUE_100,
            adaptive=True,
        )

        workday_only_checkbox = ft.Row([
            ft.Text("法定工作日（智能跳过节假日）", size=13, color=ft.Colors.GREY_800),
            workday_only_switch,
        ], spacing=12, alignment=ft.MainAxisAlignment.SPACE_BETWEEN)

        # 保存到函数属性，方便其他地方访问
        open_add_dialog.workday_only_checkbox = workday_only_checkbox
        open_add_dialog.workday_only_switch = workday_only_switch


        # 多提醒时间容器
        reminders_container = ft.Container(
            content=ft.Column([
                ft.Text("⏰ 多时段提醒", size=14, weight=ft.FontWeight.BOLD),
                reminders_list,
                ft.Divider(height=5),
                workday_only_checkbox,  # 添加工作日选项
                ft.Divider(height=5),
                ft.Row(
                    [ft.ElevatedButton(
                        "添加提醒时间",
                        on_click=lambda e: add_reminder_time(),
                        icon=ft.Icons.ADD_ALARM,
                        height=36,
                    )],
                    
                    alignment=ft.MainAxisAlignment.CENTER,  # 水平居中
                ),
            ], spacing=8),
            padding=10,
            bgcolor=ft.Colors.TRANSPARENT,  # 改为透明
            border_radius=10,
            visible=True,
        )

        # 调用日期显示切换函数，根据当前事件类型设置正确的显示
        update_date_visibility()

        # 创建顶部按钮栏
        def cancel_click(e):
            close_dialog()
            show_bottom_message("已取消")
        
        def save_click_wrapper(e):
            save_click(e)  # 调用原有的保存函数
        
        top_bar = ft.Row([
            ft.IconButton(
                icon=ft.Icons.CLOSE,
                icon_size=24,
                icon_color=ft.Colors.RED_700,
                tooltip="取消",
                on_click=cancel_click,
            ),
            ft.Text("编辑事件" if is_edit else "添加事件", size=18, weight=ft.FontWeight.BOLD, expand=True, text_align=ft.TextAlign.CENTER),
            ft.IconButton(
                icon=ft.Icons.CHECK,
                icon_size=24,
                icon_color=ft.Colors.GREEN_700,
                tooltip="保存",
                on_click=save_click_wrapper,
            ),
        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.CENTER)

        # ========== 创建可滚动的内容区域 ==========
        scrollable_content = ft.Column([
            ft.Container(height=1),
            event_type_dropdown,  # 替换原来的 event_type_container
            name_field,
            ft.Row([date_display_field], alignment=ft.MainAxisAlignment.CENTER),
            weekday_dropdown,     # 替换原来的 weekday_row
            calendar_dropdown,    # 替换原来的 calendar_container
            ft.Divider(height=5),
            ft.Text("⏰ 提醒设置", size=14, weight=ft.FontWeight.BOLD),
            reminders_container,  # 确保这一行存在
            ft.Divider(height=5),
            music_field,
            music_buttons,
            selected_file_display,
            ft.Divider(height=5),
            ft.Text("🎵 在线搜索音乐", size=14, weight=ft.FontWeight.BOLD),
            ft.Row([search_keyword_field, search_btn], spacing=8),
            search_results_dropdown,
            ft.Row([download_btn],alignment=ft.MainAxisAlignment.CENTER,),
            ft.Row([search_status],alignment=ft.MainAxisAlignment.CENTER,),
            ft.Divider(height=5),
            hint_text,
        ], spacing=15, scroll=ft.ScrollMode.AUTO)

        # ========== 整体布局：顶部固定 + 内容滚动 ==========
        dialog_content = ft.Column([
            top_bar,  # 顶部固定
            ft.Divider(height=5),
            ft.Container(
                content=scrollable_content,
                expand=True,  # 占据剩余空间
            ),
        ], spacing=10, height=500)  # 固定总高度

        # 创建容器并添加到页面
        dialog_container = ft.Container(
            content=ft.Container(
                content=dialog_content,
                bgcolor=ft.Colors.WHITE,
                padding=20,
                border_radius=12,
                #border=ft.border.all(1, ft.Colors.BLUE_200),
                shadow=ft.BoxShadow(
                    spread_radius=1,
                    blur_radius=15,
                    color=ft.Colors.BLACK12,
                ),
                expand=True,
            ),
            left=20,
            top=50,
            right=20,
            bottom=50,
        )
        
        update_date_visibility()
        page.overlay.append(dialog_container)
        update_date_visibility()
        page.update()

        # 然后再设置值 安全检查
        if is_edit and selected_event:
            # 设置事件类型
            event_type_dropdown.value = get_event_type_text(selected_event.event_type)
            # 设置历法
            calendar_dropdown.value = get_calendar_text(selected_event.calendar_type)
            # 设置星期
            if selected_event.event_type == "weekly" and selected_event.birth_date:
                weekday_dropdown.value = get_weekday_text(selected_event.birth_date)
            page.update()
    
    def group_events_by_date(events_list):
        """将同一天的事件分组"""
        grouped = {}
        for event, days_until in events_list:
            key = days_until  # 使用剩余天数作为分组键
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(event)
        return grouped
    
    def log_event_reminder(event_name, days_left, event_id=None):
        """统一处理事件提醒日志（带去重）"""
        if days_left == 0:
            day_text = "今天"
        elif days_left == 1:
            day_text = "明天"
        elif days_left == 2:
            day_text = "后天"
        else:
            day_text = f"{days_left}天后"
        
        # ========== 去重检查 ==========
        reminder_key = f"{event_id}_{days_left}_{datetime.now().strftime('%Y-%m-%d')}"
        if reminder_key in sent_reminders:
            return  # 已发送过，跳过
        
        sent_reminders.add(reminder_key)
        
        # 只在 Windows 平台打印日志
        if IS_WINDOWS:
            print(f"[事件提醒] {event_name} {day_text} 就到啦！")
        else:
            # 非 Windows 平台发送通知
            show_event_notification(event_name, "事件", days_left)

    def show_combined_reminder(events_by_day, is_today=False):
        """显示合并后的提醒弹窗"""
        print(f"[弹框调试] show_combined_reminder 被调用, is_today={is_today}")
        print(f"[弹框调试] events_by_day: {events_by_day}")

        if not events_by_day:
            print("[弹框调试] events_by_day 为空，返回")
            return
        
        # ========== 发送通知（使用统一函数，带去重） ==========
        for days, events in events_by_day.items():
            for event in events:
                if is_today:
                    log_event_reminder(event.name, 0, event.id)
                else:
                    log_event_reminder(event.name, days, event.id)

        
        def close_combined_reminder():
            try:
                if combined_container in page.overlay:
                    page.overlay.remove(combined_container)
                    page.update()
            except:
                pass

        # ========== 发送通知（使用统一函数） ==========
        for days, events in events_by_day.items():
            for event in events:
                # ========== 使用统一的日志函数 ==========
                if is_today:
                    log_event_reminder(event.name, 0, event.id)
                else:
                    log_event_reminder(event.name, days, event.id)
        
        if is_today:
            # 区分生日和事件
            birthday_events = []
            other_events = []
            
            for days, events in events_by_day.items():
                for event in events:
                    if event.event_type == "birthday":
                        birthday_events.append(event)
                    else:
                        other_events.append(event)
            
            # 构建生日列表
            events_text = []
            music_file = None
            event_name_for_music = None  # 新增：用于播放的事件名称
            event_id_for_music = None  # 新增：用于播放的事件id
            
            if birthday_events:
                events_text.append(ft.Text("🎂 生日祝福：", size=16, weight=ft.FontWeight.BOLD, color=ft.Colors.RED_700))
                for event in birthday_events:
                    month, day, year, birth_year, _ = event.get_next_date_info()
                    age = datetime.now().year - birth_year
                    calendar_icon = "☀️" if event.calendar_type == "solar" else "🌙"
                    events_text.append(ft.Text(f"  {calendar_icon} {event.name}（{age}岁）", size=14))
                    if not music_file and event.sound_file:
                        music_file = event.sound_file
                        event_name_for_music = event.name  # 保存事件名称
                        event_id_for_music = event.id      # 保存事件id
            
            if other_events:
                events_text.append(ft.Text("📅 纪念日提醒：", size=16, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700))
                for event in other_events:
                    calendar_icon = "☀️" if event.calendar_type == "solar" else "🌙"
                    events_text.append(ft.Text(f"  {calendar_icon} {event.name}", size=14))
                    if not music_file and event.sound_file:
                        music_file = event.sound_file
                        event_name_for_music = event.name  # 保存事件名称
                        event_id_for_music = event.id      # 保存事件id
            
            title = "🎉 今日提醒"
            title_color = ft.Colors.PURPLE_700
            
            # 创建美化后的内容容器
            content_column = ft.Column([
                # 顶部图标
                ft.Container(
                    content=ft.Icon(ft.Icons.CELEBRATION, size=48, color=ft.Colors.PURPLE_700),
                    padding=10,
                    bgcolor=ft.Colors.PURPLE_50,
                    border_radius=50,
                ),
                ft.Text(title, size=24, weight=ft.FontWeight.BOLD, color=title_color),
                ft.Divider(height=1, color=ft.Colors.GREY_300),
                ft.Column(events_text, spacing=8),
                ft.Divider(height=1, color=ft.Colors.GREY_300),
                ft.Row([
                    ft.ElevatedButton(
                        "关闭", 
                        on_click=lambda e: close_combined_reminder(),
                        style=ft.ButtonStyle(
                            bgcolor=ft.Colors.BLUE_700,
                            color=ft.Colors.WHITE,
                        ),
                    ),
                ], alignment=ft.MainAxisAlignment.CENTER),
            ], spacing=15, horizontal_alignment=ft.CrossAxisAlignment.CENTER)
            
        else:
            title = "🎈 事件预告"
            title_color = ft.Colors.ORANGE_700
            
            events_by_day_list = []
            music_file = None
            event_name_for_music = None      # 新增：用于播放的事件名称
            event_id_for_music = None        # 新增：用于播放的事件id
            
            for days_left in sorted(events_by_day.keys()):
                if days_left == 1:
                    day_text = "明天"
                elif days_left == 2:
                    day_text = "后天"
                else:
                    day_text = f"{days_left}天后"
                
                birthday_names = []
                event_names = []
                
                for event in events_by_day[days_left]:
                    calendar_icon = "☀️" if event.calendar_type == "solar" else "🌙"
                    if event.event_type == "birthday":
                        birthday_names.append(f"{calendar_icon} {event.name}（生日）")
                    else:
                        event_names.append(f"{calendar_icon} {event.name}")
                    if not music_file and event.sound_file:
                        music_file = event.sound_file
                        event_name_for_music = event.name  # 保存事件名称
                        event_id_for_music = event.id      # 保存事件id
                
                text_parts = []
                if birthday_names:
                    text_parts.append("🎂 " + "、".join(birthday_names))
                if event_names:
                    text_parts.append("📅 " + "、".join(event_names))
                
                month, day, year, birth_year, _ = events_by_day[days_left][0].get_next_date_info()
                events_by_day_list.append(
                    ft.Text(f"• {day_text}（{month}月{day}日）：{'，'.join(text_parts)}", size=14)
                )
            
            # 创建美化后的内容容器
            content_column = ft.Column([
                # 顶部图标
                ft.Container(
                    content=ft.Icon(ft.Icons.NOTIFICATIONS_ACTIVE, size=48, color=ft.Colors.ORANGE_700),
                    padding=10,
                    bgcolor=ft.Colors.ORANGE_50,
                    border_radius=50,
                ),
                ft.Text("🎈 事件预告", size=24, weight=ft.FontWeight.BOLD, color=ft.Colors.ORANGE_700),
                ft.Divider(height=1, color=ft.Colors.GREY_300),
                ft.Text("以下事件即将到来：", size=14, color=ft.Colors.GREY_700),
                ft.Column(events_by_day_list, spacing=8),
                ft.Text("记得提前准备哦！", size=12, color=ft.Colors.GREY_500),
                ft.Divider(height=1, color=ft.Colors.GREY_300),
                ft.Row([
                    ft.ElevatedButton(
                        "关闭", 
                        on_click=lambda e: close_combined_reminder(),
                        style=ft.ButtonStyle(
                            bgcolor=ft.Colors.BLUE_700,
                            color=ft.Colors.WHITE,
                        ),
                    ),
                ], alignment=ft.MainAxisAlignment.CENTER),
            ], spacing=15, horizontal_alignment=ft.CrossAxisAlignment.CENTER)

            # 自动播放音乐（预警事件）- 如果有音乐在播放，则跳过
            if music_file:
                with music_playing_lock:
                    if not is_playing:
                        print(f"[预警自动播放] 播放: {os.path.basename(music_file)}")
                        play_music(music_file, loop=False, event_name=event_name_for_music,event_id=event_id_for_music)  # 传递事件名称
                    else:
                        print(f"[预警自动播放] 音乐正在播放中，跳过: {os.path.basename(music_file)}")
        
        # 创建美化的对话框容器（居中、带边框和阴影）
        combined_container = ft.Container(
            content=ft.Column([
                ft.Container(expand=True),  # 上方弹性空间
                ft.Row([
                    ft.Container(expand=True),  # 左侧弹性空间
                    ft.Container(
                        content=content_column,
                        width=320,
                        padding=20,
                        bgcolor=ft.Colors.WHITE,
                        border_radius=16,
                        #border=ft.border.all(1, ft.Colors.BLUE_200),
                        shadow=ft.BoxShadow(
                            spread_radius=1,
                            blur_radius=15,
                            color=ft.Colors.BLACK12,
                        ),
                    ),
                    ft.Container(expand=True),  # 右侧弹性空间
                ]),
                ft.Container(expand=True),  # 下方弹性空间
            ]),
            expand=True,
            bgcolor=ft.Colors.BLACK26,  # 半透明背景遮罩
            on_click=lambda e: close_combined_reminder(),  # 点击背景关闭
        )
        
        page.overlay.append(combined_container)
        page.update()
        
        # 10秒后自动关闭
        threading.Timer(10.0, close_combined_reminder).start()
        
        # 自动播放音乐（仅生日当天）- 修改为所有事件当天都播放
        if is_today and music_file:
            with music_playing_lock:
                if not is_playing:
                    print(f"[事件自动播放] 播放: {os.path.basename(music_file)}")
                    # 获取第一个事件作为显示名称
                    event_name = None
                    event_id = None
                    
                    # 遍历所有今天的事件，找到第一个有音乐的事件
                    for days, events in events_by_day.items():
                        for event in events:
                            if event.sound_file:
                                event_name = event.name
                                event_id = event.id
                                break
                        if event_name:
                            break
                    
                    if event_name:
                        play_music(music_file, loop=False, event_name=event_name, event_id=event_id)
                    else:
                        play_music(music_file, loop=False)
                else:
                    print(f"[事件自动播放] 音乐正在播放中，跳过: {os.path.basename(music_file)}")
    
    def check_startup_events():
        """启动时检查今日事件和预警事件"""
        today = datetime.now().date()
        now = datetime.now()
        current_time = now.strftime("%H:%M")
        
        print(f"[启动检查] ========== 开始 ==========")
        print(f"[启动检查] 当前日期: {today}, 当前时间: {current_time}")
        
        today_events = []
        upcoming_events = []
        
        for event in events.values():
            # 跳过每日事件
            if event.event_type == "daily":
                continue
            
            # 每周事件
            if event.event_type == "weekly":
                target_weekday = int(event.birth_date) if event.birth_date else 1
                if now.isoweekday() == target_weekday:
                    # 每周事件：检查是否设置了提醒时间
                    reminder_time = None
                    if event.reminders:
                        for reminder in event.reminders:
                            if reminder.get("enabled"):
                                reminder_time = reminder.get("time")
                                break
                    
                    # 如果有提醒时间且还没到，跳过今日提醒
                    if reminder_time and reminder_time > current_time:
                        print(f"[启动检查] 每周事件今天: {event.name}，提醒时间 {reminder_time} 还没到，跳过今日提醒")
                        # 但仍然加入预警？每周事件只在当天，所以如果时间没到就不提醒
                        continue
                    
                    print(f"[启动检查] 每周事件今天: {event.name}")
                    today_events.append((event, 0))
                continue
            
            # 获取事件信息
            month, day, year, base_year, days_until = event.get_next_date_info()
            
            # 跳过已经完成的一次性事件
            if event.repeat_type == "once" and event.completed:
                continue
            
            # 检查是否设置了提醒时间
            reminder_time = None
            has_reminder = False
            if event.reminders:
                for reminder in event.reminders:
                    if reminder.get("enabled"):
                        has_reminder = True
                        reminder_time = reminder.get("time")
                        break
            
            # ========== 今日事件 ==========
            if days_until == 0:
                # 如果设置了提醒时间且还没到，跳过今日提醒
                if has_reminder and reminder_time and reminder_time > current_time:
                    print(f"[启动检查] {event.name} 今天发生，但提醒时间 {reminder_time} 还没到，跳过今日提醒")
                    continue
                
                print(f"[启动检查] 今日事件: {event.name}")
                today_events.append((event, 0))
            
            # ========== 预警事件（未来3天）：忽略提醒时间 ==========
            elif 0 < days_until <= 3:
                print(f"[启动检查] 预警事件: {event.name}, {days_until}天后")
                upcoming_events.append((event, days_until))
        
        print(f"[启动检查] 今日事件: {len(today_events)}, 预警事件: {len(upcoming_events)}")
        
        # 触发弹框
        if today_events:
            print("[启动检查] 触发今日事件弹框")
            grouped = group_events_by_date(today_events)
            show_combined_reminder(grouped, is_today=True)
            for event, _ in today_events:
                if event.repeat_type != "once":
                    event.last_remind_year = today.year
                    event.reminded_this_year = True
            _save_events_silent()
        
        if upcoming_events:
            print("[启动检查] 触发预警弹框")
            grouped = group_events_by_date(upcoming_events)
            show_combined_reminder(grouped, is_today=False)
        
        print(f"[启动检查] ========== 完成 ==========")

    def check_missed_reminders():
        """检查今天已经错过的提醒（程序启动时执行）"""
        now = datetime.now()
        current_time = now.strftime("%H:%M")
        today = now.date()
        
        print(f"[错过提醒检查] ========== 开始检查 ==========")
        print(f"[错过提醒检查] 当前时间: {current_time}")
        
        triggered_events = []
        
        for event in events.values():
            # 跳过每日事件（由 check_time_reminders 处理）
            if event.event_type == "daily":
                continue
            
            if not is_event_today(event):
                continue
            
            if not event.reminders:
                continue
            
            reminder_time = None
            for reminder in event.reminders:
                if reminder.get("enabled"):
                    reminder_time = reminder.get("time")
                    break
            
            if not reminder_time:
                continue
            
            if reminder_time <= current_time:
                notification_key = f"{event.id}_{today.strftime('%Y-%m-%d')}"
                if notification_key in sent_notifications:
                    continue
                
                print(f"[错过提醒检查] 触发错过提醒: {event.name} - {reminder_time}")
                triggered_events.append((event, reminder_time))
                sent_notifications.add(notification_key)
        
        if triggered_events:
            triggered_events.sort(key=lambda x: x[1])
            
            for event, reminder_time in triggered_events:
                # ========== 弹框 ==========
                show_reminder_popup(
                    "⏰ 错过提醒",
                    f"{event.name}\n提醒时间: {reminder_time}",
                    event
                )
                
                # ========== 播放音乐 ==========
                if event.sound_file and os.path.exists(event.sound_file):
                    async def do_play(e=event):
                        with music_playing_lock:
                            if not is_playing:
                                print(f"[错过提醒] 播放音乐: {os.path.basename(e.sound_file)}")
                                play_music_with_lock(e.sound_file, loop=False, event_name=e.name, event_id=e.id)
                    page.run_task(do_play)
        
        print(f"[错过提醒检查] ========== 检查完成 ==========")

    def check_today_birthdays_on_start():
        """启动时检查今日事件并播放音乐（只播放音乐，不弹框）"""
        debug_log("========== 启动时检查 ==========")
        today = datetime.now().date()
        now = datetime.now()
        current_time = now.strftime("%H:%M")
        debug_log(f"启动日期: {today}, 当前时间: {current_time}")
        
        # 只处理今日事件，不处理预警
        today_events = []
        
        for event in events.values():
            # 跳过每天事件
            if event.event_type == "daily":
                continue
            
            # 每周事件：检查今天是否是提醒日
            if event.event_type == "weekly":
                target_weekday = int(event.birth_date) if event.birth_date else 1
                if now.isoweekday() != target_weekday:
                    continue
            
            # 每月事件
            if event.repeat_type == "monthly":
                target_day = int(event.birth_date) if event.birth_date else 1
                if now.day != target_day:
                    continue
            
            # 检查事件是否在今天发生
            month, day, year, birth_year, days_until = event.get_next_date_info()
            if month != today.month or day != today.day:
                continue
            
            # ========== 检查是否设置了提醒时间 ==========
            has_reminder = False
            reminder_time = None
            if event.reminders:
                for reminder in event.reminders:
                    if reminder.get("enabled"):
                        has_reminder = True
                        reminder_time = reminder.get("time")
                        break
            
            # 如果设置了提醒时间且时间还没到，跳过播放
            if has_reminder and reminder_time:
                if reminder_time > current_time:
                    debug_log(f"事件: {event.name} 提醒时间 {reminder_time} 还没到，跳过启动播放")
                    continue
            
            debug_log(f"  -> 今天是 {event.name} 的事件!")
            today_events.append(event)
        
        # 只播放音乐，不弹框
        if today_events and not is_playing:
            for event in today_events:
                if event.sound_file and os.path.exists(event.sound_file):
                    debug_log(f"[启动播放] 播放: {os.path.basename(event.sound_file)}")
                    play_music_with_lock(event.sound_file, loop=False, event_name=event.name, event_id=event.id)
                    break
        
        debug_log("========== 启动检查完成 ==========")

    def show_reminder_popup(title, message, event=None):
        """显示提醒弹框"""
        def close_popup():
            if popup_container in page.overlay:
                page.overlay.remove(popup_container)
                page.update()
        
        # 弹框内容
        popup_content = ft.Container(
            content=ft.Column([
                ft.Container(
                    content=ft.Icon(ft.Icons.NOTIFICATIONS_ACTIVE, size=55, color=ft.Colors.BLUE_700),
                    padding=10,
                    bgcolor=ft.Colors.BLUE_50,
                    border_radius=50,
                ),
                ft.Text(title, size=22, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700, text_align=ft.TextAlign.CENTER),
                ft.Divider(height=1, color=ft.Colors.GREY_300),
                ft.Text(message, size=16, color=ft.Colors.GREY_700, text_align=ft.TextAlign.CENTER),
                ft.Divider(height=1, color=ft.Colors.GREY_300),
                ft.Row([
                    ft.ElevatedButton(
                        "知道了",
                        on_click=lambda e: close_popup(),
                        expand=True,
                        style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_700, color=ft.Colors.WHITE),
                    ),
                ], alignment=ft.MainAxisAlignment.CENTER),
            ], spacing=15, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            width=320,
            padding=20,
            bgcolor=ft.Colors.WHITE,
            border_radius=16,
            shadow=ft.BoxShadow(
                spread_radius=1,
                blur_radius=15,
                color=ft.Colors.BLACK12,
                offset=ft.Offset(0, 4),
            ),
        )
        
        popup_container = ft.Container(
            content=ft.Column([
                ft.Container(expand=True),
                ft.Row([
                    ft.Container(expand=True),
                    popup_content,
                    ft.Container(expand=True),
                ]),
                ft.Container(expand=True),
            ]),
            expand=True,
            bgcolor=ft.Colors.BLACK26,
            on_click=lambda e: close_popup(),
        )
        
        page.overlay.append(popup_container)
        page.update()

    def reset_all_reminders():
        """重置所有提醒标记"""
        global  reminder_flags
        print("[调试] 开始重置所有提醒标记")
        reminder_flags.clear()
        print("[调试] 重置完成")
    
    def check_time_reminders():
        """检查时间提醒"""
        global sent_notifications
        now = datetime.now()
        current_time = now.strftime("%H:%M")
        current_weekday = now.weekday() + 1
        is_workday_today = is_workday(now)
        
        #print(f"[时间提醒] ========== 开始检查 ==========")
        #print(f"[时间提醒] 当前时间: {current_time}, 当前星期: {current_weekday}")
        
        for event in events.values():
            if not event.reminders:
                continue
            
            has_enabled_reminder = False
            for reminder in event.reminders:
                if reminder.get("enabled"):
                    has_enabled_reminder = True
                    break
            
            if not has_enabled_reminder:
                continue
            
            if event.event_type == "daily" and hasattr(event, 'workday_only') and event.workday_only:
                if not is_workday_today:
                    continue
            
            for reminder in event.reminders:
                if not reminder.get("enabled"):
                    continue
                if reminder.get("time") != current_time:
                    continue
                
                if not is_event_today(event):
                    continue
                
                today = now.date()
                notification_key = f"{event.id}_{today.strftime('%Y-%m-%d')}"
                if notification_key in sent_notifications:
                    #print(f"[时间提醒] {event.name} 今天已提醒过，跳过")
                    continue
                
                #print(f"[时间提醒] 触发事件: {event.name} ({event.event_type})")
                sent_notifications.add(notification_key)
                
                # ========== 弹框 ==========
                show_reminder_popup(
                    "🔔 时间提醒",
                    f"{event.name}\n时间: {current_time}",
                    event
                )
                
                # ========== 播放音乐 ==========
                if event.sound_file and os.path.exists(event.sound_file):
                    async def do_play(e=event):
                        with music_playing_lock:
                            if not is_playing:
                                #print(f"[时间提醒] 播放音乐: {os.path.basename(e.sound_file)}")
                                play_music_with_lock(e.sound_file, loop=False, event_name=e.name, event_id=e.id)
                    page.run_task(do_play)
                break

        #print(f"[时间提醒] ========== 检查完成 ==========")

    def is_event_today(event):
        """判断事件是否在今天发生"""
        now = datetime.now()
        
        if event.event_type == "weekly":
            target_weekday = int(event.birth_date) if event.birth_date else 1
            return now.isoweekday() == target_weekday
        
        elif event.event_type == "monthly" or event.repeat_type == "monthly":
            target_day = int(event.birth_date) if event.birth_date else 1
            return now.day == target_day
        
        elif event.repeat_type == "once":
            month, day, year, _, _ = event.get_next_date_info()
            return month == now.month and day == now.day
        
        else:  # birthday, event
            month, day, year, _, _ = event.get_next_date_info()
            return month == now.month and day == now.day

    # 修改 check_events 函数，添加详细日志
    def check_events():
        """检查事件发生（处理预警事件）"""
        try:
            today = datetime.now().date()
            now = datetime.now()
            current_time = now.strftime("%H:%M")
            
            print(f"[定时检查] ========== 开始检查事件 ==========")
            print(f"[定时检查] 当前日期: {today}, 当前时间: {current_time}")
            
            today_events = []
            upcoming_events = []
            
            for event in events.values():
                # 排除每日事件
                if event.event_type == "daily":
                    continue
                
                # 每周事件：只在当天
                if event.event_type == "weekly":
                    target_weekday = int(event.birth_date) if event.birth_date else 1
                    if now.isoweekday() == target_weekday:
                        # 检查提醒时间
                        reminder_time = None
                        if event.reminders:
                            for reminder in event.reminders:
                                if reminder.get("enabled"):
                                    reminder_time = reminder.get("time")
                                    break
                        
                        if reminder_time and reminder_time > current_time:
                            print(f"[定时检查] 每周事件今天: {event.name}，提醒时间 {reminder_time} 还没到，跳过")
                            continue
                        
                        print(f"[定时检查] 每周事件今天: {event.name}")
                        today_events.append((event, 0))
                    continue
                
                # 获取事件信息
                month, day, year, base_year, days_until = event.get_next_date_info()
                
                # 一次性事件
                if event.repeat_type == "once":
                    if event.completed:
                        continue
                    if days_until == 0:
                        print(f"[定时检查] 一次性事件今天: {event.name}")
                        today_events.append((event, 0))
                        event.completed = True
                        _save_events_silent()
                    elif 0 < days_until <= 3:
                        print(f"[定时检查] 一次性事件预警: {event.name}, {days_until}天后")
                        upcoming_events.append((event, days_until))
                    continue
                
                # 每月事件
                if event.event_type == "monthly":
                    if days_until == 0:
                        print(f"[定时检查] 每月事件今天: {event.name}")
                        today_events.append((event, 0))
                    elif 0 < days_until <= 3:
                        print(f"[定时检查] 每月事件预警: {event.name}, {days_until}天后")
                        upcoming_events.append((event, days_until))
                    continue
                
                # ========== 生日/纪念日 ==========
                if days_until == 0:
                    # 检查提醒时间
                    reminder_time = None
                    if event.reminders:
                        for reminder in event.reminders:
                            if reminder.get("enabled"):
                                reminder_time = reminder.get("time")
                                break
                    
                    if reminder_time and reminder_time > current_time:
                        print(f"[定时检查] {event.name} 今天发生，但提醒时间 {reminder_time} 还没到，跳过")
                        continue
                    
                    if event.last_remind_year != today.year:
                        print(f"[定时检查] 生日/纪念日今天: {event.name}")
                        today_events.append((event, 0))
                elif 0 < days_until <= 3:
                    # ========== 预警事件：忽略提醒时间 ==========
                    print(f"[定时检查] 生日/纪念日预警: {event.name}, {days_until}天后")
                    upcoming_events.append((event, days_until))
            
            print(f"[定时检查] 今日事件: {len(today_events)}, 预警事件: {len(upcoming_events)}")
            
            if today_events:
                grouped = group_events_by_date(today_events)
                show_combined_reminder(grouped, is_today=True)
                for event, _ in today_events:
                    if event.repeat_type != "once":
                        event.last_remind_year = today.year
                        event.reminded_this_year = True
                _save_events_silent()
            
            if upcoming_events:
                print(f"[定时检查] 触发预警弹框")
                grouped = group_events_by_date(upcoming_events)
                show_combined_reminder(grouped, is_today=False)
                
            print(f"[定时检查] ========== 检查完成 ==========")
            
        except Exception as e:
            print(f"检查出错: {e}")
            import traceback
            traceback.print_exc()

    def trigger_event_reminder(event, message):
        """触发事件提醒"""
        notification_key = f"{event.id}_{datetime.now().strftime('%Y-%m-%d_%H%M')}"
        if notification_key in sent_notifications:
            return
        
        sent_notifications.add(notification_key)
        show_notification(page, f"🔔 事件提醒", message)
        
        if event.sound_file and os.path.exists(event.sound_file):
            async def do_play():
                with music_playing_lock:
                    if not is_playing:
                        print(f"[时间提醒] 播放音乐: {os.path.basename(event.sound_file)}")
                        play_music_with_lock(event.sound_file, loop=False, event_name=event.name, event_id=event.id)
                    else:
                        print(f"[时间提醒] 音乐正在播放中，跳过")
            page.run_task(do_play)

    def _save_events_silent():
        """静默保存事件（不触发任何其他操作）"""
        try:
            json_path = get_data_file_path("events.json")
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump([e.to_dict() for e in events.values()], f, ensure_ascii=False, indent=2)
            print(f"静默保存 {len(events)} 个事件")
        except Exception as e:
            print(f"静默保存失败: {e}")

    def start_background_check():
        """启动后台定时检查"""
        def check_loop():
            while True:
                try:
                    # ========== 先执行时间提醒 ==========
                    check_time_reminders()
                    # ========== 再执行事件检查 ==========
                    check_events()  # 每小时检查事件
                    time.sleep(3600)
                except Exception as e:
                    print(f"定时检查出错: {e}")
                    time.sleep(60)
        
        def time_reminder_loop():
            """时间提醒循环 - 每10分钟检查"""
            while True:
                try:
                    # ========== 只在非 Windows 平台发送保活通知 ==========
                    if not IS_WINDOWS:
                        show_notification(page, "🔔 保活通知", f"当前时间: {datetime.now().strftime('%H:%M:%S')}")      # 10分钟发个通知
                    else:
                        # Windows 平台只在控制台打印
                        print(f"[保活] 当前时间: {datetime.now().strftime('%H:%M:%S')}")
                    time.sleep(600)           # 每10分钟检查一次
                except Exception as e:
                    print(f"时间提醒循环出错: {e}")
                    time.sleep(30)
        
        # 启动两个线程
        check_thread = threading.Thread(target=check_loop, daemon=True)
        check_thread.start()
        
        time_thread = threading.Thread(target=time_reminder_loop, daemon=True)
        time_thread.start()
        
        print("后台定时检查已启动（每小时检查事件）")
        print("时间提醒检查已启动（每10分钟检查）")

    # ========== 农历日期辅助函数 ==========
    def get_lunar_date_str(year, month, day):
        """获取农历日期字符串"""
        try:
            from lunardate import LunarDate
            lunar = LunarDate.fromSolarDate(year, month, day)
            lunar_month_str = number_to_chinese_month(lunar.month)
            lunar_day_str = number_to_chinese_day(lunar.day)
            return f"{lunar_month_str}{lunar_day_str}"
        except:
            return ""
    
    def number_to_chinese_month(month):
        """月份数字转中文"""
        chinese_months = ['正月', '二月', '三月', '四月', '五月', '六月', 
                        '七月', '八月', '九月', '十月', '十一月', '十二月']
        return chinese_months[month - 1] if 1 <= month <= 12 else str(month)

    def number_to_chinese_day(day):
        """日期数字转中文"""
        chinese_days = ['初一', '初二', '初三', '初四', '初五', '初六', '初七', '初八', '初九', '初十',
                        '十一', '十二', '十三', '十四', '十五', '十六', '十七', '十八', '十九', '二十',
                        '廿一', '廿二', '廿三', '廿四', '廿五', '廿六', '廿七', '廿八', '廿九', '三十']
        return chinese_days[day - 1] if 1 <= day <= 30 else str(day)

    load_events()           # 加载事件列表

    load_accounting_data()  # 加载记账列表

    # ========== 粘贴你提供的日历测试代码 ==========
    current_year = datetime.now().year
    current_month = datetime.now().month
    today = datetime.now().date()

    # 农历和节日数据
    solar_holidays = {(1,1):"元旦", (5,1):"劳动节", (5,4):"青年节", (6,1):"儿童节", (10,1):"国庆节"}
    lunar_holidays = {(1,1):"春节", (5,5):"端午节", (8,15):"中秋节"}
    solar_terms = {
        (2,4):"立春", (3,5):"惊蛰", (3,20):"春分", (4,5):"清明", (5,5):"立夏", 
        (5,21):"小满", (6,6):"芒种", (6,21):"夏至", (7,7):"小暑", (7,23):"大暑",
        (8,7):"立秋", (8,23):"处暑", (9,8):"白露", (9,23):"秋分", (10,8):"寒露",
        (10,23):"霜降", (11,7):"立冬", (11,22):"小雪", (12,7):"大雪", (12,21):"冬至"
    }

    def get_mothers_day(year):
        first_day = datetime(year, 5, 1).weekday()
        first_sunday = 1 if first_day == 6 else 7 - first_day
        return first_sunday + 7

    def get_fathers_day(year):
        first_day = datetime(year, 6, 1).weekday()
        first_sunday = 1 if first_day == 6 else 7 - first_day
        return first_sunday + 14

    def get_lunar_str(year, month, day):
        try:
            lunar = ZhDate.from_datetime(datetime(year, month, day))
            lunar_days = ['初一','初二','初三','初四','初五','初六','初七','初八','初九','初十',
                        '十一','十二','十三','十四','十五','十六','十七','十八','十九','二十',
                        '廿一','廿二','廿三','廿四','廿五','廿六','廿七','廿八','廿九','三十']
            return lunar_days[lunar.lunar_day - 1]
        except:
            return ""

    def get_holiday_name(year, month, day):
        if month == 5 and day == get_mothers_day(year):
            return "母亲节"
        if month == 6 and day == get_fathers_day(year):
            return "父亲节"
        if (month, day) in solar_holidays:
            return solar_holidays[(month, day)]
        try:
            lunar = ZhDate.from_datetime(datetime(year, month, day))
            if (lunar.lunar_month, lunar.lunar_day) in lunar_holidays:
                return lunar_holidays[(lunar.lunar_month, lunar.lunar_day)]
        except:
            pass
        if (month, day) in solar_terms:
            return solar_terms[(month, day)]
        return None
    

    # 创建月份文本控件
    month_text = ft.Text(
        f"{current_year}年{current_month}月",
        size=20,
        color=ft.Colors.BLACK,
        weight=ft.FontWeight.BOLD,  # 添加这行
    )

    # 创建回到今天的圆形按钮（初始隐藏）
    # ========== 创建圆形返回按钮（与添加按钮样式一致） ==========
    today_circle_button = ft.Container(
        content=ft.Column(
            [
                ft.Text(
                    str(datetime.now().day),
                    size=20,
                    weight=ft.FontWeight.BOLD,
                    color=ft.Colors.BLUE_700,
                    text_align=ft.TextAlign.CENTER,
                ),
            ],
            spacing=0,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            alignment=ft.MainAxisAlignment.CENTER,
        ),
        width=50,
        height=50,
        #bgcolor=ft.Colors.GREY_100,
        #border=ft.border.Border(
            #left=ft.border.BorderSide(1, ft.Colors.GREY),
            #top=ft.border.BorderSide(1, ft.Colors.GREY),
            #right=ft.border.BorderSide(1, ft.Colors.GREY),
            #bottom=ft.border.BorderSide(1, ft.Colors.GREY),
        #),
        border_radius=25,
        #alignment="center",
        ink=True,
        on_click=lambda e: go_to_today(),
        tooltip=f"回到今天 ({datetime.now().month}月{datetime.now().day}日)",
        visible=False,
        shadow=ft.BoxShadow(
            spread_radius=1,
            blur_radius=5,
            color=ft.Colors.RED_300,
        ),
    )

    # 标题行
    title_row = ft.Row(
        [
            # 年份减按钮
            ft.IconButton(
                icon=ft.Icons.KEYBOARD_DOUBLE_ARROW_LEFT,
                icon_size=20,
                icon_color=ft.Colors.GREY_700,
                on_click=lambda e: change_year(-1),
                tooltip="上一年",
            ),
            # 月份减按钮
            ft.IconButton(
                icon=ft.Icons.KEYBOARD_ARROW_LEFT,
                icon_size=24,
                icon_color=ft.Colors.GREY_700,
                on_click=lambda e: change_month(-1),
                tooltip="上个月",
            ),
            # 月份文本
            ft.Container(
                content=month_text,
                padding=10,  # 简单的整数值
            ),
            # 月份加按钮
            ft.IconButton(
                icon=ft.Icons.KEYBOARD_ARROW_RIGHT,
                icon_size=24,
                icon_color=ft.Colors.GREY_700,
                on_click=lambda e: change_month(1),
                tooltip="下个月",
            ),
            # 年份加按钮
            ft.IconButton(
                icon=ft.Icons.KEYBOARD_DOUBLE_ARROW_RIGHT,
                icon_size=20,
                icon_color=ft.Colors.GREY_700,
                on_click=lambda e: change_year(1),
                tooltip="下一年",
            ),
        ],
        alignment=ft.MainAxisAlignment.CENTER,
        spacing=8,
    )

    # 表格
    data_table = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("   一", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("   二", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("   三", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("   四", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("   五", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("   六", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("   日", color=ft.Colors.BLACK)),
        ],
        rows=[],
        divider_thickness=0,
        column_spacing=8,  # 缩小列间距
    )
    
    def go_to_today():
        """回到当前日期"""
        global current_year, current_month, selected_date, current_date, current_view
        
        print(f"[go_to_today] 执行前 - current_year: {current_year}, current_month: {current_month}")
        
        # 获取当前日期
        today = datetime.now()
        current_year = today.year
        current_month = today.month
        
        print(f"[go_to_today] 设置后 - current_year: {current_year}, current_month: {current_month}")
        
        # 更新月份文本显示
        month_text.value = f"{current_year}年{current_month}月"
        
        # 关键修复：设置 selected_date 为今天的日期，而不是 None
        selected_date = today.date()
        current_date = today.date()
        
        # 更新日历显示（重新生成日历表格，会高亮 selected_date）
        update_calendar()
        
        # 刷新事件列表（显示全部事件）
        #current_view = "all"
        #refresh_events_list()
        refresh_current_view_by_state()
        
        # 更新日期显示
        date_display.value = today.strftime("%Y年%m月%d日")
        
        # 强制刷新页面
        page.update()
        
        show_bottom_message(f"已回到今天 {today.strftime('%Y年%m月%d日')}")

    def change_month(delta):
        global current_year, current_month, selected_date
        current_month += delta
        if current_month > 12:
            current_month = 1
            current_year += 1
        elif current_month < 1:
            current_month = 12
            current_year -= 1
        
        # 切换月份后，清除选中的日期
        selected_date = None
        
        update_calendar()

    def change_year(delta):
        """改变年份"""
        global current_year, current_month, selected_date
        current_year += delta
        # 确保年份在合理范围内（1900-2100）
        if current_year < 1900:
            current_year = 1900
        elif current_year > 2100:
            current_year = 2100
        
        # 切换年份后，清除选中的日期
        selected_date = None
        
        update_calendar()

    def update_calendar():
        global selected_date
        
        # 更新月份文本显示
        month_text.value = f"{current_year}年{current_month}月"
        
        # ========== 判断是否显示返回按钮 ==========
        today = datetime.now()
        is_current_month = (current_year == today.year and current_month == today.month)
        
        # 获取今天日期
        today_date = datetime.now().date()
        is_selected_today = (selected_date == today_date) if selected_date else False
        
        # 决定是否显示返回按钮
        if today_circle_button:
            if is_current_month:
                # 本月：只有选中了非今天的日期时才显示
                today_circle_button.visible = selected_date is not None and not is_selected_today
            else:
                # 其他月份：始终显示
                today_circle_button.visible = True
            
            #print(f"[调试] 本月: {is_current_month}, 选中日期: {selected_date}, 是今天: {is_selected_today}")
            #print(f"[调试] 按钮显示: {today_circle_button.visible}")
        
        # 更新按钮上的日期数字（显示今天的日期）
        if today_circle_button and hasattr(today_circle_button, 'content'):
            if isinstance(today_circle_button.content, ft.Text):
                today_circle_button.content.value = str(today.day)
            elif isinstance(today_circle_button.content, ft.Column):
                if today_circle_button.content.controls and len(today_circle_button.content.controls) > 0:
                    if isinstance(today_circle_button.content.controls[0], ft.Text):
                        today_circle_button.content.controls[0].value = str(today.day)
        if today_circle_button:
            today_circle_button.tooltip = f"回到今天 ({today.month}月{today.day}日)"
            #today_circle_button.update()
        
        # 清空表格并重新生成
        data_table.rows.clear()
        today_date = datetime.now().date()
        
        # 日期点击处理函数
        def on_date_click(e, year, month, day):
            global selected_date, current_date, previous_view
            selected_date = datetime(year, month, day).date()
            print(f"选中日期: {selected_date}")
            
            # 保存当前视图到 previous_view
            previous_view = current_view
            current_date = selected_date
            date_display.value = selected_date.strftime("%Y年%m月%d日")

            # 获取农历日期
            lunar_str = get_lunar_date_str(year, month, day)
            current_selected_lunar = lunar_str
            
            # 传入筛选日期刷新事件列表
            refresh_events_list(filter_date=selected_date)
            
            # 更新日历（会重新计算返回按钮的显示）
            update_calendar()
            
            show_bottom_message(f"已切换到 {selected_date.strftime('%Y年%m月%d日')}")
            
            page.update()
        
        for week in calendar.monthcalendar(current_year, current_month):
            cells = []
            for i, day in enumerate(week):
                if day == 0:
                    cells.append(ft.DataCell(ft.Text("")))
                else:
                    holiday = get_holiday_name(current_year, current_month, day)
                    lunar = get_lunar_str(current_year, current_month, day)
                    
                    # 第二行文本（纯字符串，用于当天拼接）
                    if holiday:
                        second_line_text_str = holiday
                        if holiday in ["劳动节","国庆节","春节","母亲节","父亲节"]:
                            second_line_color = ft.Colors.RED
                        else:
                            second_line_color = ft.Colors.GREEN
                        
                        second_line_text_widget = ft.Text(
                            holiday,
                            size=10,
                            weight=ft.FontWeight.BOLD if holiday in ["劳动节","国庆节","春节","母亲节","父亲节"] else ft.FontWeight.NORMAL,
                            color=second_line_color,
                        )
                    else:
                        second_line_text_str = lunar
                        second_line_color = ft.Colors.GREY_600
                        second_line_text_widget = ft.Text(lunar, size=10, color=second_line_color)
                    
                    # 公历数字颜色
                    if i == 5:
                        num_color = ft.Colors.BLACK
                    elif i == 6:
                        num_color = ft.Colors.BLACK
                    else:
                        num_color = ft.Colors.BLACK
                    
                    current_date = datetime(current_year, current_month, day).date()
                    # 判断优先级：选中日期 > 当天日期
                    is_selected = (selected_date is not None and current_date == selected_date)
                    is_today = (current_date == today_date)
                    # 判断是否有其他日期被选中（selected_date 不为 None 且不是当天）
                    has_other_selected = (selected_date is not None and selected_date != today_date)
                    
                    if is_selected:
                        # 选中的日期
                        if is_today:
                            # 选中的是当天日期：蓝色实心圆圈
                            cell_content = ft.Container(
                                content=ft.Column(
                                    [
                                        ft.Text(str(day), size=15, weight=ft.FontWeight.BOLD, color=ft.Colors.WHITE),
                                        ft.Text(second_line_text_str, size=10, color=ft.Colors.WHITE),
                                    ],
                                    spacing=0,
                                    horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                                ),
                                width=40,
                                height=40,
                                bgcolor=ft.Colors.BLUE_700,
                                border_radius=20,
                            )
                        else:
                            # 选中的是其他日期：蓝色空心圆圈
                            cell_content = ft.Container(
                                content=ft.Column(
                                    [
                                        ft.Text(str(day), size=15, weight=ft.FontWeight.BOLD, color=ft.Colors.BLACK),
                                        ft.Text(second_line_text_str, size=10, color=ft.Colors.BLACK),
                                    ],
                                    spacing=0,
                                    horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                                ),
                                width=40,
                                height=40,
                                #bgcolor=ft.Colors.GREEN_700,
                                border=ft.border.Border(
                                    left=ft.border.BorderSide(1, ft.Colors.BLUE),
                                    top=ft.border.BorderSide(1, ft.Colors.BLUE),
                                    right=ft.border.BorderSide(1, ft.Colors.BLUE),
                                    bottom=ft.border.BorderSide(1, ft.Colors.BLUE),
                                ),
                                border_radius=20,
                            )
                    elif is_today and not has_other_selected:
                        # 当天日期：只有没有被选中且没有其他日期被选中时，才显示蓝色实心圆圈
                        cell_content = ft.Container(
                            content=ft.Column(
                                [
                                    ft.Text(str(day), size=15, weight=ft.FontWeight.BOLD, color=ft.Colors.WHITE),
                                    ft.Text(second_line_text_str, size=10, color=ft.Colors.WHITE),
                                ],
                                spacing=0,
                                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                            ),
                            width=40,
                            height=40,
                            bgcolor=ft.Colors.BLUE_700,
                            border_radius=20,
                        )
                    else:
                        # 普通日期：当天日期在其他日期被选中时，也要显示加粗
                        is_bold = is_today  # 当天日期加粗，其他日期不加粗
                        cell_content = ft.Container(
                            content=ft.Column([
                                ft.Text(str(day), size=15, color=ft.Colors.BLUE if is_bold else num_color, 
                                        weight=ft.FontWeight.BOLD if is_bold else ft.FontWeight.NORMAL),
                                second_line_text_widget,
                            ], spacing=0, horizontal_alignment=ft.CrossAxisAlignment.CENTER, width=40),
                            height=40,
                        )
                    
                    cells.append(ft.DataCell(
                        cell_content,
                        on_tap=lambda e, y=current_year, m=current_month, d=day: on_date_click(e, y, m, d)
                    ))
            data_table.rows.append(ft.DataRow(cells=cells))
        
        page.update()
    
    # 创建日历容器
    calendar_widget = ft.Container(
        content=ft.Column([
            title_row,
            ft.Divider(height=5),
            data_table,
        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),  # 添加这一行
        bgcolor=None,
        padding=0,
        border_radius=10,
    )

    # 设置页面宽度自适应手机
    page.padding = 5
    page.bgcolor = ft.Colors.WHITE

    # 初始化日历
    update_calendar()
    # ========== 日历代码结束 ==========
    



    # ========== 开始添加导入事件和导出事件按钮功能 ==========
    async def export_events_async(e):
        """导出事件到Excel（兼容 Windows 和 Android）"""
        try:
            if not events:
                show_bottom_message("没有事件可导出")
                return
            
            # 创建临时文件
            temp_dir = get_data_file_path("")
            temp_file = os.path.join(temp_dir, f"events_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "事件列表"
            
            # 写入表头（添加 reminders 和 workday_only 字段）
            headers = ["事件类型", "名称", "birth_date", "历法", "重复类型", "音乐文件路径", 
                    "已提醒年份", "提醒时间(多个用|分隔)", "仅法定工作日提醒"]
            ws.append(headers)
            
            # 设置表头样式
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCE6FF", end_color="CCE6FF", fill_type="solid")
            
            # 写入事件数据
            for event in events.values():
                if event.event_type == "birthday":
                    event_type = "生日"
                elif event.event_type == "event":
                    event_type = "纪念日/事件"
                elif event.event_type == "monthly":
                    event_type = "每月提醒"
                elif event.event_type == "daily":
                    event_type = "每天提醒"
                elif event.event_type == "weekly":
                    event_type = "每周提醒"
                else:
                    event_type = "一次性事件"
                
                calendar_str = "阳历" if event.calendar_type == "solar" else "农历"
                
                if event.repeat_type == "yearly":
                    repeat_str = "每年"
                elif event.repeat_type == "monthly":
                    repeat_str = "每月"
                elif event.repeat_type == "daily":
                    repeat_str = "每天"
                elif event.repeat_type == "weekly":
                    repeat_str = "每周"
                else:
                    repeat_str = "一次性"
                
                reminded_year = event.last_remind_year if event.last_remind_year > 0 else ""
                
                # 处理提醒时间：多个用 | 分隔
                reminders_str = ""
                if hasattr(event, 'reminders') and event.reminders:
                    time_list = [r.get("time", "") for r in event.reminders if r.get("enabled")]
                    reminders_str = "|".join(time_list)
                
                # 处理法定工作日提醒
                workday_only_str = "是" if getattr(event, 'workday_only', False) else "否"
                
                ws.append([
                    event_type,
                    event.name,
                    event.birth_date,
                    calendar_str,
                    repeat_str,
                    event.sound_file,
                    reminded_year,
                    reminders_str,
                    workday_only_str,
                ])
            
            # 调整列宽
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 8
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 40
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 15
            
            # 保存临时文件
            wb.save(temp_file)
            
            # 读取文件内容为字节数组（移动端需要）
            with open(temp_file, 'rb') as f:
                file_bytes = f.read()
            
            # 创建 FilePicker
            file_picker = ft.FilePicker()
            page.services.append(file_picker)
            page.update()
            
            # 选择保存位置 - 移动端需要传递 src_bytes
            result = await file_picker.save_file(
                file_name=f"events_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                src_bytes=file_bytes,  # 移动端必需！
                dialog_title="保存Excel文件"
            )
            
            # 移除 FilePicker
            page.services.remove(file_picker)
            page.update()
            
            # 删除临时文件
            os.remove(temp_file)
            
            if result:
                show_bottom_message(f"成功导出 {len(events)} 条事件")
            else:
                show_bottom_message("已取消导出")
            
            page.update()
            
        except Exception as ex:
            show_bottom_message(f"导出失败: {str(ex)}")
            print(f"导出错误: {ex}")
            import traceback
            traceback.print_exc()


    async def import_events_async(e):
        """从Excel导入事件 - 仿照删除事件的对话框模式"""
        
        # 存储对话框容器的引用
        menu_container = None
        
        def close_menu():
            nonlocal menu_container
            if menu_container and menu_container in page.overlay:
                page.overlay.remove(menu_container)
                menu_container = None
                page.update()
        
        # 选择文件并导入（异步）
        async def select_file_and_import():
            file_picker = None
            try:
                # 创建 FilePicker
                file_picker = ft.FilePicker()
                page.services.append(file_picker)
                page.update()
                
                result = await file_picker.pick_files(
                    allow_multiple=False,
                    allowed_extensions=["xlsx", "xls"],
                    dialog_title="选择Excel文件"
                )
                
                # 移除 FilePicker
                if file_picker and file_picker in page.overlay:
                    page.services.remove(file_picker)
                page.update()
                
                if not result or len(result) == 0:
                    show_bottom_message("未选择文件")
                    return
                
                # 获取文件路径
                if hasattr(result[0], 'path'):
                    file_path = result[0].path
                elif hasattr(result[0], 'bytes'):
                    temp_dir = get_data_file_path("")
                    temp_file = os.path.join(temp_dir, f"temp_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                    with open(temp_file, 'wb') as f:
                        f.write(result[0].bytes)
                    file_path = temp_file
                else:
                    file_path = str(result[0])
                
                # 执行导入
                await do_import(file_path)
                
                # 如果是临时文件，删除
                if 'temp_file' in locals() and os.path.exists(temp_file):
                    os.remove(temp_file)
                
            except Exception as ex:
                show_bottom_message(f"导入失败: {str(ex)}")
                print(f"导入错误: {ex}")
                import traceback
                traceback.print_exc()
            finally:
                check_events._running = False
                if file_picker and file_picker in page.overlay:
                    page.overlay.remove(file_picker)
                page.update()
        
        # 包装函数：先关闭菜单，再启动选择文件
        def on_select_file():
            close_menu()  # 先关闭菜单
            asyncio.create_task(select_file_and_import())  # 然后选择文件
        
        # 取消按钮
        def on_cancel():
            close_menu()
            show_bottom_message("已取消导入")
        
        async def do_import(file_path):
            """执行导入逻辑"""
            show_bottom_message(f"正在导入: {os.path.basename(file_path)}")
            page.update()
            
            # 读取Excel文件
            wb = load_workbook(file_path)
            ws = wb.active
            
            imported_count = 0
            skipped_count = 0
            new_events = {}
            
            # 从第二行开始读取
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 5:
                    continue
                
                event_type_str = str(row[0]).strip() if row[0] else ""
                name = str(row[1]).strip() if row[1] else ""
                birth_date_raw = str(row[2]).strip() if row[2] else ""
                calendar_str = str(row[3]).strip() if row[3] else "阳历"
                repeat_str = str(row[4]).strip() if row[4] else "每年"
                sound_file = str(row[5]).strip() if row[5] else ""
                reminded_year_str = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                reminders_str = str(row[7]).strip() if len(row) > 7 and row[7] else ""  # 新增：提醒时间
                workday_only_str = str(row[8]).strip() if len(row) > 8 and row[8] else "否"  # 新增：法定工作日提醒
                
                if not name:
                    skipped_count += 1
                    continue
                
                # 清理日期字符串
                if ' ' in birth_date_raw:
                    birth_date_raw = birth_date_raw.split(' ')[0]
                
                # 转换事件类型
                if event_type_str in ["生日", "birthday"]:
                    event_type = "birthday"
                elif event_type_str in ["纪念日/事件", "event"]:
                    event_type = "event"
                elif event_type_str in ["每月提醒", "monthly"]:
                    event_type = "monthly"
                elif event_type_str in ["每天提醒", "daily"]:
                    event_type = "daily"
                elif event_type_str in ["每周提醒", "weekly"]:
                    event_type = "weekly"
                elif event_type_str in ["一次性事件", "once"]:
                    event_type = "once"
                else:
                    # 根据 repeat_str 推断
                    if repeat_str in ["每月", "monthly"]:
                        event_type = "monthly"
                    elif repeat_str in ["每天", "daily"]:
                        event_type = "daily"
                    elif repeat_str in ["每周", "weekly"]:
                        event_type = "weekly"
                    elif repeat_str in ["一次性", "once"]:
                        event_type = "once"
                    else:
                        event_type = "birthday"
                
                calendar_type = "lunar" if calendar_str in ["农历", "lunar"] else "solar"
                
                # 处理重复类型
                if repeat_str in ["每月", "monthly"]:
                    repeat_type = "monthly"
                elif repeat_str in ["每天", "daily"]:
                    repeat_type = "daily"
                elif repeat_str in ["每周", "weekly"]:
                    repeat_type = "weekly"
                elif repeat_str in ["一次性", "once"]:
                    repeat_type = "once"
                else:
                    repeat_type = "yearly"
                
                # 处理birth_date格式
                try:
                    if event_type == "monthly" or repeat_type == "monthly":
                        # 每月事件：只存日
                        day_num = int(float(birth_date_raw)) if '.' in birth_date_raw else int(birth_date_raw)
                        if 1 <= day_num <= 31:
                            birth_date = f"{day_num:02d}"
                        else:
                            skipped_count += 1
                            continue
                            
                    elif event_type == "daily" or repeat_type == "daily":
                        # 每天事件：不需要日期
                        birth_date = ""
                        
                    elif event_type == "weekly" or repeat_type == "weekly":
                        # 每周事件：存星期几 (1-7)
                        weekday = str(int(float(birth_date_raw))) if '.' in birth_date_raw else str(birth_date_raw)
                        if weekday in ["1", "2", "3", "4", "5", "6", "7"]:
                            birth_date = weekday
                        else:
                            birth_date = "1"
                            
                    elif event_type == "once" or repeat_type == "once":
                        # 一次性事件：完整日期
                        parts = birth_date_raw.split("-")
                        if len(parts) == 3:
                            year = int(parts[0])
                            month = int(parts[1])
                            day = int(parts[2])
                            datetime(year, month, day)
                            birth_date = f"{year:04d}-{month:02d}-{day:02d}"
                        else:
                            skipped_count += 1
                            continue
                        
                    else:
                        # 生日/纪念日：完整日期
                        parts = birth_date_raw.split("-")
                        if len(parts) == 3:
                            year = int(parts[0])
                            month = int(parts[1])
                            day = int(parts[2])
                            if 1 <= month <= 12 and 1 <= day <= 31:
                                birth_date = f"{year:04d}-{month:02d}-{day:02d}"
                            else:
                                raise ValueError
                        else:
                            raise ValueError
                            
                except Exception as e:
                    print(f"解析日期失败: {birth_date_raw}, 错误: {e}")
                    skipped_count += 1
                    continue
                
                # 处理提醒时间（多个用|分隔）
                reminders = []
                if reminders_str and reminders_str != "":
                    time_list = reminders_str.split("|")
                    for t in time_list:
                        t = t.strip()
                        if t and ":" in t:
                            reminders.append({"time": t, "enabled": True})
                
                # 处理法定工作日提醒
                workday_only = workday_only_str == "是"
                
                # 处理已提醒年份
                last_remind_year = 0
                if reminded_year_str and reminded_year_str.isdigit():
                    last_remind_year = int(reminded_year_str)
                
                # 生成新的事件ID
                event_id = str(int(datetime.now().timestamp() * 1000) + imported_count)
                
                # 创建事件
                new_event = Event(
                    event_id, name, birth_date, calendar_type,
                    event_type, sound_file, repeat_type,
                    reminders=reminders
                )
                new_event.workday_only = workday_only
                new_event.last_remind_year = last_remind_year
                new_event.reminded_this_year = (last_remind_year == datetime.now().year)
                new_events[event_id] = new_event
                imported_count += 1
                print(f"[导入] 成功导入: {name} (类型: {event_type})")
            
            if imported_count == 0:
                show_bottom_message(f"没有导入任何事件，跳过 {skipped_count} 行")
                return
            
            # 确认替换对话框
            confirm_dialog_container = None
            
            def close_confirm_dialog():
                nonlocal confirm_dialog_container
                if confirm_dialog_container and confirm_dialog_container in page.overlay:
                    page.overlay.remove(confirm_dialog_container)
                    confirm_dialog_container = None
                    page.update()
            
            def confirm_replace():
                close_confirm_dialog()
                events.clear()
                events.update(new_events)
                save_events(trigger_check=False)

                refresh_current_view_by_state()

                update_calendar()

                # ========== 导入成功后，立即检查今日事件 ==========
                # 直接调用，不需要 Timer
                check_events()
                check_time_reminders()
                # ========== 导入后重新检查视图 ==========
                determine_startup_view()
                
                show_bottom_message(f"成功导入 {imported_count} 条事件")
                page.update()
            
            def cancel_replace():
                close_confirm_dialog()
                show_bottom_message("已取消导入")
                page.update()
            
            confirm_content = ft.Container(
                content=ft.Column([
                    ft.Container(
                        content=ft.Icon(ft.Icons.INFO, size=55, color=ft.Colors.BLUE_700),
                        padding=10,
                        bgcolor=ft.Colors.BLUE_50,
                        border_radius=50,
                    ),
                    ft.Text("确认导入", size=22, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700, text_align=ft.TextAlign.CENTER),
                    ft.Divider(height=1, color=ft.Colors.GREY_300),
                    ft.Text(f"即将导入 {imported_count} 条事件", size=14, color=ft.Colors.GREY_700, text_align=ft.TextAlign.CENTER),
                    ft.Text(f"当前有 {len(events)} 条事件将被替换", size=12, color=ft.Colors.ORANGE_700, text_align=ft.TextAlign.CENTER),
                    ft.Divider(height=1, color=ft.Colors.GREY_300),
                    ft.Row([
                        ft.ElevatedButton(
                            "取消", 
                            on_click=lambda e: cancel_replace(), 
                            expand=True,
                            style=ft.ButtonStyle(bgcolor=ft.Colors.GREY_100, color=ft.Colors.GREY_700),
                        ),
                        ft.ElevatedButton(
                            "确认导入", 
                            on_click=lambda e: confirm_replace(), 
                            expand=True,
                            style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_700, color=ft.Colors.WHITE),
                        ),
                    ], spacing=12, alignment=ft.MainAxisAlignment.CENTER),
                ], spacing=15, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                width=320,
                padding=20,
                bgcolor=ft.Colors.WHITE,
                border_radius=16,
            )
            
            confirm_dialog_container = ft.Container(
                content=ft.Column([
                    ft.Container(expand=True),
                    ft.Row([
                        ft.Container(expand=True),
                        confirm_content,
                        ft.Container(expand=True),
                    ]),
                    ft.Container(expand=True),
                ]),
                expand=True,
                bgcolor=ft.Colors.BLACK26,
                on_click=lambda e: close_confirm_dialog(),
            )
            
            page.overlay.append(confirm_dialog_container)
            page.update()
        
        # 创建底部操作菜单
        menu_content = ft.Container(
            content=ft.Column([
                ft.Container(
                    content=ft.Icon(ft.Icons.FOLDER_OPEN, size=55, color=ft.Colors.BLUE_700),
                    padding=10,
                    bgcolor=ft.Colors.BLUE_50,
                    border_radius=50,
                ),
                ft.Text("导入事件", size=22, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700, text_align=ft.TextAlign.CENTER),
                ft.Divider(height=1, color=ft.Colors.GREY_300),
                ft.Text("请选择Excel文件", size=14, color=ft.Colors.GREY_700, text_align=ft.TextAlign.CENTER),
                ft.Text("支持格式: .xlsx, .xls", size=12, color=ft.Colors.GREY_500, text_align=ft.TextAlign.CENTER),
                ft.Divider(height=1, color=ft.Colors.GREY_300),
                ft.Row([
                    ft.ElevatedButton(
                        "选择Excel文件", 
                        on_click=lambda e: on_select_file(),  # 使用包装函数
                        expand=True,
                        style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_700, color=ft.Colors.WHITE),
                    ),
                ], alignment=ft.MainAxisAlignment.CENTER),
                ft.Row([
                    ft.ElevatedButton(
                        "取消", 
                        on_click=lambda e: on_cancel(), 
                        expand=True,
                        style=ft.ButtonStyle(bgcolor=ft.Colors.GREY_100, color=ft.Colors.GREY_700),
                    ),
                ], alignment=ft.MainAxisAlignment.CENTER),
            ], spacing=15, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            width=320,
            padding=20,
            bgcolor=ft.Colors.WHITE,
            border_radius=16,
        )
        
        menu_container = ft.Container(
            content=ft.Column([
                ft.Container(expand=True),
                ft.Row([
                    ft.Container(expand=True),
                    menu_content,
                    ft.Container(expand=True),
                ]),
                ft.Container(expand=True),
            ]),
            expand=True,
            bgcolor=ft.Colors.BLACK26,
            on_click=lambda e: close_menu(),
        )
        
        page.overlay.append(menu_container)
        page.update()

    # ========== 记账数据导入导出 ==========
    async def export_accounting_async(e):
        """导出记账数据到Excel"""
        global transactions  # 添加这行，确保使用全局变量
        try:

            print(f"[导出记账] transactions 数量: {len(transactions)}")  # 添加调试
            
            if not transactions:
                show_bottom_message("没有记账数据可导出")
                return
            
            temp_dir = get_data_file_path("")
            temp_file = os.path.join(temp_dir, f"accounting_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "记账本"
            
            # 写入表头
            headers = ["日期", "类型", "分类", "金额", "备注"]
            ws.append(headers)
            
            # 设置表头样式
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCE6FF", end_color="CCE6FF", fill_type="solid")
            
            # 写入数据
            for t in transactions:
                type_str = "收入" if t.type == "income" else "支出"
                ws.append([
                    t.date,
                    type_str,
                    t.category,
                    t.amount,
                    t.note,
                ])
            
            # 调整列宽
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 30
            
            wb.save(temp_file)
            
            with open(temp_file, 'rb') as f:
                file_bytes = f.read()
            
            file_picker = ft.FilePicker()
            page.services.append(file_picker)
            page.update()
            
            result = await file_picker.save_file(
                file_name=f"accounting_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                src_bytes=file_bytes,
                dialog_title="保存记账Excel文件"
            )
            
            page.services.remove(file_picker)
            page.update()
            os.remove(temp_file)
            
            if result:
                show_bottom_message(f"成功导出 {len(transactions)} 条记账记录")
            else:
                show_bottom_message("已取消导出")
            
            page.update()
            
        except Exception as ex:
            show_bottom_message(f"导出失败: {str(ex)}")
            print(f"导出错误: {ex}")
            import traceback
            traceback.print_exc()


    async def import_accounting_async(e):
        """从Excel导入记账数据"""
        global transactions  # 添加这行
        
        menu_container = None
        
        def close_menu():
            nonlocal menu_container
            if menu_container and menu_container in page.overlay:
                page.overlay.remove(menu_container)
                menu_container = None
                page.update()
        
        async def select_file_and_import():
            file_picker = None
            try:
                file_picker = ft.FilePicker()
                page.services.append(file_picker)
                page.update()
                
                result = await file_picker.pick_files(
                    allow_multiple=False,
                    allowed_extensions=["xlsx", "xls"],
                    dialog_title="选择记账Excel文件"
                )
                
                if file_picker and file_picker in page.overlay:
                    page.services.remove(file_picker)
                page.update()
                
                if not result or len(result) == 0:
                    show_bottom_message("未选择文件")
                    return
                
                if hasattr(result[0], 'path'):
                    file_path = result[0].path
                elif hasattr(result[0], 'bytes'):
                    temp_dir = get_data_file_path("")
                    temp_file = os.path.join(temp_dir, f"temp_accounting_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                    with open(temp_file, 'wb') as f:
                        f.write(result[0].bytes)
                    file_path = temp_file
                else:
                    file_path = str(result[0])
                
                await do_import_accounting(file_path)
                
                if 'temp_file' in locals() and os.path.exists(temp_file):
                    os.remove(temp_file)
                
            except Exception as ex:
                show_bottom_message(f"导入失败: {str(ex)}")
                print(f"导入错误: {ex}")
                import traceback
                traceback.print_exc()
            finally:
                if file_picker and file_picker in page.overlay:
                    page.overlay.remove(file_picker)
                page.update()
        
        def on_select_file():
            close_menu()
            asyncio.create_task(select_file_and_import())
        
        def on_cancel():
            close_menu()
            show_bottom_message("已取消导入")
        
        async def do_import_accounting(file_path):
            show_bottom_message(f"正在导入记账数据: {os.path.basename(file_path)}")
            page.update()
            
            wb = load_workbook(file_path)
            ws = wb.active
            
            imported_count = 0
            skipped_count = 0
            new_transactions = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 4:
                    continue
                
                date = str(row[0]).strip() if row[0] else ""
                type_str = str(row[1]).strip() if row[1] else ""
                category = str(row[2]).strip() if row[2] else ""
                amount_str = str(row[3]).strip() if row[3] else ""
                note = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                
                if not date or not category or not amount_str:
                    skipped_count += 1
                    continue
                
                try:
                    amount = float(amount_str)
                    if amount <= 0:
                        skipped_count += 1
                        continue
                except:
                    skipped_count += 1
                    continue
                
                if type_str == "收入":
                    transaction_type = "income"
                elif type_str == "支出":
                    transaction_type = "expense"
                else:
                    skipped_count += 1
                    continue
                
                transaction_id = str(int(datetime.now().timestamp() * 1000) + imported_count)
                new_transaction = Transaction(
                    id=transaction_id,
                    date=date,
                    type=transaction_type,
                    category=category,
                    amount=amount,
                    note=note,
                )
                new_transactions.append(new_transaction)
                imported_count += 1
            
            if imported_count == 0:
                show_bottom_message(f"没有导入任何记账记录，跳过 {skipped_count} 行")
                return
            
            confirm_dialog_container = None
            
            def close_confirm_dialog():
                nonlocal confirm_dialog_container
                if confirm_dialog_container and confirm_dialog_container in page.overlay:
                    page.overlay.remove(confirm_dialog_container)
                    confirm_dialog_container = None
                    page.update()
            
            def confirm_replace():
                close_confirm_dialog()
                global transactions
                transactions = new_transactions
                save_accounting_data()
                show_bottom_message(f"成功导入 {imported_count} 条记账记录")
                page.update()
            
            def cancel_replace():
                close_confirm_dialog()
                show_bottom_message("已取消导入")
                page.update()
            
            confirm_content = ft.Container(
                content=ft.Column([
                    ft.Container(
                        content=ft.Icon(ft.Icons.INFO, size=55, color=ft.Colors.BLUE_700),
                        padding=10,
                        bgcolor=ft.Colors.BLUE_50,
                        border_radius=50,
                    ),
                    ft.Text("确认导入记账数据", size=22, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700),
                    ft.Divider(),
                    ft.Text(f"即将导入 {imported_count} 条记账记录", size=14),
                    ft.Text(f"当前有 {len(transactions)} 条记录将被替换", size=12, color=ft.Colors.ORANGE_700),
                    ft.Divider(),
                    ft.Row([
                        ft.ElevatedButton("取消", on_click=lambda e: cancel_replace(), expand=True),
                        ft.ElevatedButton("确认导入", on_click=lambda e: confirm_replace(), expand=True,
                                        style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_700, color=ft.Colors.WHITE)),
                    ], spacing=12),
                ], spacing=15),
                width=320, padding=20, bgcolor=ft.Colors.WHITE, border_radius=16,
            )
            
            confirm_dialog_container = ft.Container(
                content=ft.Column([
                    ft.Container(expand=True),
                    ft.Row([ft.Container(expand=True), confirm_content, ft.Container(expand=True)]),
                    ft.Container(expand=True),
                ]),
                expand=True, bgcolor=ft.Colors.BLACK26, on_click=lambda e: close_confirm_dialog(),
            )
            
            page.overlay.append(confirm_dialog_container)
            page.update()
        
        menu_content = ft.Container(
            content=ft.Column([
                ft.Container(
                    content=ft.Icon(ft.Icons.FOLDER_OPEN, size=55, color=ft.Colors.BLUE_700),
                    padding=15,
                    bgcolor=ft.Colors.BLUE_50,
                    border_radius=50,
                    #alignment=ft.Alignment(0, 0),  # 图标居中
                ),
                ft.Text("导入记账数据", size=22, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700, text_align=ft.TextAlign.CENTER),
                ft.Divider(),
                ft.Text("请选择记账Excel文件", size=14, text_align=ft.TextAlign.CENTER),
                ft.Text("支持格式: .xlsx, .xls", size=12, color=ft.Colors.GREY_500, text_align=ft.TextAlign.CENTER),
                ft.Divider(),
                ft.Button(
                    "选择文件", 
                    on_click=lambda e: on_select_file(), 
                    expand=True,
                    style=ft.ButtonStyle(
                        bgcolor=ft.Colors.BLUE_700,
                        color=ft.Colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                ),
                ft.Button(
                    "取消", 
                    on_click=lambda e: on_cancel(), 
                    expand=True,
                    style=ft.ButtonStyle(
                        bgcolor=ft.Colors.GREY_100,
                        color=ft.Colors.GREY_700,
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                ),
            ], spacing=15, horizontal_alignment=ft.CrossAxisAlignment.CENTER),  # 水平居中
            width=320,
            padding=25,
            bgcolor=ft.Colors.WHITE,
            border_radius=20,
            shadow=ft.BoxShadow(
                spread_radius=1,
                blur_radius=15,
                color=ft.Colors.BLACK12,
                offset=ft.Offset(0, 4),
            ),
        )
        
        menu_container = ft.Container(
            content=ft.Column([
                ft.Container(expand=True),
                ft.Row([ft.Container(expand=True), menu_content, ft.Container(expand=True)]),
                ft.Container(expand=True),
            ]),
            expand=True, bgcolor=ft.Colors.BLACK26, on_click=lambda e: close_menu(),
        )
        
        page.overlay.append(menu_container)
        page.update()

    # 导入导出包装函数，增加选择菜单
    def show_export_menu(e):
        """显示导出选择菜单"""
        menu_container = None
        
        def close_menu():
            nonlocal menu_container
            if menu_container and menu_container in page.overlay:
                page.overlay.remove(menu_container)
                menu_container = None
                page.update()
        
        menu_content = ft.Container(
            content=ft.Column([
                ft.Container(
                    content=ft.Icon(ft.Icons.DOWNLOAD, size=48, color=ft.Colors.BLUE_700),
                    padding=10,
                    bgcolor=ft.Colors.BLUE_50,
                    border_radius=50,
                    #alignment=ft.Alignment(0, 0),  # 图标居中
                ),
                ft.Text("导出数据", size=20, weight=ft.FontWeight.BOLD, text_align=ft.TextAlign.CENTER),
                ft.Text("请选择要导出的数据类型", size=12, color=ft.Colors.GREY_500, text_align=ft.TextAlign.CENTER),
                ft.Divider(),
                ft.Button(
                    "📅 事件列表", 
                    on_click=lambda e: [close_menu(), asyncio.create_task(export_events_async(e))], 
                    expand=True,
                    style=ft.ButtonStyle(
                        bgcolor=ft.Colors.BLUE_700,
                        color=ft.Colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                ),
                ft.Button(
                    "💰 记账列表", 
                    on_click=lambda e: [close_menu(), asyncio.create_task(export_accounting_async(e))], 
                    expand=True,
                    style=ft.ButtonStyle(
                        bgcolor=ft.Colors.GREEN_700,
                        color=ft.Colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                ),
                ft.Divider(),
                ft.TextButton(
                    "取消", 
                    on_click=lambda e: close_menu(),
                    expand=True,
                    style=ft.ButtonStyle(color=ft.Colors.GREY_600),
                ),
            ], spacing=12, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            width=280,
            padding=20,
            bgcolor=ft.Colors.WHITE,
            border_radius=20,
            shadow=ft.BoxShadow(
                spread_radius=1,
                blur_radius=15,
                color=ft.Colors.BLACK12,
                offset=ft.Offset(0, 4),
            ),
        )
        
        menu_container = ft.Container(
            content=ft.Column([
                ft.Container(expand=True),
                ft.Row([ft.Container(expand=True), menu_content, ft.Container(expand=True)]),
                ft.Container(expand=True),
            ]),
            expand=True, bgcolor=ft.Colors.BLACK26, on_click=lambda e: close_menu(),
        )
        
        page.overlay.append(menu_container)
        page.update()


    def show_import_menu(e):
        """显示导入选择菜单"""
        menu_container = None
        
        def close_menu():
            nonlocal menu_container
            if menu_container and menu_container in page.overlay:
                page.overlay.remove(menu_container)
                menu_container = None
                page.update()
        
        menu_content = ft.Container(
            content=ft.Column([
                ft.Container(
                    content=ft.Icon(ft.Icons.UPLOAD, size=48, color=ft.Colors.BLUE_700),
                    padding=10,
                    bgcolor=ft.Colors.BLUE_50,
                    border_radius=50,
                    #alignment=ft.Alignment(0, 0),  # 图标居中
                ),
                ft.Text("导入数据", size=20, weight=ft.FontWeight.BOLD, text_align=ft.TextAlign.CENTER),
                ft.Text("请选择要导入的数据类型", size=12, color=ft.Colors.GREY_500, text_align=ft.TextAlign.CENTER),
                ft.Divider(),
                ft.Button(
                    "📅 事件列表", 
                    on_click=lambda e: [close_menu(), import_events_wrapper(e)], 
                    expand=True,
                    style=ft.ButtonStyle(
                        bgcolor=ft.Colors.BLUE_700,
                        color=ft.Colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                ),
                ft.Button(
                    "💰 记账列表", 
                    on_click=lambda e: [close_menu(), asyncio.create_task(import_accounting_async(e))], 
                    expand=True,
                    style=ft.ButtonStyle(
                        bgcolor=ft.Colors.GREEN_700,
                        color=ft.Colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                ),
                ft.Divider(),
                ft.TextButton(
                    "取消", 
                    on_click=lambda e: close_menu(),
                    expand=True,
                    style=ft.ButtonStyle(color=ft.Colors.GREY_600),
                ),
            ], spacing=12, horizontal_alignment=ft.CrossAxisAlignment.CENTER),  # 添加水平居中
            width=280,
            padding=20,
            bgcolor=ft.Colors.WHITE,
            border_radius=20,
            shadow=ft.BoxShadow(
                spread_radius=1,
                blur_radius=15,
                color=ft.Colors.BLACK12,
                offset=ft.Offset(0, 4),
            ),
        )
        
        menu_container = ft.Container(
            content=ft.Column([
                ft.Container(expand=True),
                ft.Row([ft.Container(expand=True), menu_content, ft.Container(expand=True)]),
                ft.Container(expand=True),
            ]),
            expand=True, bgcolor=ft.Colors.BLACK26, on_click=lambda e: close_menu(),
        )
        
        page.overlay.append(menu_container)
        page.update()

    # 包装函数
    def import_events_wrapper(e):
        asyncio.create_task(import_events_async(e))


    def export_events_wrapper(e):
        asyncio.create_task(export_events_async(e))

    # ========== 结束添加导入事件和导出事件按钮 ===============

    # ========== 设置音乐状态更新回调 ==========
    def set_music_state_update_callback():
        """设置音乐状态更新回调"""
        global music_state_update_callback
        
        def on_music_state_change(event_id, state):
            global current_playing_event_id, current_music_state
            #print(f"[on_music_state_change] 收到回调 - event_id: {event_id}, state: {state}")
            current_playing_event_id = event_id
            current_music_state = state
            try:
                update_current_playing_info()
            except Exception as e:
                print(f"更新播放信息失败: {e}")

            # 根据当前视图刷新对应的视图
            refresh_current_view_by_state()
            
            page.update()
        
        music_state_update_callback = on_music_state_change

    # 立即设置回调，确保在任何播放操作之前回调已就绪
    set_music_state_update_callback()

    date_display = ft.Text(value=current_date.strftime("%Y年%m月%d日"), size=24, weight=ft.FontWeight.BOLD)

    # 3. 滚动时暂停字幕动画
    def on_events_list_scroll(e):
        """事件列表滚动时暂停字幕动画"""
        if hasattr(e, 'pixels') and e.pixels is not None:
            if marquee_text._is_playing:
                marquee_text.stop()
        else:
            if not marquee_text._is_playing and current_music_state in ["playing", "paused"]:
                import threading
                def restart_marquee():
                    import time
                    time.sleep(0.5)
                    if not marquee_text._is_playing and current_music_state in ["playing", "paused"]:
                        marquee_text.start()
                threading.Thread(target=restart_marquee, daemon=True).start()

    #events_list = ft.Column(spacing=12, scroll=ft.ScrollMode.AUTO, height=400)
    #events_list = ft.Column(spacing=12)
    events_list = ft.ListView(
        spacing=12,
        padding=10,
        auto_scroll=False,
        on_scroll=on_events_list_scroll,  # 滚动时暂停字幕
    )
    
    is_android = platform.system() == "Linux"

    # 添加新的平滑滚动字幕
    marquee_text = SmoothMarqueeText(
        text="🎵 未播放",
        #width=280,
        height=60,
        speed=0.5 if is_android else 0.8,  # 可以适当降低速度
        fps=24 if is_android else 60,     # 降低帧率
        gap=None,  # None 表示自动计算，间隙 = 文本宽度
        font_size=15,
        font_weight=ft.FontWeight.BOLD,
        color=ft.Colors.BLUE_700,
        bgcolor=ft.Colors.TRANSPARENT,
        direction="rtl",
        auto_start=False,
        show_message=show_snack_bar,  # 传入显示消息的函数
    )

    # 歌曲名称滚动容器
    music_title_container = ft.Container(
        content=marquee_text,
        #width=280,
        height=60,
        #alignment="center",  # 让内容垂直居中
        border_radius=5,
        bgcolor=ft.Colors.TRANSPARENT,  # 改为透明，与系统背景一致
    )
    
    # ========== 根据平台设置滑块宽度 ==========
    #import platform
    is_android = platform.system() == "Linux"  # Android 是 Linux
    
    if is_android:
        # 手机：使用较小的固定值
        SLIDER_WIDTH = 320  # 可以根据实际手机调整
        print(f"[手机模式] SLIDER_WIDTH = {SLIDER_WIDTH}")
    else:
        # 电脑：根据页面宽度计算
        try:
            if hasattr(page, 'window_width') and page.window_width:
                page_width = page.window_width
                SLIDER_WIDTH = page_width - 60
                if SLIDER_WIDTH < 400:
                    SLIDER_WIDTH = 400
                if SLIDER_WIDTH > 600:
                    SLIDER_WIDTH = 600
            else:
                SLIDER_WIDTH = 470
        except:
            SLIDER_WIDTH = 470
        #print(f"[电脑模式] SLIDER_WIDTH = {SLIDER_WIDTH}")

    # ========== 创建进度显示容器（默认隐藏，气泡方式跟随滑块） ==========
    progress_text = ft.Text(
        "0:00", 
        size=10, 
        color=ft.Colors.WHITE,
        weight=ft.FontWeight.BOLD,
        text_align=ft.TextAlign.CENTER,
    )

    # 气泡容器
    progress_bubble = ft.Container(
        content=progress_text,
        width=70,  # 固定宽度
        height=30,  # 固定高度（小于宽度，形成椭圆形）
        bgcolor=ft.Colors.BLUE_700,
        border_radius=15,  # 高度的一半，形成椭圆
        visible=False,
        alignment=ft.Alignment(0, 0),  # 内容居中
        shadow=ft.BoxShadow(
            spread_radius=1,
            blur_radius=6,
            color=ft.Colors.BLACK26,
            offset=ft.Offset(0, 2),
        ),
    )

    # 滑块
    progress_slider = ft.Slider(
        min=0, 
        max=100, 
        value=0,
        disabled=False,
        expand=True,
        active_color=ft.Colors.BLUE_700,
        inactive_color=ft.Colors.GREY_300,
    )
    
    # 使用动态宽度
    slider_wrapper = ft.Container(
        content=ft.Row([progress_slider], alignment=ft.MainAxisAlignment.CENTER),
        width=SLIDER_WIDTH,
    )

    # ========== 节流控制 ==========
    hide_progress_timer = None

    def hide_progress_text():
        """隐藏进度文本"""
        global hide_progress_timer
        progress_bubble.visible = False
        page.update()
        hide_progress_timer = None

    def get_slider_value_position():
        """获取滑块值对应的像素位置"""
        global SLIDER_WIDTH

        # 直接使用固定宽度
        slider_width = SLIDER_WIDTH
        value_percent = progress_slider.value / 100

        # ========== 根据平台判断 ==========
        #import platform
        is_android = platform.system() == "Linux"
        
        if is_android:
            # 手机上滑块的左右边距更大
            slider_padding = 28  # 手机使用更大的 padding
            print(f"[手机] value: {value_percent:.2f}, padding: {slider_padding}")
        else:
            slider_padding = 18

        available = slider_width - slider_padding * 2
        bubble_center = slider_padding + available * value_percent
        bubble_half_width = 35
        left_pos = bubble_center - bubble_half_width
        
        if left_pos < 0:
            left_pos = 0
        if left_pos > slider_width - 70:
            left_pos = slider_width - 70
        
        return left_pos

    def on_slider_change(e):
        """滑块值改变时显示时间，2秒后自动隐藏"""
        global hide_progress_timer
        
        if hide_progress_timer:
            try:
                hide_progress_timer.cancel()
            except:
                pass
            hide_progress_timer = None
        
        # ========== 获取当前时长 ==========
        duration = current_duration
        if duration <= 0:
            if current_music_file and os.path.exists(current_music_file):
                try:
                    from mutagen.mp3 import MP3
                    duration = MP3(current_music_file).info.length
                except:
                    try:
                        from mutagen.wave import WAVE
                        duration = WAVE(current_music_file).info.length
                    except:
                        duration = 0
        
        if duration > 0:
            current_pos = (progress_slider.value / 100) * duration
            # ========== 显示完整时间：当前时间 / 总时长 ==========
            progress_text.value = f"{format_time(current_pos)} / {format_time(duration)}"
        else:
            progress_text.value = "0:00 / 0:00"
        
        # 计算并更新气泡位置
        new_left = get_slider_value_position()
        progress_bubble_container.left = new_left
        
        # 显示气泡
        progress_bubble.visible = True
        page.update()
        
        # 2秒后自动隐藏
        hide_progress_timer = threading.Timer(2.0, hide_progress_text)
        hide_progress_timer.daemon = True
        hide_progress_timer.start()
    
    # ========== 新增：拖动结束，跳转到指定位置 ==========
    def on_slider_change_end(e):
        """用户结束拖动时，跳转到指定位置"""
        global current_audio, current_position_sec, current_duration
        
        if not current_audio:
            print("[快进] 没有正在播放的音乐")
            show_snack_bar("没有正在播放的音乐")
            return
        
        # ========== 直接使用全局 current_duration ==========
        duration = current_duration

        print(f"打印duration长度： {duration}")
        
        # 如果 current_duration 为 0，尝试从音乐文件读取
        if duration <= 0 and current_music_file and os.path.exists(current_music_file):
            try:
                from mutagen.mp3 import MP3
                duration = MP3(current_music_file).info.length
                print(f"[时长] 从MP3读取: {duration}")
            except:
                try:
                    from mutagen.wave import WAVE
                    duration = WAVE(current_music_file).info.length
                    print(f"[时长] 从WAVE读取: {duration}")
                except:
                    duration = 0
        
        if duration <= 0:
            print(f"[快进] 无效的时长: {duration}")
            show_snack_bar("无法获取音乐时长")
            return
        
        # 计算目标位置（秒）
        target_position = (progress_slider.value / 100) * duration
        target_ms = int(target_position * 1000)
        
        print(f"[快进] 跳转到: {format_time(target_position)} / {format_time(duration)}")
        print(f"[快进] current_duration 值: {current_duration}")
        
        try:
            async def seek_to():
                try:
                    # 尝试 seek
                    if hasattr(current_audio, 'seek'):
                        await current_audio.seek(target_ms)
                        print(f"[快进] seek 成功: {target_ms}ms")
                    else:
                        print("[快进] 当前音频控件不支持跳转")
                        show_snack_bar("当前音频控件不支持拖动快进")
                        return
                    
                    # 更新全局位置
                    current_position_sec = target_position
                    
                    # 更新歌词显示
                    if current_lyrics:
                        update_lyrics_display(target_position, current_lyrics, lyrics_display_widgets, is_fullscreen=False)
                    
                    page.update()
                    show_snack_bar(f"已跳转到 {format_time(target_position)}")
                    #show_snack_bar(f"打印SLIDER_WIDTH宽度：{SLIDER_WIDTH}")
                    
                except Exception as ex:
                    print(f"[快进] 跳转失败: {ex}")
                    show_snack_bar(f"跳转失败: {str(ex)}")
            
            asyncio.create_task(seek_to())
            
        except Exception as ex:
            print(f"[快进] 跳转失败: {ex}")
            show_snack_bar(f"跳转失败: {str(ex)}")

    # 绑定事件
    progress_slider.on_change = on_slider_change
    progress_slider.on_change_end = on_slider_change_end  # 新增：拖动结束事件

    # 气泡容器（用于定位）
    progress_bubble_container = ft.Container(
        content=progress_bubble,
        top=2,
        left=0,
    )
    
    # 歌词显示容器
    lyrics_display_container, lyrics_display_widgets = create_lyrics_display()

    count_text = ft.Text(value=f"📊 事件总数: {len(events)}", size=12, color=ft.Colors.BLUE_700)
    
    async def async_start_marquee():
        """异步启动滚动字幕"""
        marquee_text.start()

    async def async_stop_marquee():
        """异步停止滚动字幕"""
        marquee_text.stop()

    # ========== 添加 update_current_playing_info 函数在这里 ==========
    def update_current_playing_info():
        """更新顶部当前播放信息显示"""
        global current_playing_event_id, current_music_state, marquee_text, music_section_container, playback_buttons
        
        #print(f"[update_current_playing_info] 被调用 - event_id: {current_playing_event_id}, state: {current_music_state}")
        
        # 如果有音乐正在播放（无论是否有事件），都显示音乐区域
        if current_music_state in ["playing", "paused"]:
            # 显示音乐区域
            if music_section_container:
                music_section_container.visible = True
                music_section_container.update()
            if playback_buttons:
                playback_buttons.visible = True
                playback_buttons.update()
            
            # 获取音乐名称
            if current_music_file and os.path.exists(current_music_file):
                # 直接去掉扩展名，显示完整文件名
                base_name = os.path.basename(current_music_file)
                music_name = os.path.splitext(base_name)[0]
            else:
                music_name = "未知音乐"
            
            # 判断是否是试听模式（没有事件ID或事件不存在）
            is_preview = (current_playing_event_id is None or 
                        current_playing_event_id not in events)
            
            if is_preview:
                # 试听模式
                if current_music_state == "playing":
                    full_text = f"🎵 试听中: {music_name}"
                    marquee_text.color = ft.Colors.BLUE_700
                    marquee_text.update_text(full_text)
                    page.run_task(async_start_marquee)
                else:
                    full_text = f"⏸️ 已暂停: {music_name}"
                    marquee_text.color = ft.Colors.ORANGE_700
                    marquee_text.update_text(full_text)
                    page.run_task(async_stop_marquee)
            else:
                # 正式事件
                event = events[current_playing_event_id]

                # ========== 获取事件类型显示名称 ==========
                event_type_display = {
                    "birthday": "🎉生日",
                    "event": "💝纪念日",
                    "monthly": "🔄每月",
                    "once": "🎯一次性",
                    "daily": "⏰每日",
                    "weekly": "🔁每周",
                }.get(event.event_type, "📅事件")

                # ========== 判断是今日事件还是预警事件 ==========
                today = datetime.now().date()
                month, day, year, base_year, days_until = event.get_next_date_info()

                if days_until == 0:
                    event_tag = "📌今日事件"
                    tag_color = ft.Colors.RED_700
                elif 0 < days_until <= 3:
                    event_tag = "⏰预警事件"
                    tag_color = ft.Colors.ORANGE_700
                else:
                    event_tag = "📋普通事件"
                    tag_color = ft.Colors.BLUE_700
                
                if current_music_state == "playing":
                    full_text = f"正在播放: {event_tag} {event_type_display}【{event.name}】: {music_name}"
                    marquee_text.color = tag_color
                    marquee_text.update_text(full_text)
                    page.run_task(async_start_marquee)
                else:
                    full_text = f"已暂停: {music_name}"
                    marquee_text.color = tag_color
                    marquee_text.update_text(full_text)
                    page.run_task(async_stop_marquee)
        else:
            # 停止状态，隐藏音乐区域
            if music_section_container:
                music_section_container.visible = False
                music_section_container.update()
            if playback_buttons:
                playback_buttons.visible = False
                playback_buttons.update()
            marquee_text.update_text("🎵 未播放")
            marquee_text.color = ft.Colors.GREY_600
            page.run_task(async_stop_marquee)
            if marquee_text._initialized:
                marquee_text._draw_frame()

        # ========== 新增：刷新事件列表，更新卡片上的时长 ==========
        #refresh_current_view_by_state()

        #print(f"[update_current_playing_info] UI更新完成")
    # ========== 函数添加结束 ==========

    # 创建时钟（传入 page 参数）
    #clock = AnalogClock(page, size=160)
    #page.update()  # 强制刷新页面


    # 创建日期显示
    #date_text = ft.Text(value="", size=14, color=ft.Colors.GREY_600, text_align=ft.TextAlign.CENTER)

    # 事件状态和个数展示 - 使用 TextButton 确保可点击
    date_text = ft.TextButton(
        content=ft.Text(
            value="加载中...",
            size=13,
            text_align=ft.TextAlign.CENTER,
        ),
        on_click=on_date_text_click,
        style=ft.ButtonStyle(
            color=ft.Colors.GREY_600,
            bgcolor=ft.Colors.TRANSPARENT,
            overlay_color=ft.Colors.TRANSPARENT,
            padding=5,
        ),
        tooltip="点击查看事件",
    )

    # ========== 创建音乐控制容器 ==========
    music_control_container = ft.Container(
        content=ft.Column([
            music_title_container,
            ft.Container(height=8),
            ft.Container(
                content=ft.Stack([
                    slider_wrapper,
                    progress_bubble_container,
                ], height=90),
                alignment=ft.Alignment(0, 0),
            ),
            lyrics_display_container,
        ], spacing=5, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
        padding=10,
        bgcolor=ft.Colors.TRANSPARENT,
        border_radius=10,
        visible=False,
    )

    # 创建播放控制按钮（可隐藏）
    playback_buttons = ft.Row([
        ft.TextButton("⏸️ 暂停", on_click=pause_music, tooltip="暂停音乐"),
        ft.TextButton("⏹️ 停止", on_click=lambda e: stop_music(), tooltip="停止音乐"),
    ], spacing=0, visible=False)  # 初始隐藏

    # 创建导入导出按钮（始终显示）
    import_export_buttons = ft.Row([
        ft.TextButton(
            "💰 账单", 
            on_click=lambda e: show_accounting_page(page), 
            tooltip="账单",
            style=ft.ButtonStyle(color=ft.Colors.BLUE_700,text_style=ft.TextStyle(weight=ft.FontWeight.BOLD), ),
        ),
        ft.TextButton("📥 导入", on_click=show_import_menu, tooltip="从Excel导入事件"),
        ft.TextButton("📤 导出", on_click=show_export_menu, tooltip="导出事件到Excel"),
        #ft.TextButton("🔔 通知", on_click=test_notification)
    ], spacing=0)


    # 创建音乐播放相关内容的容器
    music_section_container = ft.Column([
        music_control_container,
        ft.Divider(),
    ], spacing=8, visible=False)

    # 确保内部控件的可见性初始为 True
    music_control_container.visible = True
    #playback_buttons.visible=True

    # 创建一个变量记录日历是否显示
    calendar_visible = True

    def toggle_calendar(e):
        """切换日历显示/隐藏"""
        nonlocal calendar_visible
        calendar_visible = not calendar_visible
        calendar_widget.visible = calendar_visible
        
        # ========== 更新箭头按钮的图标 ==========
        arrow_button.icon = ft.Icons.EXPAND_LESS if calendar_visible else ft.Icons.EXPAND_MORE
        arrow_button.update()
        calendar_widget.update()
        page.update()

    # 创建箭头按钮
    arrow_button = ft.IconButton(
        icon=ft.Icons.EXPAND_LESS,
        icon_size=20,
        icon_color=ft.Colors.BLUE_700,
        on_click=toggle_calendar,
        tooltip="切换日历显示",
        padding=5,
    )

    # 箭头按钮（可点击）
    arrow_button = ft.IconButton(
        icon=ft.Icons.EXPAND_LESS,
        icon_size=20,
        icon_color=ft.Colors.BLUE_700,
        on_click=toggle_calendar,
        tooltip="切换日历显示",
        padding=5,
    )

    # 修改日历和日期文本的组合
    calendar_section = ft.Column([
        calendar_widget,
        ft.Row([
            # 箭头靠左
            arrow_button,
            # 日期文本居中
            ft.Container(
                content=date_text,
                expand=True,
                alignment=ft.Alignment(0, 0),
            ),
            # 右边添加空白占位，宽度与 arrow_button 相同
            ft.Container(
                width=arrow_button.width if hasattr(arrow_button, 'width') and arrow_button.width else 40,
                height=1,
                bgcolor=ft.Colors.TRANSPARENT,
            ),
        ], alignment=ft.MainAxisAlignment.START, spacing=5),
    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER)

    # ========== 可滚动的内容区域（其他所有内容） ==========
    scrollable_content =ft.Column(
        [
        
            # 顶部留白
            #ft.Container(height=5),
            
            # 日历和事件提醒组合
            calendar_section,
            
            ft.Divider(),  # 音乐区域上方的分割线

            # 音乐相关区域（整个区域统一控制显示/隐藏）
            music_section_container,

            # 所有按钮行（播放控制按钮 + 导入导出按钮）
            ft.Row([
                playback_buttons,
                import_export_buttons,
            ], alignment=ft.MainAxisAlignment.CENTER, spacing=0),

            ft.Divider(),
            
            # 事件列表（移除自己的滚动，让外层统一滚动）
            events_list, # 这里不再设置 scroll，让内容自然扩展
            
            ft.Divider(),
            
            # 底部信息
            ft.Container(
                content=ft.Column([
                    #ft.Divider(height=5),
                    ft.Text("💡 使用说明", size=14, weight=ft.FontWeight.BOLD),
                    ft.Text("• 点击「+」添加事件\n• 点击 💰 账单 进入账单界面\n• 各类事件当天或提前3天预警自动弹框并播放音乐\n• 启动程序自动检查今日是否有事件发生", selectable=True),
                    # ========== 修改这里：提醒服务单独一行，count_text和版本在同一行 ==========
                    ft.Row([
                        ft.Text("🔔 提醒服务运行中", size=12, color=ft.Colors.GREEN_700),
                    ], alignment=ft.MainAxisAlignment.START),
                    ft.Row([
                        ft.Text(f"📱 版本 {APP_VERSION}", size=10, color=ft.Colors.GREY_500),
                        ft.Container(expand=True),  # 弹性空间，把版本推到右边
                        count_text,
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ]),
                padding=12,
            )
        ], 
        spacing=8, 
        scroll=ft.ScrollMode.AUTO,
        on_scroll=lambda e: on_scroll_changed(e),
    )


    # 修改 main_content 的顶部部分
    main_content = ft.Column([
        # ========== 固定标题区域 ==========
        ft.Container(height=20),  # 顶部留白
        
        # 标题
        ft.Container(
            content=ft.Column([
                ft.Text(
                    "📋 事件提醒", 
                    size=20, 
                    weight=ft.FontWeight.BOLD, 
                    color=ft.Colors.BLUE_700, 
                    text_align=ft.TextAlign.CENTER,
                    width=float("inf"),  # 让文本占满宽度，才能居中
                )
            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            padding=13,
        ),

        ft.Divider(),

        # 可滚动的内容区域
        ft.Container(
            content=scrollable_content,
            expand=True,
        ),
    ], spacing=0, expand=True)

    # 创建回到顶部按钮
    scroll_top_button = ft.Container(
        content=ft.Icon(ft.Icons.ARROW_UPWARD, size=28, color=ft.Colors.BLUE_700),
        width=50,  # 与 today_circle_button 一致
        height=50,  # 与 today_circle_button 一致
        bgcolor=ft.Colors.WHITE,
        border_radius=25,  # 50/2 = 25
        ink=True,
        on_click=lambda e: asyncio.create_task(scroll_to_top(e)),
        tooltip="回到顶部",
        shadow=ft.BoxShadow(
            spread_radius=1,
            blur_radius=8,
            color=ft.Colors.BLACK12,
            offset=ft.Offset(0, 2),
        ),
        visible=False,
    )

    def on_scroll_changed(e):
        """滚动事件回调"""
        global show_scroll_top_btn
        
        # 获取滚动位置
        scroll_offset = e.pixels if hasattr(e, 'pixels') else 0
        
        # 只要滚动超过0像素（即滑动了）就显示回到顶部按钮
        if scroll_offset > 0 and not show_scroll_top_btn:
            show_scroll_top_btn = True
            scroll_top_button.visible = True
            page.update()
        elif scroll_offset == 0 and show_scroll_top_btn:
            show_scroll_top_btn = False
            scroll_top_button.visible = False
            page.update()

    async def scroll_to_top(e):
        """滚动到顶部"""
        if hasattr(scrollable_content, 'scroll_to'):
            await scrollable_content.scroll_to(offset=0, duration=500, curve=ft.AnimationCurve.EASE_IN_OUT)
            page.update()

    # 创建返回今天按钮
    today_circle_button = ft.Container(
        content=ft.Column(
            [
                ft.Text(
                    str(datetime.now().day),
                    size=18,
                    #weight=ft.FontWeight.BOLD,
                    color=ft.Colors.BLUE_700,
                    text_align="center",
                ),
            ],
            alignment=ft.MainAxisAlignment.CENTER,  # 垂直居中
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,  # 水平居中
        ),
        width=50,
        height=50,
        bgcolor=ft.Colors.WHITE,
        border_radius=25,
        ink=True,
        visible=False,
        on_click=lambda e: go_to_today(),
        tooltip=f"回到今天 ({datetime.now().month}月{datetime.now().day}日)",
        shadow=ft.BoxShadow(
            spread_radius=1,
            blur_radius=8,
            color=ft.Colors.BLACK12,
            offset=ft.Offset(0, 2),
        ),
    )

    # 创建添加事件按钮
    floating_add_button = ft.Container(
        content=ft.Icon(ft.Icons.ADD, size=28, color=ft.Colors.WHITE),
        bgcolor=ft.Colors.BLUE_700,
        border_radius=30,
        padding=14,
        ink=True,
        shadow=ft.BoxShadow(
            spread_radius=1,
            blur_radius=10,
            color=ft.Colors.BLACK26,
            offset=ft.Offset(0, 2),
        ),
    )

    def on_floating_add_click(e):
        """悬浮按钮点击事件（根据当前页面执行不同操作）"""
        global current_page
        
        if current_page == "main":
            # 主界面：添加事件
            open_add_dialog(is_edit=False)

    floating_add_button.on_click = on_floating_add_click

    # 使用 Stack 布局，返回按钮在添加按钮上方
    # 悬浮按钮组
    floating_buttons = ft.Column(
        [
            scroll_top_button,      # 回到顶部按钮（放在最上面）
            today_circle_button,
            floating_add_button,
        ],
        spacing=12,  # 按钮间距 
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
    )

    # Stack 布局
    main_stack = ft.Stack(
        [
            main_content,
            ft.Container(
                content=floating_buttons,
                right=20,
                bottom=20,
            ),
        ],
        expand=True,
    )

    # 添加到页面
    page.add(main_stack)

    if platform.system() == "Linux":
        # 延迟2秒显示后台通知（避免与启动检查冲突）
        threading.Timer(2.0, show_background_notification).start()

    page.update()

    # 刷新事件列表（页面添加后再刷新）
    refresh_events_list()

    # ========== 启动时自动选择视图 ==========
    def determine_startup_view():
        """根据事件情况决定启动时显示的视图（只切换视图，不触发提醒）"""
        global current_view
        
        today = datetime.now().date()
        has_today_event = False
        has_warning_event = False
        
        print(f"[启动视图] ========== 开始检查 ==========")
        print(f"[启动视图] 当前日期: {today}")
        
        for event in events.values():
            # 跳过每天事件和每周事件
            if event.event_type == "daily" or event.event_type == "weekly":
                continue
            
            # 检查是否设置了提醒时间，如果设置了且时间还没到，今日事件不显示
            has_reminder = False
            reminder_time = None
            if event.reminders:
                for reminder in event.reminders:
                    if reminder.get("enabled"):
                        has_reminder = True
                        reminder_time = reminder.get("time")
                        break
            
            month, day, year, base_year, days_until = event.get_next_date_info()
            
            # 如果是今日事件
            if month == today.month and day == today.day:
                # 如果设置了提醒时间且时间还没到，不视为今日事件（防止提前显示）
                if has_reminder and reminder_time:
                    now = datetime.now()
                    if reminder_time > now.strftime("%H:%M"):
                        print(f"[启动视图] {event.name} 提醒时间 {reminder_time} 还没到，今日不显示")
                        continue
                has_today_event = True
                print(f"[启动视图] ✓ 今日事件: {event.name}")
                break
            
            # 预警事件（3天内）
            if 0 < days_until <= 3:
                has_warning_event = True
                print(f"[启动视图] ✓ 预警事件: {event.name}, {days_until}天后")
        
        # 根据检查结果设置初始视图
        if has_today_event:
            current_view = "today"
            show_today_events()
            show_bottom_message("📌 今日有事件，自动切换到今日事件视图")
        elif has_warning_event:
            current_view = "three_days"
            show_three_days_events()
            show_bottom_message("🔔 未来3天有事件，自动切换到预警事件视图")
        else:
            current_view = "daily"
            show_daily_events()
            show_bottom_message("⏰ 切换到每日事件视图")
        
        # 更新下拉框的显示
        update_view_dropdown_display(current_view)
        page.update()

    # ========== 设置页面关闭回调 ==========
    def on_page_close():
        """页面关闭时清理所有通知"""
        cancel_notification(MUSIC_NOTIFICATION_ID)
        cancel_notification(EVENT_NOTIFICATION_ID)
        cancel_notification(BACKGROUND_NOTIFICATION_ID)
        print("✅ 已清理所有通知")
    
    # 设置页面关闭回调
    page.on_close = on_page_close
    
    async def update_all():
        global last_check_date, reminder_flags, current_year, current_month, selected_date, current_date, current_view, three_days_events, sent_notifications
        
        while True:
            try:
                now = datetime.now()
                current_date_today = now.date()  # 重命名避免与全局变量冲突
                
                # ========== 添加跨天检测 ==========
                if last_check_date is None:
                    last_check_date = current_date_today
                elif current_date_today != last_check_date:
                    # 日期发生了变化（跨天了）
                    print(f"[跨天检测] 日期从 {last_check_date} 变更为 {current_date_today}，立即触发事件检查")
                    
                    # ========== 1. 更新 last_check_date（必须在最前面） ==========
                    last_check_date = current_date_today

                    # ========== 2. 更新日历到当前日期 ==========
                    # 更新全局的年月变量
                    current_year = now.year
                    current_month = now.month

                    # 更新月份文本显示
                    month_text.value = f"{current_year}年{current_month}月"

                    # 更新选中的日期为今天
                    selected_date = current_date_today
                    current_date = current_date_today

                    # 重新生成日历（会高亮今天）
                    update_calendar()

                    # ========== 强制更新回到今天按钮的日期数字 ==========
                    today = datetime.now()
                    if hasattr(today_circle_button, 'content'):
                        if isinstance(today_circle_button.content, ft.Text):
                            today_circle_button.content.value = str(today.day)
                        elif isinstance(today_circle_button.content, ft.Column):
                            if today_circle_button.content.controls and len(today_circle_button.content.controls) > 0:
                                if isinstance(today_circle_button.content.controls[0], ft.Text):
                                    today_circle_button.content.controls[0].value = str(today.day)
                    today_circle_button.tooltip = f"回到今天 ({today.month}月{today.day}日)"
                    today_circle_button.update()

                    # 更新日期显示
                    date_display.value = current_date_today.strftime("%Y年%m月%d日")

                    # ========== 3. 检查是否需要跨年重置 ==========
                    if current_date_today.year != last_check_date.year:
                        print(f"[跨天检测] 检测到跨年！从 {last_check_date.year} 年到 {current_date_today.year} 年")

                        # 重置所有事件的 last_remind_year
                        for event in events.values():
                            if event.last_remind_year < current_date_today.year:
                                print(f"[跨天检测] 重置事件 {event.name} 的提醒状态 (从 {event.last_remind_year} 到 0)")
                                event.last_remind_year = 0
                                event.reminded_this_year = False
                        save_events()
                        print(f"[跨天检测] 跨年重置完成")
                    
                    #last_check_date = current_date_today
                    
                    # ========== 4. 重置临时提醒标记 ==========
                    reminder_flags.clear()
                    print(f"[跨天检测] 已重置提醒标记")

                    # ========== 5. 关键：重新计算 three_days_events ==========
                    three_days_events = []
                    for evt in events.values():
                        if evt.event_type == "daily" or evt.event_type == "weekly":
                            continue
                        month, day, year, base_year, days_until = evt.get_next_date_info()
                        if evt.repeat_type == "once" and (evt.completed or days_until < 0):
                            continue
                        if 0 < days_until <= 3:
                            three_days_events.append((evt, days_until))
                    
                    # ========== 6. 更新顶部日期文本 ==========
                    update_date_text_with_events(current_date_today, three_days_events)

                    determine_startup_view()
                    
                    # ========== 7. 刷新事件列表（根据当前视图） ==========
                    refresh_current_view_by_state()

                    # ========== 8. 立即执行事件检查 ==========
                    check_events()

                
                # 原有的更新时钟代码继续...
                #clock.update_clock()
                
                # 获取农历日期
                try:
                    lunar = LunarDate.fromSolarDate(now.year, now.month, now.day)
                    # 转换为中文显示
                    lunar_month_str = number_to_chinese_month(lunar.month)
                    lunar_day_str = number_to_chinese_day(lunar.day)
                    lunar_str = f"农历{lunar_month_str}{lunar_day_str}"
                except:
                    lunar_str = "农历计算失败"
                
                # 获取星期几
                weekdays = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
                weekday_str = weekdays[now.weekday()]
                
                # 更新显示
                #current_datetime_text.value = f"📅 当前时间: {now.strftime('%Y年%m月%d日 %H:%M:%S')}"
                
                # 更新运行时间
                #elapsed = datetime.now() - start_time
                #total_seconds = int(elapsed.total_seconds())
                #hours = total_seconds // 3600
                #minutes = (total_seconds % 3600) // 60
                #seconds = total_seconds % 60
                #run_time_text.value = f"⏱️ 运行时间: {hours:02d}:{minutes:02d}:{seconds:02d}"
                
                # 更新日期文字
                #date_text.value = f"{now.year}年{now.month:02d}月{now.day:02d}日 {weekday_str} {lunar_str} {now.strftime('%H:%M:%S')}"
                #date_text.update()
                
                # 同时更新两个控件
                #current_datetime_text.update()
                #run_time_text.update()

                # 使用新函数更新
                update_date_text_with_events(current_date_today, three_days_events)
                
                # 实时检查是否到时间触发事件提醒
                check_time_reminders()
                
                await asyncio.sleep(1)
            except Exception as e:
                print(f"更新时间出错: {e}")
                await asyncio.sleep(1)

    # 只启动一个循环
    asyncio.create_task(update_all())

    async def auto_refresh():
        """每小时自动刷新事件列表"""
        while True:
            await asyncio.sleep(60)  # 每分钟刷新一次
            # ========== 根据当前视图刷新对应的视图 ==========
            refresh_current_view_by_state()
            print(f"[自动刷新] 已刷新当前视图 ({current_view}) - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    asyncio.create_task(auto_refresh())

    # 更新顶部当前播放信息显示
    update_current_playing_info()

    # 启动后台定时检查
    debug_log("设置后台检查定时器（30秒后启动，之后每15分钟）")
    threading.Timer(1.0, start_background_check).start()

     # ========== 执行启动时检查 ==========
    # 1. 检查错过的提醒（设置了提醒时间且时间已过）
    check_missed_reminders()
    
    # 2. 检查今日事件和预警事件（未设置提醒时间）- 使用 show_combined_reminder
    check_startup_events()
    
    # 3. 设置启动视图
    determine_startup_view()
    
    # 4. 延迟执行首次检查
    debug_log("设置首次检查定时器（2秒后）")
    threading.Timer(2.0, check_events).start()

if __name__ == "__main__":
    ft.app(target=main, assets_dir="assets")